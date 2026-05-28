# utils/import_helper.py
import openpyxl
import re
from io import BytesIO
from datetime import datetime

def extract_country_from_wh_id(wh_id):
    """从库房ID提取国家代码"""
    if not wh_id:
        return ''
    # 匹配前两位大写字母作为国家代码
    match = re.match(r'^([A-Z]{2})', str(wh_id).strip().upper())
    return match.group(1) if match else ''


def validate_country_and_wh_id(country_input, wh_id):
    """
    验证国家与库房ID的一致性
    
    返回:
        (final_country, is_valid, error_message)
        - final_country: 最终使用的国家代码
        - is_valid: 是否有效
        - error_message: 错误信息（如果无效）
    """
    # 清理输入
    country = country_input.strip().upper() if country_input else ''
    wh_id_clean = wh_id.strip().upper() if wh_id else ''
    
    # 从库房ID提取国家
    wh_country = extract_country_from_wh_id(wh_id_clean) if wh_id_clean else ''
    
    # 情况1：国家和库房ID都为空
    if not country and not wh_country:
        return '', True, None
    
    # 情况2：只有国家有值，库房ID为空
    if country and not wh_country:
        return country, True, None
    
    # 情况3：只有库房ID有值，国家为空
    if not country and wh_country:
        return wh_country, True, None
    
    # 情况4：国家和库房ID都有值，检查是否一致
    if country and wh_country:
        if country == wh_country:
            return country, True, None
        else:
            return country, False, f"国家({country})与库房ID({wh_id_clean})提取的国家({wh_country})不一致，请修正"
    
    return country, True, None


def generate_import_template(template_type='user'):
    """
    生成导入模板
    template_type: 'user' 或 'wh'
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if template_type == 'user':
        ws.title = "UserImportTemp"
        headers = [
            ("country", "country", "国家代码，如 NP、LK、ID。可不填，从库房ID提取"),
            ("email", "email", "用户邮箱，必填"),
            ("name_en", "name_en", "用户姓名，必填"),
            ("role", "role", "user/admin/super_admin，默认 user"),
            ("is_partner", "is_supplier", "Y/N，默认 N"),
            ("company", "company", "可选"),
            ("department", "department", "可选"),
            ("wh_type", "wh_type", "可选"),
            ("wh_id", "wh_id", "库房编码，如 NP001。如填写，前两位需与国家一致"),
            ("wh_name_en", "wh_name_en", "可选"),
            ("employee_id", "employee_id", "可选"),
            ("phone", "phone", "可选"),
            ("birthday", "birthday", "YYYY-MM-DD 格式，可选"),
            ("admin_countries", "admin_countries", "管理员/超管必填，如 NP,LK"),
        ]
    else:
        ws.title = "WHImportTemp"
        headers = [
            ("wh_id", "wh_id", "必填，如 NP001。前两位将作为国家代码"),
            ("wh_name_cn", "wh_name_cn", "中文名称，可选"),
            ("wh_name_en", "wh_name_en", "英文名称，可选"),
            ("wh_type", "wh_type", "系统库/备件库/第三方库，可选"),
            ("country_code", "country_code", "可选，不填则从库房ID提取。如填写需与库房ID前两位一致"),
        ]
    
    # 写入表头
    for col, (cn_name, en_name, tip) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=cn_name)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # 添加注释
        if tip:
            cell.comment = openpyxl.comments.Comment(tip, "System")
    
    # 添加示例数据行
    if template_type == 'user':
        example_row = ["NP", "user@example.com", "张三", "user", "N", "公司名", "部门名", "系统库", "NP001", "Warehouse Name", "EMP001", "13800138000", "1990-01-01", ""]
    else:
        example_row = ["NP001", "加德满都中心库", "Kathmandu Central", "系统库", "NP"]
    
    for col, value in enumerate(example_row, 1):
        ws.cell(row=2, column=col, value=value)
    
    # 添加第二个示例（说明不一致会报错）
    if template_type == 'user':
        ws.cell(row=3, column=1, value="⚠️ 注意")
        ws.cell(row=3, column=9, value="LK001")
        ws.cell(row=3, column=2, value="国家(NP)与库房ID(LK001)不一致会导入失败")
    else:
        ws.cell(row=3, column=5, value="⚠️ 注意")
        ws.cell(row=3, column=1, value="NP001")
        ws.cell(row=3, column=5, value="LK")
        ws.cell(row=3, column=2, value="国家代码(LK)与库房ID(NP001)不一致会导入失败")
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_excel_to_dict(file, header_map):
    """
    解析 Excel 文件为字典列表
    返回: (data_list, error_messages)
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
    except Exception as e:
        return [], [f"文件解析失败: {str(e)}"]
    
    # 读取表头映射
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            header_str = str(cell_value).strip()
            if header_str in header_map:
                headers[col] = header_map[header_str]
    
    # 检查必填列
    required_fields = ['name_en'] if 'name_en' in header_map.values() else ['wh_id']
    missing_fields = [f for f in required_fields if f not in headers.values()]
    if missing_fields:
        return [], [f"Excel missing required columns: {', '.join(missing_fields)}"]
    
    data_list = []
    errors = []
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, field in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[field] = str(cell_value).strip() if cell_value else ''
        
        # 检查是否有有效数据（至少有一列有值）
        has_data = any(v for v in row_data.values() if v)
        if not has_data:
            continue  # 跳过空行
        
        data_list.append((row_idx, row_data))
    
    return data_list, errors

def is_row_empty(row_data):
    """检查一行数据是否为空（所有值都为空或None）"""
    for value in row_data.values():
        if value and str(value).strip():
            return False
    return True


def parse_excel_rows(file, header_map, required_fields):
    """
    解析 Excel 文件，返回有效数据行列表和错误信息
    
    Args:
        file: 上传的文件对象
        header_map: 表头映射字典 {表头文本: 字段名}
        required_fields: 必填字段列表（字段名）
    
    Returns:
        (valid_rows, errors, headers)
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
    except Exception as e:
        return [], [f"文件解析失败: {str(e)}"], {}
    
    # 读取表头行（第一行）
    headers = {}  # {列索引: 字段名}
    header_row_values = []
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            header_str = str(cell_value).strip()
            header_row_values.append(header_str)
            # ✅ 直接匹配（不区分大小写）
            matched = False
            for key, field in header_map.items():
                if header_str.lower() == key.lower():
                    headers[col] = field
                    matched = True
                    break
            # 如果没有精确匹配，尝试去除空格后匹配
            if not matched:
                clean_header = header_str.replace(' ', '').lower()
                for key, field in header_map.items():
                    if clean_header == key.replace(' ', '').lower():
                        headers[col] = field
                        break
    
    # 检查必填列
    found_fields = set(headers.values())
    missing_fields = [f for f in required_fields if f not in found_fields]
    if missing_fields:
        return [], [f"Excel缺少必填列: {', '.join(missing_fields)}，请确保表头包含这些列"], headers
    
    valid_rows = []
    errors = []
    max_row = ws.max_row
    
    for row_idx in range(2, max_row + 1):
        row_data = {}
        for col_idx, field in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[field] = str(cell_value).strip() if cell_value else ''
        
        # 检查是否为空行
        if is_row_empty(row_data):
            continue  # 跳过空行
        
        valid_rows.append((row_idx, row_data))
    
    if not valid_rows:
        errors.append("Excel文件中没有有效数据，请检查后重试")
    
    return valid_rows, errors, headers

def format_import_result(success_count, error_rows, update_count=0, total_rows=None):
    """格式化导入结果
    
    Args:
        success_count: 成功处理的总数（包括新增和更新）
        error_rows: 错误行列表
        update_count: 更新记录数（可选，默认为0）
        total_rows: 总行数（可选）
    
    Returns:
        dict: 格式化后的结果字典
    """
    new_count = success_count - update_count
    result = {
        "success": True,
        "success_count": success_count,
        "update_count": update_count,
        "new_count": new_count,
        "error_count": len(error_rows),
        "errors": error_rows[:20],  # 最多返回20条错误
        "message": f"导入完成: 成功 {success_count} 条（新增 {new_count}，更新 {update_count}），失败 {len(error_rows)} 条"
    }
    if total_rows:
        result["total"] = total_rows
    return result
    