# routes/admin_export.py
import logging
import zipfile, json, re
import openpyxl 
from datetime import datetime, timezone, timedelta, date
from . import admin_export_bp
from services.db import get_supabase
from dateutil import parser
from apscheduler.schedulers.background import BackgroundScheduler
from io import BytesIO
from flask import (
    Flask, render_template, request, redirect, url_for, 
    session, flash, jsonify, send_file, make_response
)
from supabase import create_client
from functools import wraps
from services import auth, exam, export
from services.export import find_wkhtmltopdf
from services.auth import hash_password
from config import Config
from utils.status import get_exam_status
from utils.common import (
    match_country_code, quarter_to_date_range, get_reviewer_by_country, format_admin_countries_display,
    format_countries_display,  # ✅ 新增
    format_single_country_display  # ✅ 新增（可选）
)
from utils.email_notifier import send_bilingual_notification, EmailScenario, _format_time, _send_training_notifications
from utils.permissions import (
    is_developer, get_developer_id, has_role, can_manage_role,
    can_view_user, get_admin_allowed_countries, set_admin_allowed_countries,
    parse_countries_input, developer_required, super_admin_required,
    filter_users_by_permission, apply_country_filter
)
from routes.helpers import login_required, admin_required, robust_parse_json, get_allowed_countries, parse_exam_countries
from utils.training_helpers import get_training_country_templates_status

logger = logging.getLogger(__name__)

@admin_export_bp.route('/export/pdf/<int:exam_id>')
@login_required
def export_pdf(exam_id):
    """导出 PDF 成绩单"""
    buf = export.generate_pdf(
        "演示考生", 85, [], {}, {}, "Admin"
    )
    return send_file(
        buf, 
        mimetype="application/pdf", 
        as_attachment=True, 
        download_name=f"exam_{exam_id}.pdf"
    )

@admin_export_bp.route('/admin/export/excel/<int:training_id>/<int:exam_id>')
@login_required
@admin_required
def export_bilingual_excel(training_id, exam_id):
    """导出双语 Excel 报告"""
    country = request.args.get('country', None)
    try:
        buffer, filename = export.generate_bilingual_excel(
            training_id=training_id,
            exam_id=exam_id,
            country=country
        )
        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"❌ Excel 导出失败: {e}")
        #flash(f"❌ 导出失败: {str(e)}", "danger")
        flash({'msg': 'export_error', 'params': [str(e)]}, 'danger')
        return redirect(url_for('admin_dashboard'))

@admin_export_bp.route('/api/admin/export_filtered_excel', methods=['POST'])
@login_required
@admin_required
def export_filtered_excel():
    # 复杂导出逻辑，直接复制 app.py 对应函数体
    data = request.json
    country = data.get('country', '')
    training_name = data.get('training_name', '')
    exam_name = data.get('exam_name', '')
    start_date = data.get('start_date', '')
    end_date = data.get('end_date', '')
    wh_raw = data.get('wh_id', '').strip()
    wh_id = wh_raw.split('(')[0].strip() if wh_raw else None
    lang = data.get('lang', 'zh')

    db = get_supabase()
    
    # ✅ 新增：获取当前管理员的权限范围
    allowed_countries = get_admin_allowed_countries()
    
    # ✅ 新增：如果管理员有权限限制，且没有指定国家，则使用权限范围
    if allowed_countries is not None and not country:
        # 如果管理员权限范围不为空，用于过滤用户
        final_user_ids = None
        # 后续查询时需要按权限范围过滤
    else:
        final_user_ids = None

    # 构建统一的候选用户ID列表
    user_ids_for_country = None
    user_ids_for_wh = None

    # 1. 如果指定了国家，获取该国用户ID
    if country:
        users_res = db.table("users").select("id").eq("country", country).execute()
        user_ids_for_country = [u['id'] for u in (users_res.data or [])]
        if not user_ids_for_country:
            # 没有该国用户，直接返回空文件
            wb = openpyxl.Workbook()
            wb.active.title = "空报告"
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")

    # 2. 如果指定了库房，获取该库房下的所有用户ID
    if wh_id:
        users_in_wh = db.table("users").select("id").eq("wh_id", wh_id).execute()
        user_ids_for_wh = [u['id'] for u in (users_in_wh.data or [])]
        if not user_ids_for_wh:
            wb = openpyxl.Workbook()
            wb.active.title = "空报告"
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")

    # 3. 合并国家与库房的用户范围
    if user_ids_for_country is not None and user_ids_for_wh is not None:
        final_user_ids = list(set(user_ids_for_country) & set(user_ids_for_wh))
    elif user_ids_for_country is not None:
        final_user_ids = user_ids_for_country
    elif user_ids_for_wh is not None:
        final_user_ids = user_ids_for_wh
    else:
        final_user_ids = None
    
    # ✅ 新增：如果管理员有权限范围，进一步过滤用户
    if allowed_countries is not None and final_user_ids is not None:
        # 获取这些用户中在权限范围内的
        users_in_allowed = db.table("users").select("id").in_("id", final_user_ids).in_("country", allowed_countries).execute()
        final_user_ids = [u['id'] for u in (users_in_allowed.data or [])]
    elif allowed_countries is not None:
        # 没有指定其他筛选条件，使用权限范围内的所有用户
        users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
        final_user_ids = [u['id'] for u in (users_in_allowed.data or [])]

    # 4. 查询培训（需要根据权限范围过滤）
    # ✅ 新增：培训查询也需要根据管理员权限过滤
    if allowed_countries is not None:
        # 查询权限范围内的培训（培训的 countries 字段包含允许的国家）
        # 由于 countries 是 JSON 数组，需要使用特定查询语法
        training_query = db.table("trainings").select("*")
        # 简化处理：查询所有培训，后续在 Python 中过滤
    else:
        training_query = db.table("trainings").select("*")
    
    if country:
        training_query = training_query.eq("country", country)
    if training_name:
        training_query = training_query.ilike("name", f"%{training_name}%")
    if start_date:
        training_query = training_query.gte("start_time", start_date)
    if end_date:
        training_query = training_query.lte("end_time", end_date)
    
    trainings = training_query.execute().data or []
    
    # ✅ 新增：根据管理员权限过滤培训
    if allowed_countries is not None:
        filtered_trainings = []
        for t in trainings:
            training_country = t.get('country') or t.get('countries')
            if training_country:
                # 解析国家列表
                country_list = []
                if isinstance(training_country, str):
                    try:
                        parsed = json.loads(training_country)
                        country_list = parsed if isinstance(parsed, list) else [training_country]
                    except:
                        country_list = [training_country]
                elif isinstance(training_country, list):
                    country_list = training_country
                else:
                    country_list = [str(training_country)]
                
                # 检查是否有交集
                if any(c in allowed_countries for c in country_list):
                    filtered_trainings.append(t)
        trainings = filtered_trainings

    # 5. 查询考试（类似逻辑）
    if allowed_countries is not None:
        exam_query = db.table("exams").select("*")
    else:
        exam_query = db.table("exams").select("*")
    
    if country:
        exam_query = exam_query.eq("country", country)
    if exam_name:
        exam_query = exam_query.ilike("title", f"%{exam_name}%")
    
    exams = exam_query.execute().data or []
    
    # 根据管理员权限过滤考试
    if allowed_countries is not None:
        filtered_exams = []
        for e in exams:
            exam_country = e.get('country') or e.get('countries')
            if exam_country:
                country_list = []
                if isinstance(exam_country, str):
                    try:
                        parsed = json.loads(exam_country)
                        country_list = parsed if isinstance(parsed, list) else [exam_country]
                    except:
                        country_list = [exam_country]
                elif isinstance(exam_country, list):
                    country_list = exam_country
                else:
                    country_list = [str(exam_country)]
                
                if any(c in allowed_countries for c in country_list):
                    filtered_exams.append(e)
        exams = filtered_exams

    # 6. 调用生成函数
    try:
        buffer, filename = export.generate_bilingual_excel_filtered(
            trainings=trainings,
            exams=exams,
            country=country,
            start_date=start_date,
            end_date=end_date,
            user_ids=final_user_ids,
            wh_id=wh_id,
            lang=lang
        )
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Excel 生成失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@admin_export_bp.route('/admin/export_pdf_by_result/<int:result_id>')
@login_required
@admin_required
def admin_export_pdf_by_result(result_id):
    """通过成绩记录ID直接导出PDF（精确匹配，避免取错记录）"""
    db = get_supabase()
    
    # 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).execute()
    if not result_res.data:
        #flash("成绩记录不存在", "danger")
        flash({'msg': 'result_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    result = result_res.data[0]
    
    exam_id = result['exam_id']
    user_id = result['user_id']
    
    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    if not exam_res.data:
        #flash("考试不存在", "danger")
        flash({'msg': 'exam_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    exam_data = exam_res.data[0]
    
    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    if not user_res.data:
        #flash("考生不存在", "danger")
        flash({'msg': 'student_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    user_data = user_res.data[0]
    user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生')

    # 解析 answers 和 details（递归处理双重转义）
    def robust_parse_json(value):
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, str):
                    return robust_parse_json(parsed)
                return parsed
            except:
                return {}
        return {}
    
    answers = robust_parse_json(result.get('answers'))
    details = robust_parse_json(result.get('details'))
    
    logger.info(f"导出 result_id={result_id}, answers 键数: {len(answers)}, details 键数: {len(details)}")
    
    # 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    
    # 阅卷人
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=request.args.get('reviewer')
    )
    
    # 生成 PDF
    try:
        pdf_buffer = export.generate_user_pdf(
            user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
            user_email=user_data.get('email', ''),
            exam_title=exam_data.get('title', '未命名考试'),
            score=result.get('total_score', 0),
            questions=questions,
            answers=answers,
            details=details,
            submitted_at=result.get('created_at', ''),
            reviewer=reviewer
        )
    except Exception as e:
        logger.error(f"PDF 生成失败: {e}")
        #flash(f"PDF 生成失败: {str(e)}", "danger")
        flash({'msg': 'pdf_generation_error', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    
    filename = f"Transcript_{user_name}_{exam_data.get('title', 'exam')}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

def generate_pdf_by_result_id(result_id):
    """根据 result_id 直接生成 PDF 字节流"""
    db = get_supabase()
    # 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).execute()
    if not result_res.data:
        return None
    result = result_res.data[0]
    exam_id = result['exam_id']
    user_id = result['user_id']

    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    exam_data = exam_res.data[0] if exam_res.data else {}

    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    user_data = user_res.data[0] if user_res.data else {}

    # 递归解析 answers, details
    def robust_parse_json(value):
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, str):
                    return robust_parse_json(parsed)
                return parsed
            except:
                return {}
        return {}

    answers = robust_parse_json(result.get('answers'))
    details = robust_parse_json(result.get('details'))

    # 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []

    # 阅卷人
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=None
    )

    # 生成 PDF 并返回字节数据
    pdf_buffer = export.generate_user_pdf(
        user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
        user_email=user_data.get('email', ''),
        exam_title=exam_data.get('title', '未命名考试'),
        score=result.get('total_score', 0),
        questions=questions,
        answers=answers,
        details=details,
        submitted_at=result.get('created_at', ''),
        reviewer=reviewer
    )
    return pdf_buffer.getvalue()  # 返回字节数据

@admin_export_bp.route('/api/admin/exam/batch_export_by_result', methods=['POST'])
@login_required
@admin_required
def admin_batch_export_by_result():
    """根据成绩记录ID批量导出PDF（自动去重，每个考生只导出最新一次考试）"""
    try:
        data = request.json
        result_ids = data.get('result_ids', [])
        exam_id = data.get('exam_id')
        
        if not result_ids:
            return jsonify({"success": False, "message": "请选择要导出的成绩记录"}), 400
        
        db = get_supabase()
        
        # 获取所有选中的成绩记录
        results_res = db.table("exam_results").select("*").in_("id", result_ids).is_("deleted_at", "null").execute()
        results = results_res.data or []
        
        if not results:
            return jsonify({"success": False, "message": "未找到有效的成绩记录"}), 404
        
        # ✅ 关键：按 user_id 去重，只保留每个考生最新的成绩记录
        user_latest_result = {}
        for result in results:
            user_id = result['user_id']
            created_at = result.get('created_at', '')
            
            # 如果该用户还没有记录，或者当前记录更新，则保存
            if user_id not in user_latest_result:
                user_latest_result[user_id] = result
            else:
                existing_created_at = user_latest_result[user_id].get('created_at', '')
                if created_at > existing_created_at:
                    user_latest_result[user_id] = result
        
        # 获取去重后的最新成绩记录列表
        unique_latest_results = list(user_latest_result.values())
        
        logger.info(f"批量导出: 原始选中 {len(results)} 条记录，去重后 {len(unique_latest_results)} 个考生")
        
        # 记录被过滤掉的重复记录
        if len(results) > len(unique_latest_results):
            logger.info(f"过滤掉 {len(results) - len(unique_latest_results)} 条重复记录（同一考生的旧成绩）")
        
        # 获取考试信息
        exam_res = db.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
        if not exam_res.data:
            return jsonify({"success": False, "message": "考试不存在"}), 404
        exam_data = exam_res.data
        
        # 获取所有用户信息
        user_ids = list(set(r['user_id'] for r in unique_latest_results))
        users_res = db.table("users").select("*").in_("id", user_ids).execute()
        users_dict = {u['id']: u for u in (users_res.data or [])}
        
        # 获取题目列表
        questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
        questions = questions_res.data or []
        
        from io import BytesIO
        import zipfile
        import re
        
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for result in unique_latest_results:
                user_id = result['user_id']
                user_data = users_dict.get(user_id, {})
                score = result.get('total_score', 0)
                
                # 解析答案和详情
                answers = {}
                details = {}
                
                if result.get('answers'):
                    try:
                        answers = json.loads(result['answers']) if isinstance(result['answers'], str) else result['answers']
                    except:
                        pass
                
                if result.get('details'):
                    try:
                        details = json.loads(result['details']) if isinstance(result['details'], str) else result['details']
                    except:
                        pass
                
                # 获取阅卷人
                reviewer = get_reviewer_by_country(
                    user_country=user_data.get('country'),
                    exam_reviewer=exam_data.get('reviewer'),
                    url_reviewer=None
                )
                
                user_name = user_data.get('name_cn') or user_data.get('name_en', '未知考生')
                
                # 生成PDF
                pdf_buffer = export.generate_user_pdf(
                    user_name=user_name,
                    user_email=user_data.get('email', ''),
                    exam_title=exam_data.get('title', '未命名考试'),
                    score=score,
                    questions=questions,
                    answers=answers,
                    details=details,
                    submitted_at=result.get('created_at', ''),
                    reviewer=reviewer
                )
                
                # 清理文件名中的非法字符
                safe_name = re.sub(r'[\\/*?:"<>|]', '_', user_name)
                safe_title = re.sub(r'[\\/*?:"<>|]', '_', exam_data.get('title', 'exam'))
                
                # 文件名包含得分和提交时间（用于区分多次考试）
                submitted_date = result.get('created_at', '')[:10] if result.get('created_at') else 'unknown'
                filename = f"Transcript_{safe_name}_{safe_title}_{submitted_date}_{score}.pdf"
                
                zf.writestr(filename, pdf_buffer.getvalue())
                logger.info(f"  已添加: {filename}")
        
        zip_buffer.seek(0)
        
        safe_title = re.sub(r'[\\/*?:"<>|]', '_', exam_data.get('title', 'exam'))
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'{safe_title}_考试成绩_{len(unique_latest_results)}人.zip'
        )
        
    except Exception as e:
        logger.error(f"批量导出失败: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500

@admin_export_bp.route('/admin/export_pdf/<int:exam_id>/<user_id>')
@login_required
@admin_required
def admin_export_user_pdf(exam_id, user_id):
    """管理员导出指定考生的成绩单 PDF, 好像未使用"""
    db = get_supabase()
    
    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    if not exam_res.data:
        #flash("考试不存在", "danger")
        flash({'msg': 'exam_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    exam_data = exam_res.data[0]
    
    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    if not user_res.data:
        #flash("考生不存在", "danger")
        flash({'msg': 'student_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    user_data = user_res.data[0]
    
    # 获取成绩记录（不使用 maybe_single，避免 204 异常）
    result_res = db.table("exam_results").select("*").eq("exam_id", exam_id).eq("user_id", user_id).execute()
    if not result_res.data:
        #flash("该考生尚无成绩记录", "warning")
        flash({'msg': 'no_score_record', 'params': []}, 'warning')
        return redirect(url_for('admin_dashboard'))
    result = result_res.data[0]

    raw_answers = result.get('answers')
    logger.info(f"原始 answers 类型: {type(raw_answers)}, 值前200: {str(raw_answers)[:200]}")

    raw_details = result.get('details')
    logger.info(f"原始 details 类型: {type(raw_details)}, 值前200: {str(raw_details)[:200]}")

    answers = robust_parse_json(result.get('answers'), "answers")
    details = robust_parse_json(result.get('details'), "details")

    logger.info(f"解析后 answers 类型: {type(answers)}, 键示例: {list(answers.keys())[:5] if isinstance(answers, dict) else 'not dict'}")
    logger.info(f"解析后 details 类型: {type(details)}, 键示例: {list(details.keys())[:5] if isinstance(details, dict) else 'not dict'}")

    # 在获取成绩记录后，如果 answers 为空，则尝试从 user_exam_drafts 表读取草稿答案
    if not answers:
        draft_res = db.table("user_exam_drafts").select("answers").eq("user_id", user_id).eq("exam_id", exam_id).execute()
        if draft_res.data:
            draft_answers = draft_res.data[0].get('answers')
            if draft_answers:
                if isinstance(draft_answers, str):
                    try:
                        answers = json.loads(draft_answers)
                        logger.info(f"从草稿表恢复了答案，共 {len(answers)} 条")
                    except:
                        pass
                else:
                    answers = draft_answers

    logger.info(f"解析后 answers 类型: {type(answers)}, 键示例: {list(answers.keys())[:3]}")
    logger.info(f"解析后 details 类型: {type(details)}, 键示例: {list(details.keys())[:3]}")

    # 获取题目列表（用于展示题干和标准答案）
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=request.args.get('reviewer')
    )


    # ✅ 添加调试日志
    logger.info(f"========== 阅卷人调试 ==========")
    logger.info(f"考生国家: {user_data.get('country')}")
    logger.info(f"考试表 reviewer: {exam_data.get('reviewer')}")
    logger.info(f"URL reviewer: {request.args.get('reviewer')}")
    logger.info(f"最终 reviewer: {reviewer}")
    logger.info(f"================================")

    # 生成 PDF（捕获异常）
    try:
        pdf_buffer = export.generate_user_pdf(
            user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
            user_email=user_data.get('email', ''),
            exam_title=exam_data.get('title', '未命名考试'),
            score=result.get('total_score', 0),
            questions=questions,
            answers=answers,
            details=details,
            submitted_at=result.get('created_at', ''),
            reviewer=reviewer
        )
    except Exception as e:
        logger.error(f"PDF 生成失败: {e}")
        #flash(f"PDF 生成失败: {str(e)}", "danger")
        flash({'msg': 'pdf_generation_error', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    
    filename = f"Transcript_{user_name}_{exam_data.get('title', 'exam')}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

@admin_export_bp.route('/api/admin/exam/<int:exam_id>/batch_export', methods=['POST'])
@admin_required
def admin_batch_export_pdf(exam_id):
    logger.info(f"✅ batch_export 被调用，exam_id={exam_id}")

    data = request.json
    db = get_supabase()
    user_ids = data.get('user_ids', [])
    if not user_ids:
        return jsonify({"error": "未选择考生"}), 400
    # 去重
    user_ids = list(set(user_ids))
    logger.info(f"批量导出考试 {exam_id}，考生数量: {len(user_ids)}，IDs: {user_ids}")

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for uid in user_ids:
            # 调用生成单个考生PDF的函数，返回字节流
            pdf_bytes, score, result_id = generate_single_user_pdf(exam_id, uid)
            if pdf_bytes:
                # 获取考生姓名
                user_res = db.table("users").select("name_en, email").eq("id", uid).execute()
                name = user_res.data[0].get('name_en', uid) if user_res.data else uid                
                result_id = get_latest_result_id(exam_id, uid)  # 需实现该函数
                safe_name = re.sub(r'[\\/*?:"<>|]', '_', name)
                # filename = f"{name}_{exam_id}_{result_id}.pdf"
                if score is not None:
                    filename = f"{safe_name}_({exam_id}_{result_id})_{score}.pdf"
                else:
                    filename = f"{safe_name}_({exam_id}_{result_id}).pdf"
                zf.writestr(filename, pdf_bytes)
                logger.info(f"  已添加文件: {filename}")
    zip_buffer.seek(0)
    
    # 获取考试标题用于文件名
    exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
    exam_title = exam_res.data.get('title', 'exam') if exam_res.data else 'exam'
    safe_title = re.sub(r'[\\/*?:"<>|]', '_', exam_title)
    
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=f"{safe_title}_{exam_id}_scores.zip")

def generate_single_user_pdf(exam_id, user_id):
    """为指定考试和考生生成PDF字节流，返回 (pdf_bytes, score, result_id)"""
    # logger.info(f"[批量导出] 开始生成 PDF: exam_id={exam_id}, user_id={user_id}")
    db = get_supabase()
    # 1. 获取该考生在该考试的最新成绩记录
    result_res = db.table("exam_results") \
        .select("*") \
        .eq("exam_id", exam_id) \
        .eq("user_id", user_id) \
        .order("created_at", desc=True) \
        .limit(1) \
        .execute()
    if not result_res.data:
        logger.warning(f"未找到考试 {exam_id} 用户 {user_id} 的成绩记录")
        return None, None, None

    result = result_res.data[0]
    score = result.get('total_score', 0)
    result_id = result.get('id')
    # logger.info(f"  找到成绩记录 result_id={result['id']}, score={result.get('total_score')}")

    # 2. 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    if not exam_res.data:
        return None, None, None
    exam_data = exam_res.data[0]

    # 3. 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    if not user_res.data:
        return None, None, None
    user_data = user_res.data[0]

    # 4. 解析 answers, details（递归）
    def robust_parse_json(value):
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, str):
                    return robust_parse_json(parsed)
                return parsed
            except:
                return {}
        return {}
    answers = robust_parse_json(result.get('answers'))
    details = robust_parse_json(result.get('details'))

    # 5. 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []

    # 6. 阅卷人（默认）
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=None  # 批量导出时没有 URL 参数
    )

    '''
    # ✅ 添加调试日志
    logger.info(f"========== generate_single_user_pdf 阅卷人 ==========")
    logger.info(f"考生国家: {user_data.get('country')}")
    logger.info(f"考试表 reviewer: {exam_data.get('reviewer')}")
    logger.info(f"最终 reviewer: {reviewer}")
    logger.info(f"=================================================")
    '''

    # 7. 生成PDF字节流
    pdf_buffer = export.generate_user_pdf(
        user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
        user_email=user_data.get('email', ''),
        exam_title=exam_data.get('title', '未命名考试'),
        score=score,
        questions=questions,
        answers=answers,
        details=details,
        submitted_at=result.get('created_at', ''),
        reviewer=reviewer
    )
    # 返回字节数据（BytesIO 需要 .getvalue()）
    return pdf_buffer.getvalue(), score, result_id  # 注意：generate_user_pdf 返回 BytesIO，需提取内容

def get_latest_result_id(exam_id, user_id):
    db = get_supabase()
    res = db.table("exam_results").select("id").eq("exam_id", exam_id).eq("user_id", user_id).order("created_at", desc=True).limit(1).execute()
    return res.data[0]['id'] if res.data else None

# ==================导出API参数0615==================
@admin_export_bp.route('/api/admin/export/quarters')
@login_required
@admin_required
def api_get_quarters_for_export():
    """获取季度列表（基于权限范围内的培训签到日期和考试提交日期）"""
    db = get_supabase()
    allowed_countries = get_admin_allowed_countries()
    
    quarters = set()
    
    try:
        # 1. 从培训签到记录中提取日期
        trainings_query = db.table("trainings").select("start_time, end_time")
        if allowed_countries is not None and allowed_countries:
            trainings_query = trainings_query.in_("country", allowed_countries)
        trainings_res = trainings_query.execute()
        
        for t in (trainings_res.data or []):
            for time_field in ['start_time', 'end_time']:
                time_str = t.get(time_field)
                if time_str:
                    quarter = _get_quarter_from_date(time_str)
                    if quarter:
                        quarters.add(quarter)
        
        # 2. 从考试成绩提交日期中提取季度
        # ✅ 修复：先获取所有考试，再在 Python 中过滤权限
        exams_res = db.table("exams").select("id, countries").is_("deleted_at", "null").execute()
        all_exams = exams_res.data or []
        
        # 在 Python 中过滤权限范围内的考试
        filtered_exam_ids = []
        for exam in all_exams:
            exam_countries = exam.get('countries')
            if not exam_countries:
                continue
            
            # 解析 JSONB 数组
            if isinstance(exam_countries, str):
                try:
                    import json
                    exam_countries = json.loads(exam_countries)
                except:
                    exam_countries = []
            elif not isinstance(exam_countries, list):
                exam_countries = []
            
            # 检查是否有交集
            if allowed_countries is not None and allowed_countries:
                if any(c in allowed_countries for c in exam_countries):
                    filtered_exam_ids.append(exam['id'])
            else:
                filtered_exam_ids.append(exam['id'])
        
        if filtered_exam_ids:
            # 查询这些考试的考试成绩
            results_res = db.table("exam_results").select("created_at").in_("exam_id", filtered_exam_ids).execute()
            for r in (results_res.data or []):
                quarter = _get_quarter_from_date(r.get('created_at'))
                if quarter:
                    quarters.add(quarter)
        
        # 3. 排序并返回（最近12个季度）
        sorted_quarters = sorted(list(quarters), key=lambda x: (int(x[:4]), int(x[5])))
        return jsonify(sorted_quarters[-12:])
        
    except Exception as e:
        logger.error(f"获取季度列表失败: {e}")
        return jsonify([])

def _get_quarter_from_date(date_str):
    """从日期字符串提取季度（格式：2026Q2）"""
    if not date_str:
        return None
    try:
        if isinstance(date_str, str):
            # 处理 ISO 格式
            if 'T' in date_str:
                date_str = date_str.split('T')[0]
            elif ' ' in date_str:
                date_str = date_str.split(' ')[0]
        from datetime import datetime
        date = datetime.fromisoformat(date_str) if isinstance(date_str, str) else date_str
        year = date.year
        month = date.month
        quarter = (month - 1) // 3 + 1
        return f"{year}Q{quarter}"
    except:
        return None

@admin_export_bp.route('/api/admin/export/recent_countries')
@login_required
@admin_required
def api_get_recent_countries_for_export():
    """获取最近使用的国家列表（基于权限范围内的培训和考试）"""
    db = get_supabase()
    allowed_countries = get_admin_allowed_countries()
    
    # 获取权限范围内的国家列表
    if allowed_countries is not None and allowed_countries:
        # 如果管理员有权限范围限制，只返回这些国家
        countries_res = db.table("countries").select("code, name_zh, name_en").in_("code", allowed_countries).execute()
    else:
        # 无限制，返回所有国家
        countries_res = db.table("countries").select("code, name_zh, name_en").execute()
    
    countries = countries_res.data or []
    
    # 按中文名称排序
    countries.sort(key=lambda x: x.get('name_zh', ''))
    
    return jsonify(countries[:100])  # 返回所有权限范围内的国家

@admin_export_bp.route('/api/admin/export/recent_warehouses')
@login_required
@admin_required
def api_get_recent_warehouses_for_export():
    """获取最近使用的库房列表（基于有培训签到的用户）"""
    db = get_supabase()
    allowed_countries = get_admin_allowed_countries()
    
    # 获取有培训签到的用户ID
    attend_res = db.table("training_attendances").select("user_id").execute()
    user_ids = list(set([a['user_id'] for a in (attend_res.data or [])]))
    
    if not user_ids:
        return jsonify([])
    
    # 获取这些用户的库房信息
    users_query = db.table("users").select("wh_id, wh_name_en, country").in_("id", user_ids)
    
    if allowed_countries is not None and allowed_countries:
        users_query = users_query.in_("country", allowed_countries)
    
    users_res = users_query.execute()
    
    # 去重
    seen = set()
    warehouses = []
    for u in (users_res.data or []):
        wh_id = u.get('wh_id')
        if wh_id and wh_id not in seen:
            seen.add(wh_id)
            warehouses.append({
                "wh_id": wh_id,
                "wh_name_en": u.get('wh_name_en', ''),
                "display": f"{wh_id} ({u.get('wh_name_en', '')})" if u.get('wh_name_en') else wh_id
            })
    
    return jsonify(warehouses[:20])

# =================生成综合报告=====================
# routes/admin_export.py

@admin_export_bp.route('/api/admin/export_multi_excel', methods=['POST'])
@login_required
@admin_required
def export_multi_excel():
    """
    综合导出接口（支持多培训/多考试）
    复用 generate_bilingual_excel_filtered 生成 Excel
    """
    try:
        data = request.json
        training_ids = data.get('training_ids', [])
        exam_ids = data.get('exam_ids', [])
        country = data.get('country', '')
        wh_raw = data.get('wh_id', '').strip()
        wh_id = wh_raw.split('(')[0].strip() if wh_raw else None
        lang = data.get('lang', 'zh')
        
        db = get_supabase()
        allowed_countries = get_admin_allowed_countries()
        
        # ========== 1. 获取用户ID（按国家或库房过滤） ==========
        final_user_ids = None
        
        # 按国家过滤
        user_ids_for_country = None
        if country:
            users_res = db.table("users").select("id").eq("country", country).execute()
            user_ids_for_country = [u['id'] for u in (users_res.data or [])]
            if not user_ids_for_country:
                return _create_empty_excel(f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}")
        
        # 按库房过滤
        user_ids_for_wh = None
        if wh_id:
            users_in_wh = db.table("users").select("id").eq("wh_id", wh_id).execute()
            user_ids_for_wh = [u['id'] for u in (users_in_wh.data or [])]
            if not user_ids_for_wh:
                return _create_empty_excel(f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}")
        
        # 合并国家与库房的用户范围
        if user_ids_for_country is not None and user_ids_for_wh is not None:
            final_user_ids = list(set(user_ids_for_country) & set(user_ids_for_wh))
        elif user_ids_for_country is not None:
            final_user_ids = user_ids_for_country
        elif user_ids_for_wh is not None:
            final_user_ids = user_ids_for_wh
        
        # 管理员权限范围过滤
        if allowed_countries is not None and final_user_ids is not None:
            users_in_allowed = db.table("users").select("id").in_("id", final_user_ids).in_("country", allowed_countries).execute()
            final_user_ids = [u['id'] for u in (users_in_allowed.data or [])]
        elif allowed_countries is not None:
            users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
            final_user_ids = [u['id'] for u in (users_in_allowed.data or [])]
        
        # ========== 2. 获取培训信息（根据 training_ids） ==========
        trainings = []
        if training_ids:
            trainings_res = db.table("trainings").select("*").in_("id", training_ids).is_("deleted_at", "null").execute()
            trainings = trainings_res.data or []
        
        # 如果 training_ids 为空，但 exam_ids 有值，从考试绑定关系获取培训
        if not trainings and exam_ids:
            bindings_res = db.table("training_exam_bindings").select("training_id")\
                .in_("exam_id", exam_ids)\
                .is_("deleted_at", "null")\
                .execute()
            bound_training_ids = list(set([b['training_id'] for b in (bindings_res.data or [])]))
            if bound_training_ids:
                trainings_res = db.table("trainings").select("*").in_("id", bound_training_ids).is_("deleted_at", "null").execute()
                trainings = trainings_res.data or []
        
        # ========== 3. 获取考试信息（根据 exam_ids） ==========
        exams = []
        if exam_ids:
            exams_res = db.table("exams").select("*").in_("id", exam_ids).is_("deleted_at", "null").execute()
            exams = exams_res.data or []
        
        # 如果 exam_ids 为空，但 training_ids 有值，从培训绑定关系获取考试
        if not exams and training_ids:
            bindings_res = db.table("training_exam_bindings").select("exam_id")\
                .in_("training_id", training_ids)\
                .is_("deleted_at", "null")\
                .execute()
            bound_exam_ids = list(set([b['exam_id'] for b in (bindings_res.data or [])]))
            if bound_exam_ids:
                exams_res = db.table("exams").select("*").in_("id", bound_exam_ids).is_("deleted_at", "null").execute()
                exams = exams_res.data or []
        
        # ========== 4. 权限过滤 ==========
        if allowed_countries is not None:
            # 过滤培训
            filtered_trainings = []
            for t in trainings:
                training_country = t.get('country') or t.get('countries')
                if training_country:
                    # ✅ 使用 parse_exam_countries 解析国家列表（传入字典对象）
                    country_list = parse_exam_countries(t)
                    if any(c in allowed_countries for c in country_list):
                        filtered_trainings.append(t)
            trainings = filtered_trainings
            
            # 过滤考试
            filtered_exams = []
            for e in exams:
                exam_country = e.get('country') or e.get('countries')
                if exam_country:
                    # ✅ 使用 parse_exam_countries 解析国家列表（传入字典对象）
                    country_list = parse_exam_countries(e)
                    if any(c in allowed_countries for c in country_list):
                        filtered_exams.append(e)
            exams = filtered_exams
        
        # 如果没有任何数据，返回空文件
        if not trainings and not exams:
            return _create_empty_excel(f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}")
        
        # ========== 5. 调用 generate_bilingual_excel_filtered 生成 Excel ==========
        buffer, filename = export.generate_bilingual_excel_filtered(
            trainings=trainings,
            exams=exams,
            country=country,
            start_date='',
            end_date='',
            user_ids=final_user_ids,
            wh_id=wh_id,
            lang=lang
        )
        
        # 修改文件名，标识为综合报告
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        lang_suffix = 'en' if lang == 'en' else 'zh'
        filename = f"综合报告_{timestamp}_{lang_suffix}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"综合导出失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

def _create_empty_excel(filename):
    """创建空 Excel 文件"""
    from io import BytesIO
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空报告"
    ws.cell(row=1, column=1, value="没有匹配的数据")
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
