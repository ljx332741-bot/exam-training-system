# routes/admin_user.py
import os, json, logging, uuid, secrets, string, sys, openpyxl, re
from io import BytesIO
from datetime import datetime, timezone, timedelta, date
from flask import  (
    Flask, render_template, request, redirect, url_for, 
    session, flash, jsonify, send_file, make_response, current_app as app
)
# ✅ 添加 openpyxl 样式导入
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from . import admin_user_bp
from services import auth, exam, export
from services.db import get_supabase
from services.auth import hash_password
from utils.email_notifier import send_bilingual_notification, EmailScenario, _format_time
from utils.permissions import (is_developer, apply_country_filter, can_view_user, get_admin_allowed_countries, 
    can_modify_user, parse_countries_input, filter_users_by_permission, can_resign_user, can_rehire_user
)
from routes.helpers import login_required, admin_required, get_current_user
from utils.import_helper import parse_excel_rows, validate_country_and_wh_id, generate_import_template, format_import_result
from utils.i18n_messages import I18nMessages

        
logger = logging.getLogger(__name__)

@admin_user_bp.route('/admin/users')
@login_required
@admin_required
def admin_user_list():
    # 添加调试代码，检查模板文件路径
    import os
    template_path = os.path.join(app.root_path, 'templates', 'admin', 'list_users.html')
    
    return render_template('admin/list_users.html', is_developer=is_developer())

@admin_user_bp.route('/api/admin/users/<user_id>')
@login_required
@admin_required
def api_admin_user_detail(user_id):
    """后端单用户查询接口"""
    db = get_supabase()
    
    try:
        res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
        
        # ✅ 检查是否有数据（maybe_single 可能返回 None 或空数据）
        if not res or not hasattr(res, 'data') or not res.data:
            return jsonify({"error": "用户不存在"}), 404
        
        user_data = res.data
        # ✅ 确保离职状态字段存在
        user_data['is_resign'] = user_data.get('is_resign', False)
        user_data['is_rehire'] = user_data.get('is_rehire', False)
        
        return jsonify(user_data)
    except Exception as e:
        # ✅ 捕获可能的 204 异常
        error_msg = str(e)
        if '204' in error_msg or 'Missing response' in error_msg:
            return jsonify({"error": "用户不存在"}), 404
        logger.error(f"获取用户详情失败: {e}")
        return jsonify({"error": "查询失败"}), 500

'''
@admin_user_bp.route('/api/admin/users')
@login_required
@admin_required
def api_admin_users():
    """获取用户列表（带完整权限控制 + 国家名称模糊匹配）"""
    db = get_supabase()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    country = request.args.get('country', '').strip()
    countries_param = request.args.get('countries', '').strip()
    training_id = request.args.get('training_id', '').strip()
    exam_id = request.args.get('exam_id', '').strip()
    wh_id = request.args.get('wh_id', '').strip()
    user_status = request.args.get('status', '')
    
    allowed_countries = get_admin_allowed_countries()
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    is_dev = is_developer()

    # ========== 1. 从考试/培训获取国家列表 ==========
    exam_countries = []
    if exam_id:
        try:
            exam_res = db.table("exams").select("countries, country").eq("id", int(exam_id)).execute()
            if exam_res.data:
                exam = exam_res.data[0]
                if exam.get('countries'):
                    countries_data = exam.get('countries')
                    if isinstance(countries_data, str):
                        try:
                            exam_countries = json.loads(countries_data)
                        except:
                            exam_countries = []
                    elif isinstance(countries_data, list):
                        exam_countries = countries_data
                if not exam_countries and exam.get('country'):
                    exam_countries = [exam.get('country')]
        except:
            pass

    training_countries = []
    if training_id:
        try:
            training_res = db.table("trainings").select("country, countries").eq("id", int(training_id)).execute()
            if training_res.data:
                training = training_res.data[0]
                if training.get('countries'):
                    countries_data = training.get('countries')
                    if isinstance(countries_data, str):
                        try:
                            training_countries = json.loads(countries_data)
                        except:
                            training_countries = []
                    elif isinstance(countries_data, list):
                        training_countries = countries_data
                if not training_countries and training.get('country'):
                    training_countries = [training.get('country')]
        except:
            pass

    # ========== 2. 国家模糊匹配：将输入的文本转换为匹配的国家代码列表 ==========
    matched_country_codes = []
    country_search_term = country or countries_param
    
    if country_search_term:
        # 获取所有国家列表
        countries_res = db.table("countries").select("code, name_zh, name_en").execute()
        all_countries = countries_res.data or []
        
        search_lower = country_search_term.lower().strip()
        
        for c in all_countries:
            code = (c.get('code') or '').lower()
            name_zh = (c.get('name_zh') or '').lower()
            name_en = (c.get('name_en') or '').lower()
            
            # ✅ 模糊匹配：中英文名称或代码包含搜索词
            if (search_lower in name_zh or 
                search_lower in name_en or 
                search_lower in code):
                matched_country_codes.append(c['code'])
        
        logger.info(f"国家模糊匹配: 搜索词='{country_search_term}', 匹配到 {len(matched_country_codes)} 个国家: {matched_country_codes}")
        
        # 如果没有匹配到任何国家，返回空结果
        if not matched_country_codes:
            return jsonify({
                "data": [],
                "total": 0,
                "page": page,
                "per_page": per_page
            })
    
    # ========== 3. 确定最终国家过滤列表 ==========
    final_countries = []
    
    # 优先使用模糊匹配的国家代码
    if matched_country_codes:
        final_countries = matched_country_codes
    elif exam_countries:
        final_countries = exam_countries
    elif training_countries:
        final_countries = training_countries
    
    # 与管理员权限范围取交集
    if allowed_countries is not None and final_countries:
        final_countries = [c for c in final_countries if c in allowed_countries]
    
    # 如果管理员有权限范围但没有指定国家筛选，使用权限范围
    if not final_countries and allowed_countries is not None and allowed_countries:
        final_countries = allowed_countries
    
    logger.info(f"用户列表请求 - 最终国家过滤: {final_countries}")

    # ========== 4. 基础查询 ==========
    query = db.table("users").select("*", count="exact").is_("deleted_at", "null")

    if not is_dev:
        query = query.eq("is_protected", False)
    
    if current_role != 'super_admin' and not is_dev:
        query = query.neq("role", "super_admin").neq("role", "developer")
    
    if user_status:
        query = query.eq("user_status", user_status)
    
    # 获取所有数据
    all_res = query.execute()
    all_users = all_res.data or []

    # ========== 5. 内存过滤 ==========
    exclude_resigned = request.args.get('exclude_resigned', 'true').lower() == 'true'
    
    filtered = []
    for user in all_users:
        # 姓名/邮箱/工号搜索
        if search:
            search_lower = search.lower()
            name = (user.get('name_en') or '').lower()
            email = (user.get('email') or '').lower()
            emp_id = (user.get('employee_id') or '').lower()
            if search_lower not in name and search_lower not in email and search_lower not in emp_id:
                continue

        # ✅ 国家筛选：使用匹配的国家代码列表
        if final_countries:
            user_country = user.get('country') or ''
            if user_country not in final_countries:
                continue
        
        # 库房模糊匹配
        if wh_id:
            wh_lower = wh_id.lower()
            user_wh_id = (user.get('wh_id') or '').lower()
            user_wh_name = (user.get('wh_name_en') or '').lower()
            if wh_lower not in user_wh_id and wh_lower not in user_wh_name:
                continue
        
        filtered.append(user)

    # 离职过滤
    if exclude_resigned:
        filtered = [u for u in filtered if not u.get('is_resign', False)]
        logger.info(f"排除离职人员后剩余 {len(filtered)} 人")
    
    # 权限过滤
    filtered_users = filter_users_by_permission(filtered, allowed_countries, current_user_id)
    
    # 按创建时间倒序排序
    filtered_users.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    # 内存分页
    total = len(filtered_users)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = filtered_users[start:end]

    # 确保字段存在
    for user in paginated:
        if 'is_resign' in user:
            user['is_resign'] = user.get('is_resign', False)
        if 'is_rehire' in user:
            user['is_rehire'] = user.get('is_rehire', False)

    return jsonify({
        "data": paginated,
        "total": total,
        "page": page,
        "per_page": per_page
    })
'''

@admin_user_bp.route('/api/admin/users')
@login_required
@admin_required
def api_admin_users():
    """获取用户列表（带完整权限控制 + 国家名称模糊匹配）"""
    db = get_supabase()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    country = request.args.get('country', '').strip()
    countries_param = request.args.get('countries', '').strip()
    training_id = request.args.get('training_id', '').strip()
    exam_id = request.args.get('exam_id', '').strip()
    wh_id = request.args.get('wh_id', '').strip()
    user_status = request.args.get('status', '')
    
    allowed_countries = get_admin_allowed_countries()
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    is_dev = is_developer()

    # ========== 1. 从考试/培训获取国家列表 ==========
    exam_countries = []
    if exam_id:
        try:
            exam_res = db.table("exams").select("countries, country").eq("id", int(exam_id)).execute()
            if exam_res.data:
                exam = exam_res.data[0]
                if exam.get('countries'):
                    countries_data = exam.get('countries')
                    if isinstance(countries_data, str):
                        try:
                            exam_countries = json.loads(countries_data)
                        except:
                            exam_countries = []
                    elif isinstance(countries_data, list):
                        exam_countries = countries_data
                if not exam_countries and exam.get('country'):
                    exam_countries = [exam.get('country')]
        except:
            pass

    training_countries = []
    if training_id:
        try:
            training_res = db.table("trainings").select("country, countries").eq("id", int(training_id)).execute()
            if training_res.data:
                training = training_res.data[0]
                if training.get('countries'):
                    countries_data = training.get('countries')
                    if isinstance(countries_data, str):
                        try:
                            training_countries = json.loads(countries_data)
                        except:
                            training_countries = []
                    elif isinstance(countries_data, list):
                        training_countries = countries_data
                if not training_countries and training.get('country'):
                    training_countries = [training.get('country')]
        except:
            pass

    # ========== 2. 国家模糊匹配 ==========
    matched_country_codes = []
    country_search_term = country or countries_param
    
    if country_search_term:
        countries_res = db.table("countries").select("code, name_zh, name_en").execute()
        all_countries = countries_res.data or []
        search_lower = country_search_term.lower().strip()
        
        for c in all_countries:
            code = (c.get('code') or '').lower()
            name_zh = (c.get('name_zh') or '').lower()
            name_en = (c.get('name_en') or '').lower()
            
            if (search_lower in name_zh or 
                search_lower in name_en or 
                search_lower in code):
                matched_country_codes.append(c['code'])
        
        if not matched_country_codes:
            return jsonify({
                "data": [],
                "total": 0,
                "page": page,
                "per_page": per_page
            })
    
    # ========== 3. 确定最终国家过滤列表 ==========
    final_countries = []
    
    if matched_country_codes:
        final_countries = matched_country_codes
    elif exam_countries:
        final_countries = exam_countries
    elif training_countries:
        final_countries = training_countries
    
    if allowed_countries is not None and final_countries:
        final_countries = [c for c in final_countries if c in allowed_countries]
    
    if not final_countries and allowed_countries is not None and allowed_countries:
        final_countries = allowed_countries
    
    logger.info(f"用户列表请求 - 最终国家过滤: {final_countries}")


    # ========== 4. 基础查询 ==========
    query = db.table("users").select("*", count="exact").is_("deleted_at", "null")

    if not is_dev:
        query = query.eq("is_protected", False)
    
    if user_status:
        query = query.eq("user_status", user_status)
    
    all_res = query.execute()
    all_users = all_res.data or []

    # ========== 5. 内存过滤 ==========
    exclude_resigned = request.args.get('exclude_resigned', 'true').lower() == 'true'
    
    filtered = []
    for user in all_users:
        # 姓名/邮箱/工号搜索
        if search:
            search_lower = search.lower()
            name = (user.get('name_en') or '').lower()
            email = (user.get('email') or '').lower()
            emp_id = (user.get('employee_id') or '').lower()
            if search_lower not in name and search_lower not in email and search_lower not in emp_id:
                continue

        # 国家筛选
        if final_countries:
            user_country = user.get('country') or ''
            if user_country not in final_countries:
                continue
        
        # 库房模糊匹配
        if wh_id:
            wh_lower = wh_id.lower()
            user_wh_id = (user.get('wh_id') or '').lower()
            user_wh_name = (user.get('wh_name_en') or '').lower()
            if wh_lower not in user_wh_id and wh_lower not in user_wh_name:
                continue
        
        filtered.append(user)

    # 离职过滤
    if exclude_resigned:
        filtered = [u for u in filtered if not u.get('is_resign', False)]
        logger.info(f"排除离职人员后剩余 {len(filtered)} 人")
    
    # ========== 6. ✅ 使用修改后的 filter_users_by_permission ==========
    logger.info(f"权限过滤前: {len(filtered)} 个用户")
    for u in filtered:
        logger.info(f"  用户: {u.get('name_en')}, role={u.get('role')}, country={u.get('country')}, admin_countries={u.get('admin_countries')}")

    # 这个函数现在支持同级管理员可见
    filtered_users = filter_users_by_permission(filtered, allowed_countries, current_user_id)

    logger.info(f"权限过滤后: {len(filtered_users)} 个用户")
    for u in filtered_users:
        logger.info(f"  ✅ 保留: {u.get('name_en')}, role={u.get('role')}, country={u.get('country')}")
        
    # 按创建时间倒序排序
    filtered_users.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    # 内存分页
    total = len(filtered_users)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = filtered_users[start:end]

    # 确保字段存在
    for user in paginated:
        if 'is_resign' in user:
            user['is_resign'] = user.get('is_resign', False)
        if 'is_rehire' in user:
            user['is_rehire'] = user.get('is_rehire', False)

    return jsonify({
        "data": paginated,
        "total": total,
        "page": page,
        "per_page": per_page
    })

@admin_user_bp.route('/api/admin/users', methods=['POST'])
@login_required
@admin_required
def api_admin_add_user():
    """管理员添加用户（带角色权限控制）"""
    data = request.json
    email = data.get('email', '').strip().lower() or None
    name_en = data.get('name_en', '').strip()
    role = data.get('role', 'user')
    admin_countries = data.get('admin_countries', '[]')
    user_status = data.get('user_status', 'imported')
    
    db = get_supabase()
    
    # ========== 角色权限校验 ==========
    current_role = session.get('role')
    
    # 开发者可以创建任何角色
    if not is_developer():
        # 非开发者不能创建开发者角色
        if role == 'developer':
            return jsonify({"success": False, "message": "cannot_create_developer", "params": []}), 403
        
        # 超管可以创建超管、管理员、用户
        if current_role == 'super_admin':
            if role not in ['super_admin', 'admin', 'user']:
                return jsonify({"success": False, "message": "invalid_role", "params": []}), 400
        # 管理员只能创建用户
        elif current_role == 'admin':
            if role != 'user':
                return jsonify({"success": False, "message": "admin_can_only_create_user", "params": []}), 403
        else:
            return jsonify({"success": False, "message": "permission_denied", "params": []}), 403
    
    # ========== 超管/管理员权限范围校验 ==========
    if role in ['super_admin', 'admin']:
        # 解析权限范围
        if isinstance(admin_countries, str):
            try:
                countries_list = json.loads(admin_countries)
            except:
                countries_list = parse_countries_input(admin_countries)
        else:
            countries_list = admin_countries
        
        # 权限范围不能为空
        if not countries_list or len(countries_list) == 0:
            return jsonify({
                "success": False, 
                "message": "admin_countries_required",
                "params": [role]
            }), 400
        
        # 存储为 JSON 字符串
        admin_countries_json = json.dumps(countries_list)
    else:
        admin_countries_json = None
    
    # ========== 姓名必填校验 ==========
    if not name_en:
        return jsonify({"success": False, "message": "name_cannot_empty", "params": []}), 400
    
    # ========== 邮箱条件校验 ==========
    if not email and user_status != 'imported':
        return jsonify({"success": False, "message": "mail_cannot_empty", "params": []}), 400
    
    # ========== 检查用户是否已存在 ==========
    existing_users = db.table("users").select("*").eq("name_en", name_en).execute()
    existing_users_list = existing_users.data or []

    # 分离活跃用户和已删除用户
    active_users = [u for u in existing_users_list if u.get('deleted_at') is None]
    deleted_users = [u for u in existing_users_list if u.get('deleted_at') is not None]
    
    # 检查活跃用户重复
    birthday = data.get('birthday', '') or None
    employee_id = data.get('employee_id', '').strip() or None
    
    for active in active_users:
        active_birthday = active.get('birthday')
        active_employee_id = active.get('employee_id')
        
        if (birthday and active_birthday == birthday) or (employee_id and active_employee_id == employee_id):
            return jsonify({"success": False, "message": "duplicate_user_found", "params": []}), 400
        if not birthday and not employee_id:
            return jsonify({"success": False, "message": "duplicate_user_found", "params": []}), 400
    
    # 处理已删除用户的恢复
    if deleted_users:
        existing_deleted = deleted_users[0]
        update_data = {
            "email": email,
            "user_status": user_status,
            "is_active": False if user_status == 'imported' else True,
            "deleted_at": None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "role": role,
            "admin_countries": admin_countries_json,
            "created_by": session['user_id']
        }
        
        # 更新其他字段（如果提供）
        if birthday:
            update_data["birthday"] = birthday
        if employee_id:
            update_data["employee_id"] = employee_id
        if data.get('company'):
            update_data["company"] = data.get('company')
        if data.get('department'):
            update_data["department"] = data.get('department')
        if data.get('country'):
            update_data["country"] = data.get('country')
        if data.get('phone'):
            update_data["phone"] = data.get('phone')
        if data.get('role'):
            update_data["role"] = data.get('role')
        if data.get('wh_type'):
            update_data["wh_type"] = data.get('wh_type')
        if data.get('wh_id'):
            update_data["wh_id"] = data.get('wh_id')
        if data.get('wh_name_en'):
            update_data["wh_name_en"] = data.get('wh_name_en')
        if data.get('is_partner'):
            update_data["is_partner"] = data.get('is_partner') == 'Y'
        
        try:
            db.table("users").update(update_data).eq("id", existing_deleted['id']).execute()
            logger.info(f"恢复已删除用户: {name_en}, ID: {existing_deleted['id']}")
            
            # 如果需要发送邮件通知（仅当有邮箱时）
            if email:
                temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
                password_hash = auth.hash_password(temp_password)
                db.table("users").update({"password_hash": password_hash}).eq("id", existing_deleted['id']).execute()
                try:
                    send_bilingual_notification(
                        email=email,
                        scenario=EmailScenario.USER_CREATED,
                        params={
                            "name": name_en or email,
                            "email": email,
                            "temp_password": temp_password,
                            "host_url": request.host_url,
                        },
                        host_url=request.host_url,
                        auth_module=auth
                    )
                except Exception as e:
                    logger.warning(f"发送邮件失败: {e}")
            
            return jsonify({"success": True, "user_id": existing_deleted['id'], "restored": True})
        except Exception as e:
            logger.error(f"恢复用户失败: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    # 邮箱唯一性检查
    if email:
        exist = db.table("users").select("id").eq("email", email).is_("deleted_at", "null").execute()
        if exist.data:
            return jsonify({"success": False, "message": "email_already_registered", "params": []}), 400

    # ✅ 新增：检查国家权限（如果国家被修改）
    if 'country' in data:
        new_country = data.get('country', '')
        if new_country:
            allowed_countries = get_admin_allowed_countries()
            current_role = session.get('role')
            
            # 非开发者需要检查权限
            if current_role != 'developer' and allowed_countries is not None:
                if allowed_countries and new_country not in allowed_countries:
                    return jsonify({
                        "success": False, 
                        "message": f"无权将用户国家修改为 {new_country}，请联系开发者"
                    }), 403
    
    # 生成用户ID和密码
    user_id = str(uuid.uuid4())
    temp_password = ''
    password_hash = ''
    if email:
        alphabet = string.ascii_letters + string.digits
        temp_password = ''.join(secrets.choice(alphabet) for _ in range(10))
        password_hash = auth.hash_password(temp_password)

    # 准备插入数据
    insert_data = {
        "id": user_id,
        "email": email,
        "password_hash": password_hash,
        "name_en": name_en,
        "company": data.get('company', ''),
        "department": data.get('department', ''),
        "employee_id": employee_id,
        "birthday": birthday,
        "country": data.get('country', ''),
        "phone": data.get('phone', ''),
        "role": role,
        "admin_countries": admin_countries_json,
        "user_status": user_status,
        "is_partner": data.get('is_partner', 'N') == 'Y',
        "wh_type": data.get('wh_type', ''),
        "wh_id": data.get('wh_id', ''),
        "wh_name_en": data.get('wh_name_en', ''),
        "is_active": False if user_status == 'imported' else True,
        "created_by": session['user_id'],
        "is_protected": (role == 'developer')  # 开发者账号自动保护
    }
    try:
        db.table("users").insert(insert_data).execute()
        # 发送邮件通知（仅当邮箱存在）
        if email:
            try:
                send_bilingual_notification(
                    email=email,
                    scenario=EmailScenario.USER_CREATED,
                    params={
                        "name": name_en or email,
                        "email": email,
                        "temp_password": temp_password,
                        "host_url": request.host_url,
                    },
                    host_url=request.host_url,
                    auth_module=auth
                )
            except Exception as e:
                logger.warning(f"发送邮件失败: {e}")
        
        logger.info(f"管理员添加用户: {email or '无邮箱'}, 状态: {user_status}")
        return jsonify({"success": True, "user_id": user_id, "temp_password": temp_password})
    except Exception as e:
        logger.error(f"添加用户失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/users/import', methods=['POST'])
@login_required
@admin_required
def api_admin_import_users():
    """批量导入用户（带角色权限校验）"""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "jsonify_no_file_selected", "params": []}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "jsonify_only_supports_files", "params": []}), 400
    
    # 定义表头映射（支持中英文）
    header_map = {
        '国家': 'country', '邮箱': 'email', '姓名': 'name_en', '角色': 'role',
        '服务商?': 'is_partner', '公司': 'company', '部门': 'department',
        '库房类型': 'wh_type', '库房ID': 'wh_id', '库房名称(EN)': 'wh_name_en',
        '工号': 'employee_id', '手机号': 'phone', '生日': 'birthday', '权限范围': 'admin_countries',
        'country': 'country', 'email': 'email', 'name_en': 'name_en', 'role': 'role',
        'is_partner': 'is_partner', 'company': 'company', 'department': 'department',
        'wh_type': 'wh_type', 'wh_id': 'wh_id', 'wh_name_en': 'wh_name_en',
        'employee_id': 'employee_id', 'phone': 'phone', 'birthday': 'birthday', 'admin_countries': 'admin_countries'
    }
    
    # 解析 Excel，获取有效数据行
    valid_rows, parse_errors, headers = parse_excel_rows(file, header_map, ['name_en'])
    
    if parse_errors:
        return jsonify({
            "success": False,
            "message": "文件解析失败",
            "errors": parse_errors
        }), 400
    
    db = get_supabase()
    allowed_countries = get_admin_allowed_countries()
    
    success_count = 0
    error_rows = []
    
    for row_idx, user_data in valid_rows:
        # 处理空字符串
        for field in ['birthday', 'employee_id', 'email', 'phone', 'company', 'department', 'wh_type', 'wh_id', 'wh_name_en', 'country']:
            if user_data.get(field) == '':
                user_data[field] = None
        
        if user_data.get('email'):
            user_data['email'] = user_data['email'].lower()
        
        name_en = user_data.get('name_en', '')
        if not name_en:
            # ✅ 修复：使用结构化错误
            error_rows.append(I18nMessages.format_error(
                row_idx, "name_required"
            ))
            continue
        
        # 国家与库房ID一致性校验
        country_input = user_data.get('country')
        wh_id = user_data.get('wh_id')
        final_country, is_valid, error_msg = validate_country_and_wh_id(country_input, wh_id)
        
        if not is_valid:
            # ✅ 修复：使用结构化错误
            error_rows.append(I18nMessages.format_error(
                row_idx, "country_wh_mismatch",
                country=country_input or '',
                wh_id=wh_id or ''
            ))
            continue
        
        # 权限校验
        if final_country:
            if allowed_countries is not None and final_country not in allowed_countries:
                # ✅ 修复：使用结构化错误
                error_rows.append(I18nMessages.format_error(
                    row_idx, "country_not_allowed",
                    country=final_country
                ))
                continue
        
        # 角色处理
        role = user_data.get('role', 'user').lower()
        if role not in ['user', 'admin', 'super_admin', 'developer']:
            role = 'user'
        user_data['role'] = role
        
        # 角色权限校验
        if role in ['admin', 'super_admin']:
            admin_countries_raw = user_data.get('admin_countries', '')
            if not admin_countries_raw:
                # ✅ 修复：使用结构化错误
                error_rows.append(I18nMessages.format_error(
                    row_idx, "admin_countries_required",
                    role=role
                ))
                continue
            if ',' in admin_countries_raw:
                countries_list = [c.strip().upper() for c in admin_countries_raw.split(',')]
            else:
                countries_list = [admin_countries_raw.strip().upper()]
            if allowed_countries is not None:
                invalid_countries = [c for c in countries_list if c not in allowed_countries]
                if invalid_countries:
                    # ✅ 修复：使用结构化错误
                    error_rows.append(I18nMessages.format_error(
                        row_idx, "admin_countries_invalid",
                        countries=', '.join(invalid_countries)
                    ))
                    continue
            user_data['admin_countries'] = json.dumps(countries_list)
        else:
            user_data['admin_countries'] = None
        
        # 服务商处理
        is_partner_val = user_data.get('is_partner', 'N')
        user_data['is_partner'] = is_partner_val.upper() in ('Y', 'YES', 'TRUE', '1')
        
        # 设置默认密码
        DEFAULT_PLACEHOLDER_HASH = hash_password('__IMPORTED_USER_PLACEHOLDER__')
        user_data['password_hash'] = DEFAULT_PLACEHOLDER_HASH
        
        # 设置默认值
        user_data['id'] = str(uuid.uuid4())
        user_data['user_status'] = 'imported'
        user_data['is_active'] = False
        user_data['created_by'] = session['user_id']
        user_data['is_protected'] = (role == 'developer')
        user_data['created_at'] = datetime.now(timezone.utc).isoformat()
        user_data['country'] = final_country or None
        
        # 移除空值字段
        user_data = {k: v for k, v in user_data.items() if v != '' and v is not None}
        
        # 修改错误添加部分
        try:
            # 检查是否已存在同名用户
            existing = db.table("users").select("id, deleted_at").eq("name_en", name_en).execute()
            if existing.data:
                existing_user = existing.data[0]
                if existing_user.get('deleted_at') is None:
                    # ✅ 使用 I18nMessages 格式化错误
                    error_rows.append(I18nMessages.format_error(
                        row_idx, "user_already_exists", name=name_en
                    ))
                    continue
                else:
                    # 恢复已删除的用户
                    update_data = {k: v for k, v in user_data.items() if k not in ['id', 'created_at']}
                    update_data['deleted_at'] = None
                    update_data['deleted_by'] = None
                    db.table("users").update(update_data).eq("id", existing_user['id']).execute()
                    success_count += 1
                    continue

            db.table("users").insert(user_data).execute()
            success_count += 1
            
        except Exception as e:
            # ✅ 使用 I18nMessages 格式化错误
            error_rows.append(I18nMessages.format_error(
                row_idx, "insert_failed", error=str(e)
            ))
            logger.error(f"导入用户失败第{row_idx}行: {e}")
    
    # ✅ 返回纯字符串格式的错误信息
    result = {
        "success": True,
        "success_count": success_count,
        "error_count": len(error_rows),
        "errors": error_rows  # 直接返回字符串列表
    }
    return jsonify(result)

# ✅ 添加模板下载接口
@admin_user_bp.route('/api/admin/users/import/template', methods=['GET'])
@login_required
@admin_required
def download_user_import_template():
    """下载用户导入模板"""
    try:
        buffer = generate_import_template('user')
    except ImportError:
        # 备用模板生成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UserImportTemp"
        headers = ['country', 'email', 'name_en', 'role', 'is_partner', 'company', 'department', 'wh_type', 'wh_id', 'wh_name_en', 'employee_id', 'phone', 'birthday', 'admin_countries']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"UserImportTemp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

@admin_user_bp.route('/api/admin/users/<user_id>', methods=['PUT'])
@login_required
@admin_required
def api_admin_edit_user(user_id):
    """编辑用户信息（带完整权限控制）"""
    data = request.json
    db = get_supabase()
    
    # 获取目标用户信息
    target_user_res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
    if not target_user_res.data:
        return jsonify({"success": False, "message": "user_not_found", "params": []}), 404
    target_user = target_user_res.data
    
    # ========== 权限检查 ==========
    # 开发者可以编辑任何用户
    if not is_developer():
        # 受保护账号只能本人编辑
        if target_user.get('is_protected') and user_id != session['user_id']:
            return jsonify({"success": False, "message": "protected_account", "params": []}), 403
        
        current_role = session.get('role')
        target_role = target_user.get('role', 'user')
        
        # 超管不能编辑开发者
        if target_role == 'developer':
            return jsonify({"success": False, "message": "cannot_edit_developer", "params": []}), 403
        
        # 管理员不能编辑超管
        if current_role == 'admin' and target_role in ['super_admin', 'admin']:
            return jsonify({"success": False, "message": "cannot_edit_admin", "params": []}), 403
        
        # 管理员不能修改角色字段
        if current_role == 'admin' and 'role' in data and data['role'] != target_role:
            return jsonify({"success": False, "message": "cannot_change_role", "params": []}), 403
    
    # ========== 构建更新数据 ==========
    update_data = {}
    
    # 基础字段（所有角色可修改自己）
    allowed_fields = [
        'name_en', 'company', 'department', 
        'employee_id', 'birthday', 'country', 
        'phone', 'wh_type', 'wh_id', 
        'wh_name_en', 'user_status', 'is_partner',
        'is_resign', 'resigned_at', 'is_rehire', 'rehire_at'  # ✅ 新增
        ]

    # 处理离职状态的特殊逻辑
    if 'is_resign' in data:
        is_resign_val = data['is_resign']
        if isinstance(is_resign_val, str):
            update_data['is_resign'] = is_resign_val.upper() == 'Y' or is_resign_val.lower() == 'true'
        else:
            update_data['is_resign'] = bool(is_resign_val)
        
        # 如果设置为离职，自动设置离职时间
        if update_data['is_resign'] and not data.get('resigned_at'):
            update_data['resigned_at'] = datetime.now(timezone.utc).isoformat()
        elif not update_data['is_resign']:
            # 如果设置为在职，清除离职时间（可选）
            update_data['resigned_at'] = None

    # 角色和权限范围字段（需要更高权限）
    if 'role' in data:
        new_role = data['role']
        # 开发者可以修改任何角色
        if is_developer():
            update_data['role'] = new_role
            # 如果改为开发者，自动设置保护
            if new_role == 'developer':
                update_data['is_protected'] = True
        # 超管可以修改角色（但不能改为开发者）
        elif session.get('role') == 'super_admin':
            if new_role != 'developer':
                update_data['role'] = new_role
            else:
                return jsonify({"success": False, "message": "cannot_set_developer", "params": []}), 403
    
    # 权限范围字段
    if 'admin_countries' in data:
        if is_developer() or session.get('role') == 'super_admin':
            countries_input = data['admin_countries']
            if isinstance(countries_input, str):
                try:
                    countries_list = json.loads(countries_input)
                except:
                    countries_list = parse_countries_input(countries_input)
            else:
                countries_list = countries_input
            
            if countries_list:
                update_data['admin_countries'] = json.dumps(countries_list)
            else:
                update_data['admin_countries'] = None
    
    # 邮箱字段（需要唯一性检查）
    if 'email' in data:
        email_val = data['email'].strip().lower() or None
        if email_val:
            conflict = db.table("users").select("id").eq("email", email_val).neq("id", user_id).execute()
            if conflict.data:
                return jsonify({"success": False, "message": "email_already_used", "params": []}), 400
        update_data['email'] = email_val
    
    # 普通字段
    for field in allowed_fields:
        if field in data:
            val = data[field]
            if field == 'birthday':
                val = val if val else None
            elif field == 'is_partner':
                val = True if val in ('Y', 'true', True, '1') else False
            update_data[field] = val
    
    if not update_data:
        return jsonify({"success": False, "message": "no_fields_to_update", "params": []}), 400
    
    try:
        db.table("users").update(update_data).eq("id", user_id).execute()
        
        # 如果更新的是当前用户，同步 session
        if user_id == session.get('user_id'):
            if 'role' in update_data:
                session['role'] = update_data['role']
            if 'admin_countries' in update_data:
                session['admin_countries'] = update_data['admin_countries']
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/users/<user_id>', methods=['DELETE'])
@login_required
@admin_required
def api_admin_delete_user(user_id):
    """删除用户：已导入 → 硬删除；已注册 → 软删除"""
    if user_id == session['user_id']:
        return jsonify({"success": False, "message": "不能删除自己的账号"}), 400
    db = get_supabase()

    # 获取目标用户信息（包括 is_protected）
    target_res = db.table("users").select("role, user_status, is_protected").eq("id", user_id).maybe_single().execute()
    if not target_res.data:
        return jsonify({"success": False, "message": "用户不存在"}), 404
    target = target_res.data

    if target.get('is_protected') and user_id != session['user_id']:
        return jsonify({"success": False, "message": "该账号被保护，无法删除"}), 403

    # 权限检查（开发者保护）
    current_user = get_current_user()
    if not can_modify_user(target, current_user, 'delete'):
        return jsonify({"success": False, "message": "该账号被保护，无法删除"}), 403

    target_role = target.get('role', 'user')
    current_role = session.get('role')

    # 权限控制：普通管理员不能删除管理员或超管
    if current_role != 'super_admin' and target_role in ('admin', 'super_admin'):
        return jsonify({"success": False, "message": "权限不足，无法删除管理员"}), 403

    user_status = target.get('user_status', 'registered')

    # ========= 已导入用户：硬删除 =========
    if user_status == 'imported':
        try:
            # 清理可能存在的考试分配记录（防止外键冲突）
            db.table("exam_assignments").delete().eq("user_id", user_id).execute()
            # 物理删除用户
            db.table("users").delete().eq("id", user_id).execute()
            logger.info(f"硬删除已导入用户 {user_id}")
            return jsonify({"success": True, "hard_delete": True})
        except Exception as e:
            # 如果硬删除失败（例如仍有其他关联记录），回退为软删除
            logger.error(f"硬删除失败，尝试软删除: {e}")
            try:
                db.table("users").update({"deleted_at": datetime.now(timezone.utc).isoformat()}).eq("id", user_id).execute()
                return jsonify({"success": True, "hard_delete": False, "fallback": True, "message": "用户存在关联记录，已转为软删除"})
            except Exception as soft_err:
                logger.error(f"软删除也失败: {soft_err}")
                return jsonify({"success": False, "message": str(soft_err)}), 500

    # ========= 已注册用户：软删除 =========
    try:
        db.table("users").update({"deleted_at": datetime.now(timezone.utc).isoformat()}).eq("id", user_id).execute()
        logger.info(f"软删除用户 {user_id}")
        return jsonify({"success": True, "hard_delete": False})
    except Exception as e:
        logger.error(f"软删除失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/users/deleted')
@login_required
@admin_required
def api_admin_deleted_users():
    """获取已删除用户列表（返回所有数据，由前端进行搜索和分页）"""
    db = get_supabase()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    country = request.args.get('country', '')
    
    # 查询所有已删除用户（不分页，返回全部数据）
    query = db.table("users").select("*").not_.is_("deleted_at", "null")
    
    # 只应用国家过滤（因为国家过滤比较简单）
    if country:
        query = query.eq("country", country)
    
    # 权限过滤
    allowed = get_admin_allowed_countries()
    if allowed is not None:
        if not allowed:
            return jsonify({"data": [], "total": 0, "page": page, "per_page": per_page})
        query = query.in_("country", allowed)
    
    # 按删除时间倒序
    res = query.order("deleted_at", desc=True).execute()
    all_users = res.data or []
    
    # 获取删除人姓名
    deleted_by_ids = [u.get('deleted_by') for u in all_users if u.get('deleted_by')]
    if deleted_by_ids:
        users_res = db.table("users").select("id, name_en").in_("id", deleted_by_ids).execute()
        name_map = {u['id']: u.get('name_en', '') for u in (users_res.data or [])}
        for u in all_users:
            u['deleted_by_name'] = name_map.get(u.get('deleted_by'), u.get('deleted_by', ''))
    
    # ✅ 前端内存过滤：返回所有数据，让前端进行搜索和分页
    return jsonify({
        "data": all_users, 
        "total": len(all_users), 
        "page": 1, 
        "per_page": len(all_users)
    })

@admin_user_bp.route('/api/admin/users/<user_id>/restore', methods=['POST'])
@login_required
@admin_required
def api_admin_restore_user(user_id):
    """恢复已删除用户"""
    db = get_supabase()
    
    # 检查用户是否存在且已删除
    user_res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "用户不存在"}), 404
    
    if user_res.data.get('deleted_at') is None:
        return jsonify({"success": False, "message": "用户未被删除"}), 400
    
    try:
        db.table("users").update({
            "deleted_at": None,
            "deleted_by": None
        }).eq("id", user_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/users/batch_restore', methods=['POST'])
@login_required
@admin_required
def api_admin_batch_restore_users():
    """批量恢复已删除用户"""
    data = request.json
    user_ids = data.get('user_ids', [])
    
    if not user_ids:
        return jsonify({"success": False, "message": "请选择要恢复的用户"}), 400
    
    db = get_supabase()
    success_count = 0
    fail_count = 0
    
    for user_id in user_ids:
        try:
            db.table("users").update({
                "deleted_at": None,
                "deleted_by": None
            }).eq("id", user_id).execute()
            success_count += 1
        except Exception:
            fail_count += 1
    
    return jsonify({"success": True, "success_count": success_count, "fail_count": fail_count})

@admin_user_bp.route('/api/admin/users/<user_id>/permanent', methods=['DELETE'])
@login_required
@admin_required
def api_admin_permanent_delete_user(user_id):
    """永久删除用户（硬删除）"""
    db = get_supabase()
    
    try:
        # 先删除关联数据
        db.table("exam_assignments").delete().eq("user_id", user_id).execute()
        db.table("interview_results").delete().eq("user_id", user_id).execute()
        db.table("training_attendances").delete().eq("user_id", user_id).execute()
        # 最后删除用户
        db.table("users").delete().eq("id", user_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/users/batch_permanent', methods=['POST'])
@login_required
@admin_required
def api_admin_batch_permanent_delete_users():
    """批量永久删除用户"""
    data = request.json
    user_ids = data.get('user_ids', [])
    
    if not user_ids:
        return jsonify({"success": False, "message": "请选择要删除的用户"}), 400
    
    db = get_supabase()
    success_count = 0
    fail_count = 0
    
    for user_id in user_ids:
        try:
            db.table("exam_assignments").delete().eq("user_id", user_id).execute()
            db.table("interview_results").delete().eq("user_id", user_id).execute()
            db.table("training_attendances").delete().eq("user_id", user_id).execute()
            db.table("users").delete().eq("id", user_id).execute()
            success_count += 1
        except Exception:
            fail_count += 1
    
    return jsonify({"success": True, "success_count": success_count, "fail_count": fail_count})

@admin_user_bp.route('/api/admin/users/<user_id>/reset_password', methods=['POST'])
@login_required
@admin_required
def api_admin_reset_user_password(user_id):
    """重置用户密码（生成新密码并发送邮件）"""
    db = get_supabase()
    # 获取用户信息（包括 is_protected）
    user_res = db.table("users").select("email, name_en, is_protected").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "用户不存在"}), 404
    target_user = user_res.data
    if target_user.get('is_protected') and user_id != session['user_id']:
        return jsonify({"success": False, "message": "该账号被保护，无法重置密码"}), 403

    # 权限检查（开发者保护）
    current_user = get_current_user()
    if not can_modify_user(target_user, current_user, 'reset_password'):
        return jsonify({"success": False, "message": "该账号被保护，无法重置密码"}), 403

    email = target_user.get('email')
    if not email:
        return jsonify({"success": False, "message": "该用户没有邮箱，无法重置密码"}), 400

    new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
    password_hash = auth.hash_password(new_password)
    
    try:
        db.table("users").update({"password_hash": password_hash}).eq("id", user_id).execute()
        # 发送邮件
        send_bilingual_notification(
            email=email,
            scenario=EmailScenario.PASSWORD_RESET,
            params={
                "name": target_user.get('name_en') or email,
                "new_password": new_password,
                "host_url": request.host_url,
            },
            host_url=request.host_url,
            auth_module=auth
        )
        return jsonify({"success": True, "new_password": new_password})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/refresh_permissions')
@login_required
@admin_required
def refresh_permissions():
    db = get_supabase()
    user_id = session['user_id']
    res = db.table("users").select("admin_countries, role, country").eq("id", user_id).maybe_single().execute()
    if res.data:
        user = res.data
        session['admin_countries'] = user.get('admin_countries')
        session['user_country'] = user.get('country')
        session['role'] = user.get('role')
        
        # ✅ 解析 admin_countries 返回给前端
        admin_countries = user.get('admin_countries')
        if admin_countries and isinstance(admin_countries, str):
            try:
                admin_countries = json.loads(admin_countries)
            except:
                admin_countries = []
        
        return jsonify({
            "success": True, 
            "admin_countries": admin_countries,
            "role": user.get('role'),
            "country": user.get('country')
        })
    return jsonify({"success": False, "message": "用户不存在"}), 404

@admin_user_bp.route('/api/admin/users/export', methods=['GET'])
@login_required
@admin_required
def export_users_to_excel():
    """导出用户清单到Excel（简化版）"""
    try:
        # 权限检查：只有超管或开发者可以导出
        current_role = session.get('role')
        if current_role not in ['admin', 'super_admin', 'developer']:
            logger.warning(f"用户 {session.get('user_id')} 尝试导出用户清单但权限不足")
            return jsonify({"success": False, "message": "权限不足"}), 403
        
        db = get_supabase()
        allowed_countries = get_admin_allowed_countries()
        current_role = session.get('role')
        current_user_id = session.get('user_id')
        is_dev = is_developer()
        
        # 基础查询：获取所有未删除用户
        query = db.table("users").select("*").is_("deleted_at", "null")
        
        # 角色过滤
        if not is_dev:
            query = query.eq("is_protected", False)
        
        if current_role != 'super_admin' and not is_dev:
            query = query.neq("role", "super_admin").neq("role", "developer")
        
        all_res = query.execute()
        all_users = all_res.data or []
        
        # ✅ 应用国家权限过滤（复用 filter_users_by_permission）
        filtered_users = filter_users_by_permission(all_users, allowed_countries, current_user_id)
        
        logger.info(f"导出用户: 原始 {len(all_users)} 人，权限过滤后 {len(filtered_users)} 人")
        
        # 如果没有用户，返回空Excel
        if not filtered_users:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "用户清单"
            ws.cell(row=1, column=1, value="无权限范围内的用户数据")
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"用户清单_空_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用户清单"
        
        # ✅ 表头（添加用户ID字段）
        headers = ['序号', '用户ID', '姓名', '邮箱', '国家', '角色', '状态', '在职状态']
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row_idx, user in enumerate(filtered_users, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)                    # 序号
            ws.cell(row=row_idx, column=2, value=user.get('id', ''))             # ✅ 用户ID
            ws.cell(row=row_idx, column=3, value=user.get('name_en', ''))        # 姓名
            ws.cell(row=row_idx, column=4, value=user.get('email', ''))          # 邮箱
            ws.cell(row=row_idx, column=5, value=user.get('country', ''))        # 国家
            ws.cell(row=row_idx, column=6, value=user.get('role', ''))           # 角色
            ws.cell(row=row_idx, column=7, value=user.get('user_status', ''))    # 状态
            ws.cell(row=row_idx, column=8, value='已离职' if user.get('is_resign') else '在职')  # 在职状态
        
        logger.info(f"实际写入 {len(filtered_users)} 条数据")

        # 调整列宽
        column_widths = [8, 38, 15, 25, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 保存文件
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"用户清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{len(filtered_users)}人.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"导出失败: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/api/admin/users/stats', methods=['GET'])
@login_required
@admin_required
def api_admin_users_stats():
    """获取用户统计数据（复用列表逻辑）"""
    try:
        db = get_supabase()
        current_role = session.get('role')
        is_dev = is_developer()
        allowed_countries = get_admin_allowed_countries()
        
        # ========== 与用户列表完全相同的查询逻辑 ==========
        query = db.table("users").select("*").is_("deleted_at", "null")
        
        if not is_dev:
            query = query.eq("is_protected", False)
        
        if current_role != 'super_admin' and not is_dev:
            query = query.neq("role", "super_admin").neq("role", "developer")
        
        # 执行查询
        all_res = query.execute()
        all_users = all_res.data or []
        
        # 手动过滤（与列表保持一致）
        filtered_users = []
        
        # 获取创建人信息（包括国家）
        creator_info = {}
        creator_ids = list(set([u.get('created_by') for u in all_users if u.get('created_by')]))
        if creator_ids:
            creator_res = db.table("users").select("id, country").in_("id", creator_ids).execute()
            creator_info = {c['id']: c.get('country', '') for c in (creator_res.data or [])}
        
        for user in all_users:
            user_country = user.get('country') or ''
            user_status = user.get('user_status', '')
            user_role = user.get('role', '')
            created_by = user.get('created_by')
            
            # 排除开发者
            if user_role == 'developer':
                continue
            
            # ========== 超管逻辑 ==========
            if current_role == 'super_admin':
                # 无权限范围限制，所有用户都可见
                if allowed_countries is None:
                    filtered_users.append(user)
                    continue
                
                # 有权限范围限制
                if allowed_countries:
                    # ✅ 已导入用户：根据创建者的国家判断
                    if user_status == 'imported':
                        creator_country = creator_info.get(created_by, '')
                        if creator_country and creator_country in allowed_countries:
                            filtered_users.append(user)
                        continue
                    
                    # 已注册用户：需要国家在权限范围内
                    if user_country and user_country in allowed_countries:
                        filtered_users.append(user)
                        continue
                    
                    # 无国家已注册用户：检查创建者国家
                    if not user_country:
                        creator_country = creator_info.get(created_by, '')
                        if creator_country and creator_country in allowed_countries:
                            filtered_users.append(user)
                            continue
                continue
            
            # ========== 管理员逻辑 ==========
            if current_role == 'admin':
                # 已注册用户：需要国家权限
                if user_status == 'registered':
                    if allowed_countries is not None and allowed_countries:
                        if user_country and user_country in allowed_countries:
                            filtered_users.append(user)
                    else:
                        user_session_country = session.get('user_country')
                        if user_country == user_session_country:
                            filtered_users.append(user)
                
                # 已导入用户：自己创建 或 创建者同国家
                elif user_status == 'imported':
                    if created_by == session.get('user_id'):
                        filtered_users.append(user)
                    else:
                        creator_country = creator_info.get(created_by, '')
                        if allowed_countries is not None and allowed_countries:
                            if creator_country in allowed_countries:
                                filtered_users.append(user)
                        else:
                            user_session_country = session.get('user_country')
                            if creator_country == user_session_country:
                                filtered_users.append(user)
                
                # 其他状态
                else:
                    if allowed_countries is not None and allowed_countries:
                        if user_country and user_country in allowed_countries:
                            filtered_users.append(user)
                    else:
                        user_session_country = session.get('user_country')
                        if user_country == user_session_country:
                            filtered_users.append(user)
                continue
            
            # 其他角色
            filtered_users.append(user)
        
        # 统计
        registered_count = 0
        imported_count = 0
        
        for user in filtered_users:
            user_status = user.get('user_status')
            if user_status == 'registered':
                registered_count += 1
            elif user_status == 'imported':
                imported_count += 1
        
        logger.info(f"用户统计API: 已注册={registered_count}, 已导入={imported_count}, 总计={len(filtered_users)}")
        
        return jsonify({
            "success": True,
            "registered_count": registered_count,
            "imported_count": imported_count,
            "total_count": len(filtered_users)
        })
        
    except Exception as e:
        logger.error(f"获取用户统计失败: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500

@admin_user_bp.route('/admin/users/resigned')
@login_required
@admin_required
def resigned_users_page():
    """离职人员管理页面（仅超管/开发者可见）"""
    if not is_developer() and session.get('role') != 'super_admin':
        flash("权限不足", "danger")
        return redirect(url_for('admin_dashboard'))
    return render_template('admin/resigned_users.html')

@admin_user_bp.route('/api/admin/users/<user_id>/resign', methods=['POST'])
@login_required
@admin_required
def api_admin_resign_user(user_id):
    """标记用户为离职"""
    db = get_supabase()
    operator_id = session['user_id']
    current_role = session.get('role')
    is_dev = is_developer()

    # 获取目标用户信息
    target_user_res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
    if not target_user_res.data:
        return jsonify({"success": False, "message": "user_not_found", "params": []}), 404
    
    target_user = target_user_res.data

    # ✅ 修复：构建 current_user 对象
    current_user = {
        'id': operator_id,
        'role': current_role,
        'is_developer': is_dev
    }
    # 使用权限检查函数
    if not can_resign_user(target_user, current_user):
        return jsonify({"success": False, "message": "no_permission_to_resign", "params": []}), 403
    
    # 能标记自己离职
    if user_id == operator_id:
        return jsonify({"success": False, "message": "cannot_resign_self", "params": []}), 400
    
    # 检查用户是否存在
    user_res = db.table("users").select("id, user_status, role, is_protected").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "user_not_found", "params": []}), 404
    
    user = user_res.data

    # 超管不能标记离职（除非是开发者）
    if user.get('role') == 'super_admin' and not is_developer():
        return jsonify({"success": False, "message": "cannot_resign_super_admin", "params": []}), 403
  
    # 只有已注册用户才能标记离职
    if user.get('user_status') != 'registered':
        return jsonify({"success": False, "message": "only_registered_users_can_resign", "params": []}), 400
    
    now = datetime.now(timezone.utc).isoformat()
    
    db.table("users").update({
        "is_resign": True,
        "resigned_at": now,
        "is_rehire": False,
        "rehire_at": None
    }).eq("id", user_id).execute()
    
    logger.info(f"用户 {user_id} 已标记为离职，操作人: {operator_id}")
    return jsonify({"success": True, "message": "user_resigned_success", "params": []})


@admin_user_bp.route('/api/admin/users/<user_id>/rehire', methods=['POST'])
@login_required
@admin_required
def api_admin_rehire_user(user_id):
    """恢复用户为在职状态（复职）"""
    db = get_supabase()
    operator_id = session['user_id']

    # 不能给自己复职（虽然已离职的自己理论上无法操作，但增加检查）
    if user_id == operator_id:
        return jsonify({"success": False, "message": "cannot_rehire_self", "params": []}), 400
  
    # 检查用户是否存在
    user_res = db.table("users").select("id, user_status, is_resign, role").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "user_not_found", "params": []}), 404
    
    user = user_res.data

    # 只有已离职的用户才能复职
    if not user.get('is_resign'):
        return jsonify({"success": False, "message": "user_not_resigned", "params": []}), 400
  
    # 只有已注册用户才能复职
    if user.get('user_status') != 'registered':
        return jsonify({"success": False, "message": "only_registered_users_can_rehire", "params": []}), 400
    
    now = datetime.now(timezone.utc).isoformat()
    
    db.table("users").update({
        "is_resign": False,
        "is_rehire": True,
        "rehire_at": now
        # 保留 resigned_at 用于审计
    }).eq("id", user_id).execute()
    
    logger.info(f"用户 {user_id} 已复职，操作人: {operator_id}")
    return jsonify({"success": True, "message": "user_rehired_success", "params": []})

# routes/admin_user.py - 添加离职人员管理接口

@admin_user_bp.route('/api/admin/users/resigned')
@login_required
@admin_required
def api_admin_resigned_users():
    """获取离职人员列表（仅超管/开发者可见）"""
    # 权限检查
    if not is_developer() and session.get('role') != 'super_admin':
        return jsonify({"data": [], "total": 0, "message": "权限不足"}), 403
    
    db = get_supabase()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    country = request.args.get('country', '').strip()
    status = request.args.get('status', '')  # resigned / rehired
    
    # 查询离职人员（包括已复职的）
    query = db.table("users").select("*").is_("deleted_at", "null").eq("is_resign", True)
    
    if search:
        query = query.or_(f"name_en.ilike.%{search}%,email.ilike.%{search}%")
    if country:
        query = query.eq("country", country)
    if status == 'rehired':
        query = query.eq("is_rehire", True)
    elif status == 'resigned':
        query = query.eq("is_rehire", False)
    
    # 分页
    total_res = query.execute()
    total = len(total_res.data or [])
    
    start = (page - 1) * per_page
    end = start + per_page - 1
    res = query.range(start, end).order("resigned_at", desc=True).execute()
    
    # 获取创建人姓名
    users = res.data or []
    creator_ids = [u.get('created_by') for u in users if u.get('created_by')]
    creator_names = {}
    if creator_ids:
        creator_res = db.table("users").select("id, name_en").in_("id", creator_ids).execute()
        for c in (creator_res.data or []):
            creator_names[c['id']] = c.get('name_en', '')
    
    for u in users:
        u['created_by_name'] = creator_names.get(u.get('created_by'), '')
    
    return jsonify({
        "data": users,
        "total": total,
        "page": page,
        "per_page": per_page
    })


@admin_user_bp.route('/api/admin/users/<user_id>/resigned_detail')
@login_required
@admin_required
def api_admin_resigned_user_detail(user_id):
    """获取离职人员详细信息（仅超管/开发者可见）"""
    if not is_developer() and session.get('role') != 'super_admin':
        return jsonify({"error": "权限不足"}), 403
    
    db = get_supabase()
    res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
    if not res.data:
        return jsonify({"error": "用户不存在"}), 404
    
    user = res.data
    
    # 获取创建人姓名
    if user.get('created_by'):
        creator_res = db.table("users").select("name_en").eq("id", user['created_by']).maybe_single().execute()
        if creator_res.data:
            user['created_by_name'] = creator_res.data.get('name_en', '')
    
    return jsonify(user)

@admin_user_bp.route('/api/admin/users/<user_id>/exam_history')
@login_required
@admin_required
def api_admin_user_exam_history(user_id):
    """获取用户的考试历史"""
    if not is_developer() and session.get('role') != 'super_admin':
        return jsonify([]), 403
    
    db = get_supabase()
    
    # ✅ 明确指定外键关系
    results = db.table("exam_results")\
        .select("*, exams!fk_exam_results_exam_id(title)")\
        .eq("user_id", user_id)\
        .order("created_at", desc=True)\
        .execute()
    
    data = []
    for r in (results.data or []):
        exam_data = r.get('exams', {})
        data.append({
            "result_id": r.get('id'),  # ✅ 添加 result_id
            "exam_title": exam_data.get('title', '未知考试'),
            "total_score": r.get('total_score', 0),
            "created_at": r.get('created_at'),
            "submit_method": r.get('submit_method', 'manual')
        })
    
    return jsonify(data)

@admin_user_bp.route('/api/admin/users/<user_id>/training_history')
@login_required
@admin_required
def api_admin_user_training_history(user_id):
    """获取用户的培训签到历史"""
    if not is_developer() and session.get('role') != 'super_admin':
        return jsonify([]), 403
    
    db = get_supabase()
    
    # ✅ 修复：明确指定外键关系（使用 !fk_training_attendances_training_id）
    attendances = db.table("training_attendances")\
        .select("*, trainings!fk_training_attendances_training_id(name)")\
        .eq("user_id", user_id)\
        .order("sign_time", desc=True)\
        .execute()
    
    data = []
    for a in (attendances.data or []):
        # 注意：关联的数据在 trainings 字段中
        training_data = a.get('trainings', {})
        data.append({
            "training_name": training_data.get('name', '未知培训'),
            "sign_time": a.get('sign_time'),
            "signature_url": a.get('signature_url')
        })
    
    return jsonify(data)

# routes/admin_user.py - 修改 api_admin_user_interview_history

@admin_user_bp.route('/api/admin/users/<user_id>/interview_history')
@login_required
@admin_required
def api_admin_user_interview_history(user_id):
    """获取用户的访谈历史"""
    if not is_developer() and session.get('role') != 'super_admin':
        return jsonify([]), 403
    
    db = get_supabase()
    
    # 获取访谈记录（按访谈分组聚合）
    results = db.table("interview_results")\
        .select("*, interviews!fk_interview_results_interview_id(id, title)")\
        .eq("user_id", user_id)\
        .execute()
    
    # 按访谈聚合
    interview_map = {}
    for r in (results.data or []):
        inv_id = r.get('interview_id')
        interview_data = r.get('interviews', {})
        inv_title = interview_data.get('title', '未知访谈')
        
        if inv_id not in interview_map:
            interview_map[inv_id] = {
                "interview_id": inv_id,  # ✅ 确保返回 interview_id
                "interview_title": inv_title,
                "total_questions": 0,
                "correct_count": 0,
                "submitted_at": None
            }
        interview_map[inv_id]["total_questions"] += 1
        if r.get('is_correct'):
            interview_map[inv_id]["correct_count"] += 1
        if r.get('submitted_at') and not interview_map[inv_id]["submitted_at"]:
            interview_map[inv_id]["submitted_at"] = r.get('submitted_at')
    
    return jsonify(list(interview_map.values()))

@admin_user_bp.route('/api/admin/exam/result/<int:result_id>/detail')
@login_required
@admin_required
def api_admin_exam_result_detail(result_id):
    """获取考试结果详情（JSON格式，用于模态框）"""
    db = get_supabase()
    
    # 1. 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).maybe_single().execute()
    if not result_res.data:
        return jsonify({"success": False, "message": "成绩记录不存在"}), 404
    
    result = result_res.data
    exam_id = result['exam_id']
    user_id = result['user_id']
    
    # 2. 获取用户信息
    user_res = db.table("users").select("email, name_en").eq("id", user_id).maybe_single().execute()
    user_info = user_res.data if user_res.data else {"email": "未知", "name_en": "未知"}
    
    # 3. 获取考试信息
    exam_res = db.table("exams").select("title, reviewer").eq("id", exam_id).maybe_single().execute()
    exam_info = exam_res.data if exam_res.data else {"title": "未知考试", "reviewer": ""}
    
    # 4. 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    
    # 5. 解析 JSON 字段
    answers = result.get('answers', {})
    if isinstance(answers, str):
        try:
            answers = json.loads(answers)
        except:
            answers = {}
    
    details = result.get('details', {})
    if isinstance(details, str):
        try:
            details = json.loads(details)
        except:
            details = {}
    
    # 调试日志
    logger.info(f"考试结果详情 - result_id: {result_id}")
    logger.info(f"总分: {result.get('total_score', 0)}")
    
    # 6. 构建题目详情数组
    question_details = []
    total_obtained_score = 0
    
    for q in questions:
        question_id = str(q.get('id'))  # 题目ID，如 "2085"
        q_num = q.get('num', 0)        # 题目序号，如 1
        
        # 解析选项
        options = q.get('options', {})
        if isinstance(options, str):
            try:
                options = json.loads(options)
            except:
                options = {}
        
        # ✅ 获取考生答案 - 使用 q_ 前缀的 key
        user_answer = '未作答'
        answer_key = f"q_{question_id}"  # 格式如 "q_2085"
        
        if answer_key in answers:
            user_answer = answers[answer_key]
        # 兼容其他可能的格式
        elif question_id in answers:
            user_answer = answers[question_id]
        elif str(q_num) in answers:
            user_answer = answers[str(q_num)]
        
        # 处理答案格式
        if user_answer is None or user_answer == '':
            user_answer = '未作答'
        
        # 正确答案
        correct_answer = q.get('answer', '')
        
        # 从 details 中获取得分（使用题目ID）
        obtained_score = 0
        total_score = q.get('score', 0)
        is_correct_from_detail = False
        
        if details and question_id in details:
            q_detail = details[question_id]
            if isinstance(q_detail, dict):
                obtained_score = q_detail.get('score', 0)
                is_correct_from_detail = q_detail.get('correct', False)
            elif isinstance(q_detail, (int, float)):
                obtained_score = q_detail
        
        # 判断是否正确（用于显示）
        is_correct = is_correct_from_detail
        if not is_correct:
            if q.get('type') == 'multi':
                # 多选题：比较集合
                user_set = set(str(user_answer).replace(' ', '').split(',')) if user_answer and user_answer != '未作答' else set()
                correct_set = set(str(correct_answer).replace(' ', '').split(','))
                is_correct = user_set == correct_set
            else:
                is_correct = str(user_answer).strip().upper() == str(correct_answer).strip().upper()
        
        total_obtained_score += obtained_score
        
        question_details.append({
            "num": q_num,
            "type": q.get('type', 'single'),
            "content": q.get('content') or q.get('content_en') or q.get('content_cn') or '',
            "options": options,
            "correct_answer": correct_answer,
            "user_answer": user_answer,
            "is_correct": is_correct,
            "obtained_score": obtained_score,
            "total_score": total_score
        })
    
    # 使用数据库中的总分
    final_total_score = result.get('total_score', total_obtained_score)
    
    logger.info(f"计算总分: {total_obtained_score}, 数据库总分: {result.get('total_score', 0)}")
    if questions:
        first_q = questions[0]
        first_answer = question_details[0]['user_answer'] if question_details else 'N/A'
        logger.info(f"第一题示例: 题目ID={first_q.get('id')}, 序号={first_q.get('num')}, 答案key=q_{first_q.get('id')}, 答案值={first_answer}")
    
    # 7. 返回 JSON 数据
    return jsonify({
        "success": True,
        "result_id": result_id,
        "user": {
            "name": user_info.get('name_en', '未知'),
            "email": user_info.get('email', '未知')
        },
        "exam": {
            "title": exam_info.get('title', '未知考试'),
            "reviewer": exam_info.get('reviewer', '')
        },
        "submitted_at": result.get('submitted_at') or result.get('created_at'),
        "submit_method": result.get('submit_method', 'manual'),
        "total_score": final_total_score,
        "questions": question_details
    })

