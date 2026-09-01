# routes/admin_training.py
import logging
import sys
import json
import pdfkit
import openpyxl
import traceback
from io import BytesIO
from flask import render_template, request, redirect, send_file, url_for, session, flash, jsonify, make_response
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from functools import lru_cache
from . import admin_training_bp
from config import Config
from datetime import datetime, timezone, timedelta
from services.db import get_supabase, get_supabase_admin
from services import auth
from services.export import find_wkhtmltopdf
from utils.timezone_utils import utc_string_to_local
from utils.status import get_exam_status, get_training_status
from utils.common import match_country_code, quarter_to_date_range
from utils.email_notifier import  _send_training_notifications
from utils.permissions import filter_users_by_permission, get_admin_allowed_countries, get_allowed_countries, is_developer
from utils.cache_manager import clear_training_related_cache, cache_get, invalidate_cache_on_change, training_cache, clear_all_assignment_caches
from utils.training_helpers import (
    get_training_country_templates_status, 
    _save_country_template,
    parse_training_countries,
    get_training_primary_country,
    training_has_country,
    training_matches_any_country,
    filter_trainings_by_country,
    get_training_countries_display,
    normalize_training_countries,
    parse_country_list,
    sync_binding_pass_score_to_exam,
    calculate_dynamic_status, 
    is_draft_time, 
    DRAFT_PLACEHOLDER
)

from routes.helpers import (
    login_required, 
    admin_required, 
    get_attendance_data, 
    convert_time_for_export,
    can_access_exam, 
    upload_signature, 
    parse_exam_countries
)
from utils.manage_messages import (
    log_signature_reset,
    log_training_unassign,
    log_admin_push_training,
    log_training_auto_assign
)

logger = logging.getLogger(__name__)


# ==================== 缓存装饰器（通用） ====================
def cache_result(seconds=300):
    """缓存装饰器（兼容旧代码）"""
    def decorator(func):
        cache = {}
        
        def wrapper(*args, **kwargs):
            logger.info(f"📌 请求方法: {request.method}")  # 应该是 GET
            # 只缓存 GET 请求
            if request.method != 'GET':
                return func(*args, **kwargs)
            
            # 构建缓存键
            user_id = session.get('user_id', '')
            params = request.args.to_dict() if request.method == 'GET' else {}
            param_str = json.dumps(params, sort_keys=True)
            key = f"{user_id}_{request.path}_{param_str}"
            if len(key) > 200:
                import hashlib
                key = hashlib.md5(key.encode()).hexdigest()
            
            now = datetime.now()
            if key in cache:
                result, timestamp = cache[key]
                if now - timestamp < timedelta(seconds=seconds):
                    return result
            
            result = func(*args, **kwargs)
            cache[key] = (result, now)
            return result
        
        wrapper.clear_cache = lambda: cache.clear()
        return wrapper
    return decorator

@admin_training_bp.route('/admin/trainings')
@login_required
@admin_required
def admin_trainings():
    return render_template('admin/list_trainings.html')

# ==================== 培训列表 API（使用缓存）开始 ====================
@admin_training_bp.route('/api/admin/trainings', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
@admin_required
def api_admin_trainings():
    db = get_supabase_admin()
    
    if request.method == 'GET':
        # 这里调用的是被缓存装饰的函数
        result = _get_trainings_list(db)
        logger.info(f"📤 API 返回结果: {type(result)}, keys: {result.keys() if isinstance(result, dict) else 'not dict'}")
        return jsonify(result)
    
    elif request.method == 'POST':
        result = _create_training(db)
        # 创建成功后清除缓存
        training_cache.clear('training_list')
        return result
    
    elif request.method == 'PUT':
        result = _update_training(db)
        # 更新成功后清除缓存
        training_cache.clear('training_list')
        return result
    
    elif request.method == 'DELETE':
        result = _delete_training(db)
        # 删除成功后清除缓存
        training_cache.clear('training_list')
        return result

# 将 GET 逻辑抽取为独立函数，使用缓存装饰器
@cache_get(ttl=300, prefix='training_list', include_user=True)
def _get_trainings_list(db):
    """获取培训列表（带缓存）"""
    logger.info("=" * 60)
    sys.stdout.flush()
    logger.info("🔥🔥🔥 _get_trainings_list 被调用（应该是第一次或缓存过期）")
    logger.info("📝 [GET_LIST] 开始获取培训列表")
    logger.info("=" * 60)
    sys.stdout.flush()

    # 获取过滤参数
    country_filter = request.args.get('country', '')
    name = request.args.get('name', '')
    quarter = request.args.get('quarter', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    logger.info("=" * 50)
    logger.info("调用 _get_trainings_list (缓存)")
    logger.info(f"当前用户: role={session.get('role')}, user_country={session.get('user_country')}")
    
    # 1. 获取所有培训
    query = db.table("trainings").select("*")
    if name:
        query = query.ilike("name", f"%{name}%")
    res = query.execute()
    all_trainings = res.data or []
    
    # 2. 权限过滤（使用多国家支持）
    allowed_countries = get_admin_allowed_countries()
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    is_dev = is_developer()

    if allowed_countries is not None:
        if not allowed_countries:
            return {"data": [], "total": 0, "page": page, "per_page": per_page}
        
        filtered_trainings = []
        for training in all_trainings:
            # 解析培训的国家列表
            training_countries = parse_training_countries(training)
            
            # 检查是否有国家在权限范围内
            has_permission = any(c in allowed_countries for c in training_countries)
            
            if has_permission:
                filtered_trainings.append(training)
            else:
                logger.debug(f"培训 {training.get('id')} 不在权限范围内，已过滤")
    else:
        # 超管/开发者无权限限制
        filtered_trainings = all_trainings
    
    # 3. 如果超管有权限范围，也需要过滤
    if current_role == 'super_admin' and not is_dev and allowed_countries:
        # 超管也应用权限过滤（如果有设置）
        filtered_trainings = [t for t in filtered_trainings if any(c in allowed_countries for c in parse_training_countries(t))]
    
    # 3. 国家过滤（使用多国家支持）
    if country_filter:
        filter_country_code = match_country_code(country_filter) if country_filter else None
        if filter_country_code:
            users_in_filter = db.table("users").select("id").eq("country", filter_country_code).execute()
            filter_user_ids = [u['id'] for u in (users_in_filter.data or [])] if users_in_filter.data else []
            filter_training_ids = set()
            if filter_user_ids:
                attend_filter = db.table("training_attendances").select("training_id").in_("user_id", filter_user_ids).execute()
                filter_training_ids = {a['training_id'] for a in (attend_filter.data or [])}
            
            temp_filtered = []
            for training in filtered_trainings:
                # 检查培训是否包含该国家
                if training_has_country(training, filter_country_code) or training['id'] in filter_training_ids:
                    temp_filtered.append(training)
            filtered_trainings = temp_filtered

    # 4. 季度过滤
    if quarter:
        q_start, q_end = quarter_to_date_range(quarter)
        if q_start and q_end:
            q_start_dt = datetime.fromisoformat(q_start)
            q_end_dt = datetime.fromisoformat(q_end)
            temp_filtered = []
            for training in filtered_trainings:
                start = training.get('start_time')
                end = training.get('end_time')
                if start and end:
                    try:
                        start_dt = datetime.fromisoformat(start)
                        end_dt = datetime.fromisoformat(end)
                        if start_dt <= q_end_dt and end_dt >= q_start_dt:
                            temp_filtered.append(training)
                    except:
                        pass
            filtered_trainings = temp_filtered

    # 5. 排序
    filtered_trainings.sort(key=lambda x: x.get('created_at', ''), reverse=True)

    # 6. 分页
    total = len(filtered_trainings)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated = filtered_trainings[start_idx:end_idx]

    # 7. 批量补充数据（优化：减少 N+1 查询）
    now = datetime.now(timezone.utc)
    training_ids = [t['id'] for t in paginated]
    
    # 7.1 批量获取签到数据
    attendance_map = {}
    if training_ids:
        att_res = db.table("training_attendances") \
            .select("training_id, user_id") \
            .in_("training_id", training_ids) \
            .execute()
        for att in (att_res.data or []):
            tid = att['training_id']
            if tid not in attendance_map:
                attendance_map[tid] = []
            attendance_map[tid].append(att['user_id'])
        
        # 获取所有签到用户的详细信息（包括国家）
        all_user_ids = set()
        for users in attendance_map.values():
            all_user_ids.update(users)

        # 获取用户的国家信息，用于权限过滤
        user_country_map = {}
        if all_user_ids:
            users_res = db.table("users").select("id, country").in_("id", list(all_user_ids)).execute()
            for u in (users_res.data or []):
                user_country_map[u['id']] = u.get('country')
        
        # 获取在职用户（同时过滤权限）
        active_user_ids = set()
        if all_user_ids:
            active_users = db.table("users").select("id").in_("id", list(all_user_ids)).eq("is_resign", False).execute()
            all_active_ids = {u['id'] for u in (active_users.data or [])}

            # 再根据管理员权限过滤
            if allowed_countries is not None:
                # 只保留国家在权限范围内的用户
                for uid in all_active_ids:
                    user_country = user_country_map.get(uid)
                    if user_country and user_country in allowed_countries:
                        active_user_ids.add(uid)
            else:
                active_user_ids = all_active_ids
        else:
            # 当没有用户时，设置为空集合
            active_user_ids = set()
    else:
        # 当 training_ids 为空时，设置默认值
        attendance_map = {}
        active_user_ids = set()

    # 7.2 批量获取绑定数量
    binding_counts = {}
    if training_ids:
        for tid in training_ids:
            count_res = db.table("training_exam_bindings").select("id", count="exact").eq("training_id", tid).is_("deleted_at", "null").execute()
            binding_counts[tid] = count_res.count or 0
    
    # 7.3 批量获取模板信息
    templates_by_training = {}
    if training_ids:
        tpl_res = db.table("training_country_templates").select("training_id, country, header_template").in_("training_id", training_ids).execute()
        for tpl in (tpl_res.data or []):
            tid = tpl['training_id']
            if tid not in templates_by_training:
                templates_by_training[tid] = []
            templates_by_training[tid].append(tpl)

    # 7.4 批量获取照片数量
    photo_counts = {}
    if training_ids:
        photo_res = db.table("training_photos") \
            .select("training_id", count="exact") \
            .in_("training_id", training_ids) \
            .eq("is_deleted", False) \
            .execute()
        
        # 按 training_id 分组计数
        for p in (photo_res.data or []):
            tid = p['training_id']
            photo_counts[tid] = photo_counts.get(tid, 0) + 1

    # 7.5 批量获取绑定关系（新增）
    training_bindings_map = {}
    exam_info_map = {}
    exam_stats_map = {}
    if training_ids:
        # 获取所有培训的绑定关系
        bindings_res = db.table("training_exam_bindings") \
            .select("id, training_id, exam_id, pass_score, is_auto_assign") \
            .in_("training_id", training_ids) \
            .is_("deleted_at", "null") \
            .execute()
        
        # 按 training_id 分组
        all_exam_ids = set()
        for b in (bindings_res.data or []):
            tid = b['training_id']
            if tid not in training_bindings_map:
                training_bindings_map[tid] = []
            training_bindings_map[tid].append({
                'binding_id': b['id'],  #
                'exam_id': b['exam_id'],
                'pass_score': b.get('pass_score', 85),
                'is_auto_assign': b.get('is_auto_assign', True),
                'is_required': b.get('is_required', True)
            })
            all_exam_ids.add(b['exam_id'])
        
        # 获取绑定的考试信息（标题、状态等）
        if all_exam_ids:
            exams_res = db.table("exams") \
                .select("id, title, status, countries, pass_score, start_time, end_time, duration") \
                .in_("id", list(all_exam_ids)) \
                .is_("deleted_at", "null") \
                .execute()
            for e in (exams_res.data or []):
                e['countries_list'] = parse_exam_countries(e)
                exam_info_map[e['id']] = e

            # 批量获取考试统计信息（completed_count, assigned_count）
            # 已完成人数
            completed_res = db.table("exam_results") \
                .select("exam_id, user_id") \
                .in_("exam_id", list(all_exam_ids)) \
                .is_("deleted_at", "null") \
                .execute()
            completed_counts = {}
            for r in (completed_res.data or []):
                eid = r['exam_id']
                if eid not in completed_counts:
                    completed_counts[eid] = set()
                completed_counts[eid].add(r['user_id'])
            
            # 已分配人数
            assigned_res = db.table("exam_assignments") \
                .select("exam_id, user_id") \
                .in_("exam_id", list(all_exam_ids)) \
                .is_("deleted_at", "null") \
                .execute()
            assigned_counts = {}
            for r in (assigned_res.data or []):
                eid = r['exam_id']
                if eid not in assigned_counts:
                    assigned_counts[eid] = set()
                assigned_counts[eid].add(r['user_id'])
            
            # 存储统计信息
            for eid in all_exam_ids:
                exam_stats_map[eid] = {
                    'completed_count': len(completed_counts.get(eid, set())),
                    'assigned_count': len(assigned_counts.get(eid, set()))
                }

    # 8. 组装返回数据
    for t in paginated:
        tid = t['id']
        
        # 签到人数
        user_ids = attendance_map.get(tid, [])
        signed_count = len([uid for uid in user_ids if uid in active_user_ids])
        t['signed_count'] = signed_count
        
        # 添加绑定考试信息（预加载）
        binding_exams_data = training_bindings_map.get(tid, [])
        binding_exams = []
        stats = exam_stats_map.get(eid, {})
        for b in binding_exams_data:
            exam_id = b['exam_id']
            exam_info = exam_info_map.get(exam_id)
            if exam_info:
                binding_exams.append({
                    "binding_id": b['binding_id'],
                    "exam_id": exam_id,
                    "exam_title": exam_info.get('title', f'考试 #{exam_id}'),
                    "exam_status": exam_info.get('status', 'draft'),
                    "exam_countries": exam_info.get('countries_list', []),
                    "pass_score": b.get('pass_score', 85),
                    "is_auto_assign": b.get('is_auto_assign', True),
                    "is_required": b.get('is_required', True),
                    "start_time": exam_info.get('start_time'),
                    "end_time": exam_info.get('end_time'),
                    "duration": exam_info.get('duration', 60),
                    "completed_count": stats.get('completed_count', 0),
                    "assigned_count": stats.get('assigned_count', 0)
                })
        
        t['binding_exams'] = binding_exams
        t['binding_count'] = len(binding_exams)  # 覆盖之前从 binding_counts 获取的值

        # 照片数量
        t['photo_count'] = photo_counts.get(tid, 0)

        # 添加国家显示（使用工具函数）
        t['countries_display'] = get_training_countries_display(t, allowed_countries)
        t['countries_list'] = parse_training_countries(t)  # 原始列表

        # ✅ 添加日志：打印每个培训的时间字段
        logger.info(f"📝 [GET_LIST] 培训 {t.get('id')}: start_time={t.get('start_time')}, end_time={t.get('end_time')}, dynamic_status={t.get('dynamic_status')}")
        
        # 动态状态
        start_time = t.get('start_time')
        end_time = t.get('end_time')
        if not start_time or not end_time:
            # 使用统一的状态计算函数
            t['dynamic_status'] = calculate_dynamic_status(
                t.get('start_time'), 
                t.get('end_time')
            )
        
        # 模板不一致检查
        templates = templates_by_training.get(tid, [])
        if not t.get('country') and templates:
            unique_templates = set()
            for ct in templates:
                tpl = ct.get('header_template', {})
                unique_templates.add(json.dumps(tpl, sort_keys=True))
            t['has_inconsistent_templates'] = len(unique_templates) > 1
        else:
            t['has_inconsistent_templates'] = False

    paginated.sort(key=lambda x: x.get('created_at') or '', reverse=True)
    return {
        "data": paginated,
        "total": total,
        "page": page,
        "per_page": per_page
    }

def _create_training(db):
    """创建培训"""
    data = request.json
    name = data.get('name')
    if not name:
        return jsonify({"success": False, "message": "jsonify_training_name_cannot_empty", "params": []}), 400

    # 创建培训（也需要权限校验）
    data = request.json
    name = data.get('name')
    if not name:
        return jsonify({"success": False, "message": "jsonify_training_name_cannot_empty", "params": []}), 400
    
    # 从数据库实时获取管理员的权限范围，而不是依赖 session
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    
    # 获取用户的权限范围（直接从数据库）
    user_res = db.table("users").select("admin_countries, country").eq("id", current_user_id).maybe_single().execute()
    user_data = user_res.data if user_res and user_res.data else {}
    
    # 计算允许的国家列表
    allowed = None
    if current_role == 'developer':
        # developer 无限制
        allowed = None
    elif current_role == 'super_admin':
        admin_countries = user_data.get('admin_countries')
        if admin_countries:
            try:
                allowed = json.loads(admin_countries) if isinstance(admin_countries, str) else admin_countries
            except:
                allowed = None
        # 超管如果没有设置权限范围，则无限制
    elif current_role == 'admin':
        admin_countries = user_data.get('admin_countries')
        if admin_countries:
            try:
                allowed = json.loads(admin_countries) if isinstance(admin_countries, str) else admin_countries
            except:
                allowed = None
        
        # 如果没有设置权限范围，使用用户所在国家
        if not allowed:
            user_country = user_data.get('country')
            if user_country:
                allowed = [user_country]
            else:
                allowed = []
    
    # 获取国家列表（支持多国家）
    countries_input = data.get('countries', [])
    
    # 兼容旧版本：如果传递了 country 字段
    if not countries_input and data.get('country'):
        countries_input = [data.get('country')]
    
    # 规范化国家数据
    countries_json = normalize_training_countries(countries_input)
    
    # 详细日志
    logger.info("=" * 50)
    logger.info("创建培训 POST 请求")
    logger.info(f"用户ID: {current_user_id}, 角色: {current_role}")
    logger.info(f"数据库 admin_countries: {user_data.get('admin_countries')}")
    logger.info(f"数据库 user_country: {user_data.get('country')}")
    logger.info(f"计算后 allowed: {allowed}")
    logger.info(f"请求中的 countries_json: {countries_json}")
    
    # # 权限检查
    if allowed is not None:
        if not allowed:
            logger.warning(f"管理员没有任何国家权限，禁止创建培训")
            return jsonify({"success": False, "message": "jsonify_no_country_permission", "params": []}), 403
        
        # 检查所有国家是否都在允许范围内
        countries_list = parse_country_list(countries_input)
        for c in countries_list:
            if c not in allowed:
                return jsonify({
                    "success": False, 
                    "message": "jsonify_no_authority_creat_training_this_county", "params": []
                }), 403
    else:
        logger.info("无权限限制（超管或开发者）")

    start_time = data.get('start_time')
    end_time = data.get('end_time')

    # 核心修复：区分 None 和空字符串
    if start_time == '':
        start_time = None
    if end_time == '':
        end_time = None
    
    # 只有完全没有传递时间时，才自动生成（预填模式）
    if start_time is None and end_time is None:
        if 'start_time' not in data or 'end_time' not in data:
            start_time = datetime.now(timezone.utc).isoformat()
            end_time = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()

    # 草稿模式：使用占位时间（满足 NOT NULL 约束）
    if start_time is None:
        start_time = DRAFT_PLACEHOLDER
    if end_time is None:
        end_time = DRAFT_PLACEHOLDER

    # 判断是否为草稿（使用占位时间）
    is_draft = is_draft_time(start_time) or is_draft_time(end_time)
    
    if is_draft:
        dynamic_status = 'draft'
        is_active = False
    else:
        # ✅ 添加日志：确认调用的是哪个函数
        logger.info("=" * 60)
        logger.info("🔍 [CREATE] 进入 else 分支")
        logger.info(f"🔍 [CREATE] start_time={start_time}, end_time={end_time}")
        logger.info(f"🔍 [CREATE] calculate_dynamic_status 函数对象: {calculate_dynamic_status}")
        logger.info("=" * 60)
        
        dynamic_status = calculate_dynamic_status(start_time, end_time)
        is_active = True

    res = db.table("trainings").insert({
        "name": name,
        "start_time": start_time,
        "end_time": end_time,
        "header_template": data.get('header_template', {}),
        "countries": countries_json,
        "quarter": data.get('quarter', ''),
        "is_active": is_active,
        "created_by": current_user_id,
        "dynamic_status": dynamic_status
    }).execute()
    
    logger.info(f"创建培训成功: id={res.data[0]['id']}, name={name}")

    # ✅ 添加日志：打印实际插入数据库的值
    logger.info(f"📝 [CREATE] 插入数据库: id={res.data[0]['id']}, start_time={start_time}, end_time={end_time}, dynamic_status={dynamic_status}")

    # 创建成功后清除缓存
    training_cache.clear('training_list')
    
    return jsonify({"success": True, "id": res.data[0]['id']})

def _update_training(db):
    """更新培训"""
    # 更新培训（需要权限校验）
    logger.info(f"========== PUT 请求收到 ==========")
    logger.info(f"请求数据: {request.json}")
    data = request.json
    tid = data.get('id')
    if tid is None or tid == 'None' or str(tid).lower() == 'null':
        return jsonify({"success": False, "message": "无效的培训ID"}), 400
    
    try:
        tid = int(tid)
    except (ValueError, TypeError):
        return jsonify({"success": False, "message": "培训ID必须是整数"}), 400
    
    # 获取原培训信息，检查权限
    original = db.table("trainings").select("country, countries, created_by").eq("id", tid).maybe_single().execute()
    if not original.data:
        return jsonify({"success": False, "message": "培训不存在"}), 404
    
    training_data = original.data
    created_by = training_data.get('created_by')
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    is_dev = is_developer()
    
    # 权限检查：国家权限
    allowed = get_admin_allowed_countries()
    training_countries = parse_training_countries(training_data)
    
    if allowed is not None:
        if not any(c in allowed for c in training_countries):
            return jsonify({"success": False, "message": "jsonify_no_authority_modify_training", "params": []}), 403
    
    # 权限检查：创建者（非超管/开发者只能编辑自己创建的）
    if not is_dev and current_role != 'super_admin':
        if created_by != current_user_id:
            return jsonify({
                "success": False, 
                "message": "jsonify_no_permmission_edit_item_created_by_others", 
                "params": []
            }), 403
    
    # 处理 header_template 保存
    country_code = data.get('country_code')
    header_template = data.get('header_template')
    if header_template is not None:
        if country_code:
            # 检查国家权限
            allowed = get_admin_allowed_countries()
            if allowed is not None and country_code not in allowed:
                return jsonify({"success": False, "message": "jsonify_no_authorith_set_up_header_template", "params": []}), 403
            # 调用辅助函数保存模板
            _save_country_template(db, tid, country_code, header_template)
        else:
            db.table("trainings").update({"header_template": header_template}).eq("id", tid).execute()

    # 记录是否推送（用于发送邮件）
    is_push = data.get('is_active', False)
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    user_ids = data.get('user_ids')  # 选中的用户ID列表
    
    update_data = {}
    if 'name' in data:
        update_data['name'] = data['name']
    if 'header_template' in data:
        update_data['header_template'] = data['header_template']
    if 'start_time' in data and data['start_time'] is not None:
        update_data['start_time'] = data['start_time']
    if 'end_time' in data and data['end_time'] is not None:
        update_data['end_time'] = data['end_time']
    if 'is_active' in data:
        update_data['is_active'] = data['is_active']

    # 如果有效期变化，重新计算状态
    if 'start_time' in update_data or 'end_time' in update_data:
        start = update_data.get('start_time') or original.get('start_time')
        end = update_data.get('end_time') or original.get('end_time')

        # 检查是否为占位符
        if is_draft_time(start) or is_draft_time(end):
            update_data['dynamic_status'] = 'draft'
            update_data['is_active'] = False
        elif start and end:
            update_data['dynamic_status'] = calculate_dynamic_status(start, end)
            update_data['is_active'] = True
        else:
            update_data['dynamic_status'] = 'draft'
            update_data['is_active'] = False
    
    # 处理多国家更新
    if 'countries' in data:
        countries_input = data['countries']
        if not countries_input and data.get('country'):
            countries_input = [data.get('country')]
        
        # 权限检查
        allowed = get_admin_allowed_countries()
        if allowed is not None:
            countries_list = parse_country_list(countries_input)
            for c in countries_list:
                if c not in allowed:
                    return jsonify({"success": False, "message": f"无权更新国家 {c} 的培训"}), 403
        
        countries_json = normalize_training_countries(countries_input)
        update_data['countries'] = countries_json
    
    if 'quarter' in data:
        update_data['quarter'] = data['quarter']
    
    if update_data:
        db.table("trainings").update(update_data).eq("id", tid).execute()
        logger.info(f"更新培训成功: id={tid}, 更新字段={list(update_data.keys())}")

    # 处理培训-学员分配关系（定点推送）
    user_ids = data.get('user_ids')
    if user_ids is not None:  # 注意：空数组表示清空所有分配
        # 先删除该培训的所有现有分配
        db.table("training_assignments").delete().eq("training_id", tid).execute()
        
        # 如果有指定用户，插入新的分配关系
        if len(user_ids) > 0:
            assignments = [{"training_id": tid, "user_id": uid, "created_by": session.get('user_id')} for uid in user_ids]
            db.table("training_assignments").insert(assignments).execute()
            logger.info(f"培训 {tid} 定点推送给 {len(user_ids)} 名学员")
        else:
            logger.info(f"培训 {tid} 清空了所有分配（推送给全国）")
                    
    # 在 PUT 方法中，处理 user_ids 的地方
    logger.info(f"========== 培训推送 ==========")
    logger.info(f"培训ID: {tid}")
    logger.info(f"user_ids: {user_ids}")
    logger.info(f"user_ids 类型: {type(user_ids)}")
    logger.info(f"user_ids 长度: {len(user_ids) if user_ids else 0}")

    # 如果只是保存表头（没有推送字段），可以提前返回
    if not is_push and not start_time and not end_time and user_ids is None:
        return jsonify({"success": True})
    
    # 发送培训推送邮件通知
    if data.get('is_active') and start_time and end_time:
        # 异步发送邮件，避免阻塞请求
        import threading
        thread = threading.Thread(
            target=_send_training_notifications,
            args=(tid, start_time, end_time, user_ids, request.host_url)
        )
        thread.daemon = True
        thread.start()
        logger.info(f"培训 {tid} 邮件通知已加入发送队列")
    
    return jsonify({"success": True})
    
    # 更新成功后清除缓存
    training_cache.clear('training_list')
    
    return jsonify({"success": True})

def _delete_training(db):
    """删除培训 - 支持多国家 + 创建者权限校验"""
    tid = request.args.get('id')
    if not tid:
        return jsonify({"success": False, "message": "jsonify_lack_training_id", "params": []}), 400
    
    # 获取培训信息（同时获取 countries 和 country 字段）
    original = db.table("trainings").select("country, countries, created_by").eq("id", tid).maybe_single().execute()
    if not original.data:
        return jsonify({"success": False, "message": "培训不存在"}), 404
    
    training_data = original.data
    created_by = training_data.get('created_by')
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    is_dev = is_developer()
    
    # 解析培训的国家列表（支持多国家）
    from utils.training_helpers import parse_training_countries
    training_countries = parse_training_countries(training_data)
    
    # 权限检查1：国家权限
    allowed = get_admin_allowed_countries()
    if allowed is not None:
        # 如果管理员有权限范围限制，检查培训是否在权限范围内
        if not any(c in allowed for c in training_countries):
            logger.warning(f"用户 {current_user_id} 无权删除培训 {tid}，国家 {training_countries} 不在权限范围 {allowed} 内")
            return jsonify({"success": False, "message": "jsonify_no_authorith_delete_training", "params": []}), 403

    # 权限检查2：创建者检查（非超管/开发者只能删除自己创建的）
    if not is_dev and current_role != 'super_admin':
        if created_by != current_user_id:
            logger.warning(f"用户 {current_user_id} 无权删除培训 {tid}，创建者为 {created_by}")
            return jsonify({"success": False, "message": "jsonify_no_permmission_delete_item_created_by_others", "params": []}), 403
    
    # 执行删除
    try:
        db.table("trainings").delete().eq("id", tid).execute()
        logger.info(f"✅ 删除培训成功: id={tid}, 操作人={current_user_id}, 培训国家={training_countries}")
    except Exception as e:
        logger.error(f"删除培训失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    
    # 删除成功后清除缓存
    training_cache.clear('training_list')
    
    return jsonify({"success": True})

@admin_training_bp.route('/api/admin/training/clear_cache', methods=['POST'])
@login_required
@admin_required
def clear_training_cache():
    """手动清除培训列表缓存"""
    training_cache.clear('training_list')
    return jsonify({"success": True, "message": "培训列表缓存已清除"})

@admin_training_bp.route('/api/admin/training/cache_stats', methods=['GET'])
@login_required
@admin_required
def get_training_cache_stats():
    """获取缓存统计信息"""
    stats = training_cache.get_stats()
    return jsonify(stats)
# ==================== 培训列表 API（使用缓存）结束 ====================

@admin_training_bp.route('/api/training/attendance/<int:training_id>')
@login_required
@admin_required
def api_training_attendance(training_id):
    db = get_supabase()
    admin_db = get_supabase_admin() 
    country = request.args.get('country', '')

    # 获取当前管理员的权限范围
    allowed_countries = get_admin_allowed_countries()
    
    # 获取培训基本信息
    training_res = db.table("trainings").select("*").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return jsonify({"error": "培训不存在"}), 404
    training = training_res.data
    
    # 解析培训的国家列表（支持多国家）
    training_countries = parse_training_countries(training)

    # 权限检查：培训国家与用户权限范围有交集
    if allowed_countries is not None:
        if not allowed_countries:
            return jsonify({"attendances": [], "training": training, "header_template": {}})
        # 检查培训是否在权限范围内
        has_permission = any(c in allowed_countries for c in training_countries)
        if not has_permission:
            return jsonify({"attendances": [], "training": training, "header_template": {}})
    
    # 签到记录查询
    att_res = db.table("training_attendances") \
        .select("id, user_id, signature_url, signed_name, sign_time, users(email, name_cn, name_en, department, employee_id, country, company, is_resign)") \
        .eq("training_id", training_id) \
        .execute()

    att_list = att_res.data or []

    # 过滤掉离职人员的签到记录
    att_list = [rec for rec in att_list if not rec.get('users', {}).get('is_resign', False)]
    
    # 按国家权限过滤签到记录（基于用户的国家）
    if allowed_countries is not None:
        if not allowed_countries:
            att_list = []
        else:
            filtered_list = []
            for rec in att_list:
                user = rec.get('users', {})
                user_country = user.get('country')
                if user_country and user_country in allowed_countries:
                    filtered_list.append(rec)
            att_list = filtered_list
    
    # 手动按国家过滤
    if country:
        att_list = [rec for rec in att_list if rec.get('users', {}).get('country') == country]

    attendance_list = []
    for rec in att_list or []:
        user = rec.get('users', {})
        # 防止 user 为 None
        if user is None:
            user = {}

        attendance_list.append({
            "id": rec['id'],
            "user_id": rec['user_id'],
            "department": user.get('department', ''),
            "name_cn": user.get('name_cn', ''),
            "name_en": user.get('name_en', ''),
            "employee_id": user.get('employee_id', ''),
            "signed_name": rec.get('signed_name', ''),
            "signature_url": rec.get('signature_url', ''),
            "sign_time": rec.get('sign_time'),
            "company": user.get('company', ''),
            "country": user.get('country', '')
        })

    # 获取表头模板：优先取国家模板，否则取培训主模板
    header_template = None
    if country:
        # 查询国家模板
        ct_res = admin_db.table("training_country_templates")\
            .select("header_template")\
            .eq("training_id", training_id)\
            .eq("country", country)\
            .execute()
        if ct_res.data and len(ct_res.data) > 0:
            header_template = ct_res.data[0].get('header_template', {})
            logger.info(f"✅ 加载国家模板: training_id={training_id}, country={country}")
    if not header_template:
        header_template = training.get('header_template', {})
        logger.info(f"✅ 加载主表头模板: training_id={training_id}")

    return jsonify({
        "training": training,
        "attendances": attendance_list,
        "header_template": header_template
    })

@admin_training_bp.route('/api/admin/training/<int:training_id>/country_templates_status')
@login_required
@admin_required
def training_country_templates_status(training_id):
    return jsonify(get_training_country_templates_status(training_id))

@admin_training_bp.route('/api/admin/training/<int:training_id>/country_template', methods=['GET'])
@login_required
@admin_required
def get_training_country_template(training_id):
    """1. 获取国家模板接口"""
    country = request.args.get('country')
    if not country:
        return jsonify({"error": "缺少 country 参数"}), 400
    db = get_supabase_admin()
    # 使用 execute()，不用 maybe_single()
    res = db.table("training_country_templates")\
        .select("header_template")\
        .eq("training_id", training_id)\
        .eq("country", country)\
        .execute()
    if res.data and len(res.data) > 0:
        template = res.data[0].get('header_template', {})
    else:
        template = {}
    return jsonify({"template": template})

@admin_training_bp.route('/api/admin/training/<int:training_id>/country_template', methods=['POST'])
@login_required
@admin_required
def save_training_country_template(training_id):
    """2. 保存国家模板接口 保存培训表头模板"""
    data = request.json
    country = data.get('country')
    template = data.get('template')
    
    if not training_id:
        return jsonify({"success": False, "message": "培训ID无效"}), 400
    
    if template is None:
        return jsonify({"success": False, "message": "缺少 template 参数"}), 400
    
    db = get_supabase_admin()
    
    try:
        # 情况1：没有指定国家 → 保存到培训主表的 header_template
        if not country or country == 'null' or country == 'undefined' or country == '':
            db.table("trainings").update({
                "header_template": template
            }).eq("id", training_id).execute()
            logger.info(f"保存培训主表头: training_id={training_id}")
            return jsonify({"success": True})
        
        # 情况2：指定了国家 → 保存到国家模板表
        check_res = db.table("training_country_templates")\
            .select("id")\
            .eq("training_id", training_id)\
            .eq("country", country)\
            .execute()
        
        if check_res.data and len(check_res.data) > 0:
            # 更新现有记录
            db.table("training_country_templates")\
                .update({
                    "header_template": template, 
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })\
                .eq("id", check_res.data[0]['id'])\
                .execute()
            logger.info(f"更新国家模板: training_id={training_id}, country={country}")
        else:
            # 插入新记录
            db.table("training_country_templates")\
                .insert({
                    "training_id": training_id, 
                    "country": country, 
                    "header_template": template
                })\
                .execute()
            logger.info(f"插入国家模板: training_id={training_id}, country={country}")
        
        return jsonify({"success": True})
        
    except Exception as e:
        logger.error(f"保存表头失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_training_bp.route('/api/admin/training/attendance/<int:attendance_id>/reset-signature', methods=['POST'])
@login_required
@admin_required
def admin_reset_signature(attendance_id):
    """管理员清除指定签到记录的签名，推送重新签字到学员端"""
    db = get_supabase()
    # 获取原记录，保留签到时间
    att_res = db.table("training_attendances").select("*").eq("id", attendance_id).maybe_single().execute()
    if not att_res.data:
        return jsonify({"success": False, "message": "签到记录不存在"}), 404
    
    # 清空签名相关字段，保留 sign_time
    db.table("training_attendances").update({
        "signature_url": None,
        "signed_name": None
    }).eq("id", attendance_id).execute()
    
    logger.info(f"管理员重置签到 {attendance_id} 的签名")
    return jsonify({"success": True})

@admin_training_bp.route('/api/admin/training/<int:training_id>/reset-signature/<user_id>', methods=['POST'])
@login_required
@admin_required
def admin_reset_signature_by_user(training_id, user_id):
    """
    管理员强制重置学员的签到签名（保留分配关系和签到时间）
    - 只清空 signature_url 和 signed_name
    - 保留 sign_time（签到时间）
    - 保留分配关系（training_assignments）
    """
    db = get_supabase()
    
    # 1. 检查培训是否存在
    training_res = db.table("trainings").select("name, country").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return jsonify({"success": False, "message": "培训不存在"}), 404
    
    # 2. 检查用户是否存在
    user_res = db.table("users").select("name_en").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "用户不存在"}), 404
    
    # 3. 权限检查
    allowed = get_admin_allowed_countries()
    training_country = training_res.data.get('country')
    if allowed is not None and training_country and training_country not in allowed:
        return jsonify({"success": False, "message": "无权操作此培训"}), 403
    
    # 4. 查找该用户的签到记录
    att_res = db.table("training_attendances") \
        .select("id, sign_time") \
        .eq("training_id", training_id) \
        .eq("user_id", user_id) \
        .is_("deleted_at", "null") \
        .maybe_single() \
        .execute()
    
    if not att_res.data:
        # 没有签到记录，自动创建一条（保留签到时间）
        # 注意：这种情况理论上不应该发生，因为调用此接口的前提是已有签到记录
        return jsonify({
            "success": False, 
            "message": "该学员没有签到记录，请先签到"
        }), 404
    
    attendance_id = att_res.data['id']
    sign_time = att_res.data.get('sign_time')
    
    # 5. 只清空签名相关字段，保留 sign_time
    db.table("training_attendances").update({
        "signature_url": None,
        "signed_name": None
    }).eq("id", attendance_id).execute()
    
    training_name = training_res.data.get('name', '')
    user_name = user_res.data.get('name_en', '')

    # 消息记录（独立于主流程，失败不影响返回）
    try:
        from utils.admin_messages import log_signature_reset
        log_signature_reset(
            training_id=training_id,
            training_name=training_name,
            user_id=user_id,
            user_name=user_name,
            admin_id=session.get('user_id')
        )
    except Exception as msg_err:
        # 消息记录失败只记录日志，不影响主流程
        logger.warning(f"消息记录失败（不影响主流程）: {msg_err}")

    logger.info(f"✅ 重置签名成功: 培训={training_name} ({training_id}), 学员={user_name} ({user_id}), 签到时间={sign_time}")
    
    return jsonify({
        "success": True,
        "message": f"已重置 {user_name} 的签名，签到时间保留",
        "sign_time": sign_time
    })

@admin_training_bp.route('/admin/training/<int:training_id>/attendance')
@login_required
@admin_required
def admin_training_attendance(training_id):
    return render_template('admin/list_training_attendance.html', training_id=training_id)

@admin_training_bp.route('/admin/training/<int:training_id>/attendance/print')
@login_required
@admin_required
def training_attendance_print(training_id):
    return render_template('admin/list_training_attendance.html', training_id=training_id, print_mode=True)

# routes/admin_training.py

@admin_training_bp.route('/admin/training/<int:training_id>/attendance/pdf')
@login_required
@admin_required
def download_training_attendance_pdf(training_id):
    country = request.args.get('country', '')
    lang = request.args.get('lang', 'zh')
    show_header = request.args.get('show_header', 'true').lower() == 'true'
    
    data = get_attendance_data(training_id, country)
    if not data:
        flash("培训不存在", "danger")
        return redirect(url_for('admin_dashboard'))

    # 渲染模板（不包含页眉页脚 HTML）
    html_content = render_template(
        'admin/attendance_pdf.html',
        training=data['training'],
        header=data['header_template'],
        attendances=data['attendances'],
        lang=lang
    )

    wkhtmltopdf_path = find_wkhtmltopdf()
    config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
    
    # 基础选项
    options = {
        'page-size': 'A4',
        'margin-top': '25mm',
        'margin-bottom': '25mm',
        'margin-left': '15mm',
        'margin-right': '15mm',
        'encoding': 'UTF-8',
        'enable-local-file-access': None,
        'javascript-delay': '200',
        'no-stop-slow-scripts': None,
    }
    
    # 如果显示页眉页脚，添加 wkhtmltopdf 的页眉页脚选项
    if show_header:
        if lang == 'en':
            options.update({
                # 页眉：左右分栏
                'header-left': 'ZTE',
                'header-right': 'Internal Use▲',
                'header-font-size': '9',
                'header-spacing': '8',
                'header-line': None,  # 在页眉下方添加分割线
                # 页脚：左右分栏
                'footer-left': 'All rights reserved. No distribution without prior permission of ZTE.',
                'footer-right': 'Page [page] / [topage]',
                'footer-font-size': '8',
                'footer-spacing': '8',
                'footer-line': None,  # 在页脚上方添加分割线
            })
        else:
            options.update({
                # 页眉：左右分栏
                'header-left': 'ZTE中兴',
                'header-right': '内部使用▲',
                'header-font-size': '9',
                'header-spacing': '8',
                'header-line': None,
                # 页脚：左右分栏
                'footer-left': '以上所有信息均为中兴通讯股份有限公司所有，不得外传',
                'footer-right': '页码 [page] / [topage]',
                'footer-font-size': '8',
                'footer-spacing': '8',
                'footer-line': None,
            })
    else:
        # 不显示页眉页脚时，使用正常边距
        options['margin-top'] = '15mm'
        options['margin-bottom'] = '15mm'
    
    pdf = pdfkit.from_string(html_content, False, configuration=config, options=options)
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=attendance_{training_id}.pdf'
    return response

@admin_training_bp.route('/api/admin/training/<int:training_id>/attendance_by_country')
@login_required
@admin_required
def api_training_attendance_by_country(training_id):
    db = get_supabase()
    allowed = get_allowed_countries()
    
    # 获取该培训的所有签到记录，并关联用户国家
    att_res = db.table("training_attendances") \
        .select("id, user_id, signed_name, signature_url, sign_time, users(country, name_cn, name_en, department, employee_id)") \
        .eq("training_id", training_id) \
        .execute()
    records = att_res.data or []
    
    # 国家权限过滤（仅保留允许国家的记录）
    if allowed is not None:
        if not allowed:
            return jsonify([])
        records = [r for r in records if r.get('users', {}).get('country') in allowed]
    
    # 按国家分组
    groups = {}
    for rec in records:
        user = rec.get('users', {})
        country = user.get('country') or '未指定'
        if country not in groups:
            groups[country] = {
                'country': country,
                'count': 0,
                'attendances': []
            }
        groups[country]['count'] += 1
        groups[country]['attendances'].append({
            'user_id': rec['user_id'],
            'department': user.get('department', ''),
            'name_cn': user.get('name_cn', ''),
            'name_en': user.get('name_en', ''),
            'employee_id': user.get('employee_id', ''),
            'signed_name': rec.get('signed_name', ''),
            'signature_url': rec.get('signature_url', ''),
            'sign_time': rec.get('sign_time')
        })
    return jsonify(list(groups.values()))

@admin_training_bp.route('/api/admin/training/attendance/<int:attendance_id>', methods=['DELETE'])
@login_required
@admin_required
def api_admin_delete_training_attendance(attendance_id):
    """删除培训签到记录"""
    db = get_supabase()
    user_id = session['user_id']
    
    # 1. 获取签到记录
    att_res = db.table("training_attendances").select("*").eq("id", attendance_id).execute()
    if not att_res.data:
        return jsonify({"success": False, "message": "签到记录不存在"}), 404
    
    attendance = att_res.data[0]
    training_id = attendance.get('training_id')
    
    # 2. 获取培训国家
    training_res = db.table("trainings").select("country").eq("id", training_id).maybe_single().execute()
    training_country = training_res.data.get('country') if training_res.data else None
    
    # 3. 权限检查
    allowed = get_admin_allowed_countries()
    if allowed is not None and training_country not in allowed:
        return jsonify({"success": False, "message": "无权删除此签到记录"}), 403
    
    try:
        # 软删除
        db.table("training_attendances").update({
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": user_id
        }).eq("id", attendance_id).execute()
        
        logger.info(f"培训签到记录已删除: attendance_id={attendance_id}, 操作人={user_id}")
        return jsonify({"success": True, "message": "签到记录已删除"})
    except Exception as e:
        logger.error(f"删除签到记录失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_training_bp.route('/api/admin/training/attendance/batch_delete', methods=['POST'])
@login_required
@admin_required
def api_admin_batch_delete_training_attendances():
    """批量删除培训签到记录（支持软删除和永久删除）"""
    data = request.json
    attendance_ids = data.get('ids', [])
    delete_type = data.get('delete_type', 'soft')
    
    if not attendance_ids:
        return jsonify({"success": False, "message": "请选择要删除的签到记录"}), 400
    
    db = get_supabase()
    user_id = session['user_id']
    success_count = 0
    fail_count = 0
    errors = []
    
    allowed = get_admin_allowed_countries()
    
    for att_id in attendance_ids:
        try:
            att_res = db.table("training_attendances").select("*").eq("id", att_id).execute()
            if not att_res.data:
                fail_count += 1
                errors.append(f"记录 {att_id} 不存在")
                continue
            
            attendance = att_res.data[0]
            training_id = attendance.get('training_id')
            
            training_res = db.table("trainings").select("country").eq("id", training_id).maybe_single().execute()
            training_country = training_res.data.get('country') if training_res.data else None
            
            if allowed is not None and training_country not in allowed:
                fail_count += 1
                errors.append(f"记录 {att_id}: 无权限删除")
                continue
            
            if delete_type == 'hard':
                db.table("training_attendances").delete().eq("id", att_id).execute()
            else:
                db.table("training_attendances").update({
                    "deleted_at": datetime.now(timezone.utc).isoformat(),
                    "deleted_by": user_id
                }).eq("id", att_id).execute()
            
            success_count += 1
        except Exception as e:
            fail_count += 1
            errors.append(f"记录 {att_id}: {str(e)}")
    
    return jsonify({
        "success": True,
        "success_count": success_count,
        "fail_count": fail_count,
        "errors": errors[:10]
    })

@admin_training_bp.route('/admin/training/attendance_status_view')
@login_required
@admin_required
def training_attendance_status_view():
    """培训签到人员状态查看页面"""
    return render_template('admin/list_training_status_view.html')

@admin_training_bp.route('/api/admin/training/users_with_status')
@login_required
@admin_required
def api_training_users_with_status():
    """获取权限范围内所有培训的学员签到状态（只显示培训对应国家的学员）"""
    db = get_supabase()
    admin_db = get_supabase_admin()
    
    allowed_countries = get_admin_allowed_countries()
    current_user_id = session.get('user_id')
    current_role = session.get('role')
    is_dev = is_developer()
    
    # 获取 URL 参数中的 training_id
    training_id_param = request.args.get('training_id', '').strip()
    search = request.args.get('search', '').strip()
    country = request.args.get('country', '').strip()
    wh_id = request.args.get('wh_id', '').strip()
    training_name = request.args.get('training_name', '').strip()
    
    # ========== 1. 获取权限范围内的所有培训 ==========
    trainings_query = db.table("trainings").select("id, name, country").is_("deleted_at", "null")
    
    if allowed_countries is not None:
        if not allowed_countries:
            return jsonify({"data": []})
        trainings_query = trainings_query.in_("country", allowed_countries)
    
    trainings_res = trainings_query.execute()
    trainings = trainings_res.data or []

    if training_id_param:
        try:
            training_id_int = int(training_id_param)
            allowed_training_ids = [t['id'] for t in trainings]
            if training_id_int not in allowed_training_ids:
                return jsonify({"data": []})
            trainings = [t for t in trainings if t['id'] == training_id_int]
        except ValueError:
            return jsonify({"data": []})
    
    if not trainings:
        return jsonify({"data": []})

    # ========== 2. ✅ 培训名称筛选（支持多字段组合查询）- 修复版 ==========
    if training_name:
        keywords = training_name.strip().split()
        # 模糊匹配每个关键词
        filtered_trainings = []
        for t in trainings:
            t_name = t.get('name', '').lower()
            # 所有关键词都必须匹配
            if all(k.lower() in t_name for k in keywords):
                filtered_trainings.append(t)
        trainings = filtered_trainings
    
    # 如果筛选后没有培训，直接返回空
    if not trainings:
        return jsonify({"data": []})
    
    # ========== 3. 获取培训涉及的国家列表 ==========
    training_countries = []
    for t in trainings:
        t_country = t.get('country')
        if t_country:
            training_countries.append(t_country)
    training_countries = list(set(training_countries))
    
    # ========== 4. 获取用户 ==========
    # 开发者特殊处理：先获取所有用户，后续再确保开发者自己出现在列表中
    if is_dev:
        # 开发者：获取所有用户（不受国家限制）
        users_query = db.table("users").select("*").is_("deleted_at", "null").eq("user_status", "registered").eq("is_resign", False)
    else:
        if training_countries:
            users_query = db.table("users").select("*").is_("deleted_at", "null").eq("user_status", "registered").eq("is_resign", False).in_("country", training_countries)
        else:
            if allowed_countries is not None and allowed_countries:
                users_query = db.table("users").select("*").is_("deleted_at", "null").eq("user_status", "registered").eq("is_resign", False).in_("country", allowed_countries)
            else:
                users_query = db.table("users").select("*").is_("deleted_at", "null").eq("user_status", "registered").eq("is_resign", False)

    # ========== 5. 获取所有用户（用于后续过滤） ==========
    users_res = users_query.execute()
    all_users = users_res.data or []

    # 开发者特殊处理：确保自己始终在用户列表中
    if is_dev:
        # 检查开发者是否在 all_users 中
        dev_in_list = any(u['id'] == current_user_id for u in all_users)
        if not dev_in_list:
            # 手动添加开发者自己
            dev_res = db.table("users").select("*").eq("id", current_user_id).maybe_single().execute()
            if dev_res and dev_res.data:
                all_users.append(dev_res.data)
                logger.info(f"✅ 开发者 {current_user_id} 已手动添加到用户列表")

    # ========== 6. 多字段组合查询（姓名/邮箱/库房/国家） ==========
    if search:
        keywords = search.strip().split()
        filtered_users = []
        for user in all_users:
            # 构建用户的搜索字段（姓名、邮箱、库房编码、库房名称、国家）
            searchable_fields = [
                user.get('name_en', '').lower(),
                user.get('name_cn', '').lower(),
                user.get('email', '').lower(),
                user.get('wh_id', '').lower(),
                user.get('wh_name_en', '').lower(),
                user.get('country', '').lower()
            ]
            
            # 组合字段文本
            combined_text = ' '.join(searchable_fields)
            
            # 检查所有关键词是否都在组合文本中
            if all(k.lower() in combined_text for k in keywords):
                filtered_users.append(user)
        users = filtered_users
    else:
        users = all_users
    
    # ========== 7. 额外筛选（国家、库房） ==========
    if country:
        users = [u for u in users if u.get('country', '').lower() == country.lower()]
    
    if wh_id:
        wh_lower = wh_id.lower()
        users = [u for u in users if 
                 (u.get('wh_id', '').lower() in wh_lower or wh_lower in u.get('wh_id', '').lower()) or
                 (u.get('wh_name_en', '').lower() in wh_lower or wh_lower in u.get('wh_name_en', '').lower())]
    
    # ========== 8. 权限过滤（使用新的权限逻辑） ==========
    filtered_users = []
    for user in users:  # ✅ 使用 users 而不是 all_users
        user_role = user.get('role', 'user')
        user_id = user.get('id')
        user_country = user.get('country', '')
        user_admin_countries = user.get('admin_countries')
        user_name = user.get('name_en', '')
        
        # 解析目标用户的权限范围
        target_admin_country_list = []
        if user_admin_countries:
            try:
                target_admin_country_list = json.loads(user_admin_countries) if isinstance(user_admin_countries, str) else user_admin_countries
            except:
                pass
        
        # 开发者：可以看到所有用户
        if is_dev:
            filtered_users.append(user)
            continue
        
        # ========== 超管逻辑 ==========
        if current_role == 'super_admin':
            # 超管看不到开发者
            if user_role == 'developer':
                continue
            
            # 如果当前超管没有权限范围限制，可以看到所有非开发者
            if allowed_countries is None:
                filtered_users.append(user)
                continue
            
            # 有权限范围限制
            if allowed_countries:
                # 目标用户是超管或管理员：检查 admin_countries 交集
                if user_role in ['super_admin', 'admin']:
                    if target_admin_country_list:
                        if any(c in allowed_countries for c in target_admin_country_list):
                            filtered_users.append(user)
                            continue
                    # 如果目标超管没有设置权限范围，视为全局，可见
                    elif user_role == 'super_admin' and not target_admin_country_list:
                        filtered_users.append(user)
                        continue
                    continue
                
                # 目标用户是普通用户：检查 country
                if user_role == 'user':
                    if user_country and user_country in allowed_countries:
                        filtered_users.append(user)
                        continue
                continue
            continue
        
        # ========== 管理员逻辑 ==========
        if current_role == 'admin':
            # 管理员看不到超管和开发者
            if user_role in ['super_admin', 'developer']:
                continue
            
            # 当前管理员的权限范围
            current_allowed = allowed_countries if allowed_countries else [session.get('user_country')]
            if not current_allowed:
                continue
            
            # 1. 如果是自己，允许
            if user_id == current_user_id:
                filtered_users.append(user)
                continue
            
            # 2. 目标用户是管理员：检查 admin_countries 交集
            if user_role == 'admin':
                if target_admin_country_list:
                    if any(c in current_allowed for c in target_admin_country_list):
                        filtered_users.append(user)
                        continue
                # 如果目标管理员没有设置权限范围，检查 country
                elif user_country and user_country in current_allowed:
                    filtered_users.append(user)
                    continue
                continue
            
            # 3. 目标用户是普通用户：检查 country
            if user_role == 'user':
                if user_country and user_country in current_allowed:
                    filtered_users.append(user)
                    continue
                continue
            
            continue
        
        # 其他角色
        filtered_users.append(user)
    
    users = filtered_users
    user_ids = [u['id'] for u in users]
    
    logger.info(f"权限过滤后 {len(users)} 个用户")
    
    if not user_ids:
        return jsonify({"data": []})
    
    # ========== 9. 获取培训分配记录 ==========
    training_ids = [t['id'] for t in trainings]
    assign_res = admin_db.table("training_assignments").select("training_id, user_id")\
        .in_("training_id", training_ids)\
        .execute()

    assignment_map = {}
    for a in (assign_res.data or []):
        key = f"{a['training_id']}_{a['user_id']}"
        assignment_map[key] = True
    
    # ========== 10. 获取签到记录 ==========
    att_res = db.table("training_attendances").select(
        "training_id, user_id, sign_time, signed_name, signature_url"
    ).in_("training_id", training_ids).in_("user_id", user_ids).is_("deleted_at", "null").execute()
    
    attendance_map = {}
    for att in (att_res.data or []):
        key = f"{att['training_id']}_{att['user_id']}"
        attendance_map[key] = {
            "sign_time": att.get('sign_time'),
            "signed_name": att.get('signed_name', ''),
            "signature_url": att.get('signature_url', '')
        }
    
    # ========== 11. 构建返回数据 ==========
    result = []
    for training in trainings:
        training_id = training['id']
        training_name_val = training.get('name', '')
        training_country = training.get('country', '')
        
        for user in users:
            user_country = user.get('country', '')
            user_id = user.get('id', '')
            
            # 开发者特殊处理：即使国家不匹配，也显示自己
            if is_dev and user_id == current_user_id:
                # 开发者自己，跳过国家检查
                pass
            else:
            
                # 如果培训有国家，用户国家必须匹配
                if training_country and user_country != training_country:
                    continue
            
            key = f"{training_id}_{user['id']}"

            is_assigned = key in assignment_map
            has_attendance = key in attendance_map
            attendance = attendance_map.get(key, {})
            sign_time = attendance.get('sign_time')
            signature_url = attendance.get('signature_url', '')
            
            if not is_assigned and not has_attendance:
                sign_status = 'not_assigned'
            elif not sign_time:
                sign_status = 'pending'
            elif not signature_url:
                sign_status = 'resign'
            else:
                sign_status = 'signed'
 
            result.append({
                "training_id": training_id,
                "training_name": training_name_val,
                "training_country": training_country,
                "user_id": user['id'],
                "user_country": user_country,
                "country": user_country,
                "name_en": user.get('name_en', ''),
                "email": user.get('email', ''),
                "wh_id": user.get('wh_id', ''),
                "wh_name_en": user.get('wh_name_en', ''),
                "company": user.get('company', ''),
                "sign_status": sign_status,
                "sign_time": sign_time,
                "signed_name": attendance.get('signed_name', ''),
                "signature_url": signature_url
            })
    
    result.sort(key=lambda x: (x.get('training_name', ''), x.get('name_en', '')))
    
    logger.info(f"最终返回 {len(result)} 条记录")
    return jsonify({"data": result})

@admin_training_bp.route('/api/admin/training/export_attendance_status')
@login_required
@admin_required
def export_training_attendance_status():
    """导出培训签到状态为Excel（只显示培训对应国家的学员）"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    
    db = get_supabase()
    
    # 获取当前管理员的权限范围
    allowed_countries = get_admin_allowed_countries()
    
    # 1. 获取权限范围内的所有培训
    trainings_query = db.table("trainings").select("id, name, country").is_("deleted_at", "null")
    
    if allowed_countries is not None:
        if not allowed_countries:
            return jsonify({"error": "无权限"}), 403
        trainings_query = trainings_query.in_("country", allowed_countries)
    
    trainings_res = trainings_query.execute()
    trainings = trainings_res.data or []
    
    if not trainings:
        return jsonify({"error": "没有可导出的数据"}), 404
    
    # 2. 获取培训涉及的国家列表
    training_countries = list(set([t.get('country') for t in trainings if t.get('country')]))
    
    # 3. 获取用户
    users_query = db.table("users").select("*").is_("deleted_at", "null").eq("user_status", "registered")
    
    if training_countries:
        users_query = users_query.in_("country", training_countries)
    
    # 筛选条件
    search = request.args.get('search', '')
    country = request.args.get('country', '')
    wh_id = request.args.get('wh_id', '')
    
    if search:
        users_query = users_query.or_(f"name_en.ilike.%{search}%,email.ilike.%{search}%")
    if country:
        users_query = users_query.eq("country", country)
    if wh_id:
        users_query = users_query.filter("wh_id", "ilike", f"%{wh_id}%")
    
    users_res = users_query.execute()
    users = users_res.data or []
    user_ids = [u['id'] for u in users]
    
    # 4. 获取签到记录
    attendance_map = {}
    if user_ids:
        training_ids = [t['id'] for t in trainings]
        att_res = db.table("training_attendances").select(
            "training_id, user_id, sign_time, signed_name, signature_url"
        ).in_("training_id", training_ids).in_("user_id", user_ids).is_("deleted_at", "null").execute()
        
        for att in (att_res.data or []):
            key = f"{att['training_id']}_{att['user_id']}"
            attendance_map[key] = {
                "sign_time": att.get('sign_time'),
                "signed_name": att.get('signed_name', ''),
                "signature_url": att.get('signature_url', '')
            }
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "培训签到状态"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # 表头
    headers = ["序号", "培训名称", "培训国家", "学员国家", "学员姓名", "邮箱", "库房编码", "库房名称", "公司", "签到时间", "签名姓名", "签到状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # 数据行
    status_map = {
        'signed': '已签到',
        'pending': '待签到',
        'resign': '需重新签名'
    }
    
    row_idx = 2
    for training in trainings:
        training_name = training.get('name', '')
        training_country = training.get('country', '')
        
        for user in users:
            user_country = user.get('country', '')
            
            # ✅ 关键：只导出培训国家对应的学员
            if user_country != training_country:
                continue
            
            key = f"{training['id']}_{user['id']}"
            attendance = attendance_map.get(key, {})
            sign_time = attendance.get('sign_time')
            signature_url = attendance.get('signature_url', '')
            
            if not sign_time:
                sign_status = 'pending'
            elif not signature_url:
                sign_status = 'resign'
            else:
                sign_status = 'signed'
            
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=training_name)
            ws.cell(row=row_idx, column=3, value=training_country)
            ws.cell(row=row_idx, column=4, value=user_country)
            ws.cell(row=row_idx, column=5, value=user.get('name_en', ''))
            ws.cell(row=row_idx, column=6, value=user.get('email', ''))
            ws.cell(row=row_idx, column=7, value=user.get('wh_id', ''))
            ws.cell(row=row_idx, column=8, value=user.get('wh_name_en', ''))
            ws.cell(row=row_idx, column=9, value=user.get('company', ''))
            ws.cell(row=row_idx, column=10, value=sign_time[:19] if sign_time else '')
            ws.cell(row=row_idx, column=11, value=attendance.get('signed_name', ''))
            ws.cell(row=row_idx, column=12, value=status_map.get(sign_status, '-'))
            
            row_idx += 1
    
    # 调整列宽
    column_widths = [6, 30, 12, 12, 15, 25, 15, 20, 20, 20, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"培训签到状态_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# routes/admin_training.py - 修复 search_trainings 函数

@admin_training_bp.route('/api/search/trainings')
@login_required
@admin_required
@cache_get(ttl=60, prefix='search_trainings', include_user=True)
def search_trainings():
    """模糊搜索培训名称（带权限过滤 + 级联筛选 + 绑定标记）- 支持多国家"""    
    q = request.args.get('q', '').strip()
    country = request.args.get('country', '').strip()
    warehouse = request.args.get('warehouse', '').strip()
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    exam_id = request.args.get('exam_id', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    if not q and not country and not warehouse and not start_date and not end_date:
        per_page = 8
    
    db = get_supabase()
    
    # 获取当前用户的权限范围
    allowed_countries = get_admin_allowed_countries()
    
    # ========== 1. 基础查询 ==========
    query = db.table("trainings").select("*, dynamic_status").is_("deleted_at", "null")
    
    if q:
        query = query.ilike("name", f"%{q}%")
    
    # ========== 2. 时间范围过滤 ==========
    if start_date:
        query = query.gte("end_time", start_date)
    if end_date:
        query = query.lte("start_time", end_date)
    
    # 获取所有匹配的培训
    res = query.execute()
    all_trainings = res.data or []
    
    logger.info(f"[search_trainings] 查询到 {len(all_trainings)} 条培训")
    
    # ========== 3. 权限过滤（支持多国家） ==========
    if allowed_countries is not None:
        if not allowed_countries:
            logger.info("[search_trainings] 用户无权限，返回空列表")
            return jsonify([])
        filtered_by_permission = []
        for t in all_trainings:
            t_countries = parse_training_countries(t)
            if any(c in allowed_countries for c in t_countries):
                filtered_by_permission.append(t)
        all_trainings = filtered_by_permission
        logger.info(f"[search_trainings] 权限过滤后 {len(all_trainings)} 条")
    
    # ========== 4. 国家过滤（支持多国家） ==========
    if country:
        filtered_by_country = []
        for t in all_trainings:
            t_countries = parse_training_countries(t)
            if country in t_countries:
                filtered_by_country.append(t)
        all_trainings = filtered_by_country
        logger.info(f"[search_trainings] 国家过滤后 {len(all_trainings)} 条")
    
    # ========== 5. 库房过滤 ==========
    if warehouse:
        users_res = db.table("users").select("id").eq("wh_id", warehouse).execute()
        user_ids = [u['id'] for u in (users_res.data or [])]
        if user_ids:
            assign_res = db.table("training_assignments").select("training_id").in_("user_id", user_ids).execute()
            training_ids = list(set([a['training_id'] for a in (assign_res.data or [])]))
            if training_ids:
                all_trainings = [t for t in all_trainings if t['id'] in training_ids]
            else:
                return jsonify([])
        else:
            return jsonify([])
    
    # ========== 6. 排序 ==========
    all_trainings.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    # ========== 7. 分页 ==========
    total = len(all_trainings)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated = all_trainings[start_idx:end_idx]
    
    # ========== 8. 获取绑定标记 ==========
    bound_training_ids = set()
    if exam_id:
        try:
            exam_id_int = int(exam_id)
            bindings_res = db.table("training_exam_bindings").select("training_id").eq("exam_id", exam_id_int).execute()
            bound_training_ids = set([b['training_id'] for b in (bindings_res.data or [])])
        except ValueError:
            bound_training_ids = set()

    # ========== 9. 获取签到人数统计 ==========
    # 批量获取所有培训的签到人数
    signed_counts = {}
    if paginated:
        training_ids = [t['id'] for t in paginated]
        # 获取每个培训的签到人数（只统计在职人员）
        for tid in training_ids:
            # 查询该培训的签到记录
            att_res = db.table("training_attendances")\
                .select("user_id, users!inner(is_resign)")\
                .eq("training_id", tid)\
                .is_("deleted_at", "null")\
                .execute()
            
            # 过滤离职人员
            signed_user_ids = []
            for att in (att_res.data or []):
                user = att.get('users', {})
                if user.get('is_resign') == False:
                    signed_user_ids.append(att['user_id'])
            
            signed_counts[tid] = len(set(signed_user_ids))
    
    # ========== 10. 组装返回数据（保持兼容性） ==========
    result = []
    for t in paginated:
        t_countries = parse_training_countries(t)
        # 确保返回所有原始字段 + 新增字段
        result.append({
            "id": t['id'],
            "name": t.get('name', ''),
            "dynamic_status": t.get('dynamic_status', 'draft'),
            "country": t.get('country', ''),  # 保留旧字段兼容
            "countries": t_countries,         # 新增多国家数组
            "countries_display": ', '.join(t_countries) if t_countries else t.get('country', ''),
            "created_at": t.get('created_at'),
            "start_time": t.get('start_time'),
            "end_time": t.get('end_time'),
            "dynamic_status": t.get('dynamic_status', 'draft'),
            "is_active": t.get('is_active', False),
            "is_bound": t['id'] in bound_training_ids,
            "header_template": t.get('header_template', {}),
            "signed_count": signed_counts.get(t['id'], 0)
        })
    
    logger.info(f"[search_trainings] 最终返回 {len(result)} 条数据")
    return jsonify(result)

def _get_quarter_from_date(date_str):
    """从日期字符串提取季度（格式：2026Q2）"""
    if not date_str:
        return ''
    try:
        from datetime import datetime
        if isinstance(date_str, str):
            if 'T' in date_str:
                date_str = date_str.split('T')[0]
            elif ' ' in date_str:
                date_str = date_str.split(' ')[0]
        date = datetime.fromisoformat(date_str) if isinstance(date_str, str) else date_str
        year = date.year
        month = date.month
        quarter = (month - 1) // 3 + 1
        return f"{year}Q{quarter}"
    except:
        return ''

@admin_training_bp.route('/api/admin/training/bindings/by_exam')
@login_required
@admin_required
def get_training_bindings_by_exam():
    """根据考试ID获取绑定的培训列表"""
    exam_id = request.args.get('exam_id', '')
    if not exam_id:
        return jsonify({"bindings": [], "training_ids": []})
    
    try:
        exam_id_int = int(exam_id)
        #db = get_supabase()
        db = get_supabase_admin()
        # 查询该考试绑定的培训
        bindings_res = db.table("training_exam_bindings").select("training_id").eq("exam_id", exam_id_int).execute()
        training_ids = list(set([b['training_id'] for b in (bindings_res.data or [])]))
        
        # 获取培训详细信息（可选）
        trainings = []
        if training_ids:
            trainings_res = db.table("trainings").select("id, name, country").in_("id", training_ids).execute()
            trainings = trainings_res.data or []
        
        return jsonify({
            "bindings": trainings,
            "training_ids": training_ids
        })
    except Exception as e:
        logger.error(f"获取考试绑定培训失败: {e}")
        return jsonify({"bindings": [], "training_ids": []}), 500

@admin_training_bp.route('/api/search/exams')
@login_required
@admin_required
def search_exams():
    """模糊搜索考试名称（带权限过滤）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    
    db = get_supabase()
    allowed_countries = get_allowed_countries()
    from routes.helpers import parse_exam_countries
    
    # 获取所有匹配的考试
    res = db.table("exams").select("title, countries, country").is_("deleted_at", "null").ilike("title", f"%{q}%").execute()
    all_exams = res.data or []
    
    # ✅ 权限过滤：只返回权限范围内的考试
    filtered_names = []
    for exam in all_exams:
        exam_countries = parse_exam_countries(exam)
        
        # 无权限限制
        if allowed_countries is None:
            filtered_names.append(exam['title'])
            continue
        
        # 有权限限制：检查是否有交集
        if allowed_countries:
            if any(c in allowed_countries for c in exam_countries):
                filtered_names.append(exam['title'])
    
    # 去重并限制数量
    unique_names = list(set(filtered_names))[:10]
    return jsonify(unique_names)

@admin_training_bp.route('/api/search/warehouses')
@login_required
@admin_required
def search_warehouses():
    """模糊搜索库房编码/名称（从 users 表，带权限过滤）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    
    db = get_supabase()
    allowed_countries = get_allowed_countries()
    
    # 基础查询 - 从 users 表查询
    query = db.table("users").select("wh_id, wh_name_en, country").is_("deleted_at", "null")
    
    # ✅ 权限过滤：只返回权限范围内的用户对应的库房
    if allowed_countries is not None and allowed_countries:
        query = query.in_("country", allowed_countries)
    
    # 模糊查询 wh_id 或 wh_name_en
    # 注意：Supabase 不支持 OR 条件的 ilike，需要分别查询后合并
    r1 = query.ilike("wh_id", f"%{q}%").limit(20).execute()
    r2 = query.ilike("wh_name_en", f"%{q}%").limit(20).execute()
    
    # 合并去重
    seen = set()
    suggestions = []
    
    for row in (r1.data or []) + (r2.data or []):
        wh_id = row.get('wh_id', '')
        wh_name = row.get('wh_name_en', '')
        
        if not wh_id:
            continue
        
        # 生成显示标签
        if wh_name:
            label = f"{wh_id} ({wh_name})"
        else:
            label = wh_id
        
        if label and label not in seen:
            seen.add(label)
            suggestions.append(label)
    
    return jsonify(suggestions[:10])

# routes/admin_training.py - 添加获取单个培训的接口

@admin_training_bp.route('/api/admin/trainings/<int:training_id>', methods=['GET'])
@login_required
@admin_required
def api_admin_training_detail(training_id):
    """获取单个培训详情"""
    from services.db import get_supabase_admin
    from utils.permissions import get_admin_allowed_countries
    
    db = get_supabase_admin()
    allowed_countries = get_admin_allowed_countries()
    
    # 获取培训信息
    res = db.table("trainings").select("*").eq("id", training_id).maybe_single().execute()
    if not res.data:
        return jsonify({"success": False, "message": "培训不存在"}), 404
    
    training = res.data
    
    # 权限检查
    training_country = training.get('country')
    if allowed_countries is not None and training_country and training_country not in allowed_countries:
        return jsonify({"success": False, "message": "无权访问此培训"}), 403
    
    return jsonify({"success": True, "data": training})

# ==================== 正确注册到 admin_training_bp 蓝图中 ====================
@admin_training_bp.route('/admin/training/completion_report')
@login_required
@admin_required
def completion_report_page():
    """完成度报表页面"""
    return render_template('admin/training_completion_report.html')

@admin_training_bp.route('/api/admin/training/completion_report')
@login_required
@admin_required
def get_completion_report():
    """获取培训完成度报表数据 - 实时查询版"""
    try:
        db = get_supabase_admin()
        
        # 获取当前管理员的权限范围
        allowed_countries = get_admin_allowed_countries()
        is_dev = is_developer()
        current_role = session.get('role')
        
        # 获取筛选参数
        name = request.args.get('name', '').strip()
        country_filter = request.args.get('country', '').strip()
        training_name = request.args.get('training_name', '').strip()
        training_id = request.args.get('training_id', '').strip()
        exam_name = request.args.get('exam_name', '').strip()
        status = request.args.get('status', '').strip()
        training_status_filter = request.args.get('training_status', '').strip()
        exam_status_filter = request.args.get('exam_status', '').strip()
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # ========== 1. 获取培训 ==========
        trainings_query = db.table("trainings").select("*").is_("deleted_at", "null")
        
        if training_id:
            trainings_query = trainings_query.eq("id", int(training_id))
        elif training_name:
            trainings_query = trainings_query.ilike("name", f"%{training_name}%")
        
        if start_date:
            trainings_query = trainings_query.gte("start_time", start_date)
        if end_date:
            trainings_query = trainings_query.lte("end_time", end_date)
        
        all_trainings = trainings_query.execute().data or []
        
        # 权限过滤（使用多国家）
        if not is_dev and allowed_countries:
            all_trainings = filter_trainings_by_country(all_trainings, allowed_countries)
        
        if not all_trainings:
            return jsonify({"data": [], "total": 0, "page": 1, "per_page": per_page, "summary": {}})
        
        training_ids = [t['id'] for t in all_trainings]
        trainings_dict = {t['id']: t for t in all_trainings}
        
        # ========== 2. 获取绑定关系 ==========
        bindings_result = db.table("training_exam_bindings").select("*")\
            .in_("training_id", training_ids)\
            .is_("deleted_at", "null")\
            .execute()
        bindings = bindings_result.data if bindings_result else []
        
        if not bindings:
            return jsonify({"data": [], "total": 0, "page": 1, "per_page": per_page, "summary": {}})
        
        # ========== 3. 获取考试信息 ==========
        exam_ids = list(set([b.get('exam_id') for b in bindings]))
        exams_result = db.table("exams").select("*").in_("id", exam_ids).is_("deleted_at", "null").execute()
        exams_dict = {e['id']: e for e in (exams_result.data or [])}
        
        if exam_name:
            filtered_bindings = []
            for b in bindings:
                exam = exams_dict.get(b['exam_id'])
                if exam and exam_name.lower() in exam.get('title', '').lower():
                    filtered_bindings.append(b)
            bindings = filtered_bindings
        
        if not bindings:
            return jsonify({"data": [], "total": 0, "page": 1, "per_page": per_page, "summary": {}})

        # ========== 4. 收集所有需要匹配的国家（培训 ∩ 考试） ==========
        # 构建一个映射：binding_id -> 有效国家列表
        binding_countries_map = {}
        
        for binding in bindings:
            training_id_val = binding.get('training_id')
            exam_id = binding.get('exam_id')
            
            training = trainings_dict.get(training_id_val)
            exam = exams_dict.get(exam_id)
            
            if not training or not exam:
                continue
            
            # 获取培训的国家列表
            training_countries = parse_training_countries(training)
            if not training_countries:
                continue
            
            # 获取考试的国家列表
            exam_countries = parse_exam_countries(exam)
            if not exam_countries:
                # 如果考试没有国家，使用培训的国家（向后兼容）
                exam_countries = training_countries
            
            # 计算交集：只有交集中的国家才匹配用户
            intersection_countries = [c for c in training_countries if c in exam_countries]
            
            if not intersection_countries:
                # 没有交集，跳过这个绑定
                continue
            
            binding_countries_map[binding.get('id')] = {
                'training_id': training_id_val,
                'exam_id': exam_id,
                'intersection_countries': intersection_countries,
                'training': training,
                'exam': exam
            }
        
        if not binding_countries_map:
            return jsonify({"data": [], "total": 0, "page": 1, "per_page": per_page, "summary": {}})
        
        # ========== 5. 收集所有有效国家（用于查询用户 ==========
        all_effective_countries = set()
        for info in binding_countries_map.values():
            all_effective_countries.update(info['intersection_countries'])
        
        # ========== 6. 获取用户 ==========
        if all_effective_countries:
            users_query = db.table("users").select("id, name_en, name_cn, email, country, wh_id, role")\
                .in_("country", list(all_effective_countries))\
                .eq("user_status", "registered")\
                .is_("deleted_at", "null")\
                .eq("is_resign", False)
        else:
            users_query = db.table("users").select("*")\
                .eq("user_status", "registered")\
                .is_("deleted_at", "null")\
                .eq("is_resign", False)
        
        if country_filter:
            users_query = users_query.ilike("country", f"%{country_filter}%")
        
        users_result = users_query.execute()
        all_users = users_result.data if users_result else []
        
        if name:
            name_lower = name.lower()
            filtered_users = []
            for u in all_users:
                name_en = (u.get('name_en') or '').lower()
                name_cn = (u.get('name_cn') or '').lower()
                email = (u.get('email') or '').lower()
                if name_lower in name_en or name_lower in name_cn or name_lower in email:
                    filtered_users.append(u)
            all_users = filtered_users

        # 权限过滤
        all_users = filter_users_by_permission(
            all_users, 
            allowed_countries=allowed_countries,
            current_user_id=session.get('user_id')
        )

        # 按国家分组用户
        users_by_country = {}
        for u in all_users:
            country = u.get('country')
            if country not in users_by_country:
                users_by_country[country] = []
            users_by_country[country].append(u)
        
        # ========== 7. 批量获取实时数据（直接从源表查询）==========
        
        # 7.1 获取培训分配关系
        training_assign_res = db.table("training_assignments").select("training_id, user_id")\
            .in_("training_id", training_ids)\
            .execute()
        training_assign_map = {}
        for a in (training_assign_res.data or []):
            key = f"{a['training_id']}_{a['user_id']}"
            training_assign_map[key] = True
        
        # 7.2 获取考试分配关系
        exam_assign_res = db.table("exam_assignments").select("exam_id, user_id")\
            .in_("exam_id", exam_ids)\
            .is_("deleted_at", "null")\
            .execute()
        exam_assign_map = {}
        for a in (exam_assign_res.data or []):
            key = f"{a['exam_id']}_{a['user_id']}"
            exam_assign_map[key] = True
        
        # 7.3 获取培训签到记录
        attendances_res = db.table("training_attendances").select("training_id, user_id, sign_time, signature_url")\
            .in_("training_id", training_ids)\
            .execute()
        attendance_map = {}
        for a in (attendances_res.data or []):
            key = f"{a['training_id']}_{a['user_id']}"
            attendance_map[key] = {
                'sign_time': a.get('sign_time'),
                'signature_url': a.get('signature_url')
            }
        
        # 7.4 获取考试成绩记录
        exam_results_res = db.table("exam_results").select("exam_id, user_id, total_score, created_at")\
            .in_("exam_id", exam_ids)\
            .is_("deleted_at", "null")\
            .execute()
        exam_results_map = {}
        for r in (exam_results_res.data or []):
            key = f"{r['exam_id']}_{r['user_id']}"
            exam_results_map[key] = {
                'score': r.get('total_score'),
                'submitted_at': r.get('submitted_at')
            }

        # ========== 8. 预计算培训和考试状态（使用 status.py） ==========
        training_status_cache = {}
        exam_status_cache = {}
        
        for training in all_trainings:
            training_status_cache[training['id']] = get_training_status(training)
        
        for exam in exams_dict.values():
            exam_status_cache[exam['id']] = get_exam_status(exam)
  
        # ========== 9. 组装数据（基于交集国家匹配用户）==========
        all_data = []
        
        for binding in bindings:
            binding_id = binding.get('id')
            binding_info = binding_countries_map.get(binding_id)
            
            if not binding_info:
                continue
            
            training = binding_info['training']
            exam = binding_info['exam']
            intersection_countries = binding_info['intersection_countries']
            
            training_id_val = binding_info['training_id']
            exam_id = binding_info['exam_id']
            pass_score = binding.get('pass_score', 85)

            # 使用预计算的状态
            training_status_global = training_status_cache.get(training_id_val, 'unknown')
            exam_status_global = exam_status_cache.get(exam_id, 'unknown')

            # 为每个国家获取用户并合并
            all_users_for_binding = []
            for country in intersection_countries:
                users = users_by_country.get(country, [])
                all_users_for_binding.extend(users)
            
            # 去重
            seen_user_ids = set()
            unique_users = []
            for u in all_users_for_binding:
                uid = u.get('id')
                if uid and uid not in seen_user_ids:
                    seen_user_ids.add(uid)
                    unique_users.append(u)
            
            # 角色过滤
            if current_role == 'admin':
                unique_users = [u for u in unique_users if u.get('role') not in ['super_admin', 'developer']]
            elif current_role == 'super_admin' and not is_dev:
                unique_users = [u for u in unique_users if u.get('role') != 'developer']
            
            for user in unique_users:
                user_id = user.get('id')
                if not user_id:
                    continue

                # 检查用户国家是否在交集中（额外安全）
                user_country = user.get('country')
                if user_country not in intersection_countries:
                    continue

                # 实时查询状态
                training_assign_key = f"{training_id_val}_{user_id}"
                exam_assign_key = f"{exam_id}_{user_id}"
                attendance_key = f"{training_id_val}_{user_id}"
                exam_result_key = f"{exam_id}_{user_id}"
                
                has_training_assign = training_assign_key in training_assign_map
                has_exam_assign = exam_assign_key in exam_assign_map

                # 获取签到详情（包含签名状态）
                attendance_data = attendance_map.get(attendance_key, {})
                is_signed = attendance_key in attendance_map
                sign_time = attendance_map.get(attendance_key)
                signature_url = attendance_data.get('signature_url')
                
                exam_result = exam_results_map.get(exam_result_key, {})
                is_completed = exam_result_key in exam_results_map
                score = exam_result.get('score')
                is_passed = score is not None and score >= pass_score if is_completed else False

                # 判断是否真正签到（有签名）
                is_signed_completed = is_signed and signature_url  # 有签到记录且有签名

                # 判断是否需要重新签名
                is_resign_needed = is_signed and not signature_url  # 有签到记录但无签名

                # 状态判断
                if is_resign_needed:
                    training_status_display = 'resign'  # 待重新签名
                elif is_signed_completed:
                    training_status_display = 'signed'  # 已签到
                elif has_training_assign:
                    training_status_display = 'pending'  # 待签到
                else:
                    training_status_display = 'not_assigned'  # 未推送
                
                if is_completed:
                    exam_status_display = 'completed'
                elif has_exam_assign:
                    exam_status_display = 'pending'
                elif is_signed:
                    exam_status_display = 'pending'
                else:
                    exam_status_display = 'not_assigned'

                # 状态筛选（新增培训和考试全局状态筛选）
                if training_status_filter and training_status_global != training_status_filter:
                    continue
                if exam_status_filter and exam_status_global != exam_status_filter:
                    continue
    
                # 状态筛选
                if status:
                    if status == 'completed' and not is_passed:
                        continue
                    if status == 'incomplete' and (is_completed or is_passed):
                        continue
                    if status == 'failed' and (not is_completed or is_passed):
                        continue
                    if status == 'not_signed' and is_signed:
                        continue
                    if status == 'not_exam' and (is_completed or not is_signed):
                        continue

                # 计算分配状态
                if training_status_display == 'not_assigned' and exam_status_display == 'not_assigned':
                    allocation_status = '未分配'
                elif training_status_display == 'signed' and exam_status_display in ['pending', 'completed']:
                    allocation_status = '已分配'
                else:
                    allocation_status = '部分分配'

                # 计算完成度
                if training_status_display == 'not_assigned' and exam_status_display == 'not_assigned':
                    completion_level = '未参与'
                elif training_status_display == 'signed' and exam_status_display == 'completed':
                    completion_level = '整体完成'
                else:
                    completion_level = '部分完成'
                    
                all_data.append({
                    "training_id": training_id_val,
                    "training_name": training.get('name', ''),
                    "training_country": user_country,  # 使用用户的实际国家
                    "training_start": training.get('start_time'),
                    "training_end": training.get('end_time'),
                    "training_status_global": training_status_global,
                    "allocation_status": allocation_status,      # 新增
                    "completion_level": completion_level,        # 新增
                    "exam_id": exam_id,
                    "exam_name": exam.get('title', ''),
                    "exam_status_global": exam_status_global,
                    "pass_score": pass_score,
                    "user_id": user_id,
                    "user_name": user.get('name_cn') or user.get('name_en', ''),
                    "user_email": user.get('email', ''),
                    "user_country": user.get('country', ''),
                    "wh_id": user.get('wh_id', ''),
                    "user_role": user.get('role', 'user'),
                    "is_signed": is_signed_completed,
                    "is_resign": is_resign_needed,
                    "is_completed": is_completed,
                    "is_passed": is_passed,
                    "score": score,
                    "training_status": training_status_display,
                    "exam_status": exam_status_display,
                    "has_training_assign": has_training_assign,
                    "has_exam_assign": has_exam_assign,
                    "signed_at": sign_time,
                    "completed_at": exam_result.get('submitted_at')
                })
        
        # 分页
        total = len(all_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated = all_data[start_idx:end_idx]
        
        # 统计汇总（基于唯一学员）
        unique_users = set()
        unique_users_signed = set()
        unique_users_completed = set()
        unique_users_passed = set()
        unassigned_users = set()
        pending_users = set()
        
        for d in all_data:
            user_id = d.get('user_id')
            if not user_id:
                continue
            unique_users.add(user_id)
            
            training_status = d.get('training_status')
            exam_status = d.get('exam_status')
            
            # 未分配：培训或考试任一未分配
            if training_status == 'not_assigned' or exam_status == 'not_assigned':
                unassigned_users.add(user_id)
            
            # 待处理：培训待签到 或 考试待考试
            if training_status == 'pending' or exam_status == 'pending':
                pending_users.add(user_id)
            
            if d.get('is_signed'):
                unique_users_signed.add(user_id)
            if d.get('is_completed'):
                unique_users_completed.add(user_id)
            if d.get('is_passed'):
                unique_users_passed.add(user_id)
        
        total_users_count = len(unique_users)
        passed_users_count = len(unique_users_passed)
        signed_users_count = len(unique_users_signed)
        completed_users_count = len(unique_users_completed)
        unassigned_count = len(unassigned_users)
        pending_count = len(pending_users)
        
        pass_rate = round(passed_users_count / total_users_count * 100, 1) if total_users_count > 0 else 0
        sign_rate = round(signed_users_count / total_users_count * 100, 1) if total_users_count > 0 else 0
        completion_rate = round(completed_users_count / total_users_count * 100, 1) if total_users_count > 0 else 0
        
        summary_stats = {
            "total_users": total_users_count,
            "total_signed": signed_users_count,
            "total_completed": completed_users_count,
            "total_passed": passed_users_count,
            "pass_rate": pass_rate,
            "sign_rate": sign_rate,
            "completion_rate": completion_rate,
            "unassigned_count": unassigned_count,
            "pending_count": pending_count 
        }
        
        return jsonify({
            "data": paginated,
            "total": total,
            "page": page,
            "per_page": per_page,
            "summary": summary_stats
        })
        
    except Exception as e:
        logger.error(f"报表API错误: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e), "data": [], "total": 0}), 500

@admin_training_bp.route('/api/admin/training/completion_report/export')
@login_required
@admin_required
def export_completion_report():
    """导出完成度报表 Excel"""
    try:
        db = get_supabase_admin()

        # 获取前端传递的时区
        timezone_param = request.args.get('timezone', '')
        
        # 如果传入了时区，设置到 session
        if timezone_param:
            try:
                pytz.timezone(timezone_param)
                session['user_timezone'] = timezone_param
                g.user_timezone = timezone_param
            except:
                pass
        # 获取当前管理员的权限范围
        allowed_countries = get_admin_allowed_countries()
        is_dev = is_developer()
        current_role = session.get('role')
        
        # 获取筛选参数
        name = request.args.get('name', '').strip()
        country_filter = request.args.get('country', '').strip()
        training_name = request.args.get('training_name', '').strip()
        training_id = request.args.get('training_id', '').strip()
        exam_name = request.args.get('exam_name', '').strip()
        status = request.args.get('status', '').strip()
        training_status_filter = request.args.get('training_status', '').strip()
        exam_status_filter = request.args.get('exam_status', '').strip()
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # ========== 1. 获取培训 ==========
        trainings_query = db.table("trainings").select("*").is_("deleted_at", "null")
        
        if training_id:
            trainings_query = trainings_query.eq("id", int(training_id))
        elif training_name:
            trainings_query = trainings_query.ilike("name", f"%{training_name}%")
        
        if start_date:
            trainings_query = trainings_query.gte("start_time", start_date)
        if end_date:
            trainings_query = trainings_query.lte("end_time", end_date)
        
        all_trainings = trainings_query.execute().data or []
        
        if not is_dev and allowed_countries:
            filtered_trainings = []
            for t in all_trainings:
                training_country = t.get('country')
                if training_country and training_country in allowed_countries:
                    filtered_trainings.append(t)
            all_trainings = filtered_trainings
        
        if not all_trainings:
            # 没有培训，返回空Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "培训完成度报表"
            ws.cell(row=1, column=1, value="暂无数据")
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=f'培训完成度报表_空.xlsx')
        
        training_ids = [t['id'] for t in all_trainings]
        trainings_dict = {t['id']: t for t in all_trainings}
        
        # ========== 2. 获取绑定关系 ==========
        bindings_result = db.table("training_exam_bindings").select("*")\
            .in_("training_id", training_ids)\
            .is_("deleted_at", "null")\
            .execute()
        bindings = bindings_result.data if bindings_result else []
        
        if not bindings:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "培训完成度报表"
            ws.cell(row=1, column=1, value="暂无绑定数据")
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=f'培训完成度报表_空.xlsx')
        
        # ========== 3. 获取考试信息 ==========
        exam_ids = list(set([b.get('exam_id') for b in bindings]))
        exams_result = db.table("exams").select("*").in_("id", exam_ids).is_("deleted_at", "null").execute()
        exams_dict = {e['id']: e for e in (exams_result.data or [])}
        
        if exam_name:
            filtered_bindings = []
            for b in bindings:
                exam = exams_dict.get(b['exam_id'])
                if exam and exam_name.lower() in exam.get('title', '').lower():
                    filtered_bindings.append(b)
            bindings = filtered_bindings
        
        if not bindings:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "培训完成度报表"
            ws.cell(row=1, column=1, value="未找到匹配的考试")
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=f'培训完成度报表_空.xlsx')
        
        # ========== 4. 获取用户 ==========
        all_countries = set()
        for t in all_trainings:
            country = t.get('country')
            if country:
                all_countries.add(country)
        
        if all_countries:
            users_query = db.table("users").select("id, name_en, name_cn, email, country, wh_id, role")\
                .in_("country", list(all_countries))\
                .eq("user_status", "registered")\
                .is_("deleted_at", "null")\
                .eq("is_resign", False)
        else:
            users_query = db.table("users").select("*")\
                .eq("user_status", "registered")\
                .is_("deleted_at", "null")\
                .eq("is_resign", False)
        
        if country_filter:
            users_query = users_query.ilike("country", f"%{country_filter}%")
        
        users_result = users_query.execute()
        all_users = users_result.data if users_result else []
        
        if name:
            name_lower = name.lower()
            filtered_users = []
            for u in all_users:
                name_en = (u.get('name_en') or '').lower()
                name_cn = (u.get('name_cn') or '').lower()
                email = (u.get('email') or '').lower()
                if name_lower in name_en or name_lower in name_cn or name_lower in email:
                    filtered_users.append(u)
            all_users = filtered_users

        # 权限过滤
        all_users = filter_users_by_permission(
            all_users, 
            allowed_countries=allowed_countries,
            current_user_id=session.get('user_id')
        )

        users_by_country = {}
        for u in all_users:
            country = u.get('country')
            if country not in users_by_country:
                users_by_country[country] = []
            users_by_country[country].append(u)
        
        # ========== 5. 获取实时数据 ==========
        # 5.1 培训分配
        training_assign_res = db.table("training_assignments").select("training_id, user_id")\
            .in_("training_id", training_ids)\
            .execute()
        training_assign_map = {}
        for a in (training_assign_res.data or []):
            key = f"{a['training_id']}_{a['user_id']}"
            training_assign_map[key] = True
        
        # 5.2 考试分配
        exam_assign_res = db.table("exam_assignments").select("exam_id, user_id")\
            .in_("exam_id", exam_ids)\
            .is_("deleted_at", "null")\
            .execute()
        exam_assign_map = {}
        for a in (exam_assign_res.data or []):
            key = f"{a['exam_id']}_{a['user_id']}"
            exam_assign_map[key] = True
        
        # 5.3 培训签到
        attendances_res = db.table("training_attendances").select("training_id, user_id, sign_time, signature_url")\
            .in_("training_id", training_ids)\
            .execute()
        attendance_map = {}
        for a in (attendances_res.data or []):
            key = f"{a['training_id']}_{a['user_id']}"
            attendance_map[key] = {
                'sign_time': a.get('sign_time'),
                'signature_url': a.get('signature_url')
            }
        
        # 5.4 考试成绩
        exam_results_res = db.table("exam_results").select("exam_id, user_id, total_score, created_at")\
            .in_("exam_id", exam_ids)\
            .is_("deleted_at", "null")\
            .execute()
        exam_results_map = {}
        for r in (exam_results_res.data or []):
            key = f"{r['exam_id']}_{r['user_id']}"
            exam_results_map[key] = {
                'score': r.get('total_score'),
                'created_at': r.get('created_at')
            }

        # ========== 6. 预计算培训和考试状态（复用辅助函数） ==========
        training_status_cache = {}
        exam_status_cache = {}
        
        for training in all_trainings:
            training_status_cache[training['id']] = get_training_status(training)
        
        for exam in exams_dict.values():
            exam_status_cache[exam['id']] = get_exam_status(exam)
  
        # ========== 7. 组装数据 ==========
        all_data = []
        
        for binding in bindings:
            training_id_val = binding.get('training_id')
            exam_id = binding.get('exam_id')
            pass_score = binding.get('pass_score', 85)
            
            training = trainings_dict.get(training_id_val)
            exam = exams_dict.get(exam_id)
            
            if not training or not exam:
                continue
            
            training_country = training.get('country', '')
            if not training_country:
                continue

            # 使用预计算的状态
            training_status_global = training_status_cache.get(training_id_val, 'unknown')
            exam_status_global = exam_status_cache.get(exam_id, 'unknown')

            # 状态筛选（新增培训和考试全局状态筛选）
            if training_status_filter and training_status_global != training_status_filter:
                continue
            if exam_status_filter and exam_status_global != exam_status_filter:
                continue
  
            users = users_by_country.get(training_country, [])
            
            # 角色过滤
            if current_role == 'admin':
                users = [u for u in users if u.get('role') not in ['super_admin', 'developer']]
            elif current_role == 'super_admin' and not is_dev:
                users = [u for u in users if u.get('role') != 'developer']
            
            for user in users:
                user_id = user.get('id')
                if not user_id:
                    continue
                
                training_assign_key = f"{training_id_val}_{user_id}"
                exam_assign_key = f"{exam_id}_{user_id}"
                attendance_key = f"{training_id_val}_{user_id}"
                exam_result_key = f"{exam_id}_{user_id}"
                
                has_training_assign = training_assign_key in training_assign_map
                has_exam_assign = exam_assign_key in exam_assign_map

                attendance_data = attendance_map.get(attendance_key, {})
                is_signed = attendance_key in attendance_map
                sign_time = attendance_data.get('sign_time')
                signature_url = attendance_data.get('signature_url')
                
                exam_result = exam_results_map.get(exam_result_key, {})
                is_completed = exam_result_key in exam_results_map
                score = exam_result.get('score')
                is_passed = score is not None and score >= pass_score if is_completed else False

                # 判断是否真正签到（有签名）
                is_signed_completed = is_signed and signature_url

                # 判断是否需要重新签名
                is_resign_needed = is_signed and not signature_url

                # 状态判断
                if is_resign_needed:
                    training_status_display = 'resign'
                elif is_signed_completed:
                    training_status_display = 'signed'
                elif has_training_assign:
                    training_status_display = 'pending'
                else:
                    training_status_display = 'not_assigned'
                
                if is_completed:
                    exam_status_display = 'completed'
                elif has_exam_assign:
                    exam_status_display = 'pending'
                elif is_signed:
                    exam_status_display = 'pending'
                else:
                    exam_status_display = 'not_assigned'

                # 状态筛选
                if status:
                    if status == 'completed' and not is_passed:
                        continue
                    if status == 'incomplete' and (is_completed or is_passed):
                        continue
                    if status == 'failed' and (not is_completed or is_passed):
                        continue
                    if status == 'not_signed' and is_signed:
                        continue
                    if status == 'not_exam' and (is_completed or not is_signed):
                        continue

                # ✅ 新增：计算分配状态
                if training_status_display == 'not_assigned' and exam_status_display == 'not_assigned':
                    allocation_status = '未分配'
                elif training_status_display == 'signed' and exam_status_display in ['pending', 'completed']:
                    allocation_status = '已分配'
                else:
                    allocation_status = '部分分配'
                
                # ✅ 新增：计算完成度
                if training_status_display == 'not_assigned' and exam_status_display == 'not_assigned':
                    completion_level = '未参与'
                elif training_status_display == 'signed' and exam_status_display == 'completed':
                    completion_level = '整体完成'
                else:
                    completion_level = '部分完成'
                
                all_data.append({
                    "training_name": training.get('name', ''),
                    "training_country": training_country,
                    "training_start": training.get('start_time'),
                    "training_end": training.get('end_time'),
                    "training_status_global": training_status_global,
                    "exam_name": exam.get('title', ''),
                    "exam_status_global": exam_status_global,
                    "allocation_status": allocation_status,
                    "completion_level": completion_level,
                    "pass_score": pass_score,
                    "user_name": user.get('name_cn') or user.get('name_en', ''),
                    "user_email": user.get('email', ''),
                    "user_country": user.get('country', ''),
                    "wh_id": user.get('wh_id', ''),
                    "is_signed": is_signed_completed,
                    "is_completed": is_completed,
                    "training_status_display": training_status_display,
                    "exam_status_display": exam_status_display,
                    "is_passed": is_passed,
                    "score": score if score is not None else '-',
                    "signed_at": sign_time,
                    "completed_at": exam_result.get('submitted_at')
                })
        
        # ========== 8. 创建 Excel ==========
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "培训完成度报表"
        
        headers = ['序号', '培训名称', '培训国家', '培训开始日期', '培训结束日期', '培训状态(全局)',
                   '考试名称', '考试状态(全局)', '及格分数', '学员姓名', '学员邮箱', '学员国家', 
                   '库房编码', '签到状态', '考试完成状态', '得分', 
                   '是否及格', '签到时间', '完成时间', '分配状态', '完成度']
        
        # 表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, item in enumerate(all_data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=item.get('training_name', ''))
            ws.cell(row=row_idx, column=3, value=item.get('training_country', ''))
            ws.cell(row=row_idx, column=4, value=item.get('training_start', '')[:10] if item.get('training_start') else '')
            ws.cell(row=row_idx, column=5, value=item.get('training_end', '')[:10] if item.get('training_end') else '')
            ws.cell(row=row_idx, column=6, value=item.get('training_status_global', 'draft'))
            ws.cell(row=row_idx, column=7, value=item.get('exam_name', ''))
            ws.cell(row=row_idx, column=8, value=item.get('exam_status_global', 'draft'))
            ws.cell(row=row_idx, column=9, value=item.get('pass_score', 85))
            ws.cell(row=row_idx, column=10, value=item.get('user_name', ''))
            ws.cell(row=row_idx, column=11, value=item.get('user_email', ''))
            ws.cell(row=row_idx, column=12, value=item.get('user_country', ''))
            ws.cell(row=row_idx, column=13, value=item.get('wh_id', ''))
            ws.cell(row=row_idx, column=14, value='已签到' if item.get('is_signed') else '未签到')
            ws.cell(row=row_idx, column=15, value='已完成' if item.get('is_completed') else '未完成')
            ws.cell(row=row_idx, column=20, value=item.get('allocation_status', ''))  # ✅ 新增
            ws.cell(row=row_idx, column=21, value=item.get('completion_level', ''))    # ✅ 新增
            ws.cell(row=row_idx, column=16, value=item.get('score', '-'))
            ws.cell(row=row_idx, column=17, value='及格' if item.get('is_passed') else ('不及格' if item.get('is_completed') else '-'))
            ws.cell(row=row_idx, column=18, value=convert_time_for_export(item.get('signed_at'), timezone_param))
            ws.cell(row=row_idx, column=19, value=convert_time_for_export(item.get('completed_at'), timezone_param))
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'培训完成度报表_{timestamp}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"导出Excel失败: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@admin_training_bp.route('/api/admin/training/refresh_cache', methods=['POST'])
@login_required
@admin_required
def refresh_report_cache():
    """手动刷新报表缓存 + 数据同步"""
    data = request.json or {}
    training_id = data.get('training_id')
    
    results = {
        "cache_cleared": False,
        "sync_results": {}
    }
    
    # 1. 清除缓存
    if hasattr(get_completion_report, 'cache'):
        get_completion_report.cache.clear()
        results["cache_cleared"] = True
    
    # 2. 执行数据同步
    db = get_supabase_admin()
    
    if training_id:
        # 同步指定培训
        sync_results = _sync_training_binding_data(db, training_id, session.get('user_id'))
        results["sync_results"] = sync_results
    else:
        # 同步所有有绑定关系的培训
        logger.info("开始同步所有培训数据...")
        
        # 获取所有有绑定关系的培训
        bindings_res = db.table("training_exam_bindings").select("training_id").is_("deleted_at", "null").execute()
        training_ids = list(set([b['training_id'] for b in (bindings_res.data or [])]))
        
        total_results = {
            "assigned_exams": 0,
            "created_attendances": 0,
            "created_assignments": 0,
            "errors": [],
            "synced_trainings": len(training_ids)
        }
        
        allowed_countries = get_admin_allowed_countries()
        
        for tid in training_ids:
            # 权限检查
            training_res = db.table("trainings").select("country").eq("id", tid).maybe_single().execute()
            if training_res.data:
                training_country = training_res.data.get('country')
                if allowed_countries is not None and training_country not in allowed_countries:
                    continue
                
                sync_result = _sync_training_binding_data(db, tid, session.get('user_id'))
                total_results["assigned_exams"] += sync_result.get("assigned_exams", 0)
                total_results["created_attendances"] += sync_result.get("created_attendances", 0)
                total_results["created_assignments"] += sync_result.get("created_assignments", 0)
                if sync_result.get("errors"):
                    total_results["errors"].extend(sync_result["errors"])
        
        results["sync_results"] = total_results
    
    return jsonify({
        "success": True,
        "message": "数据已刷新并同步",
        "results": results
    })

# ==================== 培训-考试绑定管理 API ====================

@admin_training_bp.route('/api/admin/training/<int:training_id>/bindings')
@login_required
@admin_required
@cache_get(ttl=300, prefix='training_bindings', include_user=True)
def get_training_bindings(training_id):
    """获取培训绑定的考试列表"""
    db = get_supabase_admin()
    
    # 获取当前用户角色和权限
    current_role = session.get('role')
    is_dev = is_developer()
    allowed_countries = get_admin_allowed_countries()
    
    # 获取培训信息（用于国家过滤）
    training_res = db.table("trainings").select("country, countries").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return jsonify({"error": "培训不存在"}), 404
    
    training = training_res.data
    
    # 解析培训的国家列表（支持多国家）
    training_countries = parse_training_countries(training)
    
    # 如果没有国家信息，使用旧字段
    if not training_countries and training.get('country'):
        training_countries = [training.get('country')]
    
    # 获取该培训下所有已签到学员（用于去重统计）
    signed_users_res = db.table("training_attendances").select("user_id").eq("training_id", training_id).execute()
    signed_user_ids = [u['user_id'] for u in (signed_users_res.data or [])]
    
    # 获取绑定关系
    bindings = db.table("training_exam_bindings").select("*")\
        .eq("training_id", training_id)\
        .is_("deleted_at", "null")\
        .order("sort_order", desc=False)\
        .execute()
    
    # 获取考试信息
    result = []
    for b in (bindings.data or []):
        exam = db.table("exams").select("id, title, countries, created_at, duration, start_time, end_time")\
            .eq("id", b['exam_id'])\
            .maybe_single()\
            .execute()

        # 获取考试状态
        if not exam.data:
            return jsonify({"error": "考试不存在"}), 404
        exam_data = exam.data
        if not can_access_exam(exam_data):
            return jsonify({"error": "无权访问此考试"}), 403
        status = get_exam_status(exam_data)

        # 解析 countries
        countries_list = []
        if exam.data:
            countries_data = exam.data.get('countries')
            if isinstance(countries_data, str):
                try:
                    countries_list = json.loads(countries_data)
                except:
                    countries_list = []
            elif isinstance(countries_data, list):
                countries_list = countries_data

        # 计算该考试已分配的用户数（去重）
        assigned_res = db.table("exam_assignments").select("user_id").eq("exam_id", b['exam_id']).execute()
        assigned_user_ids = [a['user_id'] for a in (assigned_res.data or [])]
        
        # 计算该考试已完成的用户数（去重）
        completed_res = db.table("exam_results").select("user_id").eq("exam_id", b['exam_id']).execute()
        completed_user_ids = [r['user_id'] for r in (completed_res.data or [])]
        
        # 计算该考试已完成且已签到培训的用户数（交集）
        completed_and_signed = set(completed_user_ids) & set(signed_user_ids)
        
        result.append({
            "id": b['id'],
            "training_id": b['training_id'],
            "exam_id": b['exam_id'],
            "exam_title": exam.data['title'] if exam.data else '未知考试',
            "exam_countries": countries_list,
            "pass_score": b.get('pass_score', 85),
            "is_auto_assign": b.get('is_auto_assign', True),
            "is_required": b.get('is_required', True),
            "sort_order": b.get('sort_order', 0),
            "created_at": b.get('created_at'),
            # ✅ 新增统计字段
            "exam_status": status,
            "assigned_count": len(set(assigned_user_ids)),           # 已分配人数（去重）
            "completed_count": len(set(completed_user_ids)),         # 已完成人数（去重）
            "signed_count": len(set(signed_user_ids)),               # 已签到人数（去重）
            "completed_and_signed_count": len(completed_and_signed)  # 已完成且已签到人数
        })
    
    # 获取可选考试列表（未绑定的）- 根据角色过滤
    all_exams = db.table("exams").select("id, title, created_at, countries, start_time, end_time")\
        .is_("deleted_at", "null")\
        .execute()
    
    bound_ids = [b['exam_id'] for b in (bindings.data or [])]
    available_exams = []
    
    for e in (all_exams.data or []):
        if e['id'] in bound_ids:
            continue
        
        # 解析考试的国家列表
        exam_countries = []
        countries_data = e.get('countries')
        if isinstance(countries_data, str):
            try:
                exam_countries = json.loads(countries_data)
            except:
                exam_countries = []
        elif isinstance(countries_data, list):
            exam_countries = countries_data
        
        # ========== 角色权限过滤 ==========
        can_see = False
        
        if is_dev:
            # 开发者可以看到所有考试
            can_see = True
        elif current_role == 'super_admin':
            # 超管：根据权限范围过滤
            if allowed_countries is not None and allowed_countries:
                # 检查考试国家是否在权限范围内
                if any(c in allowed_countries for c in exam_countries):
                    can_see = True
            else:
                # 无权限限制，可以看到所有
                can_see = True
        elif current_role == 'admin':
            # 管理员：根据权限范围或自己国家过滤
            if allowed_countries is not None and allowed_countries:
                if any(c in allowed_countries for c in exam_countries):
                    can_see = True
            else:
                # 无权限范围，使用用户注册国家
                user_country = session.get('user_country')
                if user_country and user_country in exam_countries:
                    can_see = True
        
        # 考试国家必须与培训国家有交集（支持多国家）
        if can_see and training_countries:
            # 检查考试国家是否与培训国家有交集
            has_intersection = any(c in training_countries for c in exam_countries)
            if not has_intersection:
                can_see = False

        if can_see:
            available_exams.append({
                "id": e['id'],
                "title": e['title'],
                "countries": exam_countries,
                "status": get_exam_status(e),  # 🔥 添加状态
                "created_at": e.get('created_at'),  # 🔥 添加创建时间
                "duration": e.get('duration'),  # 🔥 添加时长
                "question_count": 0  # 可选，可以计算或保留
            })
    
    return jsonify({
        "bindings": result,
        "available_exams": available_exams
    })

@admin_training_bp.route('/api/admin/training/bind_exam', methods=['POST'])
@login_required
@admin_required
def bind_exam_to_training():
    """绑定考试到培训"""
    logger.info("=" * 60)
    logger.info("🔥 bind_exam_to_training 被调用了！")
    logger.info("=" * 60)

    data = request.json
    logger.info(f"📥 接收到的数据: {data}")
    training_id = data.get('training_id')
    exam_id = data.get('exam_id')
    pass_score = data.get('pass_score', 85)
    is_auto_assign = data.get('is_auto_assign', True)
    is_required = data.get('is_required', True)
    sort_order = data.get('sort_order', 0)
    now = datetime.now(timezone.utc).isoformat()
    operator_id = session.get('user_id')
    db = get_supabase_admin()
    
    # ========== 参数验证 ==========
    if not training_id:
        logger.error("❌ training_id 缺失")
        return jsonify({"success": False, "message": "培训ID不能为空"}), 400
    
    if not exam_id:
        logger.error("❌ exam_id 缺失")
        return jsonify({"success": False, "message": "考试ID不能为空"}), 400
 
    # ========== 检查考试状态 ==========
    exam_check = db.table("exams").select("status, title").eq("id", exam_id).maybe_single().execute()
    if exam_check.data:
        exam_status = exam_check.data.get('status', 'draft')
        exam_title = exam_check.data.get('title', f'考试{exam_id}')
        
        if exam_status == 'draft':
            logger.warning(f"⚠️ 尝试绑定草稿考试: {exam_id}")
            # 允许绑定，但记录警告（或者可以选择阻止绑定）
            # 这里选择允许绑定，但在补分配时阻止
            pass

        # 使用考试自身的及格分数作为默认值
        default_pass_score = exam_check.data.get('pass_score', 85)
    else:
        default_pass_score = 85
    
    # 如果前端没有传递 pass_score，使用考试默认值
    pass_score = data.get('pass_score', default_pass_score)

    # ========== 权限验证 ==========
    # 检查当前用户是否有权限操作此培训
    training_check = db.table("trainings").select("country").eq("id", training_id).maybe_single().execute()
    if not training_check.data:
        return jsonify({"success": False, "message": "培训不存在"}), 404
    
    allowed_countries = get_admin_allowed_countries()
    training_country = training_check.data.get('country')
    if allowed_countries is not None and training_country and training_country not in allowed_countries:
        logger.warning(f"⚠️ 用户无权操作培训 {training_id} (国家: {training_country})")
        return jsonify({"success": False, "message": "无权操作此培训"}), 403

    training_data = training_check.data
    
    # ========== 检查是否存在软删除的记录 ==========
    soft_deleted = db.table("training_exam_bindings").select("*")\
        .eq("training_id", training_id)\
        .eq("exam_id", exam_id)\
        .not_.is_("deleted_at", "null")\
        .execute()
    
    if soft_deleted and soft_deleted.data:
        # 恢复软删除的记录
        result = db.table("training_exam_bindings").update({
            "deleted_at": None,
            "deleted_by": None,
            "pass_score": pass_score,
            "is_auto_assign": is_auto_assign,
            "is_required": is_required,
            "sort_order": sort_order,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": operator_id
        }).eq("id", soft_deleted.data[0]['id']).execute()

        if not result.data:
            return jsonify({"success": False, "message": "恢复绑定失败"}), 500
        
        binding_id = result.data[0]['id']
        logger.info(f"✅ 恢复绑定成功: binding_id={binding_id}")

        # 恢复绑定后，只有非草稿培训才激活
        start_time = training_data.get('start_time')
        end_time = training_data.get('end_time')
        current_status = training_data.get('dynamic_status', 'draft')
        
        if start_time and end_time and current_status != 'draft':
            db.table("trainings").update({
                "is_active": True
            }).eq("id", training_id).execute()
            logger.info(f"✅ 培训 {training_id} 恢复绑定后已激活 (is_active=True)")
        else:
            logger.info(f"⚠️ 培训 {training_id} 为草稿状态，恢复绑定后保持 is_active=False")
        
        # ========== 使用统一的缓存清除函数 ==========
        clear_training_related_cache(training_id)
        clear_all_assignment_caches(training_id=training_id, exam_id=exam_id)
        logger.info(f"🧹 已清除培训 {training_id} 和考试 {exam_id} 的相关缓存")
        
        # ========== 补分配考试给已签到学员 ==========
        sync_result = _auto_assign_exam_to_signed_users_safe(db, training_id, exam_id, is_auto_assign, operator_id, now)
        
        # ========== 获取考试标题 ==========
        exam_title = _get_exam_title(db, exam_id)
        
        return jsonify({
            "success": True, 
            "binding_id": binding_id, 
            "restored": True,
            "sync_result": sync_result,
            "training_id": training_id,
            "exam_id": exam_id,
            "exam_title": exam_title
        })
    
    # ========== 检查是否已有活跃绑定 ==========
    existing = db.table("training_exam_bindings").select("id")\
        .eq("training_id", training_id)\
        .eq("exam_id", exam_id)\
        .is_("deleted_at", "null")\
        .execute()
    
    if existing.data:
        logger.warning(f"⚠️ 考试 {exam_id} 已绑定到培训 {training_id}")
        return jsonify({"success": False, "message": "该考试已绑定到此培训"}), 400
    
    # ========== 创建新绑定 ==========
    result = db.table("training_exam_bindings").insert({
        "training_id": training_id,
        "exam_id": exam_id,
        "pass_score": pass_score,
        "is_auto_assign": is_auto_assign,
        "is_required": is_required,
        "sort_order": sort_order,
        "created_at": now,
        "created_by": operator_id
    }).execute()
    
    if not result.data:
        logger.error(f"❌ 绑定失败: training_id={training_id}, exam_id={exam_id}")
        return jsonify({"success": False, "message": "绑定失败"}), 500
    
    binding_id = result.data[0]['id']
    logger.info(f"✅ 绑定成功: binding_id={binding_id}")

    # 只有非草稿培训才激活
    start_time = training_data.get('start_time')
    end_time = training_data.get('end_time')
    current_status = training_data.get('dynamic_status', 'draft')
    
    if start_time and end_time and current_status != 'draft':
        db.table("trainings").update({
            "is_active": True
        }).eq("id", training_id).execute()
        logger.info(f"✅ 培训 {training_id} 已设置有效期，绑定考试后激活 (is_active=True)")
    else:
        logger.info(f"⚠️ 培训 {training_id} 为草稿状态，绑定考试后保持 is_active=False，等待管理员设置有效期")

    # ========== 使用统一的缓存清除函数 ==========
    clear_training_related_cache(training_id)
    clear_all_assignment_caches(training_id=training_id, exam_id=exam_id)
    logger.info(f"🧹 已清除培训 {training_id} 和考试 {exam_id} 的相关缓存")
    
    # ========== 获取考试标题 ==========
    exam_title = _get_exam_title(db, exam_id)

    # ========== 补分配考试给已签到学员 ==========
    sync_result = _auto_assign_exam_to_signed_users_safe(db, training_id, exam_id, is_auto_assign, operator_id, now)
    
    return jsonify({
        "success": True, 
        "binding_id": binding_id,
        "restored": False,
        "training_id": training_id,
        "exam_id": exam_id,
        "exam_title": exam_title,
        "sync_result": sync_result
    })

def _get_exam_title(db, exam_id):
    """获取考试标题"""
    try:
        res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
        if res.data:
            return res.data.get('title', f'考试{exam_id}')
        return f'考试{exam_id}'
    except:
        return f'考试{exam_id}'

def _auto_assign_exam_to_signed_users_safe(db, training_id, exam_id, is_auto_assign, operator_id, now):
    """
    为已签到学员分配考试（安全版：跳过已分配的学员）
    """
    if not is_auto_assign:
        logger.info(f"is_auto_assign=False，跳过自动分配")
        return {"assigned_count": 0, "skipped_count": 0}
    
    # 获取已签到学员
    attendances_res = db.table("training_attendances").select("user_id")\
        .eq("training_id", training_id)\
        .is_("deleted_at", "null")\
        .execute()
    signed_user_ids = [a['user_id'] for a in (attendances_res.data or [])]
    
    if not signed_user_ids:
        logger.info(f"培训 {training_id} 无已签到学员")
        return {"assigned_count": 0, "skipped_count": 0}

    # 如果有已签到学员，说明培训已经在使用了，更新 is_active
    db.table("trainings").update({
        "is_active": True
    }).eq("id", training_id).execute()
    logger.info(f"✅ 培训 {training_id} 有已签到学员，自动激活 (is_active=True)")
    
    # ✅ 检查哪些学员已有该考试的分配记录
    existing_res = db.table("exam_assignments").select("user_id")\
        .eq("exam_id", exam_id)\
        .in_("user_id", signed_user_ids)\
        .is_("deleted_at", "null")\
        .execute()
    existing_user_ids = set([r['user_id'] for r in (existing_res.data or [])])
    
    # 过滤出未分配的学员
    to_assign_user_ids = [uid for uid in signed_user_ids if uid not in existing_user_ids]
    skipped_count = len(signed_user_ids) - len(to_assign_user_ids)
    
    if not to_assign_user_ids:
        logger.info(f"考试 {exam_id} 所有已签到学员已分配，跳过")
        return {"assigned_count": 0, "skipped_count": skipped_count}
    
    # 构建插入数据
    to_assign_exams = [{
        "exam_id": exam_id,
        "user_id": uid,
        "assigned_at": now,
        "created_by": operator_id or 'system'
    } for uid in to_assign_user_ids]
    
    # 批量插入
    try:
        insert_result = db.table("exam_assignments").insert(to_assign_exams).execute()
        assigned_count = len(insert_result.data or [])
        logger.info(f"✅ 为 {assigned_count} 名已签到学员分配考试，跳过 {skipped_count} 名已分配学员")
        return {"assigned_count": assigned_count, "skipped_count": skipped_count}
    except Exception as e:
        # 如果批量插入失败，逐条插入
        logger.warning(f"批量插入失败，尝试逐条插入: {e}")
        success_count = 0
        for item in to_assign_exams:
            try:
                db.table("exam_assignments").insert(item).execute()
                success_count += 1
            except Exception as inner_e:
                if '23505' in str(inner_e) or 'duplicate' in str(inner_e).lower():
                    logger.warning(f"学员 {item['user_id']} 已有分配记录，跳过")
                    skipped_count += 1
                else:
                    logger.error(f"插入失败: {inner_e}")
        return {"assigned_count": success_count, "skipped_count": skipped_count}


def _sync_training_binding_data(db, training_id, created_by=None):
    """
    培训绑定考试后的数据同步：双向同步
    1. 已签到学员 → 自动分配考试
    2. 已完成考试学员 → 自动创建签到记录
    
    Args:
        db: 数据库连接
        training_id: 培训ID
        created_by: 创建人ID（用于记录）
    """
    now = datetime.now(timezone.utc).isoformat()
    sync_results = {
        "assigned_exams": 0,      # 已签到→分配考试 的数量
        "created_attendances": 0,  # 已完成考试→创建签到 的数量
        "created_assignments": 0,  # 已完成考试→创建培训分配 的数量
        "errors": []
    }
    
    logger.info("=" * 60)
    logger.info(f"🔄 开始培训绑定数据同步: training_id={training_id}")
    
    try:
        # ========== 1. 获取培训信息 ==========
        training_res = db.table("trainings").select("country, name, start_time, end_time, is_active").eq("id", training_id).maybe_single().execute()
        if not training_res.data:
            logger.warning(f"培训不存在: training_id={training_id}")
            return sync_results
        
        training = training_res.data
        
        # ========== 2. 获取该培训绑定的所有考试 ==========
        bindings_res = db.table("training_exam_bindings").select("exam_id, is_auto_assign")\
            .eq("training_id", training_id)\
            .is_("deleted_at", "null")\
            .execute()
        bindings = bindings_res.data or []
        
        if not bindings:
            logger.info("该培训暂无绑定考试，跳过同步")
            return sync_results
        
        exam_ids = [b['exam_id'] for b in bindings]
        auto_assign_exam_ids = [b['exam_id'] for b in bindings if b.get('is_auto_assign', True)]
        
        logger.info(f"绑定的考试: {exam_ids}")
        
        # ========== 3. 获取该培训国家下的所有用户 ==========
        training_country = training.get('country')
        if not training_country:
            logger.warning("培训未指定国家，跳过同步")
            return sync_results
        
        users_res = db.table("users").select("id, name_en, email")\
            .eq("country", training_country)\
            .eq("user_status", "registered")\
            .eq("is_resign", False)\
            .is_("deleted_at", "null")\
            .execute()
        all_users = users_res.data or []
        all_user_ids = [u['id'] for u in all_users]
        
        if not all_user_ids:
            logger.info("该国家无有效用户，跳过同步")
            return sync_results
        
        logger.info(f"培训国家 {training_country} 下的用户数: {len(all_user_ids)}")
        
        # ========== 4. 获取已签到学员 ==========
        attendances_res = db.table("training_attendances").select("user_id")\
            .eq("training_id", training_id)\
            .is_("deleted_at", "null")\
            .execute()
        signed_user_ids = {a['user_id'] for a in (attendances_res.data or [])}
        logger.info(f"已签到学员数: {len(signed_user_ids)}")
        
        # ========== 5. 获取已完成考试的学员（针对绑定的考试） ==========
        completed_users_per_exam = {}
        if exam_ids:
            results_res = db.table("exam_results").select("user_id, exam_id")\
                .in_("exam_id", exam_ids)\
                .in_("user_id", all_user_ids)\
                .is_("deleted_at", "null")\
                .execute()
            
            for r in (results_res.data or []):
                exam_id = r['exam_id']
                user_id = r['user_id']
                if exam_id not in completed_users_per_exam:
                    completed_users_per_exam[exam_id] = set()
                completed_users_per_exam[exam_id].add(user_id)
        
        # ========== 6. 获取已分配考试 ==========
        assigned_exam_map = {}
        if exam_ids:
            assign_res = db.table("exam_assignments").select("exam_id, user_id")\
                .in_("exam_id", exam_ids)\
                .in_("user_id", all_user_ids)\
                .is_("deleted_at", "null")\
                .execute()
            
            for a in (assign_res.data or []):
                key = f"{a['exam_id']}_{a['user_id']}"
                assigned_exam_map[key] = True
        
        # ========== 7. 获取培训分配记录 ==========
        training_assign_res = db.table("training_assignments").select("user_id")\
            .eq("training_id", training_id)\
            .in_("user_id", all_user_ids)\
            .execute()
        training_assigned_set = {a['user_id'] for a in (training_assign_res.data or [])}
        
        # ========== 8. 双向数据同步 ==========
        
        # 8.1 已签到但未分配考试 → 自动分配考试
        if signed_user_ids and auto_assign_exam_ids:
            to_assign_exams = []
            skipped_count = 0
            
            for exam_id in auto_assign_exam_ids:
                for user_id in signed_user_ids:
                    key = f"{exam_id}_{user_id}"
                    if key not in assigned_exam_map:
                        to_assign_exams.append({
                            "exam_id": exam_id,
                            "user_id": user_id,
                            "created_by": created_by or 'system',
                            "assigned_at": now
                        })
                    else:
                        skipped_count += 1
            
            if to_assign_exams:
                try:
                    insert_result = db.table("exam_assignments").insert(to_assign_exams).execute()
                    sync_results["assigned_exams"] = len(insert_result.data or [])
                except Exception as e:
                    # 如果批量插入失败，逐条插入
                    logger.warning(f"批量插入失败，尝试逐条插入: {e}")
                    success_count = 0
                    for item in to_assign_exams:
                        try:
                            db.table("exam_assignments").insert(item).execute()
                            success_count += 1
                        except Exception as inner_e:
                            if '23505' in str(inner_e) or 'duplicate' in str(inner_e).lower():
                                skipped_count += 1
                            else:
                                logger.error(f"插入失败: {inner_e}")
                    sync_results["assigned_exams"] = success_count
            
            sync_results["skipped_existing"] = skipped_count
            logger.info(f"✅ 为已签到学员分配考试: {sync_results['assigned_exams']} 条，跳过 {skipped_count} 条")
        
        # 8.2 已完成考试但未签到 → 自动创建签到记录
        if exam_ids:
            # 收集所有已完成考试的用户
            all_completed_users = set()
            for exam_id, users in completed_users_per_exam.items():
                all_completed_users.update(users)
            
            # 过滤出未签到的用户
            to_create_attendance = [uid for uid in all_completed_users if uid not in signed_user_ids]
            
            if to_create_attendance:
                # 创建签到记录
                attendance_records = [{
                    "training_id": training_id,
                    "user_id": uid,
                    "sign_time": now,
                    "signed_name": next((u.get('name_en', '自动签到') for u in all_users if u['id'] == uid), '自动签到'),
                    "signature_url": None  # 无签名
                } for uid in to_create_attendance]
                
                insert_result = db.table("training_attendances").insert(attendance_records).execute()
                sync_results["created_attendances"] = len(insert_result.data or [])
                logger.info(f"✅ 为已完成考试学员创建签到: {sync_results['created_attendances']} 条")
                
                # 同时创建培训分配记录
                to_create_assignment = [uid for uid in to_create_attendance if uid not in training_assigned_set]
                if to_create_assignment:
                    assignment_records = [{
                        "training_id": training_id,
                        "user_id": uid,
                        "created_by": created_by or 'system',
                        "created_at": now
                    } for uid in to_create_assignment]
                    
                    insert_result = db.table("training_assignments").insert(assignment_records).execute()
                    sync_results["created_assignments"] = len(insert_result.data or [])
                    logger.info(f"✅ 为已完成考试学员创建培训分配: {sync_results['created_assignments']} 条")
        
        logger.info(f"🔄 数据同步完成: {sync_results}")
        
    except Exception as e:
        logger.error(f"数据同步失败: {e}", exc_info=True)
        sync_results["errors"].append(str(e))
    
    return sync_results




def _auto_assign_exam_to_signed_users(db, training_id, exam_id, is_auto_assign, created_by, assigned_at):
    """
    为培训的已签到学员自动分配考试（公共函数）
    
    Args:
        db: 数据库连接
        training_id: 培训ID
        exam_id: 考试ID
        is_auto_assign: 是否自动分配
        created_by: 创建人ID
        assigned_at: 分配时间
    """
    if not is_auto_assign:
        logger.info(f"is_auto_assign=False，跳过补分配 (exam_id={exam_id})")
        return
    
    logger.info(f"进入补分配逻辑: training_id={training_id}, exam_id={exam_id}")
    
    try:
        # 1. 获取该培训下所有已签到的学员
        signed_users_res = db.table("training_attendances").select("user_id").eq("training_id", training_id).execute()
        signed_user_ids = [u['user_id'] for u in (signed_users_res.data or [])]
        logger.info(f"已签到学员数: {len(signed_user_ids)}")
        
        if not signed_user_ids:
            logger.info("该培训暂无已签到学员，跳过补分配")
            return
        
        # 2. 过滤出尚未分配该考试的用户
        assigned_res = db.table("exam_assignments").select("user_id")\
            .eq("exam_id", exam_id)\
            .in_("user_id", signed_user_ids)\
            .is_("deleted_at", "null")\
            .execute()
        assigned_set = {a['user_id'] for a in (assigned_res.data or [])}
        
        to_assign = [uid for uid in signed_user_ids if uid not in assigned_set]
        logger.info(f"需要补分配的学员数: {len(to_assign)}")
        
        if not to_assign:
            logger.info("所有已签到学员已分配该考试")
            return
        
        # 3. 批量插入考试分配
        assignments = [{
            "exam_id": exam_id,
            "user_id": uid,
            "created_by": created_by,
            "assigned_at": assigned_at
        } for uid in to_assign]
        
        insert_result = db.table("exam_assignments").insert(assignments).execute()
        logger.info(f"✅ 为已签到的 {len(insert_result.data or [])} 名学员补分配考试: exam_id={exam_id}")
        
    except Exception as e:
        logger.error(f"补分配考试失败: training_id={training_id}, exam_id={exam_id}, error={e}", exc_info=True)

@admin_training_bp.route('/api/admin/training/binding/<int:binding_id>', methods=['PUT'])
@login_required
@admin_required
def update_training_binding(binding_id):
    """更新绑定配置"""
    data = request.json
    db = get_supabase_admin()
    
    update_data = {}
    if 'pass_score' in data:
        new_pass_score = data['pass_score']
        update_data['pass_score'] = new_pass_score
        sync_exam = True

    if 'is_auto_assign' in data:
        update_data['is_auto_assign'] = data['is_auto_assign']
    if 'is_required' in data:
        update_data['is_required'] = data['is_required']
    if 'sort_order' in data:
        update_data['sort_order'] = data['sort_order']
    
    if not update_data:
        return jsonify({"success": False, "message": "没有要更新的字段"}), 400
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_data['updated_by'] = session.get('user_id')
    
    result = db.table("training_exam_bindings").update(update_data)\
        .eq("id", binding_id)\
        .is_("deleted_at", "null")\
        .execute()
    
    if result.data:
        # 如果及格分数被更新，尝试反向同步到考试表
        if sync_exam and new_pass_score is not None:
            sync_result = sync_binding_pass_score_to_exam(binding_id, new_pass_score)
            logger.info(f"绑定 {binding_id} 及格分数反向同步结果: {sync_result}")
        
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "更新失败"}), 500

@admin_training_bp.route('/api/admin/training/binding/<int:binding_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_training_binding(binding_id):
    """解除绑定（硬删除）"""
    db = get_supabase_admin()
    
    # ========== 先获取绑定信息（用于缓存清除和日志） ==========
    binding_res = db.table("training_exam_bindings").select("training_id, exam_id").eq("id", binding_id).maybe_single().execute()
    if not binding_res.data:
        return jsonify({"success": False, "message": "绑定关系不存在"}), 404
    
    training_id = binding_res.data.get('training_id')
    exam_id = binding_res.data.get('exam_id')
    
    # ========== 权限验证 ==========
    training_check = db.table("trainings").select("country").eq("id", training_id).maybe_single().execute()
    if training_check.data:
        allowed_countries = get_admin_allowed_countries()
        training_country = training_check.data.get('country')
        if allowed_countries is not None and training_country and training_country not in allowed_countries:
            return jsonify({"success": False, "message": "无权操作此培训"}), 403
    
    # ========== 硬删除记录 ==========
    result = db.table("training_exam_bindings").delete().eq("id", binding_id).execute()
    
    if result.data:
        # ========== 使用统一的缓存清除函数 ==========
        if training_id:
            clear_training_related_cache(training_id)
            if exam_id:
                clear_all_assignment_caches(training_id=training_id, exam_id=exam_id)
            else:
                clear_all_assignment_caches(training_id=training_id)
            logger.info(f"🧹 已清除培训 {training_id} 的相关缓存")
        
        logger.info(f"✅ 解除绑定成功: binding_id={binding_id}")
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "解除绑定失败"}), 500

# ==================== 培训学员分配管理 API ====================

@admin_training_bp.route('/api/admin/training/<int:training_id>/assign', methods=['POST'])
@login_required
@admin_required
def assign_training_to_users(training_id):
    """分配培训给指定学员（定点推送）"""
    data = request.json
    user_ids = data.get('user_ids', [])
    
    if not user_ids:
        return jsonify({"success": False, "message": "jsonify_please_select_users", "params": []}), 400
    
    db = get_supabase_admin()  # 需要绕过 RLS
    
    # 获取当前培训信息（用于日志）
    training_res = db.table("trainings").select("name").eq("id", training_id).maybe_single().execute()
    training_name = training_res.data.get('name', '') if training_res.data else ''
    
    assigned_count = 0
    skipped_count = 0
    
    for uid in user_ids:
        # 检查是否已分配
        existing = db.table("training_assignments").select("id").eq("training_id", training_id).eq("user_id", uid).execute()
        if not existing.data:
            db.table("training_assignments").insert({
                "training_id": training_id,
                "user_id": uid,
                "created_by": session.get('user_id')
            }).execute()
            assigned_count += 1
        else:
            skipped_count += 1
    
    logger.info(f"培训 {training_id} ({training_name}) 分配给了 {assigned_count} 名学员，跳过 {skipped_count} 名已分配")
    
    return jsonify({
        "success": True, 
        "assigned_count": assigned_count,
        "skipped_count": skipped_count,
        "message": "jsonify_assign_success",
        "params": [assigned_count]
    })


@admin_training_bp.route('/api/admin/training/<int:training_id>/unassign', methods=['POST'])
@login_required
@admin_required
def unassign_training_from_users(training_id):
    """取消分配培训"""
    data = request.json
    user_ids = data.get('user_ids', [])
    
    if not user_ids:
        return jsonify({"success": False, "message": "jsonify_please_select_users", "params": []}), 400
    
    db = get_supabase_admin()  # 需要绕过 RLS
    
    # 获取当前培训信息
    training_res = db.table("trainings").select("name").eq("id", training_id).maybe_single().execute()
    training_name = training_res.data.get('name', '') if training_res.data else ''
    
    # 删除分配记录
    result = db.table("training_assignments").delete().eq("training_id", training_id).in_("user_id", user_ids).execute()
    deleted_count = len(result.data or [])
    
    # 同时删除这些学员的签到记录（如果有）
    db.table("training_attendances").delete().eq("training_id", training_id).in_("user_id", user_ids).execute()
    
    logger.info(f"培训 {training_id} ({training_name}) 取消了 {deleted_count} 名学员的分配")

    log_training_unassign(
        training_id=training_id,
        training_name=training_name,
        user_id=user_id,
        user_name=user_name,
        admin_id=session.get('user_id')
    )
    return jsonify({
        "success": True, 
        "unassigned_count": deleted_count,
        "message": "jsonify_unassign_success",
        "params": [deleted_count]
    })


@admin_training_bp.route('/api/admin/training/<int:training_id>/resign/<user_id>', methods=['POST'])
@login_required
@admin_required
def admin_force_resign_training(training_id, user_id):
    """管理员强制学员重新签到（清空已有签到记录，保留分配关系）"""
    db = get_supabase_admin()
    
    # 检查培训是否存在
    training_res = db.table("trainings").select("name").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return jsonify({"success": False, "message": "jsonify_training_not_exist", "params": []}), 404
    
    # 检查用户是否存在
    user_res = db.table("users").select("name_en").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "jsonify_user_not_exist", "params": []}), 404
    
    training_name = training_res.data.get('name', '')
    user_name = user_res.data.get('name_en', '')
    
    # 删除该学员的签到记录（保留分配关系）
    db.table("training_attendances").delete().eq("training_id", training_id).eq("user_id", user_id).execute()
    
    logger.info(f"培训 {training_id} ({training_name}) 学员 {user_name} ({user_id}) 被强制重新签到")
    
    return jsonify({
        "success": True,
        "message": "jsonify_resign_success",
        "params": [user_name]
    })


@admin_training_bp.route('/api/admin/training/<int:training_id>/users')
@login_required
@admin_required
def get_training_users(training_id):
    """获取培训相关的所有学员及其分配/签到状态（用于状态查看页面）"""
    db = get_supabase()
    admin_db = get_supabase_admin()
    
    # 获取培训信息
    training_res = db.table("trainings").select("countries, country, name").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return jsonify({"data": []})
    
    training = training_res.data
    training_name = training.get('name', '')

    training_countries = parse_training_countries(training)
    if not training_countries:
        return jsonify({"data": []})

    # 获取当前管理员权限范围
    allowed_countries = get_admin_allowed_countries()
    
    # 权限过滤：只显示权限范围内的国家
    if allowed_countries is not None:
        filtered_countries = [c for c in training_countries if c in allowed_countries]
        if not filtered_countries:
            return jsonify({"data": []})
        training_countries = filtered_countries
    
    # 获取该国家下的所有学员
    users_res = db.table("users").select("id, name_en, email, country, wh_id, wh_name_en, company") \
        .in_("country", training_countries) \
        .eq("user_status", "registered") \
        .eq("is_resign", False) \
        .execute()
    users = users_res.data or []
    user_ids = [u['id'] for u in users]
    
    if not user_ids:
        return jsonify({"data": []})
    
    # 获取该培训的分配记录（使用 admin 客户端）
    assign_res = admin_db.table("training_assignments").select("user_id").eq("training_id", training_id).execute()
    assigned_user_ids = {a['user_id'] for a in (assign_res.data or [])}
    
    # 获取签到记录
    att_res = db.table("training_attendances").select("user_id, sign_time, signed_name, signature_url") \
        .eq("training_id", training_id) \
        .in_("user_id", user_ids) \
        .execute()
    
    attendance_map = {}
    for att in (att_res.data or []):
        attendance_map[att['user_id']] = att
    
    # 构建返回数据
    result = []
    for user in users:
        user_id = user['id']
        attendance = attendance_map.get(user_id, {})
        sign_time = attendance.get('sign_time')
        signature_url = attendance.get('signature_url', '')
        has_attendance = user_id in attendance_map
        
        # 优先检查是否有签到记录（全国推送场景）
        if has_attendance:
            # 有签到记录
            if not signature_url:
                sign_status = 'resign'  # 已签到但无签名（需重新签名）
            else:
                sign_status = 'signed'  # 已签到且有签名
        elif user_id in assigned_user_ids:
            # 有分配但无签到
            sign_status = 'pending'  # 待签到
        else:
            # 既无分配也无签到
            sign_status = 'not_assigned'
        
        result.append({
            "training_id": training_id,
            "training_name": training_res.data.get('name', ''),
            "training_country": training_country,
            "user_id": user_id,
            "user_country": user.get('country', ''),
            "name_en": user.get('name_en', ''),
            "email": user.get('email', ''),
            "wh_id": user.get('wh_id', ''),
            "wh_name_en": user.get('wh_name_en', ''),
            "company": user.get('company', ''),
            "sign_status": sign_status,
            "sign_time": sign_time,
            "signed_name": attendance.get('signed_name', ''),
            "signature_url": signature_url
        })
    
    return jsonify({"data": result})

# =====================添加补分配接口，用于已有签到的培训，再绑定考试时为已签到客户补分配考试=====================
@admin_training_bp.route('/api/admin/training/<int:training_id>/binding/<int:binding_id>/reassign', methods=['POST'])
@login_required
@admin_required
def reassign_training_binding(training_id, binding_id):
    """
    为培训的某个绑定关系，将所有已签到学员补分配考试
    """
    db = get_supabase_admin()
    current_user_id = session.get('user_id')
    now = datetime.now(timezone.utc).isoformat()
    
    logger.info("=" * 60)
    logger.info(f"📌 补分配考试: training_id={training_id}, binding_id={binding_id}")
    
    try:
        # ========== 1. 参数验证 ==========
        if not training_id or not binding_id:
            logger.error("❌ 参数缺失")
            return jsonify({
                "success": False, 
                "message": "培训ID和绑定ID不能为空"
            }), 400
        
        # ========== 2. 权限验证 ==========
        training_check = db.table("trainings").select("country, name").eq("id", training_id).maybe_single().execute()
        if not training_check.data:
            logger.error(f"❌ 培训不存在: training_id={training_id}")
            return jsonify({"success": False, "message": "培训不存在"}), 404
        
        allowed_countries = get_admin_allowed_countries()
        training_country = training_check.data.get('country')
        if allowed_countries is not None and training_country and training_country not in allowed_countries:
            logger.warning(f"⚠️ 用户无权操作培训 {training_id}")
            return jsonify({"success": False, "message": "无权操作此培训"}), 403
        
        # ========== 3. 获取绑定信息 ==========
        binding_res = db.table("training_exam_bindings").select("*")\
            .eq("id", binding_id)\
            .eq("training_id", training_id)\
            .is_("deleted_at", "null")\
            .maybe_single()\
            .execute()
        
        if not binding_res.data:
            logger.error(f"❌ 绑定关系不存在: binding_id={binding_id}")
            return jsonify({"success": False, "message": "绑定关系不存在"}), 404
        
        binding = binding_res.data
        exam_id = binding.get('exam_id')
        is_auto_assign = binding.get('is_auto_assign', True)
        
        # ========== 4. 检查自动分配开关 ==========
        if not is_auto_assign:
            logger.info(f"⚠️ 考试未启用自动分配: exam_id={exam_id}")
            return jsonify({
                "success": False, 
                "message": "该考试未启用自动分配，无需补分配"
            }), 400
        
        # ========== 5. 核心新增：检查考试状态 ==========
        exam_res = db.table("exams").select("title, status, start_time, end_time, countries")\
            .eq("id", exam_id)\
            .maybe_single()\
            .execute()
        
        if not exam_res.data:
            logger.error(f"❌ 考试不存在: exam_id={exam_id}")
            return jsonify({
                "success": False, 
                "message": f"考试 (ID: {exam_id}) 不存在"
            }), 404
        
        exam = exam_res.data
        exam_title = exam.get('title', f'考试{exam_id}')
        exam_status = exam.get('status', 'draft')
        
        # 检查考试状态 - 只有非草稿状态才允许补分配
        if exam_status == 'draft':
            logger.warning(f"⚠️ 考试 {exam_id} 是草稿状态，不允许补分配")
            return jsonify({
                "success": False,
                "message": f"考试「{exam_title}」当前为草稿状态，请先发布考试后再进行补分配",
                "exam_status": "draft",
                "exam_id": exam_id,
                "exam_title": exam_title
            }), 400
        
        # 检查考试是否已结束（可选，但建议允许补分配已结束的考试）
        # 如果考试已结束，可以允许补分配，但给出提示
        now_dt = datetime.now(timezone.utc)
        end_time = exam.get('end_time')
        is_closed = False
        if end_time:
            try:
                end_dt = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
                if now_dt > end_dt:
                    is_closed = True
            except:
                pass
        
        # ========== 6. 获取已签到学员 ==========
        signed_res = db.table("training_attendances").select("user_id")\
            .eq("training_id", training_id)\
            .is_("deleted_at", "null")\
            .execute()
        signed_user_ids = [u['user_id'] for u in (signed_res.data or [])]
        
        if not signed_user_ids:
            logger.info(f"ℹ️ 该培训暂无已签到学员: training_id={training_id}")
            return jsonify({
                "success": True,
                "message": "该培训暂无已签到学员",
                "assigned_count": 0
            })
        
        # ========== 7. 获取完整的用户信息并过滤 ==========
        users_res = db.table("users").select("id, name_en, role, country, is_resign")\
            .in_("id", signed_user_ids)\
            .eq("is_resign", False)\
            .eq("user_status", "registered")\
            .execute()
        users = users_res.data or []
        
        filtered_users = filter_users_by_permission(
            users,
            allowed_countries=get_admin_allowed_countries(),
            current_user_id=session.get('user_id')
        )
        
        filtered_user_ids = [u['id'] for u in filtered_users]
        logger.info(f"✅ 补分配：原始 {len(signed_user_ids)} 人，权限过滤后 {len(filtered_user_ids)} 人")
        
        if not filtered_user_ids:
            return jsonify({
                "success": True,
                "message": "没有符合权限条件的已签到学员",
                "assigned_count": 0
            })
        
        # ========== 8. 过滤掉已分配的学员 ==========
        assigned_res = db.table("exam_assignments").select("user_id")\
            .eq("exam_id", exam_id)\
            .in_("user_id", filtered_user_ids)\
            .is_("deleted_at", "null")\
            .execute()
        assigned_set = {a['user_id'] for a in (assigned_res.data or [])}
        
        to_assign = [uid for uid in filtered_user_ids if uid not in assigned_set]
        logger.info(f"📊 需要补分配: {len(to_assign)} 人 (已分配: {len(assigned_set)} 人)")
        
        if not to_assign:
            return jsonify({
                "success": True,
                "message": "所有已签到学员已分配该考试",
                "assigned_count": 0,
                "total_signed": len(filtered_user_ids)
            })
        
        # ========== 9. 批量插入考试分配 ==========
        BATCH_SIZE = 50
        inserted_count = 0
        total_batches = (len(to_assign) + BATCH_SIZE - 1) // BATCH_SIZE
        
        for i in range(0, len(to_assign), BATCH_SIZE):
            batch = to_assign[i:i + BATCH_SIZE]
            assignments = [{
                "exam_id": exam_id, 
                "user_id": uid,
                "created_by": current_user_id,
                "assigned_at": now
            } for uid in batch]
            
            try:
                result = db.table("exam_assignments").insert(assignments).execute()
                inserted_count += len(result.data or [])
                logger.info(f"✅ 批次 {i//BATCH_SIZE + 1}/{total_batches} 插入成功")
            except Exception as batch_error:
                logger.warning(f"⚠️ 批次插入失败，尝试逐条: {batch_error}")
                for item in assignments:
                    try:
                        check_res = db.table("exam_assignments").select("id")\
                            .eq("exam_id", item["exam_id"])\
                            .eq("user_id", item["user_id"])\
                            .is_("deleted_at", "null")\
                            .maybe_single()\
                            .execute()
                        if not check_res.data:
                            db.table("exam_assignments").insert(item).execute()
                            inserted_count += 1
                    except Exception as single_error:
                        if '23505' in str(single_error) or 'duplicate' in str(single_error).lower():
                            logger.warning(f"⚠️ 用户 {item['user_id']} 已有分配记录，跳过")
                            continue
                        logger.error(f"❌ 逐条插入失败: {single_error}")
        
        logger.info(f"✅ 补分配成功: {inserted_count} 名学员")
        
        # ========== 10. 清除缓存 ==========
        clear_training_related_cache(training_id)
        clear_all_assignment_caches(training_id=training_id, exam_id=exam_id)
        
        # ========== 11. 构建响应消息 ==========
        status_msg = ""
        if is_closed:
            status_msg = "（注意：该考试已结束）"
        
        return jsonify({
            "success": True,
            "message": f"补分配成功，共 {inserted_count} 名学员{status_msg}",
            "assigned_count": inserted_count,
            "requested_count": len(to_assign),
            "total_signed": len(filtered_user_ids),
            "exam_status": exam_status,
            "is_closed": is_closed
        })
        
    except Exception as e:
        logger.error(f"❌ 补分配失败: {e}", exc_info=True)
        return jsonify({
            "success": False,
            "message": f"补分配失败: {str(e)}"
        }), 500

@admin_training_bp.route('/api/admin/exams/status', methods=['GET'])
@login_required
@admin_required
def get_exam_status_for_reassign():
    """获取考试状态（用于补分配前的检查）"""
    exam_id = request.args.get('exam_id')
    if not exam_id:
        return jsonify({"error": "缺少 exam_id 参数"}), 400
    
    db = get_supabase()
    try:
        exam_res = db.table("exams").select("status, title").eq("id", int(exam_id)).maybe_single().execute()
        if not exam_res.data:
            return jsonify({"error": "考试不存在"}), 404
        
        return jsonify({
            "status": exam_res.data.get('status', 'draft'),
            "title": exam_res.data.get('title', '')
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_training_bp.route('/api/admin/training/assign', methods=['POST'])
@login_required
@admin_required
def assign_training_batch():
    """
    统一培训分配接口（支持单个和批量）
    与培训管理页面的分配逻辑保持一致
    """
    try:
        data = request.json
        training_id = data.get('training_id')
        user_ids = data.get('user_ids', [])
        
        if not training_id or not user_ids:
            return jsonify({"success": False, "message": "参数不完整"}), 400
        
        db = get_supabase_admin()
        
        # 1. 权限检查
        training_res = db.table("trainings").select("country").eq("id", training_id).maybe_single().execute()
        if not training_res.data:
            return jsonify({"success": False, "message": "培训不存在"}), 404
        
        if not can_access_training(training_res.data):
            return jsonify({"success": False, "message": "无权限"}), 403
        
        # 2. 批量插入分配（使用 upsert 避免重复）
        assignments = []
        now = datetime.now(timezone.utc).isoformat()
        
        for uid in user_ids:
            # 检查是否已存在
            existing = db.table("training_assignments")\
                .select("id")\
                .eq("training_id", training_id)\
                .eq("user_id", uid)\
                .is_("deleted_at", "null")\
                .maybe_single()\
                .execute()
            
            if existing.data:
                # 已存在，跳过
                continue
            
            assignments.append({
                "training_id": training_id,
                "user_id": uid,
                "assigned_at": now,
                "assigned_by": session.get('user_id')
            })
        
        if assignments:
            result = db.table("training_assignments").insert(assignments).execute()
            assigned_count = len(result.data) if result.data else 0
        else:
            assigned_count = 0
        
        # 3. 清除相关缓存
        clear_all_assignment_caches(
            training_id=training_id,
            user_ids=user_ids  # 可选，传入用户ID列表
        )
        return jsonify({
            "success": True,
            "assigned_count": assigned_count,
            "total_requested": len(user_ids),
            "message": f"成功分配 {assigned_count} 名学员"
        })
        
    except Exception as e:
        logger.error(f"分配培训失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ============== 培训完成度报表内部接口，用于考生培训分配统一接口0628==============
@admin_training_bp.route('/api/admin/training/unassign', methods=['POST'])
@login_required
@admin_required
def unassign_training_batch():
    """
    统一取消培训分配接口
    """
    try:
        data = request.json
        training_id = data.get('training_id')
        user_ids = data.get('user_ids', [])
        
        if not training_id or not user_ids:
            return jsonify({"success": False, "message": "参数不完整"}), 400
        
        db = get_supabase_admin()
        
        # 1. 权限检查
        training_res = db.table("trainings").select("country").eq("id", training_id).maybe_single().execute()
        if not training_res.data:
            return jsonify({"success": False, "message": "培训不存在"}), 404
        
        if not can_access_training(training_res.data):
            return jsonify({"success": False, "message": "无权限"}), 403
        
        # 2. 软删除分配
        now = datetime.now(timezone.utc).isoformat()
        unassigned_count = 0
        
        for uid in user_ids:
            result = db.table("training_assignments")\
                .update({
                    "deleted_at": now,
                    "deleted_by": session.get('user_id')
                })\
                .eq("training_id", training_id)\
                .eq("user_id", uid)\
                .is_("deleted_at", "null")\
                .execute()
            
            if result.data:
                unassigned_count += 1
        
        # 3. 清除相关缓存
        clear_all_assignment_caches(
            training_id=training_id,
            user_ids=user_ids
        )
        
        return jsonify({
            "success": True,
            "unassigned_count": unassigned_count,
            "message": f"成功取消 {unassigned_count} 名学员的分配"
        })
        
    except Exception as e:
        logger.error(f"取消培训分配失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_training_bp.route('/api/admin/test_r2')
@login_required
@admin_required
def test_r2_connection():
    """测试 R2 连接"""
    from services.cloudflare_r2 import get_r2_client
    
    try:
        client = get_r2_client()
        # 列出存储桶中的文件（最多5个）
        response = client.list_objects_v2(
            Bucket=current_app.config['CLOUDFLARE_R2_BUCKET'],
            MaxKeys=5
        )
        
        files = []
        for obj in response.get('Contents', []):
            files.append(obj['Key'])
        
        return jsonify({
            "success": True,
            "message": "R2 连接成功",
            "files": files[:5]
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"R2 连接失败: {str(e)}"
        }), 500

@admin_training_bp.route('/api/admin/training/<int:training_id>/push', methods=['POST'])
@login_required
@admin_required
def push_training_to_users(training_id):
    """推送培训给指定用户"""
    from services.db import get_supabase_admin
    from datetime import datetime, timezone
    
    data = request.json
    user_ids = data.get('user_ids', [])
    
    db = get_supabase_admin()
    now = datetime.now(timezone.utc).isoformat()
    
    try:
        # 获取培训信息
        training_res = db.table("trainings").select("name").eq("id", training_id).maybe_single().execute()
        if not training_res.data:
            return jsonify({"success": False, "message": "培训不存在"}), 404
        
        training_name = training_res.data.get('name', '')
        
        # 如果 user_ids 为空，推送给所有用户
        if not user_ids:
            # 获取培训的国家
            training_detail = db.table("trainings").select("countries, country").eq("id", training_id).maybe_single().execute()
            if not training_detail.data:
                return jsonify({"success": False, "message": "培训信息不完整"}), 404
            
            from utils.training_helpers import parse_training_countries
            countries = parse_training_countries(training_detail.data)
            
            if countries:
                # 获取这些国家的所有用户
                users_res = db.table("users").select("id").in_("country", countries).eq("user_status", "registered").eq("is_resign", False).execute()
                user_ids = [u['id'] for u in (users_res.data or [])]
            else:
                return jsonify({"success": False, "message": "培训未指定国家"}), 400
        
        # 为每个用户创建分配记录
        assigned_count = 0
        for uid in user_ids:
            # 检查是否已分配
            existing = db.table("training_assignments").select("id").eq("training_id", training_id).eq("user_id", uid).execute()
            if not existing.data:
                db.table("training_assignments").insert({
                    "training_id": training_id,
                    "user_id": uid,
                    "created_by": session.get('user_id'),
                    "created_at": now
                }).execute()
                assigned_count += 1
        
        logger.info(f"培训推送成功: training_id={training_id}, 推送人数={assigned_count}")
        return jsonify({
            "success": True,
            "message": f"成功推送给 {assigned_count} 名学员",
            "assigned_count": assigned_count
        })
        
    except Exception as e:
        logger.error(f"推送培训失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_training_bp.route('/api/admin/trainings/export_excel', methods=['POST'])
@login_required
@admin_required
def export_trainings_excel():
    """导出培训列表为 Excel"""
    try:
        data = request.json or {}
        name = data.get('name', '')
        country = data.get('country', '')
        quarter = data.get('quarter', '')

        user_timezone = request.cookies.get('user_timezone') or request.headers.get('X-Timezone') or 'Asia/Calcutta'
        
        db = get_supabase()
        admin_db = get_supabase_admin()
        
        # 构建查询
        query = db.table("trainings").select("*").is_("deleted_at", "null")
        if name:
            query = query.ilike("name", f"%{name}%")
        
        all_trainings = query.execute().data or []
        
        # 权限过滤
        allowed_countries = get_admin_allowed_countries()
        if allowed_countries is not None and allowed_countries:
            filtered = []
            for t in all_trainings:
                t_countries = parse_training_countries(t)
                if any(c in allowed_countries for c in t_countries):
                    filtered.append(t)
            all_trainings = filtered
        
        # 国家过滤
        if country:
            all_trainings = [t for t in all_trainings if country in parse_training_countries(t)]
        
        # 季度过滤
        if quarter:
            q_start, q_end = quarter_to_date_range(quarter)
            if q_start and q_end:
                q_start_dt = datetime.fromisoformat(q_start)
                q_end_dt = datetime.fromisoformat(q_end)
                temp = []
                for t in all_trainings:
                    start, end = t.get('start_time'), t.get('end_time')
                    if start and end:
                        try:
                            start_dt = datetime.fromisoformat(start)
                            end_dt = datetime.fromisoformat(end)
                            if start_dt <= q_end_dt and end_dt >= q_start_dt:
                                temp.append(t)
                        except:
                            pass
                all_trainings = temp
        
        # 获取培训ID列表
        training_ids = [t['id'] for t in all_trainings]
        logger.info(f"📋 导出培训ID列表: {training_ids}")
        
        # 批量获取统计信息
        signed_counts = {}
        binding_counts = {}
        photo_counts = {}
        
        if training_ids:
            # 1. 签到人数
            att_res = db.table("training_attendances") \
                .select("training_id, user_id") \
                .in_("training_id", training_ids) \
                .execute()
            for att in (att_res.data or []):
                tid = att['training_id']
                if tid not in signed_counts:
                    signed_counts[tid] = set()
                signed_counts[tid].add(att['user_id'])
            signed_counts = {tid: len(users) for tid, users in signed_counts.items()}
            logger.info(f"📋 签到统计: {signed_counts}")
            
            # 2. 绑定考试数量（修复）
            bind_res = admin_db.table("training_exam_bindings") \
                .select("training_id") \
                .in_("training_id", training_ids) \
                .is_("deleted_at", "null") \
                .execute()
            logger.info(f"📋 绑定查询结果: {bind_res.data}")
            for b in (bind_res.data or []):
                tid = b['training_id']
                binding_counts[tid] = binding_counts.get(tid, 0) + 1
            logger.info(f"📋 绑定统计: {binding_counts}")
            
            # 3. 照片数量
            photo_res = admin_db.table("training_photos") \
                .select("training_id") \
                .in_("training_id", training_ids) \
                .eq("is_deleted", False) \
                .execute()
            for p in (photo_res.data or []):
                tid = p['training_id']
                photo_counts[tid] = photo_counts.get(tid, 0) + 1
            logger.info(f"📋 照片统计: {photo_counts}")
        
        # 创建 Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "培训列表"
        
        # 表头
        headers = ['序号', '培训名称', '培训ID', '国家', '季度', '状态', '创建时间', 
                   '开始时间', '结束时间', '签到人数', '绑定考试数', '照片数']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        status_map = {
            'draft': '草稿',
            'pending': '未开始',
            'active': '进行中',
            'closed': '已关闭'
        }
        
        for idx, t in enumerate(all_trainings, 1):
            tid = t['id']
            t_countries = parse_training_countries(t)
            countries_display = ', '.join(t_countries) if t_countries else (t.get('country') or '-')
            
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=t.get('name', ''))
            ws.cell(row=row, column=3, value=tid)
            ws.cell(row=row, column=4, value=countries_display)
            ws.cell(row=row, column=5, value=_get_quarter_from_date(t.get('created_at')))
            ws.cell(row=row, column=6, value=status_map.get(t.get('dynamic_status', 'draft'), t.get('dynamic_status', '')))

            # 时间字段转换为本地24小时制
            created_at = t.get('created_at')
            ws.cell(row=row, column=7, value=utc_string_to_local(created_at, format_str='%Y-%m-%d %H:%M:%S') if created_at else '')
            start_time = t.get('start_time')
            ws.cell(row=row, column=8, value=utc_string_to_local(start_time, format_str='%Y-%m-%d %H:%M:%S') if start_time else '')
            end_time = t.get('end_time')
            ws.cell(row=row, column=9, value=utc_string_to_local(end_time, format_str='%Y-%m-%d %H:%M:%S') if end_time else '')

            ws.cell(row=row, column=10, value=signed_counts.get(tid, 0))
            ws.cell(row=row, column=11, value=binding_counts.get(tid, 0))
            ws.cell(row=row, column=12, value=photo_counts.get(tid, 0))
        
        # 调整列宽
        column_widths = [8, 35, 10, 20, 12, 12, 20, 20, 20, 14, 14, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"培训列表_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"导出培训列表失败: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500
        
