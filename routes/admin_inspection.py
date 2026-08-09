# routes/admin_inspection.py
import logging
import json
import openpyxl, pdfkit, os, zipfile
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
from flask import request, jsonify, render_template, session, flash, redirect, url_for, send_file
from . import admin_inspection_bp
from services.db import get_supabase, get_supabase_admin
from utils.common import get_reviewer_by_country
from routes.helpers import login_required, admin_required, random_pick_questions, get_allowed_countries, parse_exam_countries, can_access_exam
from utils.permissions import get_admin_allowed_countries, is_developer


logger = logging.getLogger(__name__)

@admin_inspection_bp.route('/admin/interviews')
@login_required
@admin_required
def admin_interviews_page():
    return render_template('admin/list_inspection.html')

@admin_inspection_bp.route('/api/admin/interviews', methods=['GET', 'POST', 'PUT'])
@login_required
@admin_required
def api_admin_interviews():
    """访谈列表查询、创建、更新"""
    db = get_supabase()
    if request.method == 'GET':
        name = request.args.get('name', '')
        country = request.args.get('country', '')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        allowed = get_allowed_countries()

        query = db.table("interviews").select("*", count="exact").is_("deleted_at", "null")
        if name:
            query = query.ilike("title", f"%{name}%")
        query = query.order("created_at", desc=True)

        res = query.execute()
        all_interviews = res.data or []

        # 根据权限过滤访谈
        if allowed is not None and allowed:
            exam_ids = set()
            for inv in all_interviews:
                if inv.get('exam_id'):
                    exam_ids.add(inv['exam_id'])

            exam_country_map = {}
            if exam_ids:
                exams_res = db.table("exams").select("id, country, countries").in_("id", list(exam_ids)).execute()
                for exam in (exams_res.data or []):
                    exam_country_map[exam['id']] = parse_exam_countries(exam)

            filtered = []
            for inv in all_interviews:
                exam_countries = exam_country_map.get(inv.get('exam_id'), [])
                if any(c in allowed for c in exam_countries):
                    filtered.append(inv)
            all_interviews = filtered

        total = len(all_interviews)
        start = (page - 1) * per_page
        end = start + per_page
        interviews = all_interviews[start:end]

        # 收集用户ID用于统计
        all_user_ids = set()
        for inv in interviews:
            user_res = db.table("interview_results").select("user_id").eq("interview_id", inv['id']).execute()
            all_user_ids.update(r['user_id'] for r in (user_res.data or []))

        user_country_map = {}
        if allowed is not None and all_user_ids:
            users_res = db.table("users").select("id, country").in_("id", list(all_user_ids)).execute()
            user_country_map = {u['id']: u.get('country') for u in (users_res.data or [])}

        now = datetime.now(timezone.utc)
        for inv in interviews:
            # 计算状态
            start_time = inv.get('start_time')
            end_time = inv.get('end_time')
            if not start_time or not end_time:
                inv['status'] = 'draft'
            else:
                try:
                    s = datetime.fromisoformat(start_time)
                    e = datetime.fromisoformat(end_time)
                    if now < s:
                        inv['status'] = 'created'
                    elif now > e:
                        inv['status'] = 'closed'
                    else:
                        inv['status'] = 'active'
                except:
                    inv['status'] = 'draft'

            # 统计人数
            user_res = db.table("interview_results").select("user_id").eq("interview_id", inv['id']).execute()
            user_ids = [r['user_id'] for r in (user_res.data or [])]

            if user_ids:
                active_users = db.table("users").select("id").in_("id", user_ids).eq("is_resign", False).execute()
                active_user_ids = [u['id'] for u in (active_users.data or [])]
            else:
                active_user_ids = []

            if allowed is not None:
                filtered_ids = [uid for uid in active_user_ids if user_country_map.get(uid) in allowed]
                inv['interviewee_count'] = len(set(filtered_ids))
            else:
                inv['interviewee_count'] = len(set(active_user_ids))

            # 附加考试信息 - 使用 parse_exam_countries 解析国家
            if inv.get('exam_id'):
                exam_res = db.table("exams").select("title, country, countries").eq("id", inv['exam_id']).maybe_single().execute()
                if exam_res.data:
                    inv['exam_title'] = exam_res.data.get('title', '')

                    # 使用 parse_exam_countries 解析多国家字段
                    exam_countries = parse_exam_countries(exam_res.data)

                    # 添加调试日志
                    logger.info(f"访谈 {inv['id']}: exam_id={inv['exam_id']}, exam_countries={exam_countries}")

                    # 根据管理员权限过滤国家显示
                    if allowed is not None and allowed:
                        filtered_countries = [c for c in exam_countries if c in allowed]
                        inv['country'] = ', '.join(filtered_countries) if filtered_countries else ''
                    else:
                        inv['country'] = ', '.join(exam_countries) if exam_countries else ''

                    inv['exam_countries'] = exam_countries

        return jsonify({"data": interviews, "total": total, "page": page, "per_page": per_page})

    elif request.method == 'POST':
        data = request.json
        exam_id = data.get('exam_id')
        title = data.get('title', '')
        if not title:
            exam_title = ''
            if exam_id:
                exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
                if exam_res.data:
                    exam_title = exam_res.data['title']
            title = f"Interview-{exam_title}" if exam_title else "未命名访谈"
        question_count = data.get('question_count', 5)
        reviewer = data.get('reviewer', '')
        is_draft = data.get('is_draft', False)
        feedback = data.get('feedback', '')
        start_time = data.get('start_time')
        end_time = data.get('end_time')

        # 将本地时间转为 UTC 存储
        start_time_utc = start_time if start_time else None
        end_time_utc = end_time if end_time else None

        status = 'draft' if is_draft else 'active'
        if start_time and end_time:
            status = 'active'  # 简单处理，后续根据时间自动判定

        user_ids = data.get('user_ids', [])
        if user_ids:
            # 校验所有用户均为已注册
            valid_users = db.table("users").select("id").in_("id", user_ids).eq("user_status", "registered").execute()
            if len(valid_users.data or []) != len(user_ids):
                return jsonify({"success": False, "message": "所选考生中包含未注册的用户"}), 400

        interviewee_count = len(user_ids)
        interview_insert = db.table("interviews").insert({
            "title": title,
            "exam_id": exam_id,
            "created_by": session['user_id'],
            "reviewer": reviewer,
            "feedback": feedback,
            "question_count": question_count,
            "status": status,
            "start_time": start_time_utc,
            "end_time": end_time_utc,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "interviewee_count": interviewee_count
        }).execute()
        if not interview_insert.data:
            return jsonify({"success": False, "message": "创建失败"}), 500
        new_id = interview_insert.data[0]['id']

        # 为选中的学员抽取题目
        user_ids = data.get('user_ids', [])
        for uid in user_ids:
            questions = random_pick_questions(exam_id, question_count)
            for q in questions:
                db.table("interview_results").insert({
                    "interview_id": new_id,
                    "user_id": uid,
                    "question_id": q['id']
                }).execute()
        return jsonify({"success": True, "id": new_id})

    elif request.method == 'PUT':
        data = request.json
        inv_id = data.get('id')
        db = get_supabase()
        # 获取原访谈
        orig_res = db.table("interviews").select("*").eq("id", inv_id).maybe_single().execute()
        if not orig_res.data:
            return jsonify({"success": False, "message": "访谈不存在"}), 404
        orig = orig_res.data

        # 更新基本字段
        update_data = {}
        for field in ['start_time', 'end_time', 'reviewer', 'feedback', 'exam_id', 'question_count', 'title']:
            if field in data:
                val = data[field]
                if field in ('start_time', 'end_time') and val:
                    val = val
                update_data[field] = val
        if update_data:
            db.table("interviews").update(update_data).eq("id", inv_id).execute()

        # 重新抽题（无论状态，只要提供了 user_ids 就更新人员题目）
        if 'user_ids' in data:
            # 删除该访谈的所有现有题目
            db.table("interview_results").delete().eq("interview_id", inv_id).execute()
            # 使用最新的 exam_id 和 question_count
            exam_id = data.get('exam_id', orig['exam_id'])
            question_count = data.get('question_count', orig['question_count'])
            for uid in data['user_ids']:
                questions = random_pick_questions(exam_id, question_count)
                for q in questions:
                    db.table("interview_results").insert({
                        "interview_id": inv_id,
                        "user_id": uid,
                        "question_id": q['id']
                    }).execute()
        
        return jsonify({"success": True})

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/detail')
@login_required
@admin_required
def api_get_interview_detail(interview_id):
    db = get_supabase()
    
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"success": False, "message": "访谈不存在"}), 404
    
    interview = inv_res.data
    
    # 计算状态（与列表接口保持一致）
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)
    start_time = interview.get('start_time')
    end_time = interview.get('end_time')
    
    if not start_time or not end_time:
        status = 'draft'
    else:
        try:
            s = datetime.fromisoformat(start_time)
            e = datetime.fromisoformat(end_time)
            if now < s:
                status = 'created'
            elif now > e:
                status = 'closed'
            else:
                status = 'active'
        except:
            status = 'draft'
    
    interview['status'] = status
    
    # 获取关联的考试标题
    exam_title = ''
    if interview.get('exam_id'):
        exam_res = db.table("exams").select("title").eq("id", interview['exam_id']).maybe_single().execute()
        if exam_res.data:
            exam_title = exam_res.data['title']
    
    # 获取已分配的学员ID
    user_ids_res = db.table("interview_results").select("user_id").eq("interview_id", interview_id).execute()
    user_ids = list(set([r['user_id'] for r in (user_ids_res.data or [])]))
    
    return jsonify({
        "success": True,
        "data": {
            **interview,
            "exam_title": exam_title,
            "user_ids": user_ids
        }
    })

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>', methods=['GET'])
@login_required
@admin_required
def api_get_interview(interview_id):
    """获取某个访谈的详细信息"""
    db = get_supabase()
    inv = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv.data:
        return jsonify({"error": "访谈不存在"}), 404
    # 获取已分配的学员ID
    user_ids_res = db.table("interview_results").select("user_id").eq("interview_id", interview_id).execute()
    user_ids = list(set([r['user_id'] for r in (user_ids_res.data or [])]))
    inv.data['user_ids'] = user_ids
    return jsonify(inv.data)

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/user_ids')
@login_required
@admin_required
def get_interview_user_ids(interview_id):
    """获取某个访谈的所有用户ID"""
    db = get_supabase()
    res = db.table("interview_results").select("user_id").eq("interview_id", interview_id).execute()
    ids = list(set(r['user_id'] for r in (res.data or [])))
    return jsonify({"user_ids": ids})

@admin_inspection_bp.route('/admin/interview/<int:interview_id>')
@login_required
@admin_required
def admin_interview_detail_page(interview_id):
    """3. 访谈二级菜单数据接口 访谈详情页面""" 
    return render_template('admin/list_inspection_details.html', interview_id=interview_id)

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/results')
@login_required
@admin_required
def api_interview_results(interview_id):
    """获取某个访谈的所有结果，支持筛选"""
    db = get_supabase()
    search = request.args.get('search', '')
    country = request.args.get('country', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # 基本信息
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    interview = inv_res.data

    # 查询访谈结果并关联用户信息
    query = db.table("interview_results").select("*, users(name_cn, name_en, email, country, wh_id)").eq("interview_id", interview_id)
    if country:
        query = query.eq("users.country", country)
    if search:
        query = query.or_(f"users.name_cn.ilike.%{search}%,users.name_en.ilike.%{search}%")
    
    # 分页
    start = (page - 1) * per_page
    end = start + per_page - 1
    res = query.range(start, end).order("user_id").execute()
    total = res.count if hasattr(res, 'count') else len(res.data or [])

    # 组装数据：按用户聚合
    user_results = {}
    for row in (res.data or []):
        uid = row['user_id']
        if uid not in user_results:
            user_info = row.get('users', {})
            user_results[uid] = {
                "user_id": uid,
                "name": user_info.get('name_cn') or user_info.get('name_en', ''),
                "email": user_info.get('email', ''), 
                "country": user_info.get('country', ''),
                "wh_id": user_info.get('wh_id', ''),
                "results": [],
                "submitted_at": None
            }
        user_results[uid]["results"].append({
            "question_id": row['question_id'],
            "answer": row['answer'],
            "is_correct": row['is_correct'],
            "feedback": row['feedback']
        })
        # 提交时间取该用户最近一条有答案的记录时间
        if row.get('submitted_at') and (not user_results[uid]["submitted_at"] or row['submitted_at'] > user_results[uid]["submitted_at"]):
            user_results[uid]["submitted_at"] = row['submitted_at']

    # 转换为列表并计算答对数量
    result_list = []
    for uid, data in user_results.items():
        correct_count = sum(1 for r in data['results'] if r['is_correct'])
        total_questions = len(data['results'])
        result_list.append({
            "user_id": uid,
            "name": data['name'],
            "email": data['email'],
            "country": data['country'],
            "wh_id": data['wh_id'],
            "total_questions": total_questions,
            "correct_count": correct_count,
            "submitted_at": data['submitted_at'],
            "feedback": "",  # 可后续合并
            "results": data['results']
        })

    return jsonify({
        "interview": interview,
        "results": result_list,
        "total": total,
        "page": page,
        "per_page": per_page
    })

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/resample/<user_id>', methods=['POST'])
@login_required
@admin_required
def resample_interview(interview_id, user_id):
    """4. 重新访谈接口 重新为指定用户抽题，保留历史记录"""
    db = get_supabase()
    inv_res = db.table("interviews").select("exam_id, question_count").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"success": False, "message": "访谈不存在"}), 404
    exam_id = inv_res.data['exam_id']
    count = inv_res.data['question_count']
    
    # 删除该用户在此访谈中的旧题目（仅删除未作答的？根据需求保留历史记录，这里简单起见先删除所有旧记录再插入新题）
    db.table("interview_results").delete().eq("interview_id", interview_id).eq("user_id", user_id).execute()
    
    questions = random_pick_questions(exam_id, count)
    for q in questions:
        db.table("interview_results").insert({
            "interview_id": interview_id,
            "user_id": user_id,
            "question_id": q['id']
        }).execute()
    return jsonify({"success": True})

@admin_inspection_bp.route('/api/admin/interview/preview', methods=['POST'])
@login_required
@admin_required
def interview_preview():
    """5. 预览接口（用于模态框 → 预览页） 预览访谈：返回每个被选中学员的抽题情况"""
    db = get_supabase()
    data = request.json
    exam_id = data.get('exam_id')
    
    # 检查题库是否存在
    q_check = db.table("questions").select("id").eq("exam_id", exam_id).limit(1).execute()
    if not q_check.data:
        return jsonify({"error": "该考试没有题目"}), 400

    user_ids = data.get('user_ids', [])
    question_count = data.get('question_count', 5)

    if not exam_id:
        return jsonify({"error": "exam_id 不能为空"}), 400

    logger.info(f"预览访谈: exam_id={exam_id}, users={len(user_ids)}, count={question_count}")
    
    exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
    exam_title = exam_res.data['title'] if exam_res.data else ''
    
    preview = []
    for uid in user_ids:
        user_res = db.table("users").select("name_cn, name_en").eq("id", uid).maybe_single().execute()
        user_name = ''
        if user_res.data:
            user_name = user_res.data.get('name_cn') or user_res.data.get('name_en', '')
        # 随机抽取题目
        questions = random_pick_questions(exam_id, question_count)
        # 处理题目数据，确保 options 为字典，并筛选必要字段
        questions_light = []
        for q in questions:
            # 解析 options（可能为字符串 JSON）
            opts = q.get('options', {})
            if isinstance(opts, str):
                try:
                    opts = json.loads(opts)
                except:
                    opts = {}
            # 过滤空选项
            if opts:
                opts = {k: v for k, v in opts.items() if v and v.strip()}
            questions_light.append({
                'num': q.get('num'),
                'content': q.get('content_cn') or q.get('content') or q.get('content_raw', '无题目内容'),
                'type': q.get('type', 'single'),
                'options': opts   # 前端可能需要展示选项
            })
        preview.append({
            "user_id": uid,
            "user_name": user_name,
            "questions": questions_light
        })
    
    return jsonify({"exam_title": exam_title, "preview": preview})

@admin_inspection_bp.route('/interview/take/<int:interview_id>')
@login_required
def take_interview(interview_id):
    """学员进入访谈（支持普通访谈和强制访谈）"""
    user_id = session['user_id']
    db = get_supabase()
    admin_db = get_supabase_admin()
    now = datetime.now(timezone.utc)
    
    # ========== 1. 检查是否是强制访谈 ==========
    is_force = False
    force_data = None
    
    try:
        # 使用 execute() 代替 maybe_single()
        force_result = admin_db.table("user_interview_force_records").select("*")\
            .eq("original_interview_id", interview_id)\
            .eq("user_id", user_id)\
            .is_("deleted_at", "null")\
            .execute()
        
        # 检查 result.data 是否有数据
        if force_result.data and len(force_result.data) > 0:
            is_force = True
            force_data = force_result.data[0]
        else:
            is_force = False
            
    except Exception as e:
        logger.warning(f"查询强制访谈记录失败: {e}")
        is_force = False
    
    inv = None
    
    if is_force and force_data:
        # ========== 强制访谈逻辑 ==========
        start_time = force_data.get('start_time')
        end_time = force_data.get('end_time')
        
        # 检查有效期
        if start_time:
            try:
                start_dt = datetime.fromisoformat(start_time)
                if now < start_dt:
                    flash("强制访谈尚未开始", "warning")
                    return redirect(url_for('dashboard'))
            except:
                pass

        if end_time:
            try:
                if isinstance(end_time, str):
                    end_str = end_time.replace('Z', '+00:00')
                    end_dt = datetime.fromisoformat(end_str)
                else:
                    end_dt = end_time
                
                now_utc = datetime.now(timezone.utc)
                
                if end_dt.tzinfo is None:
                    end_dt = end_dt.replace(tzinfo=timezone.utc)
                
                logger.info(f"强制访谈时间检查: now={now_utc.isoformat()}, end={end_dt.isoformat()}")
                
                if now_utc > end_dt:
                    logger.warning(f"强制访谈已过期，删除记录 {force_data['id']}")
                    admin_db.table("user_interview_force_records").update({
                        "deleted_at": now_utc.isoformat(),
                        "deleted_by": user_id
                    }).eq("id", force_data['id']).execute()
                    flash("强制访谈已过期", "warning")
                    return redirect(url_for('dashboard'))
            except Exception as e:
                logger.error(f"解析时间失败: {e}")
        
        # 获取原访谈信息
        inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
        if not inv_res.data:
            flash("访谈不存在", "danger")
            return redirect(url_for('dashboard'))
        inv = inv_res.data
        
        # 检查用户是否属于该访谈（强制访谈也需要检查）
        result = db.table("interview_results").select("interview_id").eq("interview_id", interview_id).eq("user_id", user_id).limit(1).execute()
        if not result.data:
            flash("您不在本次访谈名单中", "danger")
            return redirect(url_for('dashboard'))
        
        logger.info(f"用户 {user_id} 进入强制访谈 {interview_id}，有效期至 {end_time}")
        
    else:
        # ========== 普通访谈逻辑 ==========
        # 检查用户是否属于该访谈
        result = db.table("interview_results").select("interview_id").eq("interview_id", interview_id).eq("user_id", user_id).limit(1).execute()
        if not result.data:
            flash("您不在本次访谈名单中", "danger")
            return redirect(url_for('dashboard'))

        inv = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
        if not inv.data:
            flash("访谈不存在", "danger")
            return redirect(url_for('dashboard'))
        
        inv = inv.data
        
        # 检查普通访谈有效期
        start_time = inv.get('start_time')
        end_time = inv.get('end_time')
        
        if start_time:
            try:
                start_dt = datetime.fromisoformat(start_time)
                if now < start_dt:
                    flash("访谈尚未开始", "warning")
                    return redirect(url_for('dashboard'))
            except:
                pass
        
        if end_time:
            try:
                end_dt = datetime.fromisoformat(end_time)
                if now > end_dt:
                    flash("访谈已结束", "warning")
                    return redirect(url_for('dashboard'))
            except:
                pass
    
    # ========== 2. 获取题目（普通访谈和强制访谈共用）==========
    # 分步查询，避免外键歧义
    interview_results = db.table("interview_results") \
        .select("id, question_id, answer") \
        .eq("interview_id", interview_id) \
        .eq("user_id", user_id) \
        .execute()
    
    if not interview_results.data:
        questions = []
    else:
        question_ids = list(set([row['question_id'] for row in interview_results.data]))
        
        questions_data = db.table("questions") \
            .select("*") \
            .in_("id", question_ids) \
            .execute()
        
        question_map = {q['id']: q for q in (questions_data.data or [])}
        
        questions = []
        for row in interview_results.data:
            q = question_map.get(row['question_id'], {})
            q_copy = q.copy() if q else {}
            
            opts = q_copy.get('options', {})
            if isinstance(opts, str):
                try:
                    q_copy['options'] = json.loads(opts)
                except:
                    q_copy['options'] = {}
            
            if q_copy.get('type') == 'judge' and (not q_copy.get('options')):
                q_copy['options'] = {"A": "正确 True", "B": "错误 False"}
            
            if not isinstance(q_copy.get('options'), dict):
                q_copy['options'] = {}
        
            q_copy['interview_result_id'] = row['id']
            q_copy['user_answer'] = row.get('answer') or ''
            questions.append(q_copy)
    
    questions.sort(key=lambda x: x.get('num', 0))
    for idx, q in enumerate(questions, 1):
        q['num'] = idx
        if q.get('options'):
            q['options'] = {k: v for k, v in q['options'].items() if v.strip()}
        else:
            q['options'] = {}
    
    # 为模板添加额外信息
    interview_data = inv if inv else {}
    if is_force and force_data:
        # 强制访谈：覆盖标题和有效期显示
        interview_data = interview_data.copy() if interview_data else {}
        interview_data['title'] = force_data.get('title', interview_data.get('title', '强制访谈'))
        interview_data['start_time'] = force_data.get('start_time')
        interview_data['end_time'] = force_data.get('end_time')
        interview_data['is_force'] = True

    return render_template('exam/take_interview.html', interview=interview_data, questions=questions)

@admin_inspection_bp.route('/api/interview/<int:interview_id>/submit', methods=['POST'])
@login_required
def submit_interview(interview_id):
    """② 后端新增接口：提交学员的答案"""
    user_id = session['user_id']
    answers = request.json.get('answers', {})  # {result_id: answer}
    now_utc = datetime.now(timezone.utc).isoformat()
    db = get_supabase()

    # 获取访谈级别的 feedback
    interview_res = db.table("interviews").select("feedback").eq("id", interview_id).maybe_single().execute()
    interview_feedback = interview_res.data.get('feedback', '') if interview_res.data else ''

    for rid, ans in answers.items():
        # 获取关联题目 ID
        result = db.table("interview_results").select("question_id").eq("id", rid).eq("user_id", user_id).maybe_single().execute()
        if not result.data:
            continue
        qid = result.data['question_id']
        # 获取标准答案和题型
        q = db.table("questions").select("answer, type").eq("id", qid).maybe_single().execute()
        is_correct = False
        if q.data:
            correct_ans = q.data['answer'].upper()
            q_type = q.data['type']
            if q_type == 'multi':
                u_set = set(ans.upper().replace(' ', ''))
                c_set = set(correct_ans.replace(' ', ''))
                is_correct = (u_set == c_set)
            elif q_type == 'judge':
                norm = ans.upper()
                if norm in ('A', 'T', '√', '正确', '对'): norm = 'T'
                elif norm in ('B', 'F', '×', '错误', '错'): norm = 'F'
                correct_std = correct_ans.replace('√', 'T').replace('×', 'F')
                is_correct = (norm == correct_std)
            else:
                is_correct = (ans.strip().upper() == correct_ans)
        db.table("interview_results").update({
            "answer": ans,
            "is_correct": is_correct,
            "submitted_at": datetime.now(timezone.utc).isoformat(),
            "feedback": interview_feedback
        }).eq("id", rid).eq("user_id", user_id).execute()

    # 标记所有未提交的题目也为"已提交"（防止部分题目未回答导致状态不更新）
    # 如果用户提交时有些题目没有回答，将这些题目的 submitted_at 也设置为当前时间
    # 这样前端查询 submitted_at 时就能正确判断为"已完成"
    unsubmitted_results = db.table("interview_results") \
        .select("id") \
        .eq("interview_id", interview_id) \
        .eq("user_id", user_id) \
        .is_("submitted_at", "null") \
        .is_("deleted_at", "null") \
        .execute()
    
    for unsubmitted in (unsubmitted_results.data or []):
        db.table("interview_results").update({
            "submitted_at": now_utc,
            "answer": "Not answer"  # 可选：标记为未作答
        }).eq("id", unsubmitted['id']).execute()
    
    logger.info(f"用户 {user_id} 提交访谈 {interview_id}，共 {len(answers)} 道题，{len(unsubmitted_results.data or [])} 道未作答")
    
    return jsonify({"success": True, "message": "jsonify_interview_has_been_submitted", "params": []})

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/user/<user_id>/answers')
@login_required
@admin_required
def get_interview_user_answers(interview_id, user_id):
    """获取指定访谈中指定用户的答题详情"""
    db = get_supabase()
    # 获取访谈信息
    inv_res = db.table("interviews").select("id, title").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404

    # 分步查询
    # 第一步：获取用户的所有访谈结果
    interview_results = db.table("interview_results") \
        .select("id, question_id, answer, is_correct") \
        .eq("interview_id", interview_id) \
        .eq("user_id", user_id) \
        .execute()
    
    if not interview_results.data:
        return jsonify({"answers": []})
    
    # 第二步：收集 question_id
    question_ids = list(set([row['question_id'] for row in interview_results.data]))
    
    # 第三步：批量查询题目信息
    questions_data = db.table("questions") \
        .select("*") \
        .in_("id", question_ids) \
        .execute()
    
    # 第四步：构建映射
    question_map = {q['id']: q for q in (questions_data.data or [])}
    
    # 第五步：组装返回数据
    data = []
    for row in interview_results.data:
        q = question_map.get(row['question_id'], {})
        
        # 解析 options
        opts = q.get('options', {})
        if isinstance(opts, str):
            try:
                q['options'] = json.loads(opts)
            except:
                q['options'] = {}

        if q.get('type') == 'judge' and not q['options']:
            q['options'] = {"A": "正确 True", "B": "错误 False"}

        data.append({
            "question_num": q.get('num'),
            "question_content": q.get('content_cn') or q.get('content') or q.get('content_raw', ''),
            "question_type": q.get('type'),
            "options": q.get('options', {}),
            "user_answer": row.get('answer') or '',
            "is_correct": row.get('is_correct'),
            "result_id": row['id']
        })
    # 按题目编号排序
    data.sort(key=lambda x: x.get('question_num', 0))
    return jsonify({"answers": data})

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/user/<user_id>/delete', methods=['DELETE'])
@login_required
@admin_required
def api_admin_delete_interview_user_result(interview_id, user_id):
    """删除指定用户在指定访谈中的所有答题记录"""
    db = get_supabase()
    operator_id = session['user_id']
    
    # 获取访谈信息
    interview_res = db.table("interviews").select("*").eq("id", interview_id).execute()
    if not interview_res.data:
        return jsonify({"success": False, "message": "访谈不存在"}), 404
    
    interview = interview_res.data[0]
    exam_id = interview.get('exam_id')
    
    # 检查考试的国家权限（支持多国家）
    exam_data = None
    if exam_id:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_data = exam_res.data
    
    # 权限检查
    is_dev = is_developer()
    current_role = session.get('role')
    
    if not is_dev:
        if exam_data:
            exam_countries = parse_exam_countries(exam_data)
            allowed = get_admin_allowed_countries()
            
            if current_role == 'super_admin':
                if allowed is not None:
                    if not any(c in allowed for c in exam_countries):
                        return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
            elif current_role == 'admin':
                if allowed:
                    if not any(c in allowed for c in exam_countries):
                        return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
                else:
                    user_country = session.get('user_country')
                    if user_country not in exam_countries:
                        return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
            else:
                return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
        else:
            allowed = get_admin_allowed_countries()
            if allowed is not None and not allowed:
                return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
    
    try:
        # 软删除该用户的所有访谈答题记录
        db.table("interview_results").update({
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": operator_id
        }).eq("interview_id", interview_id).eq("user_id", user_id).execute()
        
        logger.info(f"访谈记录已删除: interview_id={interview_id}, user_id={user_id}, 操作人={operator_id}")
        return jsonify({"success": True, "message": "访谈记录已删除"})
    except Exception as e:
        logger.error(f"删除访谈记录失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/batch_delete', methods=['POST'])
@login_required
@admin_required
def api_admin_batch_delete_interview_results(interview_id):
    """批量删除访谈记录（支持软删除和永久删除）"""
    data = request.json
    user_ids = data.get('user_ids', [])
    delete_type = data.get('delete_type', 'soft')
    
    if not user_ids:
        return jsonify({"success": False, "message": "请选择要删除的访谈记录"}), 400
    
    db = get_supabase()
    operator_id = session['user_id']
    success_count = 0
    fail_count = 0
    errors = []
    
    # 获取访谈信息（权限检查）
    interview_res = db.table("interviews").select("*").eq("id", interview_id).execute()
    if not interview_res.data:
        return jsonify({"success": False, "message": "访谈不存在"}), 404
    
    interview = interview_res.data[0]
    exam_id = interview.get('exam_id')
    
    # 检查考试的国家权限（支持多国家）
    
    # 先获取考试信息
    exam_data = None
    if exam_id:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_data = exam_res.data
    
    # 权限检查
    is_dev = is_developer()
    current_role = session.get('role')
    
    # 开发者可以删除任何记录
    if not is_dev:
        # 检查是否有权访问该考试
        if exam_data:
            exam_countries = parse_exam_countries(exam_data)
            allowed = get_admin_allowed_countries()
            
            if current_role == 'super_admin':
                if allowed is not None:
                    if not any(c in allowed for c in exam_countries):
                        return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
            elif current_role == 'admin':
                if allowed:
                    if not any(c in allowed for c in exam_countries):
                        return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
                else:
                    user_country = session.get('user_country')
                    if user_country not in exam_countries:
                        return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
            else:
                return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
        else:
            # 没有关联考试，需要检查其他权限逻辑
            allowed = get_admin_allowed_countries()
            if allowed is not None and not allowed:
                return jsonify({"success": False, "message": "无权删除此访谈记录"}), 403
    
    # 执行删除
    for user_id in user_ids:
        try:
            if delete_type == 'hard':
                db.table("interview_results").delete().eq("interview_id", interview_id).eq("user_id", user_id).execute()
            else:
                db.table("interview_results").update({
                    "deleted_at": datetime.now(timezone.utc).isoformat(),
                    "deleted_by": operator_id
                }).eq("interview_id", interview_id).eq("user_id", user_id).execute()
            success_count += 1
        except Exception as e:
            fail_count += 1
            errors.append(f"用户 {user_id}: {str(e)}")
    
    return jsonify({
        "success": True,
        "success_count": success_count,
        "fail_count": fail_count,
        "errors": errors[:10]
    })

@admin_inspection_bp.route('/api/admin/interviews/<int:interview_id>', methods=['DELETE'])
@login_required
@admin_required
def api_admin_delete_interview_by_id(interview_id):
    """删除访谈（软删除）- 供 list_inspection.html 调用"""
    db = get_supabase()
    operator_id = session['user_id']
    
    # 获取访谈信息（用于权限检查）
    interview_res = db.table("interviews").select("*, exams(country), created_by").eq("id", interview_id).execute()
    if not interview_res.data:
        return jsonify({"success": False, "message": "访谈不存在"}), 404
    
    interview = interview_res.data[0]
    exam_data = interview.get('exams', {})
    created_by = interview.get('created_by')
    
    # 权限检查
    allowed = get_admin_allowed_countries()
    if allowed is not None:
        exam_country = exam_data.get('country') if isinstance(exam_data, dict) else None
        if exam_country and exam_country not in allowed:
            return jsonify({"success": False, "message": "jsonify_no_authorith_delete_project", "params": []}), 403

    # 创建者检查（非超管/开发者）
    if not is_dev and current_role != 'super_admin':
        if created_by != current_user_id:
            return jsonify({"success": False, "message": "jsonify_no_permmission_delete_item_created_by_others", "params": []}), 403
    
    try:
        now_utc = datetime.now(timezone.utc).isoformat()
        
        # 软删除访谈本身
        db.table("interviews").update({
            "deleted_at": now_utc,
            "deleted_by": operator_id
        }).eq("id", interview_id).execute()
        
        # 同时软删除该访谈下的所有答题记录
        db.table("interview_results").update({
            "deleted_at": now_utc,
            "deleted_by": operator_id
        }).eq("interview_id", interview_id).execute()
        
        logger.info(f"访谈已删除: interview_id={interview_id}, 操作人={operator_id}")
        return jsonify({"success": True, "message": "访谈已删除"})
    except Exception as e:
        logger.error(f"删除访谈失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/user/<user_id>/resample', methods=['POST'])
@login_required
@admin_required
def api_admin_resample_interview(interview_id, user_id):
    """重新为指定用户抽题（先删除旧记录，再插入新记录）"""
    db = get_supabase()
    
    # 获取访谈信息
    interview_res = db.table("interviews").select("exam_id, question_count").eq("id", interview_id).execute()
    if not interview_res.data:
        return jsonify({"success": False, "message": "interview_not_found", "params": []}), 404
    
    interview = interview_res.data[0]
    exam_id = interview['exam_id']
    question_count = interview['question_count']
    
    # 检查题库
    q_check = db.table("questions").select("id").eq("exam_id", exam_id).limit(1).execute()
    if not q_check.data:
        return jsonify({"success": False, "message": "jsonify_no_questions_in_exam", "params": []}), 400
    
    try:
        # 1. 先删除该用户在该访谈下的所有旧记录（硬删除）
        delete_result = db.table("interview_results").delete() \
            .eq("interview_id", interview_id) \
            .eq("user_id", user_id) \
            .execute()
        
        deleted_count = len(delete_result.data) if delete_result.data else 0
        logger.info(f"已删除用户 {user_id} 的 {deleted_count} 条旧访谈记录")
        
        # 2. 重新抽题
        questions = random_pick_questions(exam_id, question_count)
        inserted_count = 0
        for q in questions:
            result = db.table("interview_results").insert({
                "interview_id": interview_id,
                "user_id": user_id,
                "question_id": q['id'],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "answer": None,
                "is_correct": None,
                "submitted_at": None
            }).execute()
            if result.data:
                inserted_count += 1
        
        logger.info(f"已为用户 {user_id} 插入 {inserted_count} 条新访谈题目")
        
        # 使用翻译键名 + 参数 
        # return jsonify({"success": True, "message": f"重新抽题成功，已删除 {deleted_count} 条旧记录，新增 {inserted_count} 道题目"})
        return jsonify({
            "success": True, 
            "message": "jsonify_resample_success", "params": [deleted_count, inserted_count]
        })

    except Exception as e:
        logger.error(f"重新抽题失败: {e}")
        return jsonify({
            "success": False, 
            "message": "jsonify_resample_failed", "params": [str(e)]
        }), 500

@admin_inspection_bp.route('/api/admin/interviewee/stats')
@login_required
@admin_required
def admin_interviewee_stats():
    """访谈统计页面入口（二级菜单）"""
    # 此函数仅渲染模板，实际数据由前端 AJAX 请求 /api/admin/interview/<id>/details 获取
    return render_template('admin/list_inspection.html')

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/details')
@login_required
@admin_required
def api_interview_details(interview_id):
    """获取访谈详情数据（支持筛选和分页）"""
    db = get_supabase()
    search = request.args.get('search', '')
    country = request.args.get('country', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # 获取访谈信息
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    interview = inv_res.data

    # 获取考试标题
    exam_title = ''
    if interview.get('exam_id'):
        exam_res = db.table("exams").select("title").eq("id", interview['exam_id']).maybe_single().execute()
        if exam_res.data:
            exam_title = exam_res.data['title']

    # 获取访谈级别的反馈人
    interview_feedback = interview.get('feedback', '')

    # 查询该访谈的所有答题记录（按用户聚合）
    query = db.table("interview_results").select("*").eq("interview_id", interview_id)
    # 先获取所有记录，然后在 Python 中处理筛选和聚合
    all_res = query.execute()
    all_data = all_res.data or []

    # ===== 国家权限过滤 =====
    allowed = get_allowed_countries()
    if allowed is not None:
        if not allowed:
            # 没有允许的国家，直接返回空
            return jsonify({
                "interview": {
                    "id": interview_id,
                    "title": interview.get('title'),
                    "exam_title": exam_title,
                    "exam_id": interview.get('exam_id'),
                    "reviewer": interview.get('reviewer', ''),
                    "feedback": interview_feedback
                },
                "data": [],
                "total": 0,
                "page": page,
                "per_page": per_page
            })
        # 收集所有 user_id
        all_user_ids = list(set(r['user_id'] for r in all_data))
        # 查询这些用户的国家
        users_country_res = db.table("users").select("id, country").in_("id", all_user_ids).execute()
        allowed_ids = {u['id'] for u in (users_country_res.data or []) if u.get('country') in allowed}
        # 过滤数据
        all_data = [r for r in all_data if r['user_id'] in allowed_ids]
    # ======================
    # 批量获取用户信息
    user_ids = list(set(r['user_id'] for r in all_data))
    users_map = {}
    if user_ids:
        users_res = db.table("users").select("id, name_cn, name_en, email, country, wh_id, department, is_resign").in_("id", user_ids).eq("is_resign", False).execute()
        for u in (users_res.data or []):
            users_map[u['id']] = u

    # 过滤掉离职人员的记录
    all_data = [r for r in all_data if r['user_id'] in users_map]

    # 按用户聚合
    user_results = {}
    for row in all_data:
        uid = row['user_id']
        user_info = users_map.get(uid, {})
        # 筛选：姓名搜索
        if search:
            name = user_info.get('name_cn') or user_info.get('name_en', '')
            if search.lower() not in name.lower():
                continue
        # 筛选：国家
        if country and user_info.get('country') != country:
            continue

        if uid not in user_results:
            user_results[uid] = {
                "user_id": uid,
                "name": user_info.get('name_cn') or user_info.get('name_en', ''),
                "email": user_info.get('email'),
                "country": user_info.get('country', ''),
                "wh_id": user_info.get('wh_id', ''),
                "department": user_info.get('department', ''),
                "submitted_at": None,
                "total_questions": 0,
                "correct_count": 0,
                "feedback": "",   # 可暂合并所有 feedback
                "reviewer": interview.get('reviewer', ''),
                "interview_feedback": interview_feedback,
                "results": []
            }
        user_results[uid]["total_questions"] += 1
        if row.get('is_correct'):
            user_results[uid]["correct_count"] += 1
        if row.get('submitted_at') and (not user_results[uid]["submitted_at"] or row['submitted_at'] > user_results[uid]["submitted_at"]):
            user_results[uid]["submitted_at"] = row['submitted_at']

    # 转换为列表
    detail_list = []
    for uid, data in user_results.items():
        # 计算学员状态
        has_submitted = bool(data.get("submitted_at"))
        total_questions = data.get("total_questions", 0)
        answered_count = len([r for r in data.get("results", []) if r.get('answer')])
        
        if has_submitted or answered_count == total_questions:
            user_status = 'completed'
        else:
            # 检查访谈有效期
            start_time = interview.get('start_time')
            end_time = interview.get('end_time')
            now = datetime.now(timezone.utc).isoformat()
            
            if end_time and now > end_time:
                user_status = 'unfinished'  # 已过有效期未完成
            elif start_time and now < start_time:
                user_status = 'pending'     # 未开始
            else:
                user_status = 'ongoing'      # 进行中

        detail_list.append({
            "user_id": data["user_id"],
            "name": data["name"],
            "email": data['email'],
            "country": data["country"],
            "wh_id": data["wh_id"],
            "department": data["department"],
            "submitted_at": data["submitted_at"],
            "reviewer": data["reviewer"],
            "feedback": data["interview_feedback"],
            "total_questions": data["total_questions"],
            "correct_count": data["correct_count"],
            "feedback": data["feedback"],
            "has_submitted": bool(data["submitted_at"]),
            "user_status": user_status,
            "answered_count": answered_count,  # 已答题数量
            "can_force": user_status == 'unfinished'  # 是否可强制访谈
        })

    # 分页
    total = len(detail_list)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = detail_list[start:end]

    return jsonify({
        "interview": {
            "id": interview_id,
            "title": interview.get('title'),
            "exam_title": exam_title,
            "exam_id": interview.get('exam_id'),
            "reviewer": interview.get('reviewer', ''),
            "feedback": interview_feedback
        },
        "data": paginated,
        "total": total,
        "page": page,
        "per_page": per_page
    })

@admin_inspection_bp.route('/admin/interview/<int:interview_id>/details')
@login_required
@admin_required
def admin_interview_details_page(interview_id):
    """访谈详情页面（渲染HTML）"""
    return render_template('admin/list_inspection_details.html', interview_id=interview_id)

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/force_resample', methods=['POST'])
@login_required
@admin_required
def force_resample_interview(interview_id):
    """
    强制访谈：重新为指定用户抽题，设置2小时有效期
    """
    data = request.json
    user_ids = data.get('user_ids', [])
    
    if not user_ids:
        return jsonify({"success": False, "message": "please_select_users_for_force_interview", "params": []}), 400
    
    db = get_supabase()
    operator_id = session['user_id']
    
    # 获取原访谈信息
    interview_res = db.table("interviews").select("*").eq("id", interview_id).execute()
    if not interview_res.data:
        return jsonify({"success": False, "message": "interview_not_found", "params": []}), 404
    
    original_interview = interview_res.data[0]
    exam_id = original_interview.get('exam_id')
    question_count = original_interview.get('question_count', 5)
    reviewer = original_interview.get('reviewer', '')
    feedback = original_interview.get('feedback', '')
    title = original_interview.get('title', '访谈')
    
    # 检查 exam_id 是否存在
    if not exam_id:
        return jsonify({"success": False, "message": "interview_no_have_related_exam", "params": []}), 400
    
    # 检查题目是否存在
    q_check = db.table("questions").select("id").eq("exam_id", exam_id).limit(1).execute()
    if not q_check.data:
        return jsonify({"success": False, "message": "exam_no_have_questions", "params": []}), 400
    
    # 权限检查
    from utils.permissions import get_admin_allowed_countries
    allowed = get_admin_allowed_countries()
    if allowed is not None:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:

            exam_countries = parse_exam_countries(exam_res.data)
            if not any(c in allowed for c in exam_countries):
                return jsonify({"success": False, "message": "no_permission_for_interview", "params": []}), 403
    
    success_count = 0
    failed_count = 0
    skipped_count = 0
    errors = []
    
    # 设置强制访谈有效期：当前时间 + 2小时
    now = datetime.now(timezone.utc)
    force_start_time = now.isoformat()
    force_end_time = (now + timedelta(hours=2)).isoformat()
    
    for user_id in user_ids:
        try:
            logger.info(f"处理用户 {user_id}")
            
            # 1. 检查用户是否存在
            user_check = db.table("users").select("id, name_en").eq("id", user_id).maybe_single().execute()
            if not user_check.data:
                logger.error(f"用户 {user_id} 不存在")
                failed_count += 1
                errors.append(f"用户 {user_id}: 用户不存在")
                continue
            
            # 2. 检查用户是否已有答题记录（已完成）
            existing_results = db.table("interview_results").select("id, answer").eq("interview_id", interview_id).eq("user_id", user_id).execute()
            
            # 检查是否已完成答题
            if existing_results.data:
                all_answered = all(r.get('answer') for r in existing_results.data)
                if all_answered:
                    skipped_count += 1
                    errors.append(f"用户 {user_id} 已完成访谈，跳过")
                    logger.info(f"用户 {user_id} 已完成，跳过")
                    continue
            
            # 3. 先删除该用户的所有旧强制访谈记录（包括已软删除的）
            # 彻底清理，避免新旧记录冲突
            delete_result = db.table("user_interview_force_records").delete() \
                .eq("user_id", user_id) \
                .eq("original_interview_id", interview_id) \
                .execute()
            logger.info(f"用户 {user_id} 删除旧强制访谈记录，删除数量: {len(delete_result.data or []) if delete_result.data else 0}")
            
            # 4. 删除该用户的旧访谈结果记录
            db.table("interview_results").delete().eq("interview_id", interview_id).eq("user_id", user_id).execute()
            logger.info(f"用户 {user_id} 删除旧访谈结果记录")
            
            # 5. 重新抽题
            from routes.helpers import random_pick_questions
            questions = random_pick_questions(exam_id, question_count)
            logger.info(f"用户 {user_id} 抽取到 {len(questions)} 道题目")
            
            if not questions:
                logger.error(f"用户 {user_id} 抽题失败，没有题目")
                failed_count += 1
                errors.append(f"用户 {user_id}: 抽题失败，没有题目")
                continue
            
            for q in questions:
                db.table("interview_results").insert({
                    "interview_id": interview_id,
                    "user_id": user_id,
                    "question_id": q['id'],
                    "created_at": now.isoformat(),
                    "feedback": feedback
                }).execute()
                logger.debug(f"插入题目 {q['id']} 成功")
            
            # 6. 创建全新的强制访谈记录
            db.table("user_interview_force_records").insert({
                "user_id": user_id,
                "original_interview_id": interview_id,
                "exam_id": exam_id,
                # "title": f"{title} (强制访谈)",
                "title": title,
                "question_count": question_count,
                "reviewer": reviewer,
                "feedback": feedback,
                "start_time": force_start_time,
                "end_time": force_end_time,
                "created_at": now.isoformat(),
                "created_by": operator_id
            }).execute()
            logger.info(f"用户 {user_id} 创建新的强制访谈记录")
            
            success_count += 1
            
        except Exception as e:
            failed_count += 1
            error_msg = str(e)
            errors.append(f"用户 {user_id}: {error_msg}")
            logger.error(f"用户 {user_id} 强制访谈失败: {error_msg}", exc_info=True)
    
    logger.info(f"强制访谈完成: interview={interview_id}, 成功={success_count}, 跳过={skipped_count}, 失败={failed_count}")
    
    return jsonify({
        "success": True,
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        # "message": f"强制访谈完成: 成功 {success_count} 人，跳过已完成 {skipped_count} 人，失败 {failed_count} 人，有效期2小时",
        "message": "force_interview_complete", "params": [success_count, skipped_count, failed_count],
        "errors": errors[:10]
    })

# ============================================================
# 新增：已关闭访谈新增用户（创建强制访谈）
# ============================================================
@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/force_add_users', methods=['POST'])
@login_required
@admin_required
def force_add_users_for_closed_interview(interview_id):
    """已关闭访谈新增用户 - 创建强制访谈"""
    import sys
    from utils.permissions import get_admin_allowed_countries
    from datetime import datetime, timedelta, timezone
    
    # 强制输出函数（立即刷新）
    def log(msg):
        print(f"[FORCE_ADD] {msg}")
        sys.stdout.flush()
    
    data = request.json
    user_ids = data.get('user_ids', [])
    force_duration_hours = data.get('force_duration_hours', 24)
    
    log("=" * 60)
    log(f"🚀 开始强制访谈流程: interview_id={interview_id}")
    log(f"📋 接收到的 user_ids: {user_ids}")
    log(f"⏰ 强制时长: {force_duration_hours} 小时")
    log("=" * 60)
    
    if not user_ids:
        return jsonify({
            "success": False, 
            "message": "jsonify_please_select_users_to_add",
            "params": []
        }), 400
    
    db = get_supabase()
    admin_db = get_supabase_admin()
    operator_id = session['user_id']
    
    # 1. 获取原访谈信息
    interview_res = db.table("interviews").select("*").eq("id", interview_id).execute()
    if not interview_res.data:
        return jsonify({
            "success": False, 
            "message": "jsonify_interview_not_found",
            "params": []
        }), 404
    
    original_interview = interview_res.data[0]
    exam_id = original_interview.get('exam_id')
    question_count = original_interview.get('question_count', 5)
    reviewer = original_interview.get('reviewer', '')
    feedback = original_interview.get('feedback', '')
    title = original_interview.get('title', '访谈')
    
    log(f"📊 原访谈信息:")
    log(f"   - interview_id: {interview_id}")
    log(f"   - title: {title}")
    log(f"   - exam_id: {exam_id}")
    log(f"   - question_count (设定值): {question_count}")
    log(f"   - reviewer: {reviewer}")
    
    # 2. 检查考试和题目
    if not exam_id:
        return jsonify({
            "success": False, 
            "message": "jsonify_interview_no_exam",
            "params": []
        }), 400
    
    # 查询考试题目总数
    all_questions_res = db.table("questions").select("id").eq("exam_id", exam_id).execute()
    total_questions_in_exam = len(all_questions_res.data or [])
    
    q_check = db.table("questions").select("id").eq("exam_id", exam_id).limit(1).execute()
    if not q_check.data:
        return jsonify({
            "success": False, 
            "message": "jsonify_exam_no_questions",
            "params": []
        }), 400
    
    # 3. 权限检查
    allowed = get_admin_allowed_countries()
    if allowed is not None:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_countries = parse_exam_countries(exam_res.data)
            if not any(c in allowed for c in exam_countries):
                return jsonify({
                    "success": False, 
                    "message": "jsonify_no_permission_for_interview",
                    "params": []
                }), 403
    
    # 4. 获取已有用户列表
    existing_results = admin_db.table("interview_results") \
        .select("user_id") \
        .eq("interview_id", interview_id) \
        .execute()
    existing_user_ids = set(r['user_id'] for r in (existing_results.data or []))
    
    # 5. 过滤出真正需要新增的用户（去重 + 排除已有）
    unique_user_ids = list(set(user_ids))
    valid_user_ids = list(set(unique_user_ids) - existing_user_ids)

    # 确保没有重复
    valid_user_ids = list(set(valid_user_ids))

    if not valid_user_ids:
        return jsonify({
            "success": True,
            "message": "jsonify_all_users_already_in_interview",
            "params": [],
            "success_count": 0,
            "skipped_count": len(user_ids),
            "failed_count": 0
        })
    
    # 6. 验证用户是否存在
    users_res = db.table("users").select("id, name_en").in_("id", valid_user_ids).execute()
    existing_valid_ids = set(u['id'] for u in (users_res.data or []))
    valid_user_ids = [uid for uid in valid_user_ids if uid in existing_valid_ids]
    
    if not valid_user_ids:
        return jsonify({
            "success": False, 
            "message": "jsonify_selected_users_not_found",
            "params": []
        }), 400
    
    # 7. 设置强制访谈有效期
    now = datetime.now(timezone.utc)
    force_start_time = now.isoformat()
    force_end_time = (now + timedelta(hours=force_duration_hours)).isoformat()
    
    success_count = 0
    failed_count = 0
    failed_users = []
    processed_users = set()
    user_results = {}
    
    for user_id in valid_user_ids:
        if user_id in processed_users:
            continue
        processed_users.add(user_id)
        
        user_results[user_id] = {"user_id": user_id}
        
        try:
            # 7a. 清理旧的强制访谈记录
            admin_db.table("user_interview_force_records").delete() \
                .eq("user_id", user_id) \
                .eq("original_interview_id", interview_id) \
                .execute()
            
            # 7b. 检查并删除残留的 interview_results
            check_res = admin_db.table("interview_results").select("id") \
                .eq("interview_id", interview_id) \
                .eq("user_id", user_id) \
                .execute()
            
            before_delete_count = len(check_res.data or [])
            
            if before_delete_count > 0:
                admin_db.table("interview_results").delete() \
                    .eq("interview_id", interview_id) \
                    .eq("user_id", user_id) \
                    .execute()
            else:
                log(f"   ℹ️ 没有旧记录需要删除")
            
            # 验证删除结果
            verify_res = admin_db.table("interview_results").select("id") \
                .eq("interview_id", interview_id) \
                .eq("user_id", user_id) \
                .execute()
            after_delete_count = len(verify_res.data or [])
            
            # 7c. 抽题
            questions = random_pick_questions(exam_id, question_count)
            actual_count = len(questions)
            
            question_ids = [q.get('id') for q in questions]
            
            if not questions:
                failed_count += 1
                failed_users.append(user_id)
                user_results[user_id]["status"] = "failed"
                user_results[user_id]["error"] = "抽题失败，没有题目"
                continue
            
            # 7c1. 插入题目（带验证和重试）
            pre_check = admin_db.table("interview_results").select("id") \
                .eq("interview_id", interview_id) \
                .eq("user_id", user_id) \
                .execute()

            if len(pre_check.data or []) > 0:
                admin_db.table("interview_results").delete() \
                    .eq("interview_id", interview_id) \
                    .eq("user_id", user_id) \
                    .execute()

            inserted_count = 0
            inserted_question_ids = []

            for q in questions:
                result = admin_db.table("interview_results").insert({
                    "interview_id": interview_id,
                    "user_id": user_id,
                    "question_id": q['id'],
                    "created_at": now.isoformat(),
                    "feedback": feedback
                }).execute()
                if result.data:
                    inserted_count += 1
                    inserted_question_ids.append(q['id'])
                else:
                    log(f"   ⚠️ 插入题目 {q['id']} 失败")

            # 验证插入结果，如果不匹配则重试一次
            verify_insert_res = admin_db.table("interview_results").select("id, question_id") \
                .eq("interview_id", interview_id) \
                .eq("user_id", user_id) \
                .execute()
            actual_inserted = len(verify_insert_res.data or [])

            # 初始化 verify_retry_res 为 None
            verify_retry_res = None

            if actual_inserted != question_count:
                # 删除所有已插入的记录
                admin_db.table("interview_results").delete() \
                    .eq("interview_id", interview_id) \
                    .eq("user_id", user_id) \
                    .execute()
                
                # 重新插入所有题目
                for q in questions:
                    admin_db.table("interview_results").insert({
                        "interview_id": interview_id,
                        "user_id": user_id,
                        "question_id": q['id'],
                        "created_at": now.isoformat(),
                        "feedback": feedback
                    }).execute()
                
                # 再次验证
                verify_retry_res = admin_db.table("interview_results").select("id, question_id") \
                    .eq("interview_id", interview_id) \
                    .eq("user_id", user_id) \
                    .execute()
                retry_count = len(verify_retry_res.data or [])
                actual_inserted = retry_count

            # 根据 verify_retry_res 是否为空，选择正确的数据源
            if verify_retry_res is not None:
                inserted_question_ids = [r.get('question_id') for r in (verify_retry_res.data or [])]
            else:
                inserted_question_ids = [r.get('question_id') for r in (verify_insert_res.data or [])]
            
            # 7d. 创建强制访谈记录
            # force_title = f"{title} (强制访谈)"
            force_title = title
            admin_db.table("user_interview_force_records").insert({
                "user_id": user_id,
                "original_interview_id": interview_id,
                "exam_id": exam_id,
                "title": force_title,
                "question_count": actual_count,
                "reviewer": reviewer,
                "feedback": feedback,
                "start_time": force_start_time,
                "end_time": force_end_time,
                "created_at": now.isoformat(),
                "created_by": operator_id
            }).execute()
            
            success_count += 1
            user_results[user_id]["status"] = "success"
            user_results[user_id]["expected"] = question_count
            user_results[user_id]["actual"] = actual_count
            user_results[user_id]["question_ids"] = inserted_question_ids
            
        except Exception as e:
            failed_count += 1
            failed_users.append({"user_id": user_id, "error": str(e)})
            user_results[user_id]["status"] = "failed"
            user_results[user_id]["error"] = str(e)
            log(f"❌ 为用户 {user_id} 创建强制访谈失败: {e}")
            import traceback
            log(traceback.format_exc())
    
    # 8. 总结日志
    log("=" * 60)
    log("📊 强制访谈处理结果汇总:")
    log(f"   - 成功: {success_count} 人")
    log(f"   - 失败: {failed_count} 人")
    log("📊 每个用户的详细结果:")
    for uid, result in user_results.items():
        log(f"   - {uid}: {result}")
    log("=" * 60)
    
    if success_count > 0:
        return jsonify({
            "success": True,
            "message": "jsonify_force_interview_created_success",
            "params": [success_count, force_duration_hours],
            "success_count": success_count,
            "failed_count": failed_count,
            "failed_users": failed_users[:10],
            "force_end_time": force_end_time,
            "user_results": user_results
        })
    else:
        return jsonify({
            "success": False,
            "message": "jsonify_force_interview_creation_failed",
            "params": [],
            "failed_users": failed_users[:10]
        }), 500

# 访谈结果导出接口[Excel和PDF]

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/export_excel')
@login_required
@admin_required
def export_interview_report_excel(interview_id):
    """导出访谈结果为 Excel（包含所有学员答题详情）- 带权限隔离"""
    from utils.permissions import get_admin_allowed_countries, is_developer
    from routes.helpers import parse_exam_countries
    
    db = get_supabase()
    
    # ========== 1. 获取访谈基本信息 ==========
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    
    interview = inv_res.data
    
    # ========== 2. 权限检查 ==========
    is_dev = is_developer()
    current_role = session.get('role')
    allowed_countries = get_admin_allowed_countries()
    
    # 获取访谈关联的考试信息
    exam_id = interview.get('exam_id')
    exam_countries = []
    
    if exam_id:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_countries = parse_exam_countries(exam_res.data)
    
    # 检查当前用户是否有权限导出此访谈
    if not is_dev:
        # 超管逻辑
        if current_role == 'super_admin':
            if allowed_countries is not None and allowed_countries:
                # 检查访谈的国家是否在权限范围内
                if not any(c in allowed_countries for c in exam_countries):
                    return jsonify({"error": "无权导出此访谈数据"}), 403
        # 管理员逻辑
        elif current_role == 'admin':
            if allowed_countries is not None and allowed_countries:
                if not any(c in allowed_countries for c in exam_countries):
                    return jsonify({"error": "无权导出此访谈数据"}), 403
            else:
                # 无权限范围，使用用户注册国家
                user_country = session.get('user_country')
                if user_country not in exam_countries:
                    return jsonify({"error": "无权导出此访谈数据"}), 403
        else:
            return jsonify({"error": "无权导出此访谈数据"}), 403
    
    # ========== 3. 获取考试信息 ==========
    exam_title = ''
    if exam_id:
        exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_title = exam_res.data['title']
    
    # ========== 4. 获取所有学员的访谈答题记录 ==========
    results_res = db.table("interview_results").select("*").eq("interview_id", interview_id).execute()
    
    if not results_res.data:
        return jsonify({"error": "该访谈暂无数据"}), 404
    
    # ========== 5. 获取用户信息（带权限过滤） ==========
    user_ids = list(set([r['user_id'] for r in results_res.data]))
    
    # 获取用户信息（包括国家字段）
    users_map = {}
    if user_ids:
        users_res = db.table("users").select("id, name_cn, name_en, email, country, wh_id, department").in_("id", user_ids).execute()
        users_map = {u['id']: u for u in (users_res.data or [])}
    
    # ========== 6. 根据权限过滤用户 ==========
    filtered_user_ids = []
    
    if not is_dev:
        # 超管/管理员：只保留权限范围内的用户
        if allowed_countries is not None and allowed_countries:
            for uid, u in users_map.items():
                user_country = u.get('country', '')
                if user_country in allowed_countries:
                    filtered_user_ids.append(uid)
        else:
            # 没有权限范围限制（但超管可能有权限范围）
            if current_role == 'super_admin':
                # 超管无权限范围时，可以查看所有
                filtered_user_ids = user_ids
            else:
                # 管理员无权限范围时，只能查看自己国家的用户
                user_session_country = session.get('user_country')
                for uid, u in users_map.items():
                    if u.get('country') == user_session_country:
                        filtered_user_ids.append(uid)
    else:
        # 开发者：可以查看所有
        filtered_user_ids = user_ids
    
    # 过滤答题记录
    filtered_results = [r for r in results_res.data if r['user_id'] in filtered_user_ids]
    
    if not filtered_results:
        return jsonify({"error": "您权限范围内没有可导出的数据"}), 404
    
    # ========== 7. 获取题目信息 ==========
    question_ids = list(set([r['question_id'] for r in filtered_results if r.get('question_id')]))
    questions_map = {}
    if question_ids:
        q_res = db.table("questions").select("*").in_("id", question_ids).execute()
        for q in (q_res.data or []):
            opts = q.get('options', {})
            if isinstance(opts, str):
                try:
                    opts = json.loads(opts)
                except:
                    opts = {}
            q['options_parsed'] = opts
            questions_map[q['id']] = q
    
    # ========== 8. 按用户聚合答题数据 ==========
    user_answers = {}
    for r in filtered_results:
        uid = r['user_id']
        if uid not in user_answers:
            user_answers[uid] = {
                'user_id': uid,
                'answers': [],
                'correct_count': 0,
                'total_questions': 0,
                'submitted_at': r.get('submitted_at')
            }
        
        q = questions_map.get(r['question_id'], {})
        is_correct = r.get('is_correct', False)
        if is_correct:
            user_answers[uid]['correct_count'] += 1
        user_answers[uid]['total_questions'] += 1
        
        user_answers[uid]['answers'].append({
            'question_num': q.get('num', 0),
            'question_content': q.get('content_cn') or q.get('content') or q.get('content_raw', ''),
            'question_type': q.get('type', 'single'),
            'options': q.get('options_parsed', {}),
            'correct_answer': q.get('answer', ''),
            'user_answer': r.get('answer', '未作答'),
            'is_correct': is_correct
        })
    
    # ========== 9. 创建 Excel 工作簿 ==========
    wb = openpyxl.Workbook()
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # ========== Sheet 1: 汇总表 ==========
    ws1 = wb.active
    ws1.title = "访谈汇总"
    
    # 写入标题信息
    ws1.merge_cells('A1:H1')
    title_cell = ws1.cell(row=1, column=1, value=f"访谈结果汇总 - {interview.get('title', '')}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # 添加权限范围说明
    ws1.cell(row=2, column=1, value="导出范围:").font = Font(bold=True)
    if is_dev:
        ws1.cell(row=2, column=2, value="全部（开发者）")
    elif allowed_countries:
        ws1.cell(row=2, column=2, value=f"权限范围: {', '.join(allowed_countries)}")
    else:
        ws1.cell(row=2, column=2, value=session.get('user_country', '全部'))
    
    ws1.cell(row=3, column=1, value="访谈ID:").font = Font(bold=True)
    ws1.cell(row=3, column=2, value=interview_id)
    ws1.cell(row=3, column=3, value="考试名称:").font = Font(bold=True)
    ws1.cell(row=3, column=4, value=exam_title)
    ws1.cell(row=3, column=5, value="阅卷人:").font = Font(bold=True)
    ws1.cell(row=3, column=6, value=interview.get('reviewer', ''))
    ws1.cell(row=3, column=7, value="反馈人:").font = Font(bold=True)
    ws1.cell(row=3, column=8, value=interview.get('feedback', ''))
    
    # 汇总表头
    headers1 = ['序号', '学员姓名', '邮箱', '国家', '库房编码', '部门', '答对数量', '总题数', '正确率', '提交时间']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入汇总数据
    row_idx = 6
    for idx, (uid, data) in enumerate(user_answers.items(), 1):
        user_info = users_map.get(uid, {})
        correct_rate = f"{data['correct_count']/data['total_questions']*100:.1f}%" if data['total_questions'] > 0 else "0%"
        
        ws1.cell(row=row_idx, column=1, value=idx).border = border
        ws1.cell(row=row_idx, column=2, value=user_info.get('name_cn') or user_info.get('name_en', '')).border = border
        ws1.cell(row=row_idx, column=3, value=user_info.get('email', '')).border = border
        ws1.cell(row=row_idx, column=4, value=user_info.get('country', '')).border = border
        ws1.cell(row=row_idx, column=5, value=user_info.get('wh_id', '')).border = border
        ws1.cell(row=row_idx, column=6, value=user_info.get('department', '')).border = border
        ws1.cell(row=row_idx, column=7, value=data['correct_count']).border = border
        ws1.cell(row=row_idx, column=8, value=data['total_questions']).border = border
        ws1.cell(row=row_idx, column=9, value=correct_rate).border = border
        ws1.cell(row=row_idx, column=10, value=data.get('submitted_at', '')[:16] if data.get('submitted_at') else '-').border = border
        
        # 正确率高亮
        if data['correct_count'] == data['total_questions']:
            ws1.cell(row=row_idx, column=9).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        row_idx += 1
    
    # 调整列宽
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 16
    
    # ========== Sheet 2: 答题详情 ==========
    ws2 = wb.create_sheet("答题详情")
    
    # 答题详情表头
    headers2 = ['学员姓名', '邮箱', '国家', '题号', '题目内容', '题型', '选项', '正确答案', '学员答案', '是否正确']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入答题详情数据
    row_idx = 2
    for uid, data in user_answers.items():
        user_info = users_map.get(uid, {})
        user_name = user_info.get('name_cn') or user_info.get('name_en', '')
        user_email = user_info.get('email', '')
        user_country = user_info.get('country', '')
        
        # 按题号排序
        sorted_answers = sorted(data['answers'], key=lambda x: x.get('question_num', 0))
        
        for ans in sorted_answers:
            # 格式化选项
            options_str = ''
            if ans.get('options'):
                opts = ans['options']
                options_str = '; '.join([f"{k}: {v}" for k, v in opts.items() if v])
            
            ws2.cell(row=row_idx, column=1, value=user_name).border = border
            ws2.cell(row=row_idx, column=2, value=user_email).border = border
            ws2.cell(row=row_idx, column=3, value=user_country).border = border
            ws2.cell(row=row_idx, column=4, value=ans.get('question_num', 0)).border = border
            ws2.cell(row=row_idx, column=5, value=ans.get('question_content', '')).border = border
            ws2.cell(row=row_idx, column=6, value=ans.get('question_type', '')).border = border
            ws2.cell(row=row_idx, column=7, value=options_str).border = border
            ws2.cell(row=row_idx, column=8, value=ans.get('correct_answer', '')).border = border
            ws2.cell(row=row_idx, column=9, value=ans.get('user_answer', '')).border = border
            ws2.cell(row=row_idx, column=10, value='正确' if ans.get('is_correct') else '错误').border = border
            
            # 正确/错误高亮
            if ans.get('is_correct'):
                ws2.cell(row=row_idx, column=10).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws2.cell(row=row_idx, column=10).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row_idx += 1
    
    # 调整列宽
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions[get_column_letter(5)].width = 40
    
    # ========== 保存文件 ==========
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"访谈结果_{interview.get('title', 'interview')}_{timestamp}.xlsx"
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/export_pdf')
@login_required
@admin_required
def export_interview_pdf(interview_id):
    """导出访谈结果为 PDF（带权限隔离）"""
    db = get_supabase()
    
    # ========== 1. 获取访谈基本信息 ==========
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    
    interview = inv_res.data
    
    # ========== 2. 权限检查 ==========
    is_dev = is_developer()
    current_role = session.get('role')
    allowed_countries = get_admin_allowed_countries()
    
    # 获取访谈关联的考试信息
    exam_id = interview.get('exam_id')
    exam_countries = []
    
    if exam_id:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_countries = parse_exam_countries(exam_res.data)
    
    # 检查当前用户是否有权限导出此访谈
    if not is_dev:
        if current_role == 'super_admin':
            if allowed_countries is not None and allowed_countries:
                if not any(c in allowed_countries for c in exam_countries):
                    return jsonify({"error": "无权导出此访谈数据"}), 403
        elif current_role == 'admin':
            if allowed_countries is not None and allowed_countries:
                if not any(c in allowed_countries for c in exam_countries):
                    return jsonify({"error": "无权导出此访谈数据"}), 403
            else:
                user_country = session.get('user_country')
                if user_country not in exam_countries:
                    return jsonify({"error": "无权导出此访谈数据"}), 403
        else:
            return jsonify({"error": "无权导出此访谈数据"}), 403
    
    # ========== 3. 获取关联考试信息 ==========
    exam_title = ''
    if exam_id:
        exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_title = exam_res.data['title']
    
    # ========== 4. 获取所有学员的访谈答题记录 ==========
    results_res = db.table("interview_results").select("*").eq("interview_id", interview_id).execute()
    
    if not results_res.data:
        return jsonify({"error": "该访谈暂无数据"}), 404
    
    # ========== 5. 获取用户信息 ==========
    user_ids = list(set([r['user_id'] for r in results_res.data]))
    
    users_map = {}
    if user_ids:
        users_res = db.table("users").select("id, name_cn, name_en, email, country, wh_id, department").in_("id", user_ids).execute()
        users_map = {u['id']: u for u in (users_res.data or [])}
    
    # ========== 6. 根据权限过滤用户 ==========
    filtered_user_ids = []
    
    if not is_dev:
        if allowed_countries is not None and allowed_countries:
            for uid, u in users_map.items():
                user_country = u.get('country', '')
                if user_country in allowed_countries:
                    filtered_user_ids.append(uid)
        else:
            if current_role == 'super_admin':
                filtered_user_ids = user_ids
            else:
                user_session_country = session.get('user_country')
                for uid, u in users_map.items():
                    if u.get('country') == user_session_country:
                        filtered_user_ids.append(uid)
    else:
        filtered_user_ids = user_ids
    
    # 过滤答题记录
    filtered_results = [r for r in results_res.data if r['user_id'] in filtered_user_ids]
    
    if not filtered_results:
        return jsonify({"error": "您权限范围内没有可导出的数据"}), 404
    
    # ========== 7. 获取题目信息 ==========
    question_ids = list(set([r['question_id'] for r in filtered_results if r.get('question_id')]))
    questions_map = {}
    if question_ids:
        q_res = db.table("questions").select("*").in_("id", question_ids).execute()
        for q in (q_res.data or []):
            opts = q.get('options', {})
            if isinstance(opts, str):
                try:
                    opts = json.loads(opts)
                except:
                    opts = {}
            q['options_parsed'] = opts
            questions_map[q['id']] = q
    
    # ========== 8. 按用户聚合答题数据 ==========
    user_answers = {}
    for r in filtered_results:
        uid = r['user_id']
        if uid not in user_answers:
            user_answers[uid] = {
                'user_id': uid,
                'answers': [],
                'correct_count': 0,
                'total_questions': 0,
                'submitted_at': r.get('submitted_at')
            }
        
        q = questions_map.get(r['question_id'], {})
        is_correct = r.get('is_correct', False)
        if is_correct:
            user_answers[uid]['correct_count'] += 1
        user_answers[uid]['total_questions'] += 1
        
        user_answers[uid]['answers'].append({
            'question_num': q.get('num', 0),
            'question_content': q.get('content_cn') or q.get('content') or q.get('content_raw', ''),
            'question_type': q.get('type', 'single'),
            'options': q.get('options_parsed', {}),
            'correct_answer': q.get('answer', ''),
            'user_answer': r.get('answer', '未作答'),
            'is_correct': is_correct
        })
    
    # ========== 9. 生成 HTML 内容 ==========
    html_content = generate_interview_pdf_html(
        interview, 
        exam_title, 
        users_map, 
        user_answers,
        interview_id,
        is_dev,
        allowed_countries
    )
    
    # ========== 10. 配置 wkhtmltopdf 选项 ==========
    options = {
        'page-size': 'A4',
        'margin-top': '15mm',
        'margin-right': '15mm',
        'margin-bottom': '20mm',
        'margin-left': '15mm',
        'encoding': 'UTF-8',
        'footer-right': '第 [page] 页 / 共 [topage] 页',
        'footer-font-size': '8',
        'footer-spacing': '5',
    }
    
    try:
        wkhtmltopdf_path = os.environ.get('WKHTMLTOPDF_PATH')
        
        if not wkhtmltopdf_path or not os.path.exists(wkhtmltopdf_path):
            common_paths = [
                r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe',
                r'C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe',
                '/usr/local/bin/wkhtmltopdf',
                '/usr/bin/wkhtmltopdf',
            ]
            for path in common_paths:
                if os.path.exists(path):
                    wkhtmltopdf_path = path
                    break
        
        config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path) if wkhtmltopdf_path else None
        pdf_bytes = pdfkit.from_string(html_content, False, options=options, configuration=config)
        
        from io import BytesIO
        buffer = BytesIO(pdf_bytes)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"访谈结果_{interview.get('title', 'interview')}_{timestamp}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"PDF 生成失败: {e}")
        return jsonify({"error": f"PDF 生成失败: {str(e)}"}), 500

def generate_interview_pdf_html(interview, exam_title, users_map, user_answers, interview_id, is_dev=False, allowed_countries=None):
    """生成访谈报告的 HTML 内容"""
    
    # 计算统计信息
    total_users = len(user_answers)
    total_completed = sum(1 for u in user_answers.values() if u.get('submitted_at'))
    total_correct = sum(u['correct_count'] for u in user_answers.values())
    total_questions = sum(u['total_questions'] for u in user_answers.values())
    avg_correct_rate = f"{(total_correct / total_questions * 100):.1f}%" if total_questions > 0 else "0%"

    # 构建权限范围显示
    scope_display = '全部（开发者）' if is_dev else (', '.join(allowed_countries) if allowed_countries else session.get('user_country', '全部'))
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>访谈结果报告 - {interview.get('title', '')}</title>
        <style>
            /* ==================== 全局样式 ==================== */
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            body {{
                font-family: 'Microsoft YaHei', 'PingFang SC', Arial, sans-serif;
                font-size: 12px;
                line-height: 1.6;
                color: #333;
                padding: 20px;
                background: #fff;
            }}
            
            /* ==================== 封面样式 ==================== */
            .cover-page {{
                page-break-after: always;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                height: 100vh;
                text-align: center;
                padding: 40px;
            }}
            .cover-page .logo {{
                font-size: 48px;
                margin-bottom: 30px;
            }}
            .cover-page h1 {{
                font-size: 28px;
                color: #1a1a2e;
                margin-bottom: 10px;
            }}
            .cover-page .subtitle {{
                font-size: 16px;
                color: #666;
                margin-bottom: 40px;
            }}
            .cover-page .info-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 10px 40px;
                text-align: left;
                font-size: 14px;
                background: #f8f9fa;
                padding: 20px 40px;
                border-radius: 8px;
                border: 1px solid #e9ecef;
            }}
            .cover-page .info-grid .label {{
                font-weight: 600;
                color: #555;
            }}
            .cover-page .footer {{
                margin-top: 60px;
                color: #999;
                font-size: 12px;
                border-top: 1px solid #eee;
                padding-top: 20px;
                width: 100%;
            }}
            
            /* ==================== 报告内容样式 ==================== */
            .report-page {{
                page-break-after: always;
                padding: 20px 0;
            }}
            
            /* 摘要卡片 */
            .summary-cards {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 15px;
                margin-bottom: 25px;
            }}
            .summary-card {{
                background: #f8f9fa;
                border-radius: 8px;
                padding: 15px 20px;
                text-align: center;
                border: 1px solid #e9ecef;
            }}
            .summary-card .number {{
                font-size: 28px;
                font-weight: 700;
                color: #1a1a2e;
            }}
            .summary-card .label {{
                font-size: 12px;
                color: #888;
                margin-top: 4px;
            }}
            .summary-card.blue .number {{ color: #0d6efd; }}
            .summary-card.green .number {{ color: #198754; }}
            .summary-card.orange .number {{ color: #fd7e14; }}
            .summary-card.purple .number {{ color: #6f42c1; }}
            
            /* 表格样式 */
            .table-container {{
                margin-top: 20px;
            }}
            .table-title {{
                font-size: 16px;
                font-weight: 600;
                margin-bottom: 10px;
                color: #1a1a2e;
                border-left: 4px solid #0d6efd;
                padding-left: 10px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 11px;
            }}
            table thead th {{
                background: #0d6efd;
                color: #fff;
                padding: 8px 10px;
                text-align: left;
                border: 1px solid #0d6efd;
            }}
            table tbody td {{
                padding: 6px 10px;
                border: 1px solid #dee2e6;
            }}
            table tbody tr:nth-child(even) {{
                background: #f8f9fa;
            }}
            table tbody tr:hover {{
                background: #e7f3ff;
            }}
            
            .text-success {{ color: #198754; }}
            .text-danger {{ color: #dc3545; }}
            .text-center {{ text-align: center; }}
            
            /* 答题详情 */
            .answer-detail {{
                margin-top: 20px;
                border: 1px solid #e9ecef;
                border-radius: 8px;
                overflow: hidden;
            }}
            .answer-detail .user-header {{
                background: #e7f3ff;
                padding: 10px 15px;
                font-weight: 600;
                display: flex;
                justify-content: space-between;
                border-bottom: 1px solid #dee2e6;
            }}
            .answer-detail .user-header .score {{
                color: #0d6efd;
            }}
            .answer-item {{
                padding: 8px 15px;
                border-bottom: 1px solid #f0f0f0;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }}
            .answer-item:last-child {{
                border-bottom: none;
            }}
            .answer-item .question {{
                flex: 1;
            }}
            .answer-item .result {{
                font-weight: 600;
                margin-left: 15px;
            }}
            
            /* 页脚 */
            .page-footer {{
                text-align: center;
                color: #999;
                font-size: 10px;
                margin-top: 20px;
                padding-top: 10px;
                border-top: 1px solid #eee;
            }}
            
            /* 打印分页 */
            @media print {{
                .cover-page {{
                    height: 100vh;
                }}
                .no-print {{
                    display: none;
                }}
            }}
        </style>
    </head>
    <body>
        
        <!-- ==================== 封面页 ==================== -->
        <div class="cover-page">
            <div class="logo">📋</div>
            <h1>访谈结果报告</h1>
            <div class="subtitle">{interview.get('title', '')}</div>
            <div class="info-grid">
                <span><span class="label">访谈ID</span> {interview_id}</span>
                <span><span class="label">关联考试</span> {exam_title or '-'}</span>
                <span><span class="label">阅卷人</span> {interview.get('reviewer', '-')}</span>
                <span><span class="label">反馈人</span> {interview.get('feedback', '-')}</span>
                <span><span class="label">创建时间</span> {interview.get('created_at', '')[:16] if interview.get('created_at') else '-'}</span>
                <span><span class="label">报告生成</span> {datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
            </div>
            <div class="footer">
                <span>培训考试系统 · 访谈报告</span>
            </div>
        </div>
        
        <!-- ==================== 第二页：摘要统计 ==================== -->
        <div class="report-page">
            <h2 style="margin-bottom: 20px;">📊 统计摘要</h2>
            
            <div class="summary-cards">
                <div class="summary-card blue">
                    <div class="number">{total_users}</div>
                    <div class="label">参访人数</div>
                </div>
                <div class="summary-card green">
                    <div class="number">{total_completed}</div>
                    <div class="label">已完成</div>
                </div>
                <div class="summary-card orange">
                    <div class="number">{total_questions}</div>
                    <div class="label">答题总数</div>
                </div>
                <div class="summary-card purple">
                    <div class="number">{avg_correct_rate}</div>
                    <div class="label">平均正确率</div>
                </div>
            </div>
            
            <!-- 学员汇总表 -->
            <div class="table-container">
                <div class="table-title">📋 学员答题汇总</div>
                <table>
                    <thead>
                        <tr>
                            <th style="width: 50px;">序号</th>
                            <th>姓名</th>
                            <th>邮箱</th>
                            <th>国家</th>
                            <th>答对/总题</th>
                            <th style="width: 80px;">正确率</th>
                            <th>提交时间</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    # 按正确率排序
    sorted_users = sorted(
        user_answers.items(),
        key=lambda x: x[1]['correct_count'] / x[1]['total_questions'] if x[1]['total_questions'] > 0 else 0,
        reverse=True
    )
    
    for idx, (uid, data) in enumerate(sorted_users, 1):
        user_info = users_map.get(uid, {})
        correct_rate = f"{(data['correct_count'] / data['total_questions'] * 100):.1f}%" if data['total_questions'] > 0 else "0%"
        is_completed = bool(data.get('submitted_at'))
        submitted_display = data.get('submitted_at', '')[:16] if data.get('submitted_at') else ('未提交' if not is_completed else '-')
        
        html += f"""
                        <tr>
                            <td class="text-center">{idx}</td>
                            <td>{user_info.get('name_cn') or user_info.get('name_en', '')}</td>
                            <td>{user_info.get('email', '')}</td>
                            <td>{user_info.get('country', '')}</td>
                            <td>{data['correct_count']} / {data['total_questions']}</td>
                            <td class="text-center"><strong>{correct_rate}</strong></td>
                            <td>{submitted_display}</td>
                        </tr>
        """
    
    html += f"""
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- ==================== 后续页面：答题详情 ==================== -->
        <div class="report-page">
            <h2 style="margin-bottom: 20px;">📝 答题详情</h2>
    """
    
    # 按用户显示答题详情
    for uid, data in sorted_users:
        user_info = users_map.get(uid, {})
        user_name = user_info.get('name_cn') or user_info.get('name_en', '')
        correct_rate = f"{(data['correct_count'] / data['total_questions'] * 100):.1f}%" if data['total_questions'] > 0 else "0%"
        
        html += f"""
            <div class="answer-detail">
                <div class="user-header">
                    <span>👤 {user_name}</span>
                    <span class="score">正确率: {correct_rate} ({data['correct_count']}/{data['total_questions']})</span>
                </div>
        """
        
        # 按题号排序
        sorted_answers = sorted(data['answers'], key=lambda x: x.get('question_num', 0))
        
        for ans in sorted_answers:
            status_icon = '✅' if ans.get('is_correct') else '❌'
            status_color = 'text-success' if ans.get('is_correct') else 'text-danger'
            
            # 限制题目内容长度
            content = ans.get('question_content', '')
            if len(content) > 60:
                content = content[:60] + '...'
            
            html += f"""
                <div class="answer-item">
                    <span class="question">
                        <strong>Q{ans.get('question_num', 0)}.</strong> {content}
                        <span style="color: #999; font-size: 11px; margin-left: 8px;">[{ans.get('question_type', '')}]</span>
                        <span style="color: #666; font-size: 11px; margin-left: 8px;">
                            答案: {ans.get('user_answer', '未作答')}
                        </span>
                    </span>
                    <span class="result {status_color}">{status_icon} {'正确' if ans.get('is_correct') else '错误'}</span>
                </div>
            """
        
        html += """
            </div>
        """
    
    html += f"""
        </div>
        
        <!-- ==================== 页脚 ==================== -->
        <div class="page-footer">
            培训考试系统 · 访谈报告 · 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
        
    </body>
    </html>
    """
    
    return html

@admin_inspection_bp.route('/api/admin/interview/<int:interview_id>/export_batch_pdf', methods=['POST'])
@login_required
@admin_required
def export_batch_interview_pdf(interview_id):
    """
    批量导出访谈结果为 PDF（支持全部或选定学员）
    """
    data = request.json or {}
    user_ids = data.get('user_ids', [])  # 如果为空，则导出全部
    
    db = get_supabase()
    
    # 1. 获取访谈基本信息
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    
    interview = inv_res.data
    
    # 2. 权限检查
    is_dev = is_developer()
    current_role = session.get('role')
    allowed_countries = get_admin_allowed_countries()
    
    exam_id = interview.get('exam_id')
    exam_countries = []
    if exam_id:
        exam_res = db.table("exams").select("countries, country").eq("id", exam_id).maybe_single().execute()
        if exam_res.data:
            exam_countries = parse_exam_countries(exam_res.data)
    
    if not is_dev:
        if current_role == 'super_admin':
            if allowed_countries is not None and allowed_countries:
                if not any(c in allowed_countries for c in exam_countries):
                    return jsonify({"error": "无权导出此访谈数据"}), 403
        elif current_role == 'admin':
            if allowed_countries is not None and allowed_countries:
                if not any(c in allowed_countries for c in exam_countries):
                    return jsonify({"error": "无权导出此访谈数据"}), 403
            else:
                user_country = session.get('user_country')
                if user_country not in exam_countries:
                    return jsonify({"error": "无权导出此访谈数据"}), 403
        else:
            return jsonify({"error": "无权导出此访谈数据"}), 403
    
    # 3. 获取所有答题记录
    results_res = db.table("interview_results").select("*").eq("interview_id", interview_id).execute()
    if not results_res.data:
        return jsonify({"error": "该访谈暂无数据"}), 404
    
    # 4. 获取用户信息
    all_user_ids = list(set([r['user_id'] for r in results_res.data]))
    users_map = {}
    if all_user_ids:
        users_res = db.table("users").select("id, name_cn, name_en, email, country, wh_id, department").in_("id", all_user_ids).execute()
        users_map = {u['id']: u for u in (users_res.data or [])}
    
    # 5. 权限过滤
    filtered_user_ids = []
    if not is_dev:
        if allowed_countries is not None and allowed_countries:
            for uid, u in users_map.items():
                if u.get('country', '') in allowed_countries:
                    filtered_user_ids.append(uid)
        else:
            if current_role == 'super_admin':
                filtered_user_ids = all_user_ids
            else:
                user_session_country = session.get('user_country')
                for uid, u in users_map.items():
                    if u.get('country') == user_session_country:
                        filtered_user_ids.append(uid)
    else:
        filtered_user_ids = all_user_ids
    
    # 6. 确定要导出的用户
    if user_ids:
        # 只导出选中的用户（同时验证权限）
        target_user_ids = [uid for uid in user_ids if uid in filtered_user_ids]
    else:
        target_user_ids = filtered_user_ids
    
    if not target_user_ids:
        return jsonify({"error": "没有可导出的学员数据"}), 404
    
    # 7. 获取题目信息
    question_ids = list(set([r['question_id'] for r in results_res.data if r.get('question_id')]))
    questions_map = {}
    if question_ids:
        q_res = db.table("questions").select("*").in_("id", question_ids).execute()
        for q in (q_res.data or []):
            opts = q.get('options', {})
            if isinstance(opts, str):
                try:
                    opts = json.loads(opts)
                except:
                    opts = {}
            q['options_parsed'] = opts
            questions_map[q['id']] = q
    
    # 8. 按用户聚合答题数据
    user_answers = {}
    for r in results_res.data:
        uid = r['user_id']
        if uid not in target_user_ids:
            continue
        if uid not in user_answers:
            user_answers[uid] = {
                'user_id': uid,
                'answers': [],
                'correct_count': 0,
                'total_questions': 0,
                'submitted_at': r.get('submitted_at')
            }
        
        q = questions_map.get(r['question_id'], {})
        is_correct = r.get('is_correct', False)
        if is_correct:
            user_answers[uid]['correct_count'] += 1
        user_answers[uid]['total_questions'] += 1
        
        user_answers[uid]['answers'].append({
            'question_num': q.get('num', 0),
            'question_content': q.get('content_cn') or q.get('content') or q.get('content_raw', ''),
            'question_type': q.get('type', 'single'),
            'options': q.get('options_parsed', {}),
            'correct_answer': q.get('answer', ''),
            'user_answer': r.get('answer', '未作答'),
            'is_correct': is_correct
        })
    
    # 9. 配置 wkhtmltopdf
    options = {
        'page-size': 'A4',
        'margin-top': '15mm',
        'margin-right': '15mm',
        'margin-bottom': '20mm',
        'margin-left': '15mm',
        'encoding': 'UTF-8',
        'footer-right': '第 [page] 页 / 共 [topage] 页',
        'footer-font-size': '8',
        'footer-spacing': '5',
    }
    
    wkhtmltopdf_path = os.environ.get('WKHTMLTOPDF_PATH')
    if not wkhtmltopdf_path or not os.path.exists(wkhtmltopdf_path):
        common_paths = [
            r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe',
            r'C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe',
            '/usr/local/bin/wkhtmltopdf',
            '/usr/bin/wkhtmltopdf',
        ]
        for path in common_paths:
            if os.path.exists(path):
                wkhtmltopdf_path = path
                break
    
    config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path) if wkhtmltopdf_path else None
    
    # 10. 为每个用户生成 PDF
    pdf_files = []
    for uid, data in user_answers.items():
        user_info = users_map.get(uid, {})
        user_name = user_info.get('name_cn') or user_info.get('name_en', '未知')
        
        # 生成单个用户的 HTML
        html_content = generate_interview_detail_pdf_html(
            interview, 
            user_info,
            data,
            exam_id,
            interview_id
        )
        
        try:
            pdf_bytes = pdfkit.from_string(html_content, False, options=options, configuration=config)
            pdf_files.append({
                'user_name': user_name,
                'user_id': uid,
                'pdf_bytes': pdf_bytes
            })
        except Exception as e:
            logger.error(f"为用户 {user_name} 生成 PDF 失败: {e}")
            continue
    
    if not pdf_files:
        return jsonify({"error": "PDF 生成失败"}), 500
    
    # 11. 如果只有一个用户，直接返回 PDF
    if len(pdf_files) == 1:
        buffer = BytesIO(pdf_files[0]['pdf_bytes'])
        buffer.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"访谈_{interview.get('title', 'interview')}_{pdf_files[0]['user_name']}_{timestamp}.pdf"
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    # 12. 多个用户，打包成 ZIP
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for pdf in pdf_files:
            filename = f"{pdf['user_name']}_{pdf['user_id']}.pdf"
            zip_file.writestr(filename, pdf['pdf_bytes'])
    
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"访谈结果_{interview.get('title', 'interview')}_{timestamp}.zip"
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename
    )

def generate_interview_detail_pdf_html(interview, user_info, user_data, exam_id, interview_id):
    """
    生成单个学员的访谈详情 PDF HTML（以查看详情布局为准）
    """
    from datetime import datetime
    
    user_name = user_info.get('name_cn') or user_info.get('name_en', '未知')
    user_email = user_info.get('email', '')
    user_country = user_info.get('country', '')
    user_wh_id = user_info.get('wh_id', '')
    
    total_questions = user_data['total_questions']
    correct_count = user_data['correct_count']
    correct_rate = f"{(correct_count / total_questions * 100):.1f}%" if total_questions > 0 else "0%"
    
    # 格式化提交时间
    submitted_at = user_data.get('submitted_at', '')
    if submitted_at:
        try:
            from dateutil import parser
            dt = parser.parse(submitted_at)
            submitted_display = dt.strftime('%Y-%m-%d %H:%M')
        except:
            submitted_display = submitted_at[:16] if len(submitted_at) > 16 else submitted_at
    else:
        submitted_display = '未提交'
    
    # 生成答题详情 HTML
    answers_html = ''
    sorted_answers = sorted(user_data['answers'], key=lambda x: x.get('question_num', 0))
    
    for idx, ans in enumerate(sorted_answers, 1):
        status_icon = '✅' if ans.get('is_correct') else '❌'
        status_color = 'color: #198754;' if ans.get('is_correct') else 'color: #dc3545;'
        status_text = '正确Correct' if ans.get('is_correct') else '错误Wrong'
        
        # 题型标签
        type_labels = {
            'single': '单选题Single',
            'multi': '多选题Multi',
            'judge': '判断题Judge'
        }
        type_badge = type_labels.get(ans.get('question_type', ''), ans.get('question_type', ''))
        
        # 选项显示
        options_html = ''
        if ans.get('options') and ans.get('question_type') != 'judge':
            opts = ans['options']
            for k, v in opts.items():
                if v:
                    options_html += f'<div style="padding: 2px 0;">{k}. {v}</div>'
        elif ans.get('question_type') == 'judge':
            options_html = '<div>正确 True | 错误 False</div>'
        
        answers_html += f"""
        <div style="border: 1px solid #e9ecef; border-radius: 6px; padding: 12px 15px; margin-bottom: 10px; page-break-inside: avoid;">
            <div style="display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap;">
                <div style="flex: 1; min-width: 200px;">
                    <div style="font-weight: 600; margin-bottom: 4px;">
                        <span style="color: #0d6efd;">Q{ans.get('question_num', 0)}.</span> 
                        {ans.get('question_content', '')}
                        <span style="background: #e9ecef; padding: 1px 8px; border-radius: 10px; font-size: 11px; margin-left: 8px; color: #555;">{type_badge}</span>
                    </div>
                    <div style="margin-left: 20px; font-size: 13px; color: #555;">
                        {options_html}
                    </div>
                    <div style="margin-top: 6px; font-size: 13px;">
                        <span style="color: #0d6efd; font-weight: 500;">正确答案Correct Answer：</span>
                        <span style="color: #198754;">{ans.get('correct_answer', '')}</span>
                        <span style="margin-left: 15px; color: #0d6efd; font-weight: 500;">学员答案CA：</span>
                        <span style="font-weight: 500;">{ans.get('user_answer', '未作答No answer')}</span>
                    </div>
                </div>
                <div style="font-weight: 700; font-size: 16px; {status_color} padding: 4px 12px; border-radius: 4px; white-space: nowrap;">
                    {status_icon} {status_text}
                </div>
            </div>
        </div>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>访谈答题详情 - {user_name}</title>
        <style>
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            body {{
                font-family: 'Microsoft YaHei', 'PingFang SC', Arial, sans-serif;
                font-size: 12px;
                line-height: 1.6;
                color: #333;
                padding: 20px;
                background: #fff;
            }}
            .header {{
                border-bottom: 2px solid #0d6efd;
                padding-bottom: 15px;
                margin-bottom: 20px;
            }}
            .header h1 {{
                font-size: 20px;
                color: #1a1a2e;
            }}
            .header .subtitle {{
                color: #666;
                font-size: 14px;
                margin-top: 4px;
            }}
            
            /* ===== 用户信息 - Table 布局 ===== */
            .user-info-table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                table-layout: fixed;
                background: #f8f9fa;
                border-radius: 8px;
                overflow: hidden;
            }}
            .user-info-table td {{
                padding: 10px 16px;
                text-align: center;
                width: 25%;
                border: none;
                vertical-align: middle;
            }}
            .user-info-table td:not(:last-child) {{
                border-right: 1px solid #e9ecef;
            }}
            .user-info-table .label {{
                display: block;
                color: #888;
                font-size: 11px;
                font-weight: 400;
                letter-spacing: 0.3px;
                text-transform: uppercase;
                margin-bottom: 2px;
            }}
            .user-info-table .value {{
                display: block;
                font-weight: 600;
                font-size: 14px;
                color: #1a1a2e;
            }}
            
            /* ===== 统计摘要 - Table 布局 ===== */
            .summary-table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                table-layout: fixed;
            }}
            .summary-table td {{
                background: #e7f3ff;
                padding: 10px 16px;
                text-align: center;
                width: 25%;
                border: none;
            }}
            .summary-table td:first-child {{
                border-radius: 6px 0 0 6px;
            }}
            .summary-table td:last-child {{
                border-radius: 0 6px 6px 0;
            }}
            .summary-table .number {{
                font-size: 20px;
                font-weight: 700;
                color: #0d6efd;
                display: block;
                line-height: 1.3;
            }}
            .summary-table .label {{
                font-size: 11px;
                color: #666;
                display: block;
                margin-top: 2px;
            }}
            .summary-table .green .number {{ color: #198754; }}
            .summary-table .orange .number {{ color: #fd7e14; }}
            .summary-table .time-text {{
                font-size: 13px;
                color: #6c757d;
                font-weight: 500;
            }}
            
            .answers-section {{
                margin-top: 15px;
            }}
            .answers-section h3 {{
                font-size: 15px;
                color: #1a1a2e;
                border-left: 3px solid #0d6efd;
                padding-left: 10px;
                margin-bottom: 12px;
            }}
            .page-footer {{
                text-align: center;
                color: #999;
                font-size: 10px;
                margin-top: 30px;
                padding-top: 10px;
                border-top: 1px solid #eee;
            }}
            @media print {{
                body {{ padding: 15px; }}
                .page-break {{ page-break-after: avoid; }}
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>📋 访谈答题详情</h1>
            <div class="subtitle">{interview.get('title', '')} (访谈ID: {interview_id})</div>
        </div>
        
        <!-- 用户信息 - Table 布局 -->
        <table class="user-info-table">
            <tr>
                <td>
                    <span class="label">姓名Name</span>
                    <span class="value">{user_name}</span>
                </td>
                <td>
                    <span class="label">邮箱Mail</span>
                    <span class="value">{user_email}</span>
                </td>
                <td>
                    <span class="label">国家Country</span>
                    <span class="value">{user_country}</span>
                </td>
                <td>
                    <span class="label">库房编码WH Code</span>
                    <span class="value">{user_wh_id or '-'}</span>
                </td>
            </tr>
        </table>
        
        <!-- 统计摘要 - Table 布局 -->
        <table class="summary-table">
            <tr>
                <td>
                    <span class="number">{total_questions}</span>
                    <span class="label">总题数Total Q</span>
                </td>
                <td class="green">
                    <span class="number">{correct_count}</span>
                    <span class="label">答对数量Correct answer</span>
                </td>
                <td class="orange">
                    <span class="number">{correct_rate}</span>
                    <span class="label">正确率Correct rate</span>
                </td>
                <td>
                    <span class="number time-text">{submitted_display}</span>
                    <span class="label">提交时间Submitted</span>
                </td>
            </tr>
        </table>
        
        <div class="answers-section">
            <h3>📝 答题明细Answer details</h3>
            {answers_html}
        </div>
        
        <div class="page-footer">
            培训考试系统ETS · 访谈详情Interview details · 生成时间Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
    </body>
    </html>
    """
    
    return html

