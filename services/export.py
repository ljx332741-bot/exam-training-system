# services/export.py
import json, pandas as pd
from io import BytesIO
import subprocess
import os
import sys
import re
# from weasyprint import HTML
from config import Config
from services.db import get_supabase
import pdfkit
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import logging


logger = logging.getLogger(__name__)

def find_wkhtmltopdf():
    """
    智能查找 wkhtmltopdf 可执行文件路径
    优先级：1. 环境变量 WKHTMLTOPDF_PATH  2. 常见安装路径  3. PATH 中的 wkhtmltopdf
    """
    # 1. 环境变量
    env_path = os.environ.get('WKHTMLTOPDF_PATH')
    if env_path and os.path.exists(env_path):
        return env_path

    # 2. 常见安装路径（Windows）
    common_paths = [
        r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe",
        r"C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe",
        r"D:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe",
        r"D:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe",
        r"C:\wkhtmltopdf\bin\wkhtmltopdf.exe",
    ]
    for path in common_paths:
        if os.path.exists(path):
            return path

    # 3. Linux/Mac 常见位置
    if sys.platform.startswith('linux'):
        linux_paths = ['/usr/local/bin/wkhtmltopdf', '/usr/bin/wkhtmltopdf']
        for path in linux_paths:
            if os.path.exists(path):
                return path
    elif sys.platform == 'darwin':
        mac_paths = ['/usr/local/bin/wkhtmltopdf', '/opt/homebrew/bin/wkhtmltopdf']
        for path in mac_paths:
            if os.path.exists(path):
                return path

    # 4. 尝试从 PATH 中查找
    import shutil
    path_cmd = shutil.which('wkhtmltopdf')
    if path_cmd:
        return path_cmd

    raise RuntimeError(
        "未找到 wkhtmltopdf 可执行文件。请安装 wkhtmltopdf 并设置环境变量 WKHTMLTOPDF_PATH。"
        "下载地址：https://wkhtmltopdf.org/downloads.html"
    )

def generate_user_pdf(user_name, user_email, exam_title, score, questions, answers, details, submitted_at, reviewer=""):
    """
    生成考生成绩单 PDF（健壮版：从 details 累加总分、兼容 answers 缺失、判断题选项固定格式）
    """

    logger = logging.getLogger(__name__)

    # 1. 按题型分组
    grouped = defaultdict(list)
    for q in questions:
        q_type = q.get('type', 'single')
        grouped[q_type].append(q)

    type_desc = {
        'single': ('📌 单选题 (每题{}分，共计{}分)', '📌 Single Choice ({} points each, {} points total)'),
        'multi':  ('📌 多选题 (每题{}分，共计{}分)', '📌 Multiple Choice ({} points each, {} points total)'),
        'judge':  ('📌 判断题 (每题{}分，共计{}分)', '📌 True/False ({} points each, {} points total)')
    }

    # 2. 计算总分（优先使用传入的 score，若为0则从 details 累加）
    total_score = score if score else 0
    if total_score == 0 and details:
        total_score = sum(d.get('score', 0) for d in details.values())
        logger.info(f"从 details 累加总分: {total_score}")

    # 3. 构建 HTML
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{exam_title} - 成绩单</title>
        <style>
            body {{ font-family: 'Microsoft YaHei', 'SimHei', Arial, sans-serif; margin: 40px; line-height: 1.5; }}
            .exam-title {{ font-size: 22px; font-weight: bold; text-align: center; margin-bottom: 10px; color: #0d6efd; }}
            .info {{ text-align: center; margin-bottom: 10px; font-size: 16px; }}
            .reviewer {{ text-align: center; margin-bottom: 10px; font-size: 14px; }}
            .score-container {{ text-align: center; margin: 10px 0; }}
            .score-text {{ font-size: 20px; font-weight: bold;}}
            .score {{ font-family: 'Segoe Script', 'Comic Sans MS', cursive; font-size: 32px; font-weight: bold; color: #198754; }}
            .section-title {{ background: #e9ecef; padding: 8px; margin: 20px 0 10px; font-weight: bold; font-size: 16px; }}
            .question {{ margin: 20px 0 10px; }}
            .question-stem {{ font-weight: bold; }}
            .options {{ margin-left: 20px; margin-top: 5px; }}
            .option {{ margin: 5px 0; }}
            .judge-options {{ display: flex; align-items: center; font-family: Arial, sans-serif; }}
            .judge-option {{ display: inline-block; margin: 0 10px; }}
            .judge-option-separate {{ font-size: 18px; color: #999;  margin: 0 10px;}}
            .user-answer {{ margin-top: 8px; font-size: 13px; color: #0d6efd; }}
            .correct-mark {{ color: #198754; font-weight: bold; font-family: 'Segoe UI Symbol', 'Arial Unicode MS', sans-serif; }}
            .incorrect-mark {{ color: #dc3545; font-weight: bold; font-family: 'Segoe UI Symbol', 'Arial Unicode MS', sans-serif; }}
            .footer {{ margin-top: 40px; font-size: 12px; color: #6c757d; text-align: right; }}
            .print-hide {{ display: none; }}
        </style>
    </head>
    <body>
        <div class="exam-title">{exam_title}</div>
        <div class="info">考试日期 Date：{submitted_at[:10] if submitted_at else ''} | 考生姓名 Name：{user_name}</div>
        <div class="score-container">
            <span class="score-text">总分Score：</span>
            <span class="score">{total_score}</span>
        </div>
        <div class="reviewer">阅卷人Grader (姓名Name+ID）：{reviewer}</div>
    """

    for q_type, q_list in grouped.items():
        if not q_list:
            continue

        # 获取每题分值（取第一个题目的分值作为参考    
        score_per_q = q_list[0].get('score', 5)

        # 计算该题型题目数量和总分
        q_count = len(q_list)
        type_total_score = q_count * score_per_q

        cn_desc = type_desc.get(q_type, ('📌 题目', 'Questions'))[0].format(score_per_q, type_total_score)
        en_desc = type_desc.get(q_type, ('📌 题目', 'Questions'))[1].format(score_per_q, type_total_score)
        html += f'<div class="section-title">{cn_desc}<br><span style="font-size:14px;">{en_desc}</span></div>'
        
        for q in q_list:
            # 解析 options
            options = q.get('options', {})
            if isinstance(options, str):
                try:
                    options = json.loads(options)
                except:
                    options = {}

            qid = str(q['id'])

            if not isinstance(answers, dict):
                try:
                    answers = json.loads(answers) if isinstance(answers, str) else {}
                except:
                    answers = {}
            #---------------------

            # 获取考生答案（兼容 q_ 前缀和纯数字键）
            user_answer_raw = answers.get(f'q_{qid}', answers.get(qid, ''))
            if not user_answer_raw:
                user_answer_raw = "未作答"

            detail = details.get(qid, {})
            is_correct = detail.get('correct', False)
            mark = '[√]' if is_correct else '[×]'

            # 题干
            stem = q.get('content_cn') or q.get('content', '')
            if not stem:
                stem = q.get('content_raw', '')
            html += f'<div class="question"><span class="question-stem">{q.get("num", "")}. {stem}</span></div>'

            # 选项处理（判断题特殊处理）
            if q_type == 'judge':
                # 固定显示“正确 True”和“错误 False”
                html += f'''
                <div class="judge-options">
                    <span class="judge-option">正确 True</span>
                    <span class="judge-option-separate">|</span>
                    <span class="judge-option">错误 False</span>
                </div>
                '''
                # 考生答案显示 T 或 F（标准化）
                if user_answer_raw.upper() == 'T':
                    user_display = 'T'
                elif user_answer_raw.upper() == 'F':
                    user_display = 'F'
                else:
                    user_display = user_answer_raw
            else:
                # 单选/多选：正常显示选项
                sorted_keys = sorted([k for k in options.keys() if k and options[k].strip()])
                for opt_key in sorted_keys:
                    opt_text = options[opt_key]
                    html += f'<div class="option">{opt_key}. {opt_text}</div>'
                user_display = user_answer_raw

            # 显示考生答案及对错
            html += f'''
            <div class="user-answer">
                考生答案CA：{user_display} 
                <span class="{'correct-mark' if is_correct else 'incorrect-mark'}">{mark}</span>
            </div>
            '''

    html += """
        <div class="footer print-hide">本成绩单由在线考试系统自动生成，具有考试凭证效力。</div>
    </body>
    </html>
    """

    # 4. 调用 wkhtmltopdf 生成 PDF（与之前相同，省略具体实现，请保留原有子进程调用代码）
    try:
        wkhtmltopdf_path = find_wkhtmltopdf()
        logger.info(f"使用 wkhtmltopdf 路径: {wkhtmltopdf_path}")
    except RuntimeError as e:
        logger.error(str(e))
        raise

    # 使用 pdfkit 生成 PDF
    options = [
        '--page-size', 'A4',
        '--margin-top', '15mm',
        '--margin-right', '15mm',
        '--margin-bottom', '20mm',
        '--margin-left', '15mm',
        '--encoding', 'UTF-8',
        '--enable-local-file-access'
    ]

    # 查找 wkhtmltopdf 路径
    os.environ['WKHTMLTOPDF_CMD'] = wkhtmltopdf_path
    # 设置字体目录（Windows）
    os.environ['FONTCONFIG_PATH'] = r'C:\Windows\Fonts'

    # 调用子进程生成 PDF
    try:
        proc = subprocess.run(
            [wkhtmltopdf_path] + options + ['-', '-'],  # 输入 stdin，输出 stdout
            input=html.encode('utf-8'),
            capture_output=True,
            timeout=30  # 30 秒超时
        )
        if proc.returncode == 0:
            return BytesIO(proc.stdout)
        else:
            error_msg = proc.stderr.decode('utf-8', errors='ignore')
            logger.error(f"wkhtmltopdf 错误: {error_msg}")
            raise RuntimeError(f"PDF 生成失败: {error_msg}")
    except subprocess.TimeoutExpired:
        logger.error("wkhtmltopdf 执行超时（30秒）")
        raise RuntimeError("PDF 生成超时，请稍后重试")
    except Exception as e:
        logger.error(f"子进程调用失败: {e}")
        raise

def generate_pdf(user_name, score, questions, answers, details, reviewer, lang_pref='both'):
    """
    lang_pref: 'both' | 'cn' | 'en'
    """
    html = f"""
    <style>
      .q-cn {{ font-family: 'Microsoft YaHei', sans-serif; }}
      .q-en {{ font-family: 'Segoe UI', Arial, sans-serif; color: #555; font-size: 0.95em; }}
      .q-en:empty {{ display: none; }}
    </style>
    """
    
    for q in questions:
        qid = str(q["id"])
        u_ans = answers.get(qid, "未作答")
        is_correct = details.get(qid, {}).get("correct", False)
        mark = "✅" if is_correct else "❌"
        
        # 题干双语渲染
        content_html = f'<span class="q-cn">{q.get("content_cn") or q["content_raw"]}</span>'
        if lang_pref == 'both' and q.get("content_en"):
            content_html += f'<br><span class="q-en">{q["content_en"]}</span>'
        elif lang_pref == 'en' and q.get("content_en"):
            content_html = f'<span class="q-en">{q["content_en"]}</span>'
        
        # 选项双语渲染
        options_html = ""
        for key in ["A","B","C","D"]:
            if key in q.get("options", {}):
                opt_raw = q["options"][key]
                opt_en = q.get("options_struct", {}).get(key, {}).get("en", "")
                opt_html = f'<strong>{key}.</strong> {opt_raw}'
                if lang_pref == 'both' and opt_en:
                    opt_html += f'<br><span class="q-en">{opt_en}</span>'
                elif lang_pref == 'en' and opt_en:
                    opt_html = f'<strong>{key}.</strong> {opt_en}'
                options_html += f"<div>{opt_html}</div>"
        
        html += f"""
        <div class="q-block">
          <strong>{q.get('num')}. </strong>{content_html} ({q.get('score')}分) <span class="tag {'tag-ok' if is_correct else 'tag-err'}">{mark}</span><br>
          <div class="ms-3 mt-2">{options_html}</div>
          <small class="text-muted">考生答案：{u_ans} | 标准答案：{q.get('answer')}</small>
        </div>
        """
    pdf_bytes = HTML(string=html).write_pdf()
    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return buf

def generate_bilingual_excel(training_id: int, exam_id: int, country: str = None):
    """
    生成双语Excel报告（培训信息+知识测评）
    
    Args:
        training_id: 培训ID
        exam_id: 考试ID
        country: 国家筛选（可选）
    
    Returns:
        BytesIO: Excel文件流
    """
    db = get_supabase()
    
    # ==================== 创建工作簿 ====================
    wb = openpyxl.Workbook()
    
    # 清理工作表标题中的非法字符
    def clean_sheet_title(title):
        illegal_chars = r'[\\/*?:\[\]]'
        return re.sub(illegal_chars, ' ', title).strip()

    # ==================== Sheet 1: 培训信息汇总 ====================
    ws1 = wb.active
    ws1.title = clean_sheet_title("培训信息汇总Training Summary")
    
    # 表头（中英双语）
    headers1 = [
        ("序号\nNO.", "NO."),
        ("库房类型\nWH Type", "WH Type"),
        ("库房编码\nWH ID", "WH ID"),
        ("库房名称(CN)\nWarehouse Name (CN)", "Warehouse Name (CN)"),
        ("库房名称(EN)\nWarehouse Name (EN)", "Warehouse Name (EN)"),
        ("立项编号\nProject Initiation No.", "Project Initiation No."),
        ("培训名称\nTraining Name", "Training Name"),
        ("培训语种\nTraining Language", "Training Language"),
        ("主讲人+ID\nCourse Lecturer", "Course Lecturer"),
        ("培训日期\nTraining Date", "Training Date"),
        ("培训课时\nTraining Hours", "Training Hours"),
        ("培训对象(第三方服务商)\nTrainee (Third-party)", "Trainee (Third-party)"),
        ("第三方服务商数量\nNumber of Providers", "Number of Providers"),
        ("参培人数\nNumber of Trainees", "Number of Trainees"),
        ("备注\nRemark", "Remark"),
    ]
    
    # 写入表头
    for col, (cn_header, en_header) in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=f"{cn_header}")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # 查询培训签到数据
    signs_query = db.table("training_attendances").select("""
        id,
        user_id,
        training_id,
        signed_at,
        users!inner (
            id,
            name_cn,
            name_en,
            company,
            email
        )
    """).eq("training_id", training_id)
    
    if country:
        signs_query = signs_query.eq("country", country)
    
    signs_res = signs_query.execute()
    signs = signs_res.data or []
    
    # 查询培训元数据（管理员录入）
    training_meta = db.table("trainings").select("*").eq("id", training_id).single().execute()
    meta = training_meta.data or {}
    
    # 按库房分组统计
    warehouse_stats = {}
    for sign in signs:
        user = sign.get("users", {})
        # wh_type = user.get("company", "Unknown")  # 假设company存库房类型
        wh_type = user.get("系统库/备件库")
        wh_id = sign.get('wh_id', '')  # 示例生成逻辑
        wh_name_cn = user.get("wh_name_en", "")
        wh_name_en = user.get("wh_name_en", "")
        
        key = (wh_type, wh_id)
        if key not in warehouse_stats:
            warehouse_stats[key] = {
                "wh_name_cn": wh_name_en,
                "wh_name_en": wh_name_en,
                "trainees": set(),
                "providers": set(),
                "sign_dates": []
            }
        
        warehouse_stats[key]["trainees"].add(sign.get("user_id"))
        warehouse_stats[key]["sign_dates"].append(sign.get("signed_at"))
        
        # 假设第三方服务商判断逻辑（示例）
        if "third-party" in user.get("email", "").lower():
            warehouse_stats[key]["providers"].add(sign.get("user_id"))
    
    # 写入数据行
    for row_idx, ((wh_type, wh_id), stats) in enumerate(warehouse_stats.items(), 2):
        trainee_count = len(stats["trainees"])
        provider_count = len(stats["providers"])
        
        row_data = [
            row_idx - 1,  # 序号
            wh_type,  # 库房类型
            wh_id,  # 库房编码
            stats["wh_name_cn", ""],  # 库房名称(EN)
            stats["wh_name_en"],  # 库房名称(EN)
            meta.get("project_no", ""),  # 立项编号
            meta.get("training_name", ""),  # 培训名称
            meta.get("language", ""),  # 培训语种
            meta.get("lecturer", ""),  # 主讲人
            min(stats["sign_dates"])[:10] if stats["sign_dates"] else "",  # 培训日期
            meta.get("duration", ""),  # 培训课时
            "第三方服务商" if provider_count > 0 else "内部员工",  # 培训对象
            provider_count,  # 第三方服务商数量
            trainee_count,  # 参培人数
            "",  # 备注
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
    
    # 调整列宽
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15
    
    # ==================== Sheet 2: 知识测评汇总表 ====================
    ws2 = wb.create_sheet(title=clean_sheet_title("知识测评汇总表Assessment Summary"))
    
    # 查询考试题目数量（动态列数）
    questions_res = db.table("questions").select("id").eq("exam_id", exam_id).execute()
    total_questions = len(questions_res.data or [])
    
    # 表头（中英双语 + 动态NO.列）
    headers2_fixed = [
        ("序号\nNO.", "NO."),
        ("国家\nCountry", "Country"),
        ("库房类型\nWH Type", "WH Type"),
        ("库房编码\nWH ID", "WH ID"),
        ("库房名称\nWH Name", "WH Name"),
        ("考试人员姓名\nExam Person Name", "Exam Person Name"),
        ("考试日期\nDate", "Date"),
        ("考试语种\nLanguage", "Language"),
        ("成绩\nScore", "Score"),
        ("阅卷人(姓名+ID)\nGrading Personnel", "Grading Personnel"),
    ]
    
    # 动态生成 NO.1 ~ NO.N 列
    dynamic_headers = [(f"NO.{i}\n答题情况", f"NO.{i}") for i in range(1, total_questions + 1)]
    
    headers2 = headers2_fixed + dynamic_headers + [("备注\nRemark", "Remark")]
    
    # 写入表头
    for col, (cn_header, en_header) in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=cn_header)
        
        # 不同区域不同背景色
        if col <= len(headers2_fixed):
            fill_color = "4472C4"  # 蓝色（固定字段）
        elif col <= len(headers2_fixed) + total_questions:
            fill_color = "A5A5A5"  # 灰色（答题情况）
        else:
            fill_color = "FFC000"  # 黄色（备注）
        
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # 查询考试成绩数据
    results_res = db.table("exam_results").select("""
        id,
        user_id,
        exam_id,
        total_score,
        details,
        answers,
        custom1,
        custom2,
        custom3,
        custom4,
        custom5,
        created_at,
        users!inner (
            id,
            name_cn,
            name_en,
            company,
            email
        )
    """).eq("exam_id", exam_id).execute()
    
    results = results_res.data or []
    
    # 写入数据行
    for row_idx, result in enumerate(results, 2):
        user = result.get("users", {})
        details_raw = result.get("details", {})  # 获取原始字段
        
        if isinstance(details_raw, str):
            try:
                details = json.loads(details_raw)
            except:
                details = {}
        else:
            details = details_raw
        # 生成 Y/N 数组（按题号顺序）
        answer_status = []
        for q_num in range(1, total_questions + 1):
            q_key = str(q_num)  # 假设details key为题号
            is_correct = details.get(q_key, {}).get("correct", False)
            answer_status.append("Y" if is_correct else "N")
        
        row_data = [
            row_idx - 1,  # 序号
            country or "Unknown",  # 国家
            user.get("company", ""),  # 库房类型
            user.get('wh_id', ''),  # 库房编码
            user.get("wh_name_cn", ""),  # 库房名称
            user.get("wh_name_en", ""),  # 库房名称
            user.get("name_cn", "") or user.get("name_en", ""),  # 考试人员姓名
            result.get("created_at", "")[:10],  # 考试日期
            meta.get("language", "English"),  # 考试语种
            result.get("total_score", 0),  # 成绩
            result.get("custom1", ""),  # 阅卷人（假设存custom1）
        ] + answer_status + [
            result.get("custom5", "")  # 备注
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Y/N 列特殊样式
            if len(headers2_fixed) <= col_idx <= len(headers2_fixed) + total_questions:
                if value == "Y":
                    cell.font = Font(color="008000", bold=True)  # 绿色
                elif value == "N":
                    cell.font = Font(color="FF0000", bold=True)  # 红色
    
    # 调整列宽
    for col in range(1, len(headers2) + 1):
        if col <= len(headers2_fixed):
            ws2.column_dimensions[get_column_letter(col)].width = 15
        elif col <= len(headers2_fixed) + total_questions:
            ws2.column_dimensions[get_column_letter(col)].width = 8  # Y/N列窄一些
        else:
            ws2.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== 输出文件 ====================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"培训报告_{training_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buffer, filename

# services/export.py - 修改 generate_bilingual_excel_filtered 函数

def generate_bilingual_excel_filtered(trainings, exams, country, start_date, end_date, user_ids=None, wh_id=None, lang='zh'):
    """
    生成双语Excel报告，支持按国家、库房、培训名称、考试名称筛选
    增强版：增加业务单位、合并库房名称、考试年份测评次数
    """
    if not isinstance(trainings, list): trainings = []
    if not isinstance(exams, list): exams = []

    db = get_supabase()
    wb = openpyxl.Workbook()
    clean_title = lambda s: re.sub(r'[\\/*?:\[\]]', ' ', s).strip()
    
    # ============================================================
    # ========== 工作表1：培训信息汇总（增强版） ==========
    # ============================================================
    ws1 = wb.active
    ws1.title = clean_title("培训信息汇总 Training Summary")

    # ✅ 新表头：增加"业务单位"，合并库房名称，调整顺序
    headers1 = [
        ("序号\nNO.", "NO."),
        ("业务单位\nBusiness Unit", "Business Unit"),      # ✅ 新增
        ("国家\nCountry", "Country"),                       # ✅ 从后面移到这里
        ("库房类型\nWH Type", "WH Type"),
        ("库房编码\nWH ID", "WH ID"),
        ("库房名称\nWH Name", "WH Name"),                  # ✅ 合并 CN/EN
        ("立项编号\nProject Initiation No.", "Project Initiation No."),
        ("培训名称\nTraining Name", "Training Name"),
        ("培训语种\nTraining Language", "Training Language"),
        ("主讲人+ID\nCourse Lecturer", "Course Lecturer"),
        ("培训日期\nTraining Date", "Training Date"),
        ("培训课时\nTraining Hours", "Training Hours"),
        ("培训对象(第三方服务商)\nTrainee (Third-party)", "Trainee (Third-party)"),
        ("第三方服务商数量\nNumber of Providers", "Number of Providers"),
        ("参培人数\nNumber of Trainees", "Number of Trainees"),
        ("备注\nRemark", "Remark"),
    ]
    
    # 写入表头
    for col, (cn_header, en_header) in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=cn_header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

    # 构建 training_id -> meta 映射
    training_metas = {}
    training_names = {}
    for t in trainings:
        training_metas[t['id']] = t.get('header_template', {})
        training_names[t['id']] = t.get('name', '')

    row_idx = 2

    for training in trainings:
        tid = training['id']
        query = db.table("training_attendances").select("*, users(*)").eq("training_id", tid).is_("deleted_at", "null")
        
        if user_ids is not None:
            query = query.in_("user_id", user_ids)
        signs = query.execute().data or []
        if not signs:
            continue

        meta = training_metas.get(tid, {})

        if wh_id:
            # 固定库房，所有签到合并为一个组
            groups = {'default': {
                'trainees': set(),
                'partner_companies': set(),
                'sign_dates': [],
                'wh_info': {},
                'departments': set()  # ✅ 新增：收集部门
            }}
            for sign in signs:
                user = sign.get('users', {})
                groups['default']['trainees'].add(sign['user_id'])
                groups['default']['sign_dates'].append(sign.get('sign_time') or sign.get('signed_at'))
                if user.get('is_partner'):
                    company = user.get('company', '').strip()
                    if company:
                        groups['default']['partner_companies'].add(company)
                if not groups['default']['wh_info']:
                    groups['default']['wh_info'] = {
                        'wh_type': user.get('wh_type', ''),
                        'wh_id': user.get('wh_id', ''),
                        'wh_name_en': user.get('wh_name_en', '')
                    }
                # ✅ 收集部门
                dept = user.get('department', '')
                if dept:
                    groups['default']['departments'].add(dept)
            
            for g_key, g in groups.items():
                partner_str = '/'.join(sorted(g['partner_companies'])) if g['partner_companies'] else '内部员工'
                partner_count = len(g['partner_companies'])
                training_date = ""
                if g['sign_dates']:
                    first_date = min(g['sign_dates'])
                    training_date = first_date[:10] if first_date else ""
                
                # ✅ 业务单位：取部门列表的第一个（或合并显示）
                dept_str = '/'.join(sorted(g['departments'])) if g['departments'] else ''
                
                row_data = [
                    row_idx - 1,                                      # 序号
                    dept_str,                                         # ✅ 业务单位
                    '',                                               # ✅ 国家（固定库房时可能为空）
                    g['wh_info'].get('wh_type', ''),                  # 库房类型
                    g['wh_info'].get('wh_id', ''),                    # 库房编码
                    g['wh_info'].get('wh_name_en', ''),               # ✅ 库房名称（合并）
                    meta.get("project_no", ""),                       # 立项编号
                    training_names.get(tid, ''),                      # 培训名称
                    meta.get("language", ""),                         # 培训语种
                    meta.get("lecturer", ""),                         # 主讲人
                    training_date,                                    # 培训日期
                    meta.get("duration", "2"),                        # 培训课时
                    partner_str,                                      # 培训对象
                    partner_count,                                    # 第三方服务商数量
                    len(g['trainees']),                               # 参培人数
                    "",                                               # 备注
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                         top=Side(style="thin"), bottom=Side(style="thin"))
                row_idx += 1
        else:
            # 按国家分组
            country_groups = defaultdict(lambda: {
                'trainees': set(),
                'partner_companies': set(),
                'sign_dates': [],
                'wh_info': {},
                'departments': set()  # ✅ 新增：收集部门
            })
            for sign in signs:
                user = sign.get('users', {})
                cty = user.get('country', 'Unknown')
                grp = country_groups[cty]
                grp['trainees'].add(sign['user_id'])
                grp['sign_dates'].append(sign.get('sign_time') or sign.get('signed_at'))
                if user.get('is_partner'):
                    company = user.get('company', '').strip()
                    if company:
                        grp['partner_companies'].add(company)
                if not grp['wh_info']:
                    grp['wh_info'] = {
                        'wh_type': user.get('wh_type', ''),
                        'wh_id': user.get('wh_id', ''),
                        'wh_name_en': user.get('wh_name_en', '')
                    }
                # ✅ 收集部门
                dept = user.get('department', '')
                if dept:
                    grp['departments'].add(dept)
            
            for cty, grp in country_groups.items():
                partner_str = '/'.join(sorted(grp['partner_companies'])) if grp['partner_companies'] else '内部员工'
                partner_count = len(grp['partner_companies'])
                training_date = ""
                if grp['sign_dates']:
                    first_date = min(grp['sign_dates'])
                    training_date = first_date[:10] if first_date else ""
                
                # ✅ 业务单位：取部门列表的第一个（或合并显示）
                dept_str = '/'.join(sorted(grp['departments'])) if grp['departments'] else ''
                
                row_data = [
                    row_idx - 1,                                      # 序号
                    dept_str,                                         # ✅ 业务单位
                    cty,                                              # ✅ 国家
                    grp['wh_info'].get('wh_type', ''),                # 库房类型
                    grp['wh_info'].get('wh_id', ''),                  # 库房编码
                    grp['wh_info'].get('wh_name_en', ''),             # ✅ 库房名称（合并）
                    meta.get("project_no", ""),                       # 立项编号
                    training_names.get(tid, ''),                      # 培训名称
                    meta.get("language", ""),                         # 培训语种
                    meta.get("lecturer", ""),                         # 主讲人
                    training_date,                                    # 培训日期
                    meta.get("duration", "2"),                        # 培训课时
                    partner_str,                                      # 培训对象
                    partner_count,                                    # 第三方服务商数量
                    len(grp['trainees']),                             # 参培人数
                    "",                                               # 备注
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                         top=Side(style="thin"), bottom=Side(style="thin"))
                row_idx += 1

    # 调整列宽
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # ============================================================
    # ========== 工作表2：知识测评汇总表（增强版） ==========
    # ============================================================
    ws2 = wb.create_sheet(title=clean_title("知识测评汇总表 Assessment Summary"))

    # 收集所有考试题目及最大列数
    exam_questions_dict = {}
    max_questions = 0
    for exam in exams:
        q_res = db.table("questions").select("id, num").eq("exam_id", exam['id']).order("num").execute()
        questions = q_res.data or []
        exam_questions_dict[exam['id']] = questions
        if len(questions) > max_questions:
            max_questions = len(questions)

    # ✅ 计算年份显示（用于测评次数列标题）
    exam_years = set()
    for exam in exams:
        created_at = exam.get('created_at')
        if created_at:
            try:
                year = datetime.fromisoformat(created_at.replace('Z', '+00:00')).year
                exam_years.add(str(year))
            except:
                pass
    year_display = '/'.join(sorted(exam_years)) if exam_years else ''

    # ✅ 新表头：增加"业务单位"，原"考试名称"改为"年份年测评次数"，在备注后面新增"考试名称"
    headers2_fixed = [
        ("序号\nNO.", "NO."),
        ("业务单位\nBusiness Unit", "Business Unit"),           # ✅ 新增
        ("国家\nCountry", "Country"),
        ("库房类型\nWH Type", "WH Type"),
        ("库房编码\nWH ID", "WH ID"),
        ("库房名称\nWH Name", "WH Name"),
        ("考试人员姓名\nExam Person Name", "Exam Person Name"),
        ("考试日期\nDate", "Date"),
        ("考试语种\nLanguage", "Language"),
        ("成绩\nScore", "Score"),
        ("阅卷人(姓名+ID)\nGrading Personnel", "Grading Personnel"),
    ]

    # ✅ 列位置计算：
    # 固定列数 = len(headers2_fixed) 
    # 测评次数列（在阅卷人之后，动态列之前）
    # 动态列（答题情况统计）在测评次数之后
    # 备注列在动态列之后
    # 考试名称列在备注之后
    
    fixed_col_count = len(headers2_fixed)
    exam_freq_col = fixed_col_count + 1                # ✅ 测评次数（阅卷人之后，动态列之前）
    dynamic_start_col = fixed_col_count + 2            # ✅ 动态列开始位置
    remark_col = fixed_col_count + 2 + max_questions   # ✅ 备注列（动态列之后）
    exam_name_col = remark_col + 1                     # ✅ 考试名称（备注之后）
    total_header_cols = exam_name_col

    # ========== 第一行表头 ==========
    # 固定列
    for idx, (cn_header, en_header) in enumerate(headers2_fixed, 1):
        col_letter = get_column_letter(idx)
        merge_range = f"{col_letter}1:{col_letter}2"
        ws2.merge_cells(merge_range)
        cell = ws2.cell(row=1, column=idx, value=cn_header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 测评次数列（在阅卷人之后，动态列之前）
    freq_letter = get_column_letter(exam_freq_col)
    ws2.merge_cells(f"{freq_letter}1:{freq_letter}2")
    exam_freq_header = f"{year_display}年测评次数\nExam Frequency" if year_display else "测评次数\nExam Frequency"
    cell_freq = ws2.cell(row=1, column=exam_freq_col, value=exam_freq_header)
    cell_freq.font = Font(bold=True, color="FFFFFF")
    cell_freq.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell_freq.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 动态列第一行合并 - "答题情况统计Needback"
    if max_questions > 0:
        start_col_letter = get_column_letter(dynamic_start_col)
        end_col_letter = get_column_letter(dynamic_start_col + max_questions - 1)
        merge_range = f"{start_col_letter}1:{end_col_letter}1"
        ws2.merge_cells(merge_range)
        cell_detail = ws2.cell(
            row=1, 
            column=dynamic_start_col, 
            value="答题情况统计Needback(题目答对填写 Y，答错则填写 N\If the test answer is correct fill 'Y', otherwise fill 'N')"
        )
        cell_detail.font = Font(bold=True, color="FFFFFF", size=10)
        cell_detail.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        cell_detail.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 添加注释到"答题情况统计Needback"单元格
        comment_text = "题目答对填写 Y，答错则填写 N\If the test answer is correct fill 'Y', otherwise fill 'N'"
        comment = Comment(comment_text, "系统提示")
        comment.visible = True
        cell_detail.comment = comment

    # 备注列（在动态列之后）
    remark_letter = get_column_letter(remark_col)
    ws2.merge_cells(f"{remark_letter}1:{remark_letter}2")
    cell_remark = ws2.cell(row=1, column=remark_col, value="备注\nRemark")
    cell_remark.font = Font(bold=True, color="FFFFFF")
    cell_remark.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell_remark.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 考试名称列（在备注之后）
    exam_name_letter = get_column_letter(exam_name_col)
    ws2.merge_cells(f"{exam_name_letter}1:{exam_name_letter}2")
    cell_exam_name = ws2.cell(row=1, column=exam_name_col, value="考试名称\nExam Name")
    cell_exam_name.font = Font(bold=True, color="FFFFFF")
    cell_exam_name.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell_exam_name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ========== 第二行表头（动态列子标题题号列）==========
    for i in range(1, max_questions + 1):
        col_idx = dynamic_start_col + i - 1
        col_letter = get_column_letter(col_idx)
        cell = ws2.cell(row=2, column=col_idx, value=f"NO.{i}")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ========== 统一样式与边框 ==========
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in range(1, 3):
        for col in range(1, total_header_cols + 1):
            c = ws2.cell(row=row, column=col)
            c.border = thin_border
            if c.font == Font():
                c.font = Font(bold=True, color="FFFFFF")
            if c.fill == PatternFill(fill_type=None):
                if col <= fixed_col_count:
                    c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                elif col == exam_freq_col:
                    c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                elif dynamic_start_col <= col <= dynamic_start_col + max_questions - 1:
                    c.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
                elif col == remark_col:
                    c.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif col == exam_name_col:
                    c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ========== 数据行 ==========
    row_idx2 = 3
    
    # ✅ 统计每个用户在每个考试中的参与次数（用于测评次数列）
    user_exam_count = defaultdict(lambda: defaultdict(int))
    for exam in exams:
        results_query = db.table("exam_results").select("*").eq("exam_id", exam['id'])
        if user_ids is not None:
            results_query = results_query.in_("user_id", user_ids)
        results_all = results_query.execute().data or []
        for r in results_all:
            uid = r['user_id']
            user_exam_count[uid][exam['id']] += 1

    for exam in exams:
        exam_title = exam.get('title', '')
        results_query = db.table("exam_results").select("*").eq("exam_id", exam['id'])
        if user_ids is not None:
            results_query = results_query.in_("user_id", user_ids)
        results_all = results_query.execute().data or []
        
        # 去重：每个用户只保留最新一次成绩
        latest_map = {}
        for r in results_all:
            uid = r['user_id']
            if uid not in latest_map or r.get('created_at', '') > latest_map[uid].get('created_at', ''):
                latest_map[uid] = r
        results = list(latest_map.values())

        user_ids_set = list(set([r['user_id'] for r in results]))
        users_map = {}
        if user_ids_set:
            users_res = db.table("users").select("*").in_("id", user_ids_set).execute()
            for u in users_res.data or []:
                users_map[u['id']] = u

        for result in results:
            user = users_map.get(result['user_id'], {})
            details_raw = result.get('details', {})
            if isinstance(details_raw, str):
                try:
                    details = json.loads(details_raw)
                except:
                    details = {}
            else:
                details = details_raw

            exam_questions = exam_questions_dict.get(exam['id'], [])
            answer_status = []
            for q in exam_questions:
                q_id = str(q['id'])
                q_detail = details.get(q_id) or details.get(f'q_{q_id}')
                if isinstance(q_detail, dict):
                    is_correct = q_detail.get('correct', False)
                else:
                    is_correct = False
                answer_status.append("Y" if is_correct else "N")
            while len(answer_status) < max_questions:
                answer_status.append("")

            # ✅ 获取业务单位（部门）
            dept = user.get('department', '')
            
            # ✅ 获取该用户在该考试中的参与次数（用于测评次数列）
            exam_count = user_exam_count.get(result['user_id'], {}).get(exam['id'], 1)

            # ✅ row_data 顺序与表头列一一对应
            row_data = [
                row_idx2 - 2,                                         # 1. 序号
                dept,                                                 # 2. 业务单位
                user.get('country', ''),                              # 3. 国家
                user.get('wh_type', ''),                              # 4. 库房类型
                user.get('wh_id', ''),                                # 5. 库房编码
                user.get('wh_name_en', ''),                           # 6. 库房名称
                user.get('name_cn', '') or user.get('name_en', ''),   # 7. 考试人员姓名
                result.get('created_at', '')[:10],                    # 8. 考试日期
                exam.get('language', 'English'),                      # 9. 考试语种
                result.get('total_score', 0),                         # 10. 成绩
                exam.get('reviewer', ''),                             # 11. 阅卷人
                exam_count,                                           # 12. ✅ 测评次数（在阅卷人之后）
            ] + answer_status + [                                     # 13. 动态列 (NO.1~N)
                result.get('custom5', ''),                            # 14. ✅ 备注
                exam_title,                                           # 15. ✅ 考试名称
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx2, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                # ✅ 动态列（答题情况统计）的 Y/N 着色
                if dynamic_start_col <= col_idx <= dynamic_start_col + max_questions - 1:
                    if value == "Y":
                        cell.font = Font(color="008000", bold=True)
                    elif value == "N":
                        cell.font = Font(color="FF0000", bold=True)
            row_idx2 += 1

    # 调整列宽
    for col in range(1, total_header_cols + 1):
        if col <= fixed_col_count:
            ws2.column_dimensions[get_column_letter(col)].width = 15
        elif col == exam_freq_col:
            ws2.column_dimensions[get_column_letter(col)].width = 16
        elif dynamic_start_col <= col <= dynamic_start_col + max_questions - 1:
            ws2.column_dimensions[get_column_letter(col)].width = 8
        elif col == remark_col:
            ws2.column_dimensions[get_column_letter(col)].width = 20
        elif col == exam_name_col:
            ws2.column_dimensions[get_column_letter(col)].width = 30

    # ============================================================
    # ========== 工作表3：访谈检查结果（增强版） ==========
    # ============================================================
    if exams:
        ws3 = wb.create_sheet(title=clean_title("访谈检查结果 Interview Results"))
        exam_ids = [ex['id'] for ex in exams]
        interviews_query = db.table("interviews").select("*").in_("exam_id", exam_ids).is_("deleted_at", "null").order("created_at", desc=True).execute()
        interviews = interviews_query.data or []

        if interviews:
            # ✅ 新表头：增加"业务单位"，调整备注位置
            headers3 = [
                ("序号\nNO.", "NO."),
                ("业务单位\nBusiness Unit", "Business Unit"),      # ✅ 新增
                ("国家\nCountry", "Country"),
                ("库房类型\nWH Type", "WH Type"),
                ("库房编码\nWH ID", "WH ID"),
                ("库房名称\nWH Name", "WH Name"),
                ("访谈人员姓名\nInterviewee Name", "Interviewee Name"),
                ("检查时间\nCheck Time", "Check Time"),
                ("检查人员\nInspector", "Inspector"),
                ("访谈问题数量\nTotal Questions", "Total Questions"),
                ("答对问题数量\nCorrect Answers", "Correct Answers"),
                ("备注\nRemark", "Remark"),                       # ✅ 移到反馈人前面
                ("反馈人\nFeedback Person", "Feedback Person"),
                ("访谈名称\nInterview Title", "Interview Title"),
            ]
            
            for col, (cn_header, en_header) in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col, value=cn_header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                
                # ✅ 在"访谈问题数量 Total Questions"列添加注释
                if cn_header == "访谈问题数量\nTotal Questions":
                    cell.comment = openpyxl.comments.Comment("访谈问题不少于3个", "系统提示")

            row_idx3 = 2
            for interview in interviews:
                all_rows = db.table("interview_results").select("*").eq("interview_id", interview['id']).execute().data or []
                if user_ids is not None:
                    all_rows = [r for r in all_rows if r['user_id'] in user_ids]
                if not all_rows:
                    continue

                user_ids_set = list(set(r['user_id'] for r in all_rows))
                users_map = {}
                if user_ids_set:
                    users_res = db.table("users").select("id, name_cn, name_en, country, wh_type, wh_id, wh_name_en, department").in_("id", user_ids_set).execute()
                    for u in users_res.data or []:
                        users_map[u['id']] = u

                user_stats = {}
                for row in all_rows:
                    uid = row['user_id']
                    if uid not in user_stats:
                        user_stats[uid] = {
                            'total': 0, 
                            'correct': 0, 
                            'submitted_at': None, 
                            'feedback': '',
                            'interview_title': interview.get('title', '')
                        }
                    user_stats[uid]['total'] += 1
                    if row.get('is_correct'):
                        user_stats[uid]['correct'] += 1
                    if row.get('submitted_at') and (not user_stats[uid]['submitted_at'] or row['submitted_at'] > user_stats[uid]['submitted_at']):
                        user_stats[uid]['submitted_at'] = row['submitted_at']
                    if row.get('feedback') and not user_stats[uid]['feedback']:
                        user_stats[uid]['feedback'] = row['feedback']

                reviewer = interview.get('reviewer', '')
                for uid, stats in user_stats.items():
                    user = users_map.get(uid, {})
                    dept = user.get('department', '')  # ✅ 业务单位
                    
                    row_data = [
                        row_idx3 - 1,                                     # 序号
                        dept,                                             # ✅ 业务单位
                        user.get('country', ''),                          # 国家
                        user.get('wh_type', ''),                          # 库房类型
                        user.get('wh_id', ''),                            # 库房编码
                        user.get('wh_name_en') or user.get('name_cn') or user.get('name_en', ''),  # 库房名称
                        user.get('name_cn') or user.get('name_en', ''),   # 访谈人员姓名
                        stats['submitted_at'][:10] if stats['submitted_at'] else '',  # 检查时间
                        reviewer,                                         # 检查人员
                        stats['total'],                                   # 访谈问题数量
                        stats['correct'],                                 # 答对问题数量
                        stats.get('feedback', ''),                        # ✅ 备注（移到反馈人前面）
                        stats.get('feedback', ''),                        # 反馈人（暂时取同样值，可根据实际调整）
                        stats.get('interview_title', ''),                 # 访谈名称
                    ]
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws3.cell(row=row_idx3, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                             top=Side(style="thin"), bottom=Side(style="thin"))
                    row_idx3 += 1

            for col in range(1, len(headers3) + 1):
                ws3.column_dimensions[get_column_letter(col)].width = 15

    # ============================================================
    # ========== 保存文件 ==========
    # ============================================================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if lang == 'en':
        filename = f"Training_Exam_Comprehensive_Report_{timestamp}.xlsx"
    else:
        filename = f"培训考试综合报告_{timestamp}.xlsx"
    return buffer, filename
