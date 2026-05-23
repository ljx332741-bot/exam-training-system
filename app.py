# app.py - 完整重构版 | 修复：装饰器顺序/路由语法/引号/日志
import os
import json
import uuid
import logging
import sys
import traceback
import atexit
import zipfile
import pytz
import pdfkit
import openpyxl
import random
import base64, re
from flask import make_response
from datetime import datetime, timezone, timedelta
from functools import wraps
from dateutil import parser
from apscheduler.schedulers.background import BackgroundScheduler
from io import BytesIO
from flask import (
    Flask, render_template, request, redirect, url_for, 
    session, flash, jsonify, send_file
)
from supabase import create_client
from datetime import datetime, timezone
from flask import jsonify, request
from functools import wraps
from services.db import get_supabase
from services import auth, exam, export
from services.export import find_wkhtmltopdf
from config import Config
from utils.status import get_exam_status
from utils.common import match_country_code, quarter_to_date_range, get_reviewer_by_country
from utils.permissions import (
    is_developer, get_developer_id, has_role, can_manage_role,
    can_view_user, get_admin_allowed_countries, set_admin_allowed_countries,
    parse_countries_input, developer_required, super_admin_required,
    apply_country_filter
)
from utils.training_helpers import get_training_country_templates_status
from dotenv import load_dotenv
load_dotenv() # 加载 .env 文件中的环境变量 这行代码不会影响生产环境（Render 上没有 .env 文件）

print("DEVELOPER_USER_ID from env =", os.environ.get('DEVELOPER_USER_ID'))
# ================= 1. 日志配置（在 app 创建之前） =================
logging.basicConfig(
    level=logging.DEBUG,
    format='[%(asctime)s] %(levelname)s in %(module)s: %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('exam_debug.log', encoding='utf-8', mode='a')
    ]
)
DEFAULT_LOCAL_TIMEZONE = os.environ.get('LOCAL_TIMEZONE', 'Asia/Kathmandu')
logger = logging.getLogger(__name__)
logger.info("🚀 Flask 应用启动，日志级别: DEBUG")

# 从环境变量读取所有配置（本地开发从 .env 读取，生产环境从 Render 读取）
HOST = os.environ.get('HOST', '127.0.0.1')
PORT = int(os.environ.get('PORT', 5000))
DEBUG = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'

# ================= 2. 🔑 关键：先创建 Flask 应用实例 =================
app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY
app.debug = DEBUG  # 🔥 根据环境配置决定是否调试

# ==================== 自定义 Jinja2 过滤器 ====================
@app.template_filter('join_options')
def join_options_filter(options):
    """将选项字典格式化为字符串"""
    if not options:
        return ''
    if isinstance(options, str):
        try:
            options = json.loads(options)
        except:
            return options
    if not isinstance(options, dict):
        return ''
    
    parts = []
    for key in sorted(options.keys()):
        value = options.get(key, '')
        if value and str(value).strip():
            parts.append(f"{key}. {value}")
    return '; '.join(parts)

@app.template_filter('format_options')
def format_options_filter(options):
    """与 join_options 相同，保持兼容"""
    return join_options_filter(options)

@app.template_filter('utc_to_local')
def utc_to_local_filter(utc_string):
    """UTC 转本地时间字符串"""
    if not utc_string:
        return ''
    try:
        from datetime import datetime
        import pytz
        # 解析时间
        if utc_string.endswith('Z'):
            utc_string = utc_string.replace('Z', '+00:00')
        dt = datetime.fromisoformat(utc_string)
        # 转换为本地时间（使用系统本地时区）
        local_tz = pytz.timezone('Asia/Shanghai')  # 或从配置读取
        if dt.tzinfo is None:
            dt = pytz.UTC.localize(dt)
        local_dt = dt.astimezone(local_tz)
        return local_dt.strftime('%Y-%m-%dT%H:%M')
    except Exception as e:
        print(f"时间转换错误: {e}")
        return ''

@app.template_filter('format_datetime_local')
def format_datetime_local(utc_string):
    """与 utc_to_local_filter 功能相同"""
    return utc_to_local_filter(utc_string)

def init_developer_account():
    """初始化开发者账号（如果不存在）"""
    dev_id = os.environ.get('DEVELOPER_USER_ID')
    dev_email = os.environ.get('DEVELOPER_EMAIL', 'dev@example.com')
    
    if not dev_id:
        logger.warning("未配置 DEVELOPER_USER_ID，开发者功能不可用")
        return
    
    db = get_supabase()
    existing = db.table("users").select("id").eq("id", dev_id).execute()
    
    if not existing.data:
        # 创建开发者账号
        dev_data = {
            "id": dev_id,
            "email": dev_email,
            "name_en": "System Developer",
            "role": "developer",
            "user_status": "registered",
            "is_active": True,
            "is_protected": True,  # 开发者账号受保护
            "password_hash": auth.hash_password(os.environ.get('DEVELOPER_PASSWORD', 'ChangeMe123!')),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        try:
            db.table("users").insert(dev_data).execute()
            logger.info(f"开发者账号已创建: {dev_email}")
        except Exception as e:
            logger.error(f"创建开发者账号失败: {e}")
    else:
        # 确保开发者账号受保护
        db.table("users").update({"is_protected": True}).eq("id", dev_id).execute()
        logger.info(f"开发者账号已存在: {dev_id}")

# 在 app 创建后调用
with app.app_context():
    init_developer_account()

# 启动时打印配置（便于调试）
print(f"⚙️ 启动配置: host={HOST}, port={PORT}, debug={DEBUG}")
# ================= 全局函数 =================
@app.route('/health')
def health_check():
    # 返回一个简单的 "OK" 和 200 状态码即可
    return "OK", 200

def get_allowed_countries():
    """返回当前管理员允许的国家代码列表，若为超管或未设限制则返回 None"""
    if session.get('role') == 'super_admin':
        return None
    data = session.get('admin_countries')
    if not data:
        return None
    if isinstance(data, str):
        try:
            return json.loads(data)
        except:
            return []
    return data

def local_to_utc(local_time_str, local_tz=None):
    """
    将本地时间字符串（格式：YYYY-MM-DDTHH:MM）转换为 UTC ISO 字符串。
    如果未传时区，则使用环境变量 LOCAL_TIMEZONE 或默认 Asia/Kathmandu。
    """
    if not local_time_str:
        return None
    local_dt = datetime.fromisoformat(local_time_str)
    tz_name = local_tz if local_tz else DEFAULT_LOCAL_TIMEZONE
    local_tz_obj = pytz.timezone(tz_name)
    local_dt_aware = local_tz_obj.localize(local_dt)
    utc_dt = local_dt_aware.astimezone(timezone.utc)
    return utc_dt.isoformat()

def robust_parse_json(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, str):
                return robust_parse_json(parsed)
            return parsed
        except:
            return {}
    return {}

def get_training_status(training):
    start = training.get('start_time')
    end = training.get('end_time')
    now = datetime.now(timezone.utc)
    if not start or not end:
        return 'draft'      # 草稿
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end)
    if now < start_dt:
        return 'pending'    # 未推送
    elif now > end_dt:
        return 'closed'     # 已关闭
    else:
        return 'active'     # 已推送/进行中

def random_pick_questions(exam_id, count):
    """1. 随机抽题工具函数（放在文件顶部或合适位置） 从指定考试中随机抽取 count 道题目，返回题目列表"""
    db = get_supabase()
    q_res = db.table("questions").select("*").eq("exam_id", exam_id).execute()
    questions = q_res.data or []
    if len(questions) <= count:
        return questions
    return random.sample(questions, count)

def get_default_reviewer_by_country(country_code):
    """根据国家代码获取默认阅卷人（取该国第一个管理员）"""
    if not country_code:
        return None
    
    db = get_supabase()
    # 查询该国第一个管理员（role = 'admin' 或 'super_admin'）
    # 取姓名 + 工号格式
    res = db.table("users") \
        .select("name_en, employee_id") \
        .eq("country", country_code) \
        .eq("user_status", "registered") \
        .in_("role", ["admin", "super_admin"]) \
        .limit(1) \
        .execute()
    
    if res.data and len(res.data) > 0:
        admin = res.data[0]
        name = admin.get('name_en', '')
        emp_id = admin.get('employee_id', '')
        if name and emp_id:
            return f"{name} ({emp_id})"
        elif name:
            return name
    return None

# ================= 后台定时任务：自动提交超时考试 =================
def auto_submit_timeout_exams():
    """扫描所有已超时但未提交的考试，执行自动提交"""
    with app.app_context():
        db = get_supabase()
        now = datetime.now(timezone.utc)  # 使用带时区的 UTC 时间
        logger.debug("⏳ 开始扫描超时考试...")

        status_res = db.table("user_exam_status").select("*").eq("is_submitted", False).execute()
        if not status_res.data:
            logger.debug("没有未提交的考试")
            return
            
        for status in status_res.data:
            user_id = status['user_id']
            exam_id = status['exam_id']
            started_at_str = status.get('started_at')
            if not started_at_str:
                continue
            
            # 获取考试时长
            exam_info = db.table("exams").select("duration").eq("id", exam_id).maybe_single().execute()
            duration_minutes = exam_info.data.get("duration", Config.DEFAULT_EXAM_DURATION) if exam_info.data else Config.DEFAULT_EXAM_DURATION
            total_seconds = duration_minutes * 60

            # 计算是否超时
            try:
                start_dt = datetime.fromisoformat(started_at_str.replace('Z', '+00:00'))
                elapsed = (now - start_dt).total_seconds()
            except Exception as e:
                logger.warning(f"时间解析失败: {e}")
                continue
            
            if elapsed < total_seconds:
                continue
            
            logger.info(f"⏰ 检测到超时考试：用户 {user_id}，考试 {exam_id}，超时 {int(elapsed - total_seconds)} 秒")
            
            # --- 1. 读取草稿答案 ---
            answers = {}
            draft_res = db.table("user_exam_drafts").select("answers").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
            # 检查 data 是否为字典且包含 answers 字段
            if draft_res is not None and hasattr(draft_res, 'data') and draft_res.data and isinstance(draft_res.data, dict):
                raw = draft_res.data['answers']
                try:
                    if isinstance(raw, str):
                        answers = json.loads(raw)      # 直接使用原始字典
                    else:
                        answers = raw
                    logger.info(f"📝 读取到草稿答案：用户 {user_id}，题目数 {len(answers)}，内容：{answers}")
                except json.JSONDecodeError as e:
                    logger.error(f"草稿答案 JSON 解析失败: {e}，原始数据: {raw}")
            else:
                logger.warning(f"未找到草稿答案：用户 {user_id}，考试 {exam_id}")
            
            logger.debug(f"draft_res.data = {draft_res.data}")
            # --- 2. 执行评分 ---
            try:
                grade = exam.auto_grade(answers, exam_id)
                logger.info(f"📊 自动评分结果：总分 {grade['total']}")
            except Exception as e:
                logger.error(f"评分失败: {e}", exc_info=True)
                continue
            
            # --- 3. 保存成绩 ---
            try:
                customs = {}
                exam.save_result(user_id, exam_id, answers, grade['total'], grade['details'], customs)
            except Exception as e:
                logger.error(f"保存成绩失败: {e}", exc_info=True)
                continue
            
            # --- 4. 更新状态为已提交 ---
            try:
                existing = db.table("user_exam_status").select("id").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
                update_data = {
                    "is_submitted": True,
                    "submitted_at": now.isoformat(),
                    "reset_at": None
                }
                if existing.data:
                    db.table("user_exam_status").update(update_data).eq("id", existing.data['id']).execute()
                else:
                    update_data.update({"user_id": user_id, "exam_id": exam_id, "started_at": started_at_str})
                    db.table("user_exam_status").insert(update_data).execute()
                
                logger.info(f"✅ 自动提交成功：用户 {user_id}，得分 {grade['total']}")
            except Exception as e:
                logger.error(f"状态更新失败: {e}", exc_info=True)
            
            # --- 5. 清理草稿（可选） ---
            try:
                db.table("user_exam_drafts").delete().eq("user_id", user_id).eq("exam_id", exam_id).execute()
            except:
                pass

# 创建调度器实例
scheduler = BackgroundScheduler()
# 从环境变量读取扫描间隔（由 launcher.py 设置）
SCAN_INTERVAL = int(os.environ.get('EXAM_SCAN_INTERVAL', 60))
# 添加定时任务
scheduler.add_job(func=auto_submit_timeout_exams, trigger="interval", seconds=SCAN_INTERVAL)
# 启动调度器
scheduler.start()
# 确保程序退出时关闭调度器
atexit.register(lambda: scheduler.shutdown())

# ================= 3. 全局异常处理器（必须在 app 创建之后！） =================
@app.errorhandler(Exception)
def handle_all_exceptions(e):
    # ✅ 移除 flush=True（print_exc 不支持该参数）
    print(f"\n❌❌❌ GLOBAL ERROR: {type(e).__name__}: {e}", file=sys.stderr, flush=True)
    print(f"📋 Stack trace:", file=sys.stderr, flush=True)
    traceback.print_exc(file=sys.stderr)  # ✅ 移除 flush
    print(f"❌❌❌ END ERROR ❌❌❌\n", file=sys.stderr, flush=True)
    logger.error(f"未捕获异常: {type(e).__name__}: {e}", exc_info=True)
    return f"500 Error: {type(e).__name__} - 查看终端获取详情", 500
    raise e   # ✅ 调试用

# ================= 4. 启动校验 =================
try:
    Config.check()
except RuntimeError as e:
    logger.warning(f"⚠️ 配置警告: {e}")

# ================= 5. 路由装饰器 =================
def login_required(f):
    """登录校验装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """管理员权限装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ('admin', 'super_admin'):
            #flash('权限不足，仅限管理员访问', 'danger')
            flash({'msg': 'permission_denied_admin', 'params': []}, 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# ================= 6. 调试路由（诊断用） =================
@app.route('/debug/ping')
def debug_ping():
    """测试路由连通性"""
    print(f"\n🔥 /debug/ping 被访问！🔥\n", flush=True)
    return "✅ PONG - 路由系统正常"

@app.route('/debug/upload', methods=['POST'])
def debug_upload():
    """测试文件上传"""
    print(f"\n🔥 /debug/upload 被调用！🔥\n", flush=True)
    if 'file' not in request.files:
        print("❌ 无 file 字段", file=sys.stderr, flush=True)
        return jsonify({"error": "No file"}), 400
    file = request.files['file']
    print(f"📄 收到: {file.filename}, size={file.content_length}", flush=True)
    return jsonify({"ok": True, "filename": file.filename})

@app.route('/test')
def test_route():
    """简单测试路由"""
    print(f"\n🔥 /test 被访问！🔥\n", flush=True)
    return "✅ OK - 路由系统正常"

# ================= 7. 认证页面路由 =================
@app.route('/register')
def register():
    """渲染注册页面"""
    return render_template('auth/register.html')

@app.route('/api/check-name')
def check_name():
    """检查用户名是否已存在"""
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    db = get_supabase()
    res = db.table("users").select("name_en") \
        .eq("user_status", "imported") \
        .ilike("name_en", f"%{q}%") \
        .limit(10).execute()
    names = list(dict.fromkeys(r['name_en'] for r in (res.data or []) if r.get('name_en')))
    return jsonify(names)

@app.route('/api/countries')
def api_countries():
    """所有国家选择处改为下拉框（动态加载）"""
    db = get_supabase()
    res = db.table("countries").select("code, name_zh, name_en").execute()
    return jsonify(res.data)

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    db = get_supabase()
    user_id = session['user_id']
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update_info':
            # 更新基本信息
            update_data = {
                'birthday': request.form.get('birthday', ''),
                'name_en': request.form.get('name_en', ''),
                'company': request.form.get('company', ''),
                'department': request.form.get('department', ''),
                'employee_id': request.form.get('employee_id', ''),
                'country': request.form.get('country', ''),
                'phone': request.form.get('phone', ''),
                'birthday': request.form.get('birthday', '')  or None
            }
            db.table('users').update(update_data).eq('id', user_id).execute()
            #flash('个人信息已更新', 'success')
            flash({'msg': 'profile_updated', 'params': []}, 'success')

        elif action == 'change_password':
            old_pwd = request.form.get('old_password')
            new_pwd = request.form.get('new_password')
            confirm_pwd = request.form.get('confirm_password')
            
            if new_pwd != confirm_pwd:
                #flash('两次输入的新密码不一致', 'danger')
                flash({'msg': 'password_mismatch', 'params': []}, 'danger')
                return redirect(url_for('profile'))
            if len(new_pwd) < 6:
                #flash('密码长度至少6位', 'danger')
                flash({'msg': 'password_too_short', 'params': []}, 'danger')
                return redirect(url_for('profile'))
            
            # 验证原密码
            user_res = db.table('users').select('password_hash').eq('id', user_id).execute()
            if not user_res.data or not auth.check_password(old_pwd, user_res.data[0]['password_hash']):
                #flash('原密码错误', 'danger')
                flash({'msg': 'wrong_password', 'params': []}, 'danger')
                return redirect(url_for('profile'))
            
            # 更新密码
            new_hash = auth.hash_password(new_pwd)
            db.table('users').update({'password_hash': new_hash}).eq('id', user_id).execute()
            #flash('密码修改成功，请重新登录', 'success')
            flash({'msg': 'password_changed', 'params': []}, 'success')
            session.clear()
            return redirect(url_for('login'))
        
        return redirect(url_for('profile'))
    
    # GET 请求：显示表单
    user_res = db.table('users').select('*').eq('id', user_id).single().execute()
    user = user_res.data
    return render_template('auth/profile.html', user=user)

# ================= 8. 认证 API =================
@app.route('/api/send-otp', methods=['POST'])
def api_send_otp():
    """发送邮箱验证码"""
    email = request.json.get('email')
    if not email:
        return jsonify({"success": False, "message": "缺少邮箱"}), 400
    try:
        auth.send_otp(email)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"❌ send_otp 失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/register', methods=['POST'])
def api_register():
    """用户注册"""
    d = request.json
    email = d.get('email', '').strip().lower()
    name_en = d.get('name_en', '').strip()
    password = d.get('password', '')
    birthday = d.get('birthday', '')               # 格式 YYYY-MM-DD
    is_partner_val = d.get('is_partner', 'N')
    otp = d.get('otp', '')
    
    # 1. 邮箱全局唯一性检查
    db = get_supabase()
    if db.table("users").select("id").eq("email", email).execute().data:
        # 邮箱已注册
        # return jsonify({"success": False, "message": "该邮箱已注册，请直接登录或更换邮箱"}), 400
        return jsonify({"success": False, "message": "email_already_registered", "params": []}), 400
    
    country = d.get('country', '').strip().upper()
    if not country:
        return jsonify({"success": False, "message": "jsonify_country_required", "params": []}), 400

    # 2. 姓名 + 出生日期精确匹配 imported 用户（且尚未设置邮箱）
    base_query = db.table("users").select("*") \
        .eq("name_en", name_en) \
        .eq("user_status", "imported") \
        .is_("email", "null") \
        .is_("deleted_at", "null")   # ✅ 防止匹配到已删除用户

    pool = base_query.execute()
    users = pool.data or []

    # 在 Python 中过滤生日条件
    if birthday:
        # 提供生日：要求导入记录的生日为空或者与提供的生日一致
        matched_users = [u for u in users if u.get('birthday') is None or u.get('birthday') == birthday]
    else:
        # 未提供生日：要求导入记录的生日为空
        matched_users = [u for u in users if u.get('birthday') is None]
    
    count = len(matched_users)
    target = matched_users[0] if matched_users else None

    if count == 0:
        if birthday:
            # 姓名与生日不匹配
            #return jsonify({"success": False, "message": "姓名与出生日期不匹配，请核对或联系管理员"}), 403
            return jsonify({"success": False, "message": "name_birthday_mismatch", "params": []}), 403
        else:
            # 姓名未匹配
            #return jsonify({"success": False, "message": "姓名未匹配到预授权名单，请联系管理员"}), 403
            return jsonify({"success": False, "message": "name_not_matched", "params": []}), 403

    if count > 1:
        # 多条匹配
        #return jsonify({"success": False, "message": "该姓名和出生日期对应多条预授权记录，请通知管理员修正数据"}), 403
        return jsonify({"success": False, "message": "multiple_imported_records", "params": []}), 403

    # 3. 姓名匹配通过后，再验证 OTP
    if not auth.verify_otp(email, otp):
        # 验证码无效或已过期
        # return jsonify({"success": False, "message": "验证码无效或已过期"})
        return jsonify({"success": False, "message": "otp_invalid", "params": []})

    # 4. 更新记录，完成注册
    update_fields = {
        "email": email,
        "password_hash": auth.hash_password(password),
        "user_status": "registered",
        "is_active": True,
        "is_partner": True if is_partner_val.upper() == 'Y' else False,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "country": country
    }
    # 如果前端传了出生日期，可更新（确保一致），但预授权已有生日则不必改
    if birthday:
        update_fields["birthday"] = birthday

    db.table("users").update(update_fields).eq("id", target['id']).execute()

    # 5. 自动登录
    session.update({
        "user_id": target['id'],
        "user_email": email,
        "role": target.get('role', 'user'),
        "admin_countries": target.get('admin_countries', '')  # ✅ 新增
    })
    return jsonify({"success": True, "redirect": url_for('index')})

@app.route('/api/reset-password', methods=['POST'])
def api_reset_password():
    """密码重置，用户自身重置"""
    d = request.json
    if not auth.verify_otp(d.get('email'), d.get('otp')):
        return jsonify({"success": False, "message": "otp_invalid", "params": []})
    db = get_supabase()
    db.table("users").update({
        "password_hash": auth.hash_password(d['password'])
    }).eq("email", d['email']).execute()
    return jsonify({"success": True})

# ================= 9. 页面路由 =================
@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录时存储用户国家到 session"""
    if request.method == 'POST':
        email = request.form['email']
        pwd = request.form['password']
        db = get_supabase()
        user = None
        try:
            res = db.table("users").select("*").eq("email", email).is_("deleted_at", "null").maybe_single().execute()
            if res is not None and hasattr(res, 'data'):
                user = res.data
            elif isinstance(res, dict):
                user = res
        except Exception as e:
            logger.error(f"登录查询异常: {e}")

        if user and auth.check_password(pwd, user.get('password_hash', '')):
            # 解析 admin_countries
            admin_countries = user.get('admin_countries', '')
            if admin_countries and isinstance(admin_countries, str):
                try:
                    json.loads(admin_countries)
                except:
                    admin_countries = json.dumps([])
            
            # ✅ 新增：存储用户国家到 session
            user_country = user.get('country')
            # ✅ 添加调试日志
            logger.info(f"用户登录: id={user['id']}, role={user.get('role')}, country={user_country}, admin_countries={admin_countries}")
            
            session.update({
                "user_id": user['id'],
                "user_email": email,
                "role": user.get('role', 'user'),
                "admin_countries": admin_countries,
                "is_protected": user.get('is_protected', False),
                "user_country": user_country  # ✅ 新增
            })
            # ✅ 打印 session 内容
            logger.info(f"Session 已设置: user_country={session.get('user_country')}, role={session.get('role')}")

            flash({'msg': 'login_success', 'params': []}, 'success')
            return redirect(url_for('index'))
        else:
            flash({'msg': 'invalid_email_or_password', 'params': []}, 'danger')
            return render_template('auth/login.html')
    return render_template('auth/login.html')

@app.route('/logout')
def logout():
    """退出登录"""
    session.clear()
    #flash('已安全退出', 'info')
    flash({'msg': 'logout_success', 'params': []}, 'info')
    return redirect(url_for('login'))

@app.route('/')
def index():
    # 如果用户已登录，展示宣贯首页
    if 'user_id' in session:
        return render_template('index.html')
    # 未登录则重定向到登录页面
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    db = get_supabase()
    user_id = session['user_id']
    now = datetime.now(timezone.utc)

    try:
        # 获取当前用户被分配的考试ID
        assigned_exam_ids = set()
        assign_res = db.table("exam_assignments").select("exam_id").eq("user_id", user_id).execute()
        if assign_res.data:
            assigned_exam_ids = {a['exam_id'] for a in assign_res.data}
        logger.debug(f"用户 {user_id} 被分配的考试ID: {assigned_exam_ids}")

        # 获取所有考试
        exams_res = db.table("exams").select("*").execute()
        exams = []

        for ex in exams_res.data or []:
            exam_id = ex['id']

            # 检查考试是否有任何分配记录
            any_assign = db.table("exam_assignments").select("id").eq("exam_id", exam_id).limit(1).execute()
            has_any_assignment = len(any_assign.data or []) > 0

            # 可见性判断
            if has_any_assignment and exam_id not in assigned_exam_ids:
                continue

            # 状态过滤
            status = get_exam_status(ex)
            if status not in ('created', 'active'):
                continue

            # 6. 统计题目数量
            q_count = db.table("questions").select("*", count="exact").eq("exam_id", exam_id).execute()
            ex['questions_count'] = q_count.count if hasattr(q_count, 'count') else len(q_count.data or [])

            # 7. 获取当前用户在该考试中的状态（是否开始、是否已提交、剩余时间）
            status_res = db.table("user_exam_status").select(
                "exam_id, started_at, is_submitted, reset_at"
            ).eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
            status_data = status_res.data if status_res and status_res.data else {}

            ex['user_status'] = {}
            ex['user_status']['is_submitted'] = status_data.get('is_submitted', False)
            ex['user_status']['started'] = False
            ex['user_status']['remaining'] = None

            if not status_data.get('is_submitted') and status_data.get('started_at'):
                total_seconds = ex.get('duration', 60) * 60
                try:
                    start_dt = datetime.fromisoformat(status_data['started_at'])
                    elapsed = (now - start_dt).total_seconds()
                    remaining = max(0, total_seconds - int(elapsed))
                    ex['user_status']['started'] = True
                    ex['user_status']['remaining'] = remaining
                except Exception as e:
                    logger.warning(f"解析 started_at 失败: {e}")

            # 8. 判断是否可进入考试（是否已到开始时间）
            start_time = ex.get('start_time')
            can_enter = False
            if start_time:
                try:
                    start_dt = datetime.fromisoformat(start_time)
                    can_enter = now >= start_dt
                except:
                    can_enter = False
            ex['can_enter'] = can_enter

            # 添加考试成绩总分字段（用于前端显示“满分”）
            # 注意：题库总分可能需要从 questions 表动态计算，这里简化，默认100或从题库累加。
            # 可选：从考试配置读取 total_score，若无则动态计算
            total_score = 0
            q_score_res = db.table("questions").select("score").eq("exam_id", exam_id).execute()
            for q in (q_score_res.data or []):
                total_score += q.get('score', 0)
            ex['total_score'] = total_score if total_score else 100

            exams.append(ex)

        # 9. 查询用户的考试成绩（最近5条）
        # 查询用户的考试成绩（只显示有效考试）
        results_res = db.table("exam_results").select("*").eq("user_id", user_id).order("created_at", desc=True).limit(5).execute()
        results = []
        if results_res.data:
            exam_ids = list(set(r['exam_id'] for r in results_res.data))
            # 批量获取未被软删除的考试信息
            valid_exams = db.table("exams").select("id, title").in_("id", exam_ids).is_("deleted_at", "null").execute()
            valid_exam_map = {ex['id']: ex['title'] for ex in (valid_exams.data or [])}
            
            for r in results_res.data:
                if r['exam_id'] in valid_exam_map:
                    r['exam_title'] = valid_exam_map[r['exam_id']]
                    results.append(r)

        # 10. 获取用户姓名
        user_info_res = db.table("users").select("name_cn, name_en, email").eq("id", user_id).single().execute()
        user_info = user_info_res.data
        user_name = user_info.get('name_cn') or user_info.get('name_en') or user_info.get('email')

        return render_template(
            'exam/dashboard.html',
            exams=exams,
            results=results,
            training_open=True,
            user_signed_in=False,
            user_name=user_name
        )
    except Exception as e:
        logger.error(f"dashboard 异常: {e}", exc_info=True)
        return render_template(
            'exam/dashboard.html',
            exams=[],
            results=[],
            training_open=True,
            user_signed_in=False
        )


@app.route('/exam/take/<int:exam_id>')
@login_required
def take_exam(exam_id):
    try:
        db = get_supabase()
        user_id = session['user_id']
        now = datetime.now(timezone.utc)

        # 1. 获取考试基本信息（有效期和时长）
        exam_info = db.table("exams").select("start_time, end_time, duration").eq("id", exam_id).maybe_single().execute()
        if not exam_info.data:
            flash("考试不存在", "danger")
            return redirect(url_for('dashboard'))
        
        exam = exam_info.data
        start_time = exam.get('start_time')
        end_time = exam.get('end_time')
        duration_minutes = exam.get('duration', 60)

        # 2. 检查考试是否已结束
        if end_time:
            end_dt = datetime.fromisoformat(end_time)
            if now > end_dt:
                flash({'msg': 'exam_ended', 'params': []}, 'warning')
                return redirect(url_for('dashboard'))

        # 3. 检查考试是否未开始
        if start_time:
            start_dt = datetime.fromisoformat(start_time)
            if now < start_dt:
                flash({'msg': 'exam_not_started', 'params': []}, 'warning')
                return redirect(url_for('dashboard'))

        # 4. 检查是否已交卷
        status = db.table("user_exam_status").select("*").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
        if status and status.data and status.data.get("is_submitted"):
            flash("您已完成本场考试，无法再次进入。", "warning")
            return redirect(url_for('dashboard'))
        
        total_seconds = duration_minutes * 60
        
        # 5. 处理开始时间和重置标志
        started_at = None
        reset_timer = False
        reset_token = ''

        if status and status.data:
            started_at = status.data.get("started_at")
            submitted_at = status.data.get("submitted_at")
            reset_at = status.data.get("reset_at")
            
            if reset_at and (not submitted_at or reset_at > submitted_at):
                reset_timer = True
                reset_token = reset_at
            else:
                reset_token = ''
            
            if not started_at or reset_timer:
                started_at = now.isoformat()
                db.table("user_exam_status").update({
                    "started_at": started_at,
                    "reset_at": None
                }).eq("user_id", user_id).eq("exam_id", exam_id).execute()
        else:
            started_at = now.isoformat()
            db.table("user_exam_status").insert({
                "user_id": user_id,
                "exam_id": exam_id,
                "started_at": started_at,
                "is_submitted": False
            }).execute()
        
        # 6. 剩余时间计算
        if started_at:
            try:
                start_dt = datetime.fromisoformat(started_at)
                elapsed = (now - start_dt).total_seconds()
            except:
                elapsed = 0
            remaining = max(0, total_seconds - int(elapsed))
        else:
            remaining = total_seconds
        
        # 7. 超时处理
        if remaining <= 0:
            if not (status and status.data and status.data.get("is_submitted")):
                logger.info(f"⏰ 考试 {exam_id} 用户 {user_id} 超时，自动提交")
                try:
                    # 尝试从草稿表获取答案
                    draft = db.table("user_exam_drafts").select("answers").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
                    answers = {}
                    if draft and draft.data:
                        raw = draft.data.get('answers')
                        try:
                            answers = json.loads(raw) if isinstance(raw, str) else raw
                        except:
                            answers = {}
                    
                    grade = exam.auto_grade(answers, exam_id)
                    customs = {}
                    exam.save_result(user_id, exam_id, answers, grade['total'], grade['details'], customs)
                    
                    existing = db.table("user_exam_status").select("id").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
                    update_data = {
                        "is_submitted": True,
                        "submitted_at": datetime.now(timezone.utc).isoformat(),
                        "reset_at": None
                    }
                    if existing and existing.data:
                        db.table("user_exam_status").update(update_data).eq("id", existing.data['id']).execute()
                    else:
                        update_data.update({"user_id": user_id, "exam_id": exam_id, "started_at": started_at})
                        db.table("user_exam_status").insert(update_data).execute()
                    
                    logger.info(f"✅ 超时自动提交完成，得分 {grade['total']}")
                    flash({'msg': 'flash_time_expired_exam_automatically', 'params': []}, "info")
                except Exception as e:
                    logger.error(f"❌ 超时自动提交失败: {e}")
                    flash({'msg': 'flash_time_expired_exam_automatic_failed', 'params': []}, "danger")
            else:
                flash({'msg': 'flash_completed_exam_cannot_reenter', 'params': []}, "warning")
            return redirect(url_for('dashboard'))

        # 8. 查询题目
        qs = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
        questions = qs.data or []
        for q in questions:
            if isinstance(q.get('options'), str):
                try:
                    q['options'] = json.loads(q['options'])
                except:
                    q['options'] = {}
        
        # 9. 获取用户信息
        user_info_res = db.table("users").select("name_en, email").eq("id", user_id).single().execute()
        user_info = user_info_res.data
        user_display_name = f"{user_info.get('name_en', '')} ({session.get('user_email', '')})" if user_info.get('name_en') else session.get('user_email', 'User')

        logger.info(f"考试 {exam_id} 用户 {user_id} 进入，剩余 {remaining} 秒")
        return render_template(
            'exam/take.html',
            exam_id=exam_id,
            questions=questions,
            duration_minutes=duration_minutes,
            server_remaining_seconds=remaining,
            reset_timer=reset_timer,
            reset_token=reset_token,
            user_id=session['user_id'],
            user_display_name=user_display_name
        )
    except Exception as e:
        logger.error(f"take_exam 发生异常: {e}", exc_info=True)
        flash({'msg': 'flash_loading_failed_try_later', 'params': []}, "danger")
        return redirect(url_for('dashboard'))

@app.route('/api/exam/draft', methods=['POST'])
@login_required
def save_exam_draft():
    """保存考试草稿（答案）"""
    try:
        db = get_supabase()
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "无效的请求数据"}), 400

        exam_id = data.get('exam_id')
        answers = data.get('answers', {})
        if not exam_id:
            return jsonify({"success": False, "message": "缺少考试ID"}), 400

        user_id = session['user_id']
        if not user_id:
            return jsonify({"success": False, "message": "未登录"}), 401

        # 转为 JSON 字符串存入数据库
        answers_json = json.dumps(answers) if isinstance(answers, dict) else json.dumps({})
        
        # 检查是否已存在草稿记录
        try:
            existing = db.table("user_exam_drafts").select("id").eq("user_id", user_id).eq("exam_id", int(exam_id)).maybe_single().execute()
        except Exception as e:
            logger.warning(f"草稿查询失败，假定不存在: {e}")
            existing = None

        now = datetime.utcnow().isoformat()

        # 判断是否存在有效数据
        if existing and hasattr(existing, 'data') and existing.data:
            # 更新现有记录
            db.table("user_exam_drafts").update({
                    "answers": answers_json,
                    "updated_at": now}).eq("id", existing.data['id']).execute()
            logger.info(f"✏️ 草稿已更新：用户 {user_id}，考试 {exam_id}")
        else:
            # 插入新记录
            db.table("user_exam_drafts").insert({
                    "user_id": user_id,
                    "exam_id": int(exam_id),
                    "answers": answers_json,
                    "updated_at": now}).execute()
            logger.info(f"✅ 草稿已创建：用户 {user_id}，考试 {exam_id}")
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"❌ 保存草稿失败: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/exam/submit/<int:exam_id>', methods=['POST'])
@login_required
def submit_exam(exam_id):
    db = get_supabase()
    user_id = session['user_id']
    logger.info(f"📥 收到交卷请求：用户 {user_id}，考试 {exam_id}")

    # 在提交前增加超时校验（可选）
    exam_info = db.table("exams").select("duration").eq("id", exam_id).maybe_single().execute()
    duration_minutes = exam_info.data.get("duration", 60) if exam_info.data else 60
    total_seconds = duration_minutes * 60

    # 获取当前 UTC 时间（aware）
    now = datetime.now(timezone.utc)

    # 解析 started_at（假设数据库存储的是 ISO 8601 带时区格式）
    status = db.table("user_exam_status").select("started_at").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
    if status.data and status.data.get("started_at"):
        start_dt = datetime.fromisoformat(status.data['started_at'])
        elapsed = (now - start_dt).total_seconds()
        if elapsed > total_seconds:
            #flash("考试时间已超时，无法提交。", "danger")
            flash({'msg': 'exam_timeout', 'params': []}, 'danger')
            return redirect(url_for('dashboard'))

    # 检查是否已交卷
    try:
        existing = db.table("user_exam_status").select("id").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
        if existing.data and existing.data.get("is_submitted"):
            #flash('您已完成此考试，不能重复提交', 'warning')
            flash({'msg': 'already_submitted', 'params': []}, 'warning')
            return redirect(url_for('dashboard'))
    except Exception as e:
        logger.warning(f"状态检查失败: {e}")

    # 解析答案
    answers = {}
    for key, values in request.form.to_dict(flat=False).items():
        if key.startswith('q_'):
            # 多选题 checkbox 会返回列表，单选 radio 返回单个值
            if len(values) == 1:
                answers[key] = values[0]
            else:
                answers[key] = ''.join(sorted(values))  # 多选合并为字符串，如 "ABC"
    
    logger.info(f"📥 考生提交答案：{answers}")

    # 评分
    try:
        grade = exam.auto_grade(answers, exam_id)
        logger.info(f"📊 评分结果：总分={grade['total']}，详情={grade['details']}")
    except Exception as e:
        logger.error(f"❌ 评分失败: {e}")
        #flash('评分过程出错，请重试', 'danger')
        flash({'msg': 'grading_error', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))

    # 保存成绩
    customs = {f"c{i}": request.form.get(f"custom{i}", "") for i in range(1, 6)}
    try:
        exam.save_result(user_id, exam_id, answers, grade['total'], grade['details'], customs)
        logger.info(f"💾 成绩保存成功")
        
        # 备份答案到草稿表（防止 exam_results 中 answers 丢失）
        try:
            answers_json = json.dumps(answers)
            draft_res = db.table("user_exam_drafts").select("id").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
            if draft_res and draft_res.data:
                db.table("user_exam_drafts").update({"answers": answers_json}).eq("id", draft_res.data['id']).execute()
            else:
                db.table("user_exam_drafts").insert({"user_id": user_id, "exam_id": exam_id, "answers": answers_json}).execute()
            logger.info("✅ 答案已备份到草稿表")
        except Exception as e:
            logger.warning(f"备份草稿失败: {e}")
    except Exception as e:
        logger.error(f"❌ 成绩保存失败: {e}")
        #flash('成绩保存失败，请联系管理员', 'danger')
        flash({'msg': 'save_score_failed', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))

    # 标记已提交（先查询是否存在，再更新或插入）
    try:
        existing = db.table("user_exam_status") \
            .select("id") \
            .eq("user_id", user_id) \
            .eq("exam_id", exam_id) \
            .maybe_single().execute()

        update_data = {
            "is_submitted": True,
            "submitted_at": datetime.now().isoformat(),
            "reset_at": None
        }

        if existing.data:
            # 已有记录，更新
            db.table("user_exam_status") \
                .update(update_data) \
                .eq("id", existing.data['id']) \
                .execute()
            logger.info(f"✅ 考试状态已更新：用户 {user_id}，考试 {exam_id}")
        else:
            # 无记录，插入
            update_data.update({
                "user_id": user_id,
                "exam_id": exam_id
            })
            db.table("user_exam_status").insert(update_data).execute()
            logger.info(f"✅ 考试状态已插入：用户 {user_id}，考试 {exam_id}")
    except Exception as e:
        logger.error(f"❌ 状态写入失败: {e}")
        # 可选：给用户一个警告，但成绩已保存，可不阻断流程

    db.table("user_exam_drafts").delete().eq("user_id", user_id).eq("exam_id", exam_id).execute()
    flash(f'✅ 交卷成功！得分：{grade["total"]}', 'success')
    return redirect(url_for('dashboard'))

# ================= 10. 管理员路由 =================
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    """管理员仪表盘"""
    db = get_supabase()
    from utils.permissions import get_admin_allowed_countries, is_developer
    
    # 获取当前管理员的权限范围
    allowed_countries = get_admin_allowed_countries()
    is_dev = is_developer()
    
    logger.info(f"admin_dashboard: role={session.get('role')}, allowed_countries={allowed_countries}")
    
    # ==================== 1. 用户统计 ====================
    # 已注册用户数量（user_status = 'registered'）
    registered_query = db.table("users").select("id", count="exact")\
        .eq("user_status", "registered")\
        .is_("deleted_at", "null")
    
    # 已导入用户数量（user_status = 'imported'）
    imported_query = db.table("users").select("id", count="exact")\
        .eq("user_status", "imported")\
        .is_("deleted_at", "null")
    
    # 应用国家权限过滤
    if allowed_countries is not None:
        if not allowed_countries:
            registered_count = 0
            imported_count = 0
        else:
            registered_query = registered_query.in_("country", allowed_countries)
            imported_query = imported_query.in_("country", allowed_countries)
            registered_count = registered_query.execute().count or 0
            imported_count = imported_query.execute().count or 0
    else:
        registered_count = registered_query.execute().count or 0
        imported_count = imported_query.execute().count or 0
    
    logger.info(f"用户统计: 已注册={registered_count}, 已导入={imported_count}")
    
    # ==================== 2. 考试统计 ====================
    # 获取所有考试（需要根据权限过滤）
    exams_query = db.table("exams").select("*").is_("deleted_at", "null")
    
    if allowed_countries is not None and allowed_countries:
        # 获取允许国家下的用户ID（用于通过考试分配关联）
        users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
        allowed_user_ids = [u['id'] for u in (users_in_allowed.data or [])] if users_in_allowed.data else []
        
        # 查询分配了允许国家考生的考试ID
        allowed_exam_ids = set()
        if allowed_user_ids:
            assign_res = db.table("exam_assignments").select("exam_id").in_("user_id", allowed_user_ids).execute()
            allowed_exam_ids = {a['exam_id'] for a in (assign_res.data or [])}
        
        # 获取所有考试
        all_exams = exams_query.execute().data or []
        
        # 过滤：考试自身 country 在允许列表中 或 考试ID在 allowed_exam_ids 中
        filtered_exams = []
        for exam in all_exams:
            if exam.get('country') in allowed_countries:
                filtered_exams.append(exam)
            elif exam['id'] in allowed_exam_ids:
                filtered_exams.append(exam)
    else:
        filtered_exams = exams_query.execute().data or []
    
    # 统计考试状态
    exam_stats = {'draft': 0, 'created': 0, 'active': 0, 'closed': 0}
    for exam in filtered_exams:
        status = get_exam_status(exam)
        if status in exam_stats:
            exam_stats[status] += 1
    
    exams_total = len(filtered_exams)
    
    # 统计已完成考试数量（有成绩记录且考试在权限范围内）
    completed_query = db.table("exam_results").select("exam_id", count="exact").execute()
    completed_exam_ids = set([r['exam_id'] for r in (completed_query.data or [])])
    exams_completed = len([e for e in filtered_exams if e['id'] in completed_exam_ids])
    
    logger.info(f"考试统计: 总数={exams_total}, 已完成={exams_completed}, 草稿={exam_stats['draft']}, 进行中={exam_stats['active']}, 已关闭={exam_stats['closed']}")
    
    # ==================== 3. 培训统计 ====================
    trainings_query = db.table("trainings").select("*").is_("deleted_at", "null")
    
    if allowed_countries is not None and allowed_countries:
        # 获取允许国家下的用户ID
        users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
        allowed_user_ids = [u['id'] for u in (users_in_allowed.data or [])] if users_in_allowed.data else []
        
        # 查询存在允许国家学员签到的培训ID
        allowed_training_ids = set()
        if allowed_user_ids:
            attend_res = db.table("training_attendances").select("training_id").in_("user_id", allowed_user_ids).execute()
            allowed_training_ids = {a['training_id'] for a in (attend_res.data or [])}
        
        # 获取所有培训
        all_trainings = trainings_query.execute().data or []
        
        # 过滤
        filtered_trainings = []
        for training in all_trainings:
            if training.get('country') in allowed_countries:
                filtered_trainings.append(training)
            elif training['id'] in allowed_training_ids:
                filtered_trainings.append(training)
    else:
        filtered_trainings = trainings_query.execute().data or []
    
    trainings_count = len(filtered_trainings)
    
    # 统计签到总人次（只统计权限范围内的）
    total_attendances = 0
    if allowed_countries is not None and allowed_countries:
        # 查询权限范围内用户的签到记录
        users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
        allowed_user_ids = [u['id'] for u in (users_in_allowed.data or [])] if users_in_allowed.data else []
        if allowed_user_ids:
            attend_count = db.table("training_attendances").select("id", count="exact").in_("user_id", allowed_user_ids).execute()
            total_attendances = attend_count.count or 0
    else:
        attend_count = db.table("training_attendances").select("id", count="exact").execute()
        total_attendances = attend_count.count or 0
    
    # 今日签到数量
    from datetime import date
    today = date.today().isoformat()
    if allowed_countries is not None and allowed_countries:
        if allowed_user_ids:
            signins_today = db.table("training_attendances").select("id", count="exact")\
                .in_("user_id", allowed_user_ids)\
                .gte("sign_time", today).execute().count or 0
        else:
            signins_today = 0
    else:
        signins_today = db.table("training_attendances").select("id", count="exact")\
            .gte("sign_time", today).execute().count or 0
    
    logger.info(f"培训统计: 培训数={trainings_count}, 签到总人次={total_attendances}, 今日签到={signins_today}")
    
    # ==================== 4. 访谈统计 ====================
    interviews_query = db.table("interviews").select("*").is_("deleted_at", "null")
    
    if allowed_countries is not None and allowed_countries:
        # 获取允许国家下的用户ID
        users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
        allowed_user_ids = [u['id'] for u in (users_in_allowed.data or [])] if users_in_allowed.data else []
        
        # 查询存在允许国家学员的访谈ID
        allowed_interview_ids = set()
        if allowed_user_ids:
            interview_res = db.table("interview_results").select("interview_id").in_("user_id", allowed_user_ids).execute()
            allowed_interview_ids = {i['interview_id'] for i in (interview_res.data or [])}
        
        # 获取所有访谈
        all_interviews = interviews_query.execute().data or []
        
        # 过滤
        filtered_interviews = []
        for interview in all_interviews:
            # 通过关联考试的国家来判断
            exam_id = interview.get('exam_id')
            if exam_id:
                exam_res = db.table("exams").select("country").eq("id", exam_id).maybe_single().execute()
                if exam_res.data and exam_res.data.get('country') in allowed_countries:
                    filtered_interviews.append(interview)
                elif interview['id'] in allowed_interview_ids:
                    filtered_interviews.append(interview)
            elif interview['id'] in allowed_interview_ids:
                filtered_interviews.append(interview)
    else:
        filtered_interviews = interviews_query.execute().data or []
    
    interviewee_count = len(filtered_interviews)
    logger.info(f"访谈统计: 访谈数={interviewee_count}")
    
    # ==================== 5. 题库统计 ====================
    # 题库统计也需要根据权限过滤（只统计权限范围内考试的题目）
    try:
        if allowed_countries is not None and allowed_countries:
            # 获取权限范围内的考试ID
            exams_in_allowed = [e['id'] for e in filtered_exams]
            if exams_in_allowed:
                questions_count = db.table("questions").select("id", count="exact")\
                    .in_("exam_id", exams_in_allowed).execute().count or 0
            else:
                questions_count = 0
        else:
            questions_count = db.table("questions").select("id", count="exact").execute().count or 0
    except:
        questions_count = 0
    
    logger.info(f"题库统计: 题目数={questions_count}")
    
    # ==================== 6. 获取考试列表（用于前端表格和下拉框）====================
    now = datetime.now(timezone.utc)
    exams_for_table = []
    exams_for_selector = []
    
    for exam in filtered_exams:
        status = get_exam_status(exam)
        exam['status'] = status
        
        # 统计应考/实考人数（只统计权限范围内的用户）
        exam_id = exam['id']
        
        # 获取该考试分配的考生（需要权限过滤）
        assign_query = db.table("exam_assignments").select("user_id").eq("exam_id", exam_id)
        if allowed_countries is not None and allowed_countries:
            # 只统计权限范围内的考生
            if allowed_user_ids:
                assign_query = assign_query.in_("user_id", allowed_user_ids)
                assigned_count = assign_query.execute().count or 0
            else:
                assigned_count = 0
        else:
            assigned_count = assign_query.execute().count or 0
        
        # 获取已提交的考生（需要权限过滤）
        submitted_query = db.table("exam_results").select("user_id", count="exact").eq("exam_id", exam_id)
        if allowed_countries is not None and allowed_countries:
            if allowed_user_ids:
                submitted_query = submitted_query.in_("user_id", allowed_user_ids)
                submitted_count = submitted_query.execute().count or 0
            else:
                submitted_count = 0
        else:
            submitted_count = submitted_query.execute().count or 0
        
        exam['assigned_count'] = assigned_count
        exam['submitted_count'] = submitted_count
        
        # 仪表盘表格显示：草稿、已创建、进行中
        if status in ["draft", "created", "active"]:
            exam['dynamic_status'] = status
            exams_for_table.append(exam)
        
        # 考生考试状态下拉框显示：进行中
        if status in ["active"]:
            exams_for_selector.append(exam)
    
    # ==================== 7. 获取培训签到开关状态 ====================
    try:
        config_res = db.table("system_config").select("value").eq("key", "training_open").execute()
        sign_in_open = config_res.data[0].get('value', 'false').lower() == 'true' if config_res.data else False
    except:
        sign_in_open = False
    
    # ==================== 8. 组装统计数据 ====================
    stats = {
        "users": registered_count,           # 已注册用户数
        "users_imported": imported_count,    # 已导入用户数（新增）
        "exams_total": exams_total,
        "exams_completed": exams_completed,
        "exam_draft": exam_stats.get('draft', 0),
        "exam_active": exam_stats.get('active', 0),
        "exam_closed": exam_stats.get('closed', 0),
        "trainings_count": trainings_count,
        "total_attendances": total_attendances,
        "signins_today": signins_today,
        "questions": questions_count
    }
    
    logger.info(f"最终统计: {stats}")
    
    return render_template(
        'admin/dashboard.html',
        signs=[],  # 保留兼容性
        exams_table=exams_for_table,
        exams_selector=exams_for_selector,
        stats=stats,
        sign_in_open=sign_in_open,
        questions_count=questions_count,
        signins_today=signins_today,
        total_attendances=total_attendances,
        trainings_count=trainings_count,
        interviewee_count=interviewee_count
    )

@app.route('/admin/reset_exam/<int:exam_id>/<user_id>', methods=['POST'])
@login_required
@admin_required
def admin_reset_exam(exam_id, user_id):
    """重置指定考生的考试状态让其可以再次考试"""
    db = get_supabase()
    try:
        # 查询现有记录
        reset_at = datetime.now(timezone.utc).isoformat()
        existing = db.table("user_exam_status").select("id").eq("user_id", user_id).eq("exam_id", exam_id).maybe_single().execute()
        
        if existing.data:
            # 更新现有记录
            db.table("user_exam_status").update({
                "is_submitted": False,
                "reset_at": reset_at,
                "started_at": None,   # 清除开始时间，让考生下次进入重新计时
                "submitted_at": None
            }).eq("id", existing.data['id']).execute()
        else:
            # 插入新记录
            db.table("user_exam_status").insert({
                "user_id": user_id,
                "exam_id": int(exam_id),
                "is_submitted": False,
                "reset_at": reset_at
            }).execute()
        
        # 🧹 删除草稿记录，防止服务器端残留旧答案
        db.table("user_exam_drafts").delete().eq("user_id", user_id).eq("exam_id", int(exam_id)).execute()

        logger.info(f"✅ 重置成功（已清除草稿）：用户 {user_id}，考试 {exam_id}")
        return jsonify({"success": True, "reset_token": reset_at})  # 返回时间戳作为重置标识
    except Exception as e:
        logger.error(f"❌ 重置失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/admin/exam/delete/<int:exam_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_exam(exam_id):
    """支持软删除和永久删除"""
    permanent = request.args.get('permanent', 'false').lower() == 'true'
    db = get_supabase()
    
    try:
        if permanent:
            # 永久删除：删除考试及其关联题目
            db.table("questions").delete().eq("exam_id", exam_id).execute()
            db.table("exams").delete().eq("id", exam_id).execute()
            flash(f"考试 ID {exam_id} 已永久删除", "success")
        else:
            # 软删除：标记 deleted_at
            db.table("exams").update({"deleted_at": datetime.now(timezone.utc).isoformat()}).eq("id", exam_id).execute()
            flash(f"考试 ID {exam_id} 已移至回收站（软删除）", "info")
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        logger.error(f"删除考试失败: {e}")
        flash("删除失败", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/exam/restore/<int:exam_id>', methods=['POST'])
@login_required
@admin_required
def restore_exam(exam_id):
    db = get_supabase()
    db.table("exams").update({"deleted_at": None}).eq("id", exam_id).execute()
    flash(f"考试 ID {exam_id} 已从回收站恢复", "success")
    return redirect(url_for('admin_exams_page'))

@app.route('/admin/import', methods=['GET', 'POST'])
@login_required      # 🔧 测试阶段注释，通过后取消注释
@admin_required      # 🔧 测试阶段注释，通过后取消注释
def admin_import():
    """Word 题库导入页面"""
    print(f"\n🔥🔥🔥 admin_import 被调用！method={request.method} 🔥🔥🔥\n", flush=True)
    
    if request.method == 'POST' and 'docx_file' in request.files:
        file = request.files['docx_file']
        logger.info(f"📄 收到文件: {file.filename}, size={file.content_length}")
        
        if not file.filename.endswith('.docx'):
            logger.warning(f"❌ 文件格式错误: {file.filename}")
            #flash('❌ 仅支持 .docx 格式', 'danger')
            flash({'msg': 'only_docx', 'params': []}, 'danger')
            return redirect(request.url)
        
        import tempfile, os
        tmp_path = None
        
        try:
            # Windows 兼容的临时文件处理
            with tempfile.NamedTemporaryFile(
                delete=False, suffix='.docx', dir=os.getenv('TEMP', '/tmp')
            ) as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name
            logger.info(f"💾 临时文件已保存: {tmp_path}")
            
            # 调用双语解析函数
            exam_title, qs = exam.parse_docx_bilingual(tmp_path, exam_id=0)  # 不再需要 exam_id 参数
            # 如果解析出的标题无效（空或默认值），则使用文件名
            if not exam_title or exam_title == '未命名考试':
                exam_title = os.path.splitext(os.path.basename(file.filename))[0]
                logger.info(f"使用文件名作为考试标题: {exam_title}")
            logger.info(f"✅ 解析成功: 返回 {len(qs)} 道题目")
            
            if not qs:
                logger.warning("⚠️ 解析结果为空")
                #flash('⚠️ 未识别到有效题目，请检查 Word 格式', 'warning')
                flash({'msg': 'no_valid_question', 'params': []}, 'warning')
                return render_template('admin/import.html')
            
            logger.info("🔄 跳转预览页")
            return render_template('admin/import_preview.html', questions=qs, exam_title=exam_title)
            
        except AttributeError as e:
            logger.error(f"❌ AttributeError: {e}")
            logger.error(f"💡 请确认 services/exam.py 中存在 parse_docx_bilingual 函数")
            #flash('❌ 系统错误: 解析函数未找到', 'danger')
            flash({'msg': 'parse_func_missing', 'params': []}, 'danger')
        except FileNotFoundError as e:
            logger.error(f"❌ 文件未找到: {e}")
            #flash('❌ 临时文件创建失败', 'danger')
            flash({'msg': 'temp_file_failed', 'params': []}, 'danger')
        except PermissionError as e:
            logger.error(f"❌ 权限错误: {e}")
            #flash('❌ 文件访问权限不足', 'danger')
            flash({'msg': 'file_permission_denied', 'params': []}, 'danger')
        except Exception as e:
            logger.error(f"❌ 未知异常: {type(e).__name__}: {e}")
            logger.error(f"📋 完整堆栈:\n{traceback.format_exc()}")
            #flash(f'❌ 解析失败: {str(e)}', 'danger')
            flash({'msg': 'parse_error', 'params': [str(e)]}, 'danger')          # 带参数
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                    logger.debug(f"🗑️ 已清理临时文件: {tmp_path}")
                except Exception:
                    pass
        return redirect(request.url)
    
    # GET 请求：渲染上传表单
    logger.info("📄 渲染 import.html")
    return render_template('admin/import.html')

@app.route('/api/admin/users/import', methods=['POST'])
@login_required
@admin_required
def api_admin_import_users():
    """批量导入用户（带角色权限校验）"""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "jsonify_no_file_selected", "params": []}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "jsonify_only_supports_files", "params": []}), 400
    
    db = get_supabase()
    current_role = session.get('role')
    is_dev = is_developer()
    
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"success": False, "message": "file_parse_error", "params": [str(e)]}), 400
    
    # 字段映射
    column_map = {
        'country': 'country',
        'email': 'email',
        'name_en': 'name_en',
        'role': 'role',
        'is_partner': 'is_partner',
        'company': 'company',
        'department': 'department',
        'wh_type': 'wh_type',
        'wh_id': 'wh_id',
        'wh_name_en': 'wh_name_en',
        'employee_id': 'employee_id',
        'phone': 'phone',
        'birthday': 'birthday',
        'admin_countries': 'admin_countries'
    }
    
    # 读取表头
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(str(cell_value).strip().lower() if cell_value else '')
    
    # 必须列
    if 'name_en' not in headers:
        return jsonify({"success": False, "message": "Excel missing required column: name_en"}), 400
    
    success_count = 0
    error_rows = []
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            if not header:
                continue
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                row_data[header] = ''
            else:
                row_data[header] = str(cell_value).strip()
        
        # 字段映射
        user_data = {}
        for excel_field, db_field in column_map.items():
            if excel_field in row_data:
                user_data[db_field] = row_data[excel_field]

        # ========== 处理空字符串字段，转为 None ==========
        if 'birthday' in user_data:
            birthday_val = user_data['birthday']
            if birthday_val == '' or birthday_val is None:
                user_data['birthday'] = None
        
        if 'employee_id' in user_data:
            if user_data['employee_id'] == '':
                user_data['employee_id'] = None
        
        if 'email' in user_data:
            email_val = user_data['email'].strip().lower() if user_data['email'] else ''
            user_data['email'] = email_val if email_val else None
        
        if 'phone' in user_data:
            if user_data['phone'] == '':
                user_data['phone'] = None
        
        if 'company' in user_data:
            if user_data['company'] == '':
                user_data['company'] = None
        
        if 'department' in user_data:
            if user_data['department'] == '':
                user_data['department'] = None
        
        if 'wh_type' in user_data:
            if user_data['wh_type'] == '':
                user_data['wh_type'] = None
        
        if 'wh_id' in user_data:
            if user_data['wh_id'] == '':
                user_data['wh_id'] = None
        
        if 'wh_name_en' in user_data:
            if user_data['wh_name_en'] == '':
                user_data['wh_name_en'] = None
        
        if 'country' in user_data:
            if user_data['country'] == '':
                user_data['country'] = None
        
        # 姓名必填
        name_en = user_data.get('name_en', '')
        if not name_en:
            error_rows.append(f"第{row_idx}行: 姓名不能为空")
            continue
        
        # ========== 角色权限校验 ==========
        role_raw = user_data.get('role', '')
        role = role_raw.lower() if role_raw else 'user'
        user_data['role'] = role
        
        if not is_dev and role == 'developer':
            error_rows.append(f"第{row_idx}行: 只有开发者可以创建开发者账号")
            continue
        
        if current_role == 'super_admin' and not is_dev:
            if role not in ['super_admin', 'admin', 'user']:
                error_rows.append(f"第{row_idx}行: 无效角色 '{role}'")
                continue
        elif current_role == 'admin' and not is_dev:
            if role != 'user':
                error_rows.append(f"第{row_idx}行: 管理员只能导入普通用户角色")
                continue
        
        # ========== 超管/管理员权限范围校验 ==========
        if role in ['super_admin', 'admin']:
            admin_countries_raw = user_data.get('admin_countries', '')
            if not admin_countries_raw:
                error_rows.append(f"第{row_idx}行: {role}角色的权限范围不能为空")
                continue
            
            from utils.permissions import parse_countries_input
            countries_list = parse_countries_input(admin_countries_raw)
            if not countries_list:
                error_rows.append(f"第{row_idx}行: 权限范围解析失败，请填写有效的国家代码或名称")
                continue
            
            user_data['admin_countries'] = json.dumps(countries_list)
        else:
            user_data['admin_countries'] = None
        
        # ========== 重复检查 ==========
        birthday = user_data.get('birthday')
        employee_id = user_data.get('employee_id')
        
        existing_by_name = db.table("users").select("id, birthday, employee_id").eq("name_en", name_en).execute()
        existing_users = existing_by_name.data or []
        
        is_duplicate = False
        for existing in existing_users:
            existing_birthday = existing.get('birthday')
            existing_employee_id = existing.get('employee_id')
            
            if birthday and existing_birthday == birthday:
                is_duplicate = True
                break
            if employee_id and existing_employee_id == employee_id:
                is_duplicate = True
                break
            if not birthday and not employee_id:
                is_duplicate = True
                break
        
        if is_duplicate:
            error_rows.append(f"第{row_idx}行: 用户已存在（姓名重复或生日/工号重复）")
            continue
        
        # ========== 处理 is_partner ==========
        if 'is_partner' in user_data:
            val = str(user_data['is_partner']).upper()
            user_data['is_partner'] = val in ('Y', 'YES', 'TRUE', '1')
        else:
            user_data['is_partner'] = False
        
        # ========== ✅ 关键修复：设置默认密码哈希 ==========
        # 为 imported 用户设置一个默认的密码哈希（不可登录）
        # 使用一个固定的占位符哈希，表示该用户尚未注册
        from services.auth import hash_password
        DEFAULT_PLACEHOLDER_HASH = hash_password('__IMPORTED_USER_PLACEHOLDER__')
        
        user_data['password_hash'] = DEFAULT_PLACEHOLDER_HASH
        
        # ========== 设置默认值 ==========
        user_data['id'] = str(uuid.uuid4())
        user_data['user_status'] = 'imported'
        user_data['is_active'] = False
        user_data['created_by'] = session['user_id']
        user_data['is_protected'] = (role == 'developer')
        
        # 确保 created_at 有值
        from datetime import datetime, timezone
        user_data['created_at'] = datetime.now(timezone.utc).isoformat()
        
        # ========== 移除值为空字符串的字段 ==========
        user_data = {k: v for k, v in user_data.items() if v != ''}
        
        try:
            db.table("users").insert(user_data).execute()
            success_count += 1
        except Exception as e:
            error_rows.append(f"第{row_idx}行: 插入失败 - {str(e)}")
            logger.error(f"导入用户失败第{row_idx}行: {e}")
    
    result = {
        "success": True,
        "total": ws.max_row - 1,
        "success_count": success_count,
        "error_count": len(error_rows),
        "errors": error_rows[:20]
    }
    
    response = jsonify(result)
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response

@app.route('/admin/import/save', methods=['POST'])
@login_required
@admin_required
def admin_import_save():
    logger.info(f"/admin/import/save 收到请求，questions 长度: {len(request.json.get('questions', []))}")
    data = request.json.get('questions', [])
    if not data:
        return jsonify({"success": False, "message": "无数据"})

    exam_title = request.args.get('title', '未命名考试')
    is_draft = request.args.get('draft', 'false').lower() == 'true'
    db = get_supabase()
    country_code = request.json.get('country_code', '')  # 新增

    try:
        # ✅ 1. 创建新考试记录（不指定 ID，让数据库自增）
        exam_insert = db.table("exams").insert({
            "title": exam_title,
            "is_active": not is_draft,     # 草稿不激活，正常考试激活
            "country": country_code,   # 存储国家代码
            "status": "draft" if is_draft else "active"   # 草稿状态为 draft
        }).execute()
        
        if not exam_insert.data:
            raise Exception("创建考试失败")
        new_exam_id = exam_insert.data[0]['id']   # 获取自增生成的 ID
        
        # ✅ 2. 插入题目，关联新考试 ID
        for q in data:
            q['exam_id'] = new_exam_id
            q['options'] = json.dumps(q.get('options', {}))

        res = db.table("questions").insert(data).execute()
        logger.info(f"✅ 成功创建考试「{exam_title}」ID={new_exam_id}，插入 {len(res.data)} 道题目")
        return jsonify({"success": True, "exam_id": new_exam_id})
    except Exception as e:
        logger.error(f"❌ 导入保存失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/admin/exam/<int:exam_id>/copy', methods=['POST'])
@login_required
@admin_required
def copy_exam(exam_id):
    """复制考试及其所有题目"""
    data = request.json
    new_title = data.get('new_title')
    if not new_title:
        return jsonify({"success": False, "message": "新考试名称不能为空"}), 400

    db = get_supabase()

    # 1. 获取原考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
    if not exam_res.data:
        return jsonify({"success": False, "message": "原考试不存在"}), 404
    exam = exam_res.data

    # 2. 创建新考试（复制除自增ID和时间戳外的字段）
    new_exam_data = {
        "title": new_title,
        "duration": exam.get("duration", 60),
        "is_active": exam.get("is_active", False),
        "status": exam.get("status", "draft"),
        "start_time": exam.get("start_time"),
        "end_time": exam.get("end_time"),
        "quarter": exam.get("quarter"),
        "created_by": session['user_id'],  # 当前管理员
        "reviewer": exam.get("reviewer"),
        "exam_country" : exam.get('country', ''), #0426
        "country": exam.get("country", '')
    }
    # 移除 None 值，避免数据库报错
    new_exam_data = {k: v for k, v in new_exam_data.items() if v is not None}
    insert_res = db.table("exams").insert(new_exam_data).execute()
    if not insert_res.data:
        return jsonify({"success": False, "message": "复制考试失败"}), 500
    new_exam_id = insert_res.data[0]['id']

    # 3. 复制原考试的所有题目
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).execute()
    if questions_res.data:
        new_questions = []
        for q in questions_res.data:
            new_q = {
                "exam_id": new_exam_id,
                "num": q.get("num"),
                "type": q.get("type"),
                "content": q.get("content"),
                "content_cn": q.get("content_cn"),
                "content_en": q.get("content_en"),
                "content_raw": q.get("content_raw"),
                "options": q.get("options"),
                "answer": q.get("answer"),
                "score": q.get("score")
            }
            new_questions.append(new_q)
        db.table("questions").insert(new_questions).execute()

    return jsonify({"success": True, "new_id": new_exam_id})

@app.route('/admin/exam/copy_preview/<int:exam_id>')
@login_required
@admin_required
def copy_exam_preview(exam_id):
    """拷贝考试预览页（可编辑）"""
    db = get_supabase()
    # 获取原考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
    if not exam_res.data:
        #flash("考试不存在", "danger")
        flash({'msg': 'exam_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_exams_page'))
    exam = exam_res.data
    # 获取原考试的所有题目
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    for q in questions:
        # 解析 options 字符串为字典
        if isinstance(q.get('options'), str):
            try:
                q['options'] = json.loads(q['options'])
            except:
                q['options'] = {}
    # 获取用户输入的新考试名称（从 query 参数）
    new_title = request.args.get('new_title', exam['title'] + '_copy')
    return render_template('admin/import_preview.html', 
                           questions=questions,
                           exam_title=new_title,
                           copy_mode=False,        # 拷贝是新建，非编辑
                           original_exam_id=exam_id,
                           return_url=url_for('admin_exams_page'), # 拷贝后返回的页面有疑问
                           exam_country=exam.get('country', ''),
                           exam_country_name=''   # 可选
                           ) 

@app.route('/admin/exam/edit/<int:exam_id>')
@login_required
@admin_required
def edit_exam_preview(exam_id):
    """编辑考试（跳转到预览页，可修改题目）"""
    db = get_supabase()
    # 获取原考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
    if not exam_res.data:
        flash("考试不存在", "danger")
        return redirect(url_for('admin_dashboard'))
    exam = exam_res.data
    
    # ✅ 新增：检查考试状态，决定可编辑范围
    status = get_exam_status(exam)
    
    # 已关闭的考试不能编辑
    if status == 'closed':
        flash("已关闭的考试不能编辑", "warning")
        return redirect(url_for('admin_exams_page'))
    
    # 进行中的考试：只能编辑基本信息，不能编辑题目
    can_edit_questions = status in ['draft', 'created']
    
    # 获取原考试的所有题目
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    for q in questions:
        if isinstance(q.get('options'), str):
            try:
                q['options'] = json.loads(q['options'])
            except:
                q['options'] = {}

    # 获取考试的国家，并尝试转换为显示名称（可选，但方便前端显示）
    country_code = exam.get('country', '')
    country_name = ''
    if country_code:
        # 可以从国家表查询（如果有）
        try:
            c_res = db.table("countries").select("name_zh").eq("code", country_code).maybe_single().execute()
            if c_res.data:
                country_name = c_res.data['name_zh']
        except:
            pass
    # 传递原考试名称和原考试国家到预览页
    return render_template(
        'admin/import_preview.html',
        questions=questions,
        exam_title=exam['title'],
        edit_mode=True,
        original_exam_id=exam_id,
        return_url=url_for('admin_dashboard'),
        exam_country=country_code,
        exam_country_name=country_name,
        exam_status=status,                     # ✅ 新增：考试状态
        can_edit_questions=can_edit_questions,  # ✅ 新增：是否可编辑题目
        exam_duration=exam.get('duration', 60), # ✅ 新增：考试时长
        exam_reviewer=exam.get('reviewer', '')  # ✅ 新增：阅卷人
        )

@app.route('/admin/exam/<int:exam_id>/update_full', methods=['PUT'])
@login_required
@admin_required
def update_exam_full(exam_id):
    """更新现有考试的信息和题目（支持部分更新）"""
    data = request.json
    db = get_supabase()

    # 获取原考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
    if not exam_res.data:
        return jsonify({"success": False, "message": "考试不存在"}), 404
    
    original = exam_res.data
    current_status = get_exam_status(original)
    
    # 已关闭的考试不能编辑
    if current_status == 'closed':
        return jsonify({"success": False, "message": "已关闭的考试不能编辑"}), 403
    
    # 构建更新数据
    update_data = {}
    
    # 基本信息（所有状态都可更新）
    if 'title' in data:
        update_data['title'] = data['title']
    if 'duration' in data:
        update_data['duration'] = data['duration']
    if 'reviewer' in data:
        update_data['reviewer'] = data['reviewer']
        logger.info(f"更新考试阅卷人: {data['reviewer']}")
    if 'country_code' in data:
        update_data['country'] = data['country_code']
    
    # 有效期（只有草稿和未开始状态可更新）
    if current_status in ['draft', 'created']:
        if 'start_time' in data:
            update_data['start_time'] = data['start_time']
        if 'end_time' in data:
            update_data['end_time'] = data['end_time']
    
    # 更新考试基本信息
    if update_data:
        db.table("exams").update(update_data).eq("id", exam_id).execute()
    
    # 题目更新（只有草稿和未开始状态，且前端要求更新时）
    can_update_questions = current_status in ['draft', 'created']
    should_update_questions = 'questions' in data and data['questions'] and not data.get('skip_questions', False)
    
    if can_update_questions and should_update_questions:
        questions = data.get('questions', [])
        if questions:
            # 删除原题目
            db.table("questions").delete().eq("exam_id", exam_id).execute()
            # 插入新题目
            for q in questions:
                q['exam_id'] = exam_id
                q['options'] = json.dumps(q.get('options', {}))
            db.table("questions").insert(questions).execute()
            logger.info(f"更新了 {len(questions)} 道题目")
    
    # 重新计算状态
    if update_data.get('start_time') and update_data.get('end_time'):
        now = datetime.now(timezone.utc)
        start_dt = datetime.fromisoformat(update_data['start_time'])
        end_dt = datetime.fromisoformat(update_data['end_time'])
        if now < start_dt:
            db.table("exams").update({"status": "created"}).eq("id", exam_id).execute()
        elif now > end_dt:
            db.table("exams").update({"status": "closed"}).eq("id", exam_id).execute()
        else:
            db.table("exams").update({"status": "active", "is_active": True}).eq("id", exam_id).execute()
    
    return jsonify({"success": True, "exam_id": exam_id})

@app.route('/api/admin/exam/<int:exam_id>/settings', methods=['POST'])
@login_required
@admin_required
def admin_exam_settings(exam_id):
    """管理员仪表盘设置考试：有效期、时长、考生分配（清除原有分配，重新分配）"""
    data = request.json
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    duration = data.get('duration')
    user_ids = data.get('user_ids', [])
    db = get_supabase()
    
    update_data = {}
    if start_time:
        update_data['start_time'] = start_time
    if end_time:
        update_data['end_time'] = end_time
    if duration:
        update_data['duration'] = duration
    
    # 更新考试信息
    if update_data:
        db.table("exams").update(update_data).eq("id", exam_id).execute()
    
    # 更新考生分配（先删除旧关联，再插入新关联）
    db.table("exam_assignments").delete().eq("exam_id", exam_id).execute()
    for uid in user_ids:
        db.table("exam_assignments").insert({"exam_id": exam_id, "user_id": uid}).execute()
    
    # 发送邮件通知（如果有效期已设置且状态变为已创建/进行中）
    if start_time and end_time:
        # 获取考试标题
        exam_res = db.table("exams").select("title").eq("id", exam_id).execute()
        exam_title = exam_res.data[0]['title'] if exam_res.data else "考试"
        for uid in user_ids:
            user_res = db.table("users").select("email, name_en").eq("id", uid).execute()
            if user_res.data:
                email = user_res.data[0]['email']
                name = user_res.data[0].get('name_en', '用户')
                subject = f"考试通知：{exam_title}"
                body = f"您好 {name}，您有一场考试《{exam_title}》，有效期从 {start_time} 到 {end_time}，请登录系统参加。"
                try:
                    auth.send_email(email, subject, body)
                except Exception as e:
                    logger.warning(f"发送邮件失败: {e}")
    
    return jsonify({"success": True})

    if start_time and end_time:
        now = datetime.now(timezone.utc)
        start_dt = datetime.fromisoformat(start_time)
        end_dt = datetime.fromisoformat(end_time)
        if now < start_dt:
            status = "created"
        elif now > end_dt:
            status = "closed"
        else:
            status = "active"
    else:
        status = "draft"
    update_data['status'] = status

@app.route('/admin/exam/<int:exam_id>/duration', methods=['POST'])
@login_required
@admin_required
def update_exam_duration(exam_id):
    """后端 API 支持更新时长"""
    data = request.get_json()
    duration = data.get('duration')
    if not duration or not isinstance(duration, int) or duration <= 0:
        return jsonify({"success": False, "message": "无效的时长"}), 400
    db = get_supabase()
    try:
        db.table("exams").update({"duration": duration}).eq("id", exam_id).execute()
        logger.info(f"考试 {exam_id} 时长更新为 {duration} 分钟")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/admin/exam_status/<int:exam_id>')
@login_required
@admin_required
def admin_exam_status(exam_id):
    db = get_supabase()
    allowed = get_allowed_countries()

    # 获取所有用户（只限制国家权限）
    query = db.table("users").select("id, email, name_en, country").is_("deleted_at", "null")
    if allowed is not None:
        if not allowed:
            return jsonify([])
        query = query.in_("country", allowed)
    users_res = query.execute()
    users = users_res.data or []

    # 获取该考试的分配关系
    assign_res = db.table("exam_assignments").select("user_id").eq("exam_id", exam_id).execute()
    assigned_user_ids = {a['user_id'] for a in (assign_res.data or [])}

    # 获取该考试的考试状态（开始时间、是否提交等）
    status_res = db.table("user_exam_status").select("user_id, started_at, is_submitted, submitted_at").eq("exam_id", exam_id).execute()
    status_dict = {}
    for s in (status_res.data or []):
        status_dict[s['user_id']] = s
    
    # 获取成绩记录（仅已提交的才有成绩）
    results_res = db.table("exam_results").select("id, user_id, total_score").eq("exam_id", exam_id).execute()
    results_dict = {r['user_id']: {'result_id': r['id'], 'score': r['total_score']} for r in (results_res.data or [])}
    
    data = []
    for u in users:
        uid = u['id']
        # 判断考试状态
        if uid not in assigned_user_ids:
            exam_status = 'not_assigned'   # 未推送
        else:
            user_status = status_dict.get(uid, {})
            if user_status.get('is_submitted'):
                exam_status = 'submitted'   # 已提交
            elif user_status.get('started_at'):
                exam_status = 'in_progress' # 考试中
            else:
                exam_status = 'pending'     # 待考试（已推送但未开始）

        result_info = results_dict.get(uid, {})
        data.append({
            "user_id": uid,
            "email": u.get('email'),
            "name": u.get('name_en') or u.get('email'),
            "is_submitted": status_dict.get(uid, {}).get('is_submitted', False),
            "submitted_at": status_dict.get(uid, {}).get('submitted_at'),
            "score": result_info.get('score'),
            "result_id": result_info.get('result_id'),
            "exam_status": exam_status   # 新增字段
        })
    return jsonify(data)

@app.route('/admin/result/<int:result_id>')
@login_required
@admin_required
def admin_result_detail(result_id):
    """管理查看考生考试详情"""
    db = get_supabase()
    # 1. 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).maybe_single().execute()
    if not result_res.data:
        #flash("成绩记录不存在", "danger")
        flash({'msg': 'result_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    
    result = result_res.data
    exam_id = result['exam_id']
    user_id = result['user_id']

    # 2. 获取用户信息
    user_res = db.table("users").select("email, name_en").eq("id", user_id).maybe_single().execute()
    user_info = user_res.data if user_res.data else {"email": "未知", "name_en": "未知"}
    
    # 3. 获取考试信息
    exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
    exam_title = exam_res.data.get("title", "未知考试") if exam_res.data else "未知考试"
    
    # 将关联信息附加到 result 对象（模板中会使用 result.users.email 等形式）
    result['users'] = user_info
    result['exams'] = {"title": exam_title}
    
    # 4. 获取题目列表
    questions = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    
    # 5. 解析 JSON 字段（answers 和 details）
    answers = result.get('answers', {})
    if isinstance(answers, str):
        answers = json.loads(answers)
    details = result.get('details', {})
    if isinstance(details, str):
        details = json.loads(details)
    
    return render_template(
        'admin/result_detail.html',
        result=result,
        questions=questions.data or [],
        answers=answers,
        details=details
    )

@app.route('/exam/result/<int:result_id>')
@login_required
def exam_result_detail(result_id):
    """学员查看自己的考试详情"""
    db = get_supabase()
    # 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).maybe_single().execute()
    if not result_res.data:
        #flash("成绩记录不存在", "danger")
        flash({'msg': 'result_not_found', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))
    result = result_res.data
    if result['user_id'] != session['user_id']:
        #flash("无权访问", "danger")
        flash({'msg': 'access_denied', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))
    
    exam_id = result['exam_id']
    user_id = result['user_id']
    
    # 获取用户信息
    user_res = db.table("users").select("email, name_en").eq("id", user_id).maybe_single().execute()
    user_info = user_res.data if user_res.data else {"email": "未知", "name_en": "未知"}
    
    # 获取考试信息
    exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
    exam_title = exam_res.data.get("title", "未知考试") if exam_res.data else "未知考试"
    
    # 附加信息
    result['users'] = user_info
    result['exams'] = {"title": exam_title}
    
    # 获取题目列表
    questions = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    
    # 解析 answers 和 details
    def deep_parse(val):
        if not val:
            return {}
        if isinstance(val, dict):
            return val
        if isinstance(val, str):
            try:
                parsed = json.loads(val)
                if isinstance(parsed, str):
                    return deep_parse(parsed)
                return parsed
            except:
                return {}
        return {}
    
    answers = deep_parse(result.get('answers'))
    details = deep_parse(result.get('details'))
    
    return render_template(
        'exam/result_detail.html',  # 可以复用 admin/result_detail.html，但需调整布局
        result=result,
        questions=questions.data or [],
        answers=answers,
        details=details
    )

# ================= 11. 导出路由 =================
@app.route('/export/pdf/<int:exam_id>')
@login_required
def export_pdf(exam_id):
    """导出 PDF 成绩单"""
    buf = export.generate_pdf(
        "演示考生", 85, [], {}, {}, "Admin"
    )
    return send_file(
        buf, 
        mimetype="application/pdf", 
        as_attachment=True, 
        download_name=f"exam_{exam_id}.pdf"
    )

@app.route('/admin/export/excel/<int:training_id>/<int:exam_id>')
@login_required
@admin_required
def export_bilingual_excel(training_id, exam_id):
    """导出双语 Excel 报告"""
    country = request.args.get('country', None)
    try:
        buffer, filename = export.generate_bilingual_excel(
            training_id=training_id,
            exam_id=exam_id,
            country=country
        )
        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"❌ Excel 导出失败: {e}")
        #flash(f"❌ 导出失败: {str(e)}", "danger")
        flash({'msg': 'export_error', 'params': [str(e)]}, 'danger')
        return redirect(url_for('admin_dashboard'))

@app.route('/api/admin/export_filtered_excel', methods=['POST'])
@login_required
@admin_required
def export_filtered_excel():
    data = request.json
    country = data.get('country', '')
    training_name = data.get('training_name', '')
    exam_name = data.get('exam_name', '')
    start_date = data.get('start_date', '')
    end_date = data.get('end_date', '')
    wh_raw = data.get('wh_id', '').strip()
    wh_id = wh_raw.split('(')[0].strip() if wh_raw else None

    db = get_supabase()

    # 构建统一的候选用户ID列表
    user_ids_for_country = None
    user_ids_for_wh = None

    # 1. 如果指定了国家，获取该国用户ID
    if country:
        users_res = db.table("users").select("id").eq("country", country).execute()
        user_ids_for_country = [u['id'] for u in (users_res.data or [])]
        if not user_ids_for_country:
            # 没有该国用户，直接返回空文件
            wb = openpyxl.Workbook()
            wb.active.title = "空报告"
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")

    # 2. 如果指定了库房，获取该库房下的所有用户ID
    if wh_id:
        users_in_wh = db.table("users").select("id").eq("wh_id", wh_id).execute()
        user_ids_for_wh = [u['id'] for u in (users_in_wh.data or [])]
        if not user_ids_for_wh:
            wb = openpyxl.Workbook()
            wb.active.title = "空报告"
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f"空报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")

    # 3. 合并国家与库房的用户范围
    if user_ids_for_country is not None and user_ids_for_wh is not None:
        final_user_ids = list(set(user_ids_for_country) & set(user_ids_for_wh))
    elif user_ids_for_country is not None:
        final_user_ids = user_ids_for_country
    elif user_ids_for_wh is not None:
        final_user_ids = user_ids_for_wh
    else:
        final_user_ids = None

    # 4. 查询培训（带保险回退）
    if country:
        training_query = db.table("trainings").select("*").eq("country", country)
        if training_name:
            training_query = training_query.ilike("name", f"%{training_name}%")
        if start_date:
            training_query = training_query.gte("start_time", start_date)
        if end_date:
            training_query = training_query.lte("end_time", end_date)
        trainings = training_query.execute().data or []
        if not trainings:
            logger.info(f"按培训国家过滤后无结果，回退为不过滤培训国家")
            training_query = db.table("trainings").select("*")
            if training_name:
                training_query = training_query.ilike("name", f"%{training_name}%")
            if start_date:
                training_query = training_query.gte("start_time", start_date)
            if end_date:
                training_query = training_query.lte("end_time", end_date)
            trainings = training_query.execute().data or []
    else:
        training_query = db.table("trainings").select("*")
        if training_name:
            training_query = training_query.ilike("name", f"%{training_name}%")
        if start_date:
            training_query = training_query.gte("start_time", start_date)
        if end_date:
            training_query = training_query.lte("end_time", end_date)
        trainings = training_query.execute().data or []

    # 5. 查询考试（带保险回退）
    if country:
        exam_query = db.table("exams").select("*").eq("country", country)
        if exam_name:
            exam_query = exam_query.ilike("title", f"%{exam_name}%")
        exams = exam_query.execute().data or []
        if not exams:
            logger.info(f"按考试国家过滤后无结果，回退为不过滤考试国家")
            exam_query = db.table("exams").select("*")
            if exam_name:
                exam_query = exam_query.ilike("title", f"%{exam_name}%")
            exams = exam_query.execute().data or []
    else:
        exam_query = db.table("exams").select("*")
        if exam_name:
            exam_query = exam_query.ilike("title", f"%{exam_name}%")
        exams = exam_query.execute().data or []

    # 6. 调用生成函数
    try:
        buffer, filename = export.generate_bilingual_excel_filtered(
            trainings=trainings,
            exams=exams,
            country=country,
            start_date=start_date,
            end_date=end_date,
            user_ids=final_user_ids,
            wh_id=wh_id
        )
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Excel 生成失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/admin/export_pdf_by_result/<int:result_id>')
@login_required
@admin_required
def admin_export_pdf_by_result(result_id):
    """通过成绩记录ID直接导出PDF（精确匹配，避免取错记录）"""
    db = get_supabase()
    
    # 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).execute()
    if not result_res.data:
        #flash("成绩记录不存在", "danger")
        flash({'msg': 'result_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    result = result_res.data[0]
    
    exam_id = result['exam_id']
    user_id = result['user_id']
    
    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    if not exam_res.data:
        #flash("考试不存在", "danger")
        flash({'msg': 'exam_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    exam_data = exam_res.data[0]
    
    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    if not user_res.data:
        #flash("考生不存在", "danger")
        flash({'msg': 'student_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    user_data = user_res.data[0]
    user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生')

    # 解析 answers 和 details（递归处理双重转义）
    def robust_parse_json(value):
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, str):
                    return robust_parse_json(parsed)
                return parsed
            except:
                return {}
        return {}
    
    answers = robust_parse_json(result.get('answers'))
    details = robust_parse_json(result.get('details'))
    
    logger.info(f"导出 result_id={result_id}, answers 键数: {len(answers)}, details 键数: {len(details)}")
    
    # 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    
    # 阅卷人
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=request.args.get('reviewer')
    )
    
    # 生成 PDF
    try:
        pdf_buffer = export.generate_user_pdf(
            user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
            user_email=user_data.get('email', ''),
            exam_title=exam_data.get('title', '未命名考试'),
            score=result.get('total_score', 0),
            questions=questions,
            answers=answers,
            details=details,
            submitted_at=result.get('created_at', ''),
            reviewer=reviewer
        )
    except Exception as e:
        logger.error(f"PDF 生成失败: {e}")
        #flash(f"PDF 生成失败: {str(e)}", "danger")
        flash({'msg': 'pdf_generation_error', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    
    filename = f"Transcript_{user_name}_{exam_data.get('title', 'exam')}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

def generate_pdf_by_result_id(result_id):
    """根据 result_id 直接生成 PDF 字节流"""
    db = get_supabase()
    # 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).execute()
    if not result_res.data:
        return None
    result = result_res.data[0]
    exam_id = result['exam_id']
    user_id = result['user_id']

    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    exam_data = exam_res.data[0] if exam_res.data else {}

    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    user_data = user_res.data[0] if user_res.data else {}

    # 递归解析 answers, details
    def robust_parse_json(value):
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, str):
                    return robust_parse_json(parsed)
                return parsed
            except:
                return {}
        return {}

    answers = robust_parse_json(result.get('answers'))
    details = robust_parse_json(result.get('details'))

    # 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []

    # 阅卷人
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=None
    )

    # 生成 PDF 并返回字节数据
    from services import export
    pdf_buffer = export.generate_user_pdf(
        user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
        user_email=user_data.get('email', ''),
        exam_title=exam_data.get('title', '未命名考试'),
        score=result.get('total_score', 0),
        questions=questions,
        answers=answers,
        details=details,
        submitted_at=result.get('created_at', ''),
        reviewer=reviewer
    )
    return pdf_buffer.getvalue()  # 返回字节数据

@app.route('/api/admin/exam/batch_export_by_result', methods=['POST'])
@login_required
@admin_required
def admin_batch_export_by_result():
    db = get_supabase()
    data = request.json
    result_ids = data.get('result_ids', [])
    exam_id = data.get('exam_id')
    if not result_ids:
        return jsonify({"error": "未选择成绩记录"}), 400

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for rid in result_ids:
            pdf_bytes = generate_pdf_by_result_id(rid)
            if pdf_bytes:
                # 获取考生姓名用于文件名
                result_res = db.table("exam_results").select("user_id").eq("id", rid).execute()
                if result_res.data:
                    user_id = result_res.data[0]['user_id']
                    user_res = db.table("users").select("name_en").eq("id", user_id).execute()
                    name = user_res.data[0].get('name_en', user_id) if user_res.data else str(user_id)
                    filename = f"{name}_{rid}.pdf"
                    zf.writestr(filename, pdf_bytes)
                else:
                    # 降级文件名
                    zf.writestr(f"result_{rid}.pdf", pdf_bytes)
            else:
                logger.warning(f"无法生成 PDF，result_id={rid}")
    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"exam_{exam_id}_selected_scores.zip"
    )

@app.route('/exam/export_pdf/<int:result_id>')
@login_required
def exam_export_pdf(result_id):
    """学员导出自己的考试PDF"""
    db = get_supabase()
    # 获取成绩记录
    result_res = db.table("exam_results").select("*").eq("id", result_id).execute()
    if not result_res.data:
        #flash("成绩记录不存在", "danger")
        flash({'msg': 'result_not_found', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))
    result = result_res.data[0]
    if result['user_id'] != session['user_id']:
        #flash("无权访问", "danger")
        flash({'msg': 'access_denied', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))
    
    exam_id = result['exam_id']
    user_id = result['user_id']
    
    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    exam_data = exam_res.data[0] if exam_res.data else {}
    
    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    user_data = user_res.data[0] if user_res.data else {}
    user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生')

    # 解析 answers 和 details
    def deep_parse(val):
        if not val:
            return {}
        if isinstance(val, dict):
            return val
        if isinstance(val, str):
            try:
                parsed = json.loads(val)
                if isinstance(parsed, str):
                    return deep_parse(parsed)
                return parsed
            except:
                return {}
        return {}
    
    answers = deep_parse(result.get('answers'))
    details = deep_parse(result.get('details'))
    
    # 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=request.args.get('reviewer')
    )
    
    try:
        pdf_buffer = export.generate_user_pdf(
            user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
            user_email=user_data.get('email', ''),
            exam_title=exam_data.get('title', '未命名考试'),
            score=result.get('total_score', 0),
            questions=questions,
            answers=answers,
            details=details,
            submitted_at=result.get('created_at', ''),
            reviewer=reviewer
        )
    except Exception as e:
        logger.error(f"PDF生成失败: {e}")
        #flash("PDF生成失败", "danger")
        flash({'msg': 'pdf_generation_error', 'params': []}, 'danger')
        return redirect(url_for('dashboard'))
    
    filename = f"Transcript_{user_name}_{exam_data.get('title', 'exam')}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/export_pdf/<int:exam_id>/<user_id>')
@login_required
@admin_required
def admin_export_user_pdf(exam_id, user_id):
    """管理员导出指定考生的成绩单 PDF, 好像未使用"""
    db = get_supabase()
    
    # 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    if not exam_res.data:
        #flash("考试不存在", "danger")
        flash({'msg': 'exam_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    exam_data = exam_res.data[0]
    
    # 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    if not user_res.data:
        #flash("考生不存在", "danger")
        flash({'msg': 'student_not_found', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    user_data = user_res.data[0]
    
    # 获取成绩记录（不使用 maybe_single，避免 204 异常）
    result_res = db.table("exam_results").select("*").eq("exam_id", exam_id).eq("user_id", user_id).execute()
    if not result_res.data:
        #flash("该考生尚无成绩记录", "warning")
        flash({'msg': 'no_score_record', 'params': []}, 'warning')
        return redirect(url_for('admin_dashboard'))
    result = result_res.data[0]

    raw_answers = result.get('answers')
    logger.info(f"原始 answers 类型: {type(raw_answers)}, 值前200: {str(raw_answers)[:200]}")

    raw_details = result.get('details')
    logger.info(f"原始 details 类型: {type(raw_details)}, 值前200: {str(raw_details)[:200]}")

    # 解析 answers 和 details（处理双重转义）
    def robust_parse_json(value, field_name=""):
        """解析可能被多层转义的 JSON 字符串"""
        if not value:
            logger.warning(f"{field_name} 为空或None")
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            # 打印前200字符
            logger.debug(f"{field_name} 原始字符串前200: {value[:200]}")
            try:
                parsed = json.loads(value)
                # 如果解析后仍是字符串，递归解析
                if isinstance(parsed, str):
                    logger.debug(f"{field_name} 第一次解析后仍是字符串，递归解析")
                    return robust_parse_json(parsed, field_name)
                return parsed
            except Exception as e:
                logger.error(f"{field_name} 解析失败: {e}")
                return {}
        return {}

    answers = robust_parse_json(result.get('answers'), "answers")
    details = robust_parse_json(result.get('details'), "details")

    logger.info(f"解析后 answers 类型: {type(answers)}, 键示例: {list(answers.keys())[:5] if isinstance(answers, dict) else 'not dict'}")
    logger.info(f"解析后 details 类型: {type(details)}, 键示例: {list(details.keys())[:5] if isinstance(details, dict) else 'not dict'}")

    # 在获取成绩记录后，如果 answers 为空，则尝试从 user_exam_drafts 表读取草稿答案
    if not answers:
        draft_res = db.table("user_exam_drafts").select("answers").eq("user_id", user_id).eq("exam_id", exam_id).execute()
        if draft_res.data:
            draft_answers = draft_res.data[0].get('answers')
            if draft_answers:
                if isinstance(draft_answers, str):
                    try:
                        answers = json.loads(draft_answers)
                        logger.info(f"从草稿表恢复了答案，共 {len(answers)} 条")
                    except:
                        pass
                else:
                    answers = draft_answers

    logger.info(f"解析后 answers 类型: {type(answers)}, 键示例: {list(answers.keys())[:3]}")
    logger.info(f"解析后 details 类型: {type(details)}, 键示例: {list(details.keys())[:3]}")

    # 获取题目列表（用于展示题干和标准答案）
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []
    
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=request.args.get('reviewer')
    )


    # ✅ 添加调试日志
    logger.info(f"========== 阅卷人调试 ==========")
    logger.info(f"考生国家: {user_data.get('country')}")
    logger.info(f"考试表 reviewer: {exam_data.get('reviewer')}")
    logger.info(f"URL reviewer: {request.args.get('reviewer')}")
    logger.info(f"最终 reviewer: {reviewer}")
    logger.info(f"================================")

    # 生成 PDF（捕获异常）
    try:
        pdf_buffer = export.generate_user_pdf(
            user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
            user_email=user_data.get('email', ''),
            exam_title=exam_data.get('title', '未命名考试'),
            score=result.get('total_score', 0),
            questions=questions,
            answers=answers,
            details=details,
            submitted_at=result.get('created_at', ''),
            reviewer=reviewer
        )
    except Exception as e:
        logger.error(f"PDF 生成失败: {e}")
        #flash(f"PDF 生成失败: {str(e)}", "danger")
        flash({'msg': 'pdf_generation_error', 'params': []}, 'danger')
        return redirect(url_for('admin_dashboard'))
    
    filename = f"Transcript_{user_name}_{exam_data.get('title', 'exam')}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

# ================= 12. 补充 API 路由 =================
@app.route('/api/training/signin', methods=['POST'])
@login_required
def api_signin():
    """培训签到"""
    db = get_supabase()
    exist = db.table("training_signs").select("id").eq(
        "user_id", session['user_id']
    ).eq("training_id", 1).execute()
    if exist.data:
        return jsonify({"success": False, "message": "今日已签到"})
    db.table("training_signs").insert({
        "user_id": session['user_id'], 
        "training_id": 1
    }).execute()
    return jsonify({"success": True})

@app.route('/favicon.ico')
def favicon():
    return '', 204 

@app.route('/.well-known/appspecific/com.chrome.devtools.json')
def devtools():
    return '', 204

@app.route('/shutdown', methods=['POST'])
def shutdown():
    """关闭服务器（仅用于开发环境）"""
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        raise RuntimeError('Not running with the Werkzeug Server')
    func()
    return 'Server shutting down...'

# ================= 培训签到核心 API =================
def upload_signature(signature_base64, training_id, user_id):
    """共用统一的签名上传函数"""
    try:
        header, encoded = signature_base64.split(',', 1)
        img_data = base64.b64decode(encoded)
        storage_path = f"signatures/{training_id}/{user_id}.png"
        
        # 使用 service_role key 创建客户端
        storage_client = create_client(
            Config.SUPABASE_URL,
            os.environ.get('SUPABASE_SERVICE_KEY', Config.SUPABASE_KEY)
        )
        supabase_storage = storage_client.storage.from_("signatures")
        
        # 先删除旧文件（避免覆盖权限问题）
        try:
            supabase_storage.remove([storage_path])
            logger.info(f"已删除旧签名: {storage_path}")
        except Exception as e:
            logger.debug(f"删除旧签名（可能不存在）: {e}")
        
        # 上传新文件
        supabase_storage.upload(storage_path, img_data, {"content-type": "image/png"})
        public_url = supabase_storage.get_public_url(storage_path)
        logger.info(f"签名上传成功: {public_url}")
        return public_url
        
    except Exception as e:
        logger.error(f"上传签名失败: {e}")
        raise e

@app.route('/api/trainings/available')
@login_required
def api_available_trainings():
    """获取当前用户可签到的培训列表（未过期的、未签到的或已签到的）"""
    db = get_supabase()
    user_id = session['user_id']
    now = datetime.now(timezone.utc).isoformat()
    
    # 查询所有有效期内的培训
    '''
    trainings_res = db.table("trainings") \
        .select("*") \
        .eq("is_active", True) \
        .lt("start_time", now) \
        .gt("end_time", now) \
        .execute()
    '''
    trainings_res = db.table("trainings") \
        .select("*") \
        .eq("is_active", True) \
        .execute()

    trainings = trainings_res.data or []
    
    # 查询用户已签到记录
    att_res = db.table("training_attendances") \
        .select("id, training_id, sign_time, signed_name, signature_url") \
        .eq("user_id", user_id) \
        .execute()
    signed_dict = {a['training_id']: a for a in (att_res.data or [])}
    
    result = []
    for t in trainings:
        signed_info = signed_dict.get(t['id'])
        signed = signed_info is not None
        needs_resign = False
        if signed:
            # 如果已签到但签名URL为空，则需要重新签字
            if not signed_info.get('signature_url'):
                needs_resign = True

        # 判断培训是否在有效期内（字符串可直接比较 ISO 格式）
        start = t.get('start_time')
        end = t.get('end_time')
        in_period = False
        if start and end:
            in_period = start < now < end
        
        # 显示条件：①有效期内一律显示；②已签到且需要重新签字（即使已过期）
        if in_period or (signed and needs_resign):
            result.append({
                "id": t['id'],
                "name": t['name'],
                "start_time": t['start_time'],
                "end_time": t['end_time'],
                "signed": signed_info is not None,
                "sign_time": signed_info['sign_time'] if signed_info else None,
                "signed_name": signed_info['signed_name'] if signed_info else None,
                "needs_resign": needs_resign  # ✅ 新增字段
            })
    return jsonify(result)

@app.route('/api/training/sign', methods=['POST'])
@login_required
def api_training_sign():
    """提交手写签名和姓名"""
    data = request.get_json()
    training_id = data.get('training_id')
    signature_base64 = data.get('signature')   # 格式: "data:image/png;base64,xxxx"
    signed_name = data.get('name', '').strip()
    
    if not training_id or not signature_base64:
        return jsonify({"success": False, "message": "缺少必要参数"}), 400
    
    db = get_supabase()
    user_id = session['user_id']

    # 检查用户是否已注册且激活
    user_res = db.table("users").select("is_active, user_status").eq("id", user_id).maybe_single().execute()
    if not user_res.data or not user_res.data.get('is_active') or user_res.data.get('user_status') != 'registered':
        return jsonify({"success": False, "message": "用户未完成注册"}), 400

    # 检查是否已签到
    try:
        exist_res = db.table("training_attendances") \
            .select("id") \
            .eq("training_id", training_id) \
            .eq("user_id", user_id) \
            .maybe_single() \
            .execute()
        # 防御性判断
        if exist_res is not None and hasattr(exist_res, 'data') and exist_res.data:
            return jsonify({"success": False, "message": "您已签到过本培训"}), 400
    except Exception as e:
        logger.error(f"查询签到记录失败: {e}")
        return jsonify({"success": False, "message": "签到检查失败"}), 500
        
    # 检查培训是否在有效期内
    now = datetime.now(timezone.utc).isoformat()
    try:
        training_res = db.table("trainings") \
            .select("start_time, end_time") \
            .eq("id", training_id) \
            .maybe_single() \
            .execute()
        if training_res is None or not hasattr(training_res, 'data') or not training_res.data:
            return jsonify({"success": False, "message": "培训不存在"}), 404
        training = training_res.data
        if now < training['start_time']:
            return jsonify({"success": False, "message": "签到尚未开始"}), 400
        if now > training['end_time']:
            return jsonify({"success": False, "message": "签到已结束"}), 400
    except Exception as e:
        logger.error(f"查询培训信息失败: {e}")
        return jsonify({"success": False, "message": "培训验证失败"}), 500
    
    # 保存签名图片到 Supabase Storage
    try:
        public_url = upload_signature(signature_base64, training_id, user_id)
    except Exception as e:
        return jsonify({"success": False, "message": "签名保存失败"}), 500
    
    # 插入签到记录
    try:
        db.table("training_attendances").insert({
            "training_id": training_id,
            "user_id": user_id,
            "signature_url": public_url,
            "signed_name": signed_name,
            "sign_time": datetime.now(timezone.utc).isoformat()
        }).execute()
    except Exception as e:
        logger.error(f"保存签到记录失败: {e}")
        return jsonify({"success": False, "message": "数据保存失败"}), 500
    
    return jsonify({"success": True, "sign_time": datetime.now(timezone.utc).isoformat()})

@app.route('/api/training/attendance/<int:training_id>')
@login_required
@admin_required
def api_training_attendance(training_id):
    db = get_supabase()
    country = request.args.get('country', '')

    # 获取培训基本信息
    training_res = db.table("trainings").select("*").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return jsonify({"error": "培训不存在"}), 404
    training = training_res.data
    
    # 签到记录查询
    att_res = db.table("training_attendances") \
        .select("id, user_id, signature_url, signed_name, sign_time, users(email, name_cn, name_en, department, employee_id, country, company)") \
        .eq("training_id", training_id) \
        .execute()

    att_list = att_res.data or []
    # 手动按国家过滤
    if country:
        att_list = [rec for rec in att_list if rec.get('users', {}).get('country') == country]

    attendance_list = []
    for rec in att_list or []:
        user = rec.get('users', {})
        # 防止 user 为 None
        if user is None:
            user = {}

        attendance_list.append({
            "id": rec['id'],  # ✅ 新增签到记录ID
            "user_id": rec['user_id'],
            "department": user.get('department', ''),
            "name_cn": user.get('name_cn', ''),
            "name_en": user.get('name_en', ''),
            "employee_id": user.get('employee_id', ''),
            "signed_name": rec.get('signed_name', ''),
            "signature_url": rec.get('signature_url', ''),
            "sign_time": rec.get('sign_time'),
            "company": user.get('company', ''),
            "country": user.get('country', '')
        })

    # 获取表头模板：优先取国家模板，否则取培训主模板
    header_template = None
    if country:
        # 查询国家模板
        ct_res = db.table("training_country_templates")\
            .select("header_template")\
            .eq("training_id", training_id)\
            .eq("country", country)\
            .execute()
        if ct_res.data and len(ct_res.data) > 0:
            header_template = ct_res.data[0].get('header_template')
    if not header_template:
        header_template = training.get('header_template', {})

    return jsonify({
        "training": training,
        "attendances": attendance_list,
        "header_template": header_template
    })

@app.route('/api/admin/training/attendance/<int:attendance_id>/reset-signature', methods=['POST'])
@login_required
@admin_required
def admin_reset_signature(attendance_id):
    """管理员清除指定签到记录的签名，推送重新签字到学员端"""
    db = get_supabase()
    # 获取原记录，保留签到时间
    att_res = db.table("training_attendances").select("*").eq("id", attendance_id).maybe_single().execute()
    if not att_res.data:
        return jsonify({"success": False, "message": "签到记录不存在"}), 404
    
    # 清空签名相关字段，保留 sign_time
    db.table("training_attendances").update({
        "signature_url": None,
        "signed_name": None
    }).eq("id", attendance_id).execute()
    
    logger.info(f"管理员重置签到 {attendance_id} 的签名")
    return jsonify({"success": True})

@app.route('/api/training/resign', methods=['POST'])
@login_required
def api_resign_training():
    """学员重新提交签名（管理员重置后）"""
    data = request.get_json()
    training_id = data.get('training_id')
    signature_base64 = data.get('signature')
    signed_name = data.get('name', '').strip()
    
    if not training_id or not signature_base64:
        return jsonify({"success": False, "message": "缺少必要参数"}), 400
    
    db = get_supabase()
    user_id = session['user_id']
    
    # 查找该培训的签到记录（必须存在）
    exist_res = db.table("training_attendances") \
        .select("id, signature_url") \
        .eq("training_id", training_id) \
        .eq("user_id", user_id) \
        .maybe_single() \
        .execute()
    
    if not exist_res.data:
        return jsonify({"success": False, "message": "签到记录不存在"}), 404
    
    # 检查是否真的需要重新签字（签名已被清空）
    if exist_res.data.get('signature_url'):
        return jsonify({"success": False, "message": "签名已存在，无需重新签字"}), 400
    
    # 上传新签名（先删除旧文件，再上传，避免 Duplicate 或 update 不兼容）
    try:
        public_url = upload_signature(signature_base64, training_id, user_id)
    except Exception as e:
        return jsonify({"success": False, "message": "签名保存失败"}), 500
    
    # 更新签到记录
    db.table("training_attendances").update({
        "signature_url": public_url,
        "signed_name": signed_name
    }).eq("id", exist_res.data['id']).execute()
    
    return jsonify({"success": True})

@app.route('/api/admin/training/<int:training_id>/country_templates_status')
@login_required
@admin_required
def training_country_templates_status(training_id):
    """供前端获取国家模板状态（用于一级菜单“表头录入”按钮的禁用状态判断）"""
    status = get_training_country_templates_status(training_id)
    return jsonify(status)

@app.route('/api/admin/training/<int:training_id>/country_template', methods=['GET'])
@login_required
@admin_required
def get_training_country_template(training_id):
    """1. 获取国家模板接口"""
    country = request.args.get('country')
    if not country:
        return jsonify({"error": "缺少 country 参数"}), 400
    db = get_supabase()
    # 使用 execute()，不用 maybe_single()
    res = db.table("training_country_templates")\
        .select("header_template")\
        .eq("training_id", training_id)\
        .eq("country", country)\
        .execute()
    if res.data and len(res.data) > 0:
        template = res.data[0].get('header_template', {})
    else:
        template = {}
    return jsonify({"template": template})

@app.route('/api/admin/training/<int:training_id>/country_template', methods=['POST'])
@login_required
@admin_required
def save_training_country_template(training_id):
    """2. 保存国家模板接口 保存培训表头模板"""
    data = request.json
    country = data.get('country')
    template = data.get('template')
    
    if not training_id:
        return jsonify({"success": False, "message": "培训ID无效"}), 400
    
    if template is None:
        return jsonify({"success": False, "message": "缺少 template 参数"}), 400
    
    db = get_supabase()
    
    try:
        # 情况1：没有指定国家 → 保存到培训主表的 header_template
        if not country or country == 'null' or country == 'undefined' or country == '':
            db.table("trainings").update({
                "header_template": template
            }).eq("id", training_id).execute()
            logger.info(f"保存培训主表头: training_id={training_id}")
            return jsonify({"success": True})
        
        # 情况2：指定了国家 → 保存到国家模板表
        check_res = db.table("training_country_templates")\
            .select("id")\
            .eq("training_id", training_id)\
            .eq("country", country)\
            .execute()
        
        if check_res.data and len(check_res.data) > 0:
            # 更新现有记录
            db.table("training_country_templates")\
                .update({
                    "header_template": template, 
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })\
                .eq("id", check_res.data[0]['id'])\
                .execute()
            logger.info(f"更新国家模板: training_id={training_id}, country={country}")
        else:
            # 插入新记录
            db.table("training_country_templates")\
                .insert({
                    "training_id": training_id, 
                    "country": country, 
                    "header_template": template
                })\
                .execute()
            logger.info(f"插入国家模板: training_id={training_id}, country={country}")
        
        return jsonify({"success": True})
        
    except Exception as e:
        logger.error(f"保存表头失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/admin/training/<int:training_id>/attendance/print')
@login_required
@admin_required
def training_attendance_print(training_id):
    country = request.args.get('country', '')
    # 复用签到详情模板，但添加一个打印标志
    return render_template('admin/list_training_attendance.html', training_id=training_id, country=country, print_mode=True)

def get_attendance_data(training_id, country=''):
    """
    获取培训签到数据，供 PDF 生成或其他内部调用使用
    返回包含 training、attendances、header_template 的字典
    """
    db = get_supabase()

    # 1. 获取培训基本信息
    training_res = db.table("trainings").select("*").eq("id", training_id).maybe_single().execute()
    if not training_res.data:
        return None
    training = training_res.data

    # 2. 获取签到记录（含关联用户信息）
    att_res = db.table("training_attendances") \
        .select("id, user_id, signature_url, signed_name, sign_time, users(email, name_cn, name_en, department, employee_id, country, company)") \
        .eq("training_id", training_id) \
        .execute()

    att_list = att_res.data or []

    # 按国家过滤（如果有）
    if country:
        att_list = [rec for rec in att_list if rec.get('users', {}).get('country') == country]

    # 构建签到列表
    attendance_list = []
    for rec in att_list:
        user = rec.get('users', {})
        if user is None:
            user = {}
        attendance_list.append({
            "id": rec['id'],
            "user_id": rec['user_id'],
            "department": user.get('department', ''),
            "name_cn": user.get('name_cn', ''),
            "name_en": user.get('name_en', ''),
            "employee_id": user.get('employee_id', ''),
            "signed_name": rec.get('signed_name', ''),
            "signature_url": rec.get('signature_url', ''),
            "sign_time": rec.get('sign_time'),
            "company": user.get('company', ''),
            "country": user.get('country', '')
        })

    # 3. 获取表头模板（优先国家模板，其次培训主模板）
    header_template = None
    if country:
        ct_res = db.table("training_country_templates") \
            .select("header_template") \
            .eq("training_id", training_id) \
            .eq("country", country) \
            .execute()
        if ct_res.data and len(ct_res.data) > 0:
            header_template = ct_res.data[0].get('header_template')

    if not header_template:
        header_template = training.get('header_template', {})

    return {
        "training": training,
        "attendances": attendance_list,
        "header_template": header_template
    }

@app.route('/admin/training/<int:training_id>/attendance/pdf')
@login_required
@admin_required
def download_training_attendance_pdf(training_id):
    country = request.args.get('country', '')
    data = get_attendance_data(training_id, country)
    if not data:
        flash("培训不存在", "danger")
        return redirect(url_for('admin_dashboard'))

    html_content = render_template('admin/attendance_pdf.html',
                                    training=data['training'],        # ✅ 新增
                                    header=data['header_template'],
                                    attendances=data['attendances'])

    # 配置 wkhtmltopdf 路径（根据实际安装位置修改）
    wkhtmltopdf_path = find_wkhtmltopdf()   # 自动查找（支持环境变量 WKHTMLTOPDF_PATH）
    config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
    pdf = pdfkit.from_string(html_content, False, configuration=config,
                             options={
                                'page-size': 'A4',
                                'margin-top': '10mm',
                                'margin-bottom': '10mm',
                                'margin-left': '10mm',
                                'margin-right': '10mm',
                                'encoding': 'UTF-8',
                                'enable-local-file-access': None,
                                # 可选：避免因网络图片慢而超时
                                'javascript-delay': '200',
                                'no-stop-slow-scripts': None,
                             })
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=attendance_{training_id}.pdf'
    return response

# ================= 访谈管理 API =================
@app.route('/admin/interviews')
@login_required
@admin_required
def admin_interviews_page():
    """2. 访谈 CRUD 路由 访谈管理一级菜单页面"""
    return render_template('admin/list_inspection.html')

@app.route('/api/admin/interviews', methods=['GET', 'POST', 'PUT'])
@login_required
@admin_required
def api_admin_interviews():
    """访谈列表查询、创建、更新"""
    db = get_supabase()
    if request.method == 'GET':
        name = request.args.get('name', '')
        country = request.args.get('country', '')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        # 基础查询（分页）
        query = db.table("interviews").select("*", count="exact").is_("deleted_at", "null")
        if name:
            query = query.ilike("title", f"%{name}%")
        query = query.order("created_at", desc=True)
        start = (page - 1) * per_page
        end = start + per_page - 1
        res = query.range(start, end).execute()
        total = res.count if hasattr(res, 'count') else len(res.data or [])

        interviews = res.data or []   # 当前页的访谈列表

        # 获取管理员国家权限
        allowed = get_allowed_countries()

        # 收集当前页所有访谈涉及的用户ID，用于批量查询国家
        all_user_ids = set()
        for inv in interviews:
            user_res = db.table("interview_results").select("user_id").eq("interview_id", inv['id']).execute()
            all_user_ids.update(r['user_id'] for r in (user_res.data or []))

        # 批量查询用户国家（仅当有国家权限限制时）
        user_country_map = {}
        if allowed is not None and all_user_ids:
            users_res = db.table("users").select("id, country").in_("id", list(all_user_ids)).execute()
            user_country_map = {u['id']: u.get('country') for u in (users_res.data or [])}

        now = datetime.now(timezone.utc)
        for inv in interviews:
            # 动态计算状态
            start_time = inv.get('start_time')
            end_time = inv.get('end_time')
            if not start_time or not end_time:
                inv['status'] = 'draft'
            else:
                try:
                    s = datetime.fromisoformat(start_time)
                    e = datetime.fromisoformat(end_time)
                    if now < s:
                        inv['status'] = 'created'
                    elif now > e:
                        inv['status'] = 'closed'
                    else:
                        inv['status'] = 'active'
                except:
                    inv['status'] = 'draft'

            # 统计去重人数（应用国家权限过滤）
            user_res = db.table("interview_results").select("user_id").eq("interview_id", inv['id']).execute()
            user_ids = [r['user_id'] for r in (user_res.data or [])]
            if allowed is not None:
                filtered_ids = [uid for uid in user_ids if user_country_map.get(uid) in allowed]
                inv['interviewee_count'] = len(set(filtered_ids))
            else:
                inv['interviewee_count'] = len(set(user_ids))

            # 附加考试信息
            exam_info = {}
            if inv.get('exam_id'):
                exam_res = db.table("exams").select("title, country").eq("id", inv['exam_id']).maybe_single().execute()
                if exam_res.data:
                    exam_info = exam_res.data
            inv['exam_title'] = exam_info.get('title', '')
            inv['country'] = exam_info.get('country', '')

        return jsonify({"data": interviews, "total": total, "page": page, "per_page": per_page})

    elif request.method == 'POST':
        data = request.json
        exam_id = data.get('exam_id')
        title = data.get('title', '')
        if not title:
            exam_title = ''
            if exam_id:
                exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
                if exam_res.data:
                    exam_title = exam_res.data['title']
            title = f"访谈-{exam_title}" if exam_title else "未命名访谈"
        question_count = data.get('question_count', 5)
        reviewer = data.get('reviewer', '')
        is_draft = data.get('is_draft', False)
        start_time = data.get('start_time')
        end_time = data.get('end_time')

        # ✅ 将本地时间转为 UTC 存储
        start_time_utc = start_time if start_time else None
        end_time_utc = end_time if end_time else None

        status = 'draft' if is_draft else 'active'
        if start_time and end_time:
            status = 'active'  # 简单处理，后续根据时间自动判定

        user_ids = data.get('user_ids', [])
        if user_ids:
            # 校验所有用户均为已注册
            valid_users = db.table("users").select("id").in_("id", user_ids).eq("user_status", "registered").execute()
            if len(valid_users.data or []) != len(user_ids):
                return jsonify({"success": False, "message": "所选考生中包含未注册的用户"}), 400

        interviewee_count = len(user_ids)
        interview_insert = db.table("interviews").insert({
            "title": title,
            "exam_id": exam_id,
            "created_by": session['user_id'],
            "reviewer": reviewer,
            "question_count": question_count,
            "status": status,
            "start_time": start_time_utc,       # 存储 UTC 时间
            "end_time": end_time_utc,           # 存储 UTC 时间
            "created_at": datetime.now(timezone.utc).isoformat(),
            "interviewee_count": interviewee_count
        }).execute()
        if not interview_insert.data:
            return jsonify({"success": False, "message": "创建失败"}), 500
        new_id = interview_insert.data[0]['id']

        # 为选中的学员抽取题目
        user_ids = data.get('user_ids', [])
        for uid in user_ids:
            questions = random_pick_questions(exam_id, question_count)
            for q in questions:
                db.table("interview_results").insert({
                    "interview_id": new_id,
                    "user_id": uid,
                    "question_id": q['id']
                }).execute()
        return jsonify({"success": True, "id": new_id})

    elif request.method == 'PUT':
        data = request.json
        inv_id = data.get('id')
        db = get_supabase()
        # 获取原访谈
        orig_res = db.table("interviews").select("*").eq("id", inv_id).maybe_single().execute()
        if not orig_res.data:
            return jsonify({"success": False, "message": "访谈不存在"}), 404
        orig = orig_res.data

        # 更新基本字段
        update_data = {}
        for field in ['start_time', 'end_time', 'reviewer', 'exam_id', 'question_count', 'title']:
            if field in data:
                val = data[field]
                if field in ('start_time', 'end_time') and val:
                    val = val     # ✅ 转为 UTC
                update_data[field] = val
        if update_data:
            db.table("interviews").update(update_data).eq("id", inv_id).execute()
            logger.info(f"访谈 {inv_id} 字段已更新: {list(update_data.keys())}")

        # 重新抽题（无论状态，只要提供了 user_ids 就更新人员题目）
        if 'user_ids' in data:
            # 删除该访谈的所有现有题目
            db.table("interview_results").delete().eq("interview_id", inv_id).execute()
            logger.info(f"已清除访谈 {inv_id} 的旧题目")
            # 使用最新的 exam_id 和 question_count
            exam_id = data.get('exam_id', orig['exam_id'])
            question_count = data.get('question_count', orig['question_count'])
            for uid in data['user_ids']:
                questions = random_pick_questions(exam_id, question_count)
                for q in questions:
                    db.table("interview_results").insert({
                        "interview_id": inv_id,
                        "user_id": uid,
                        "question_id": q['id']
                    }).execute()
            logger.info(f"已为 {len(data['user_ids'])} 名学员重新抽题")
        
        return jsonify({"success": True})

@app.route('/api/admin/interview/<int:interview_id>', methods=['GET'])
@login_required
@admin_required
def api_get_interview(interview_id):
    """获取某个访谈的详细信息"""
    db = get_supabase()
    inv = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv.data:
        return jsonify({"error": "访谈不存在"}), 404
    # 获取已分配的学员ID
    user_ids_res = db.table("interview_results").select("user_id").eq("interview_id", interview_id).execute()
    user_ids = list(set([r['user_id'] for r in (user_ids_res.data or [])]))
    inv.data['user_ids'] = user_ids
    return jsonify(inv.data)

@app.route('/api/admin/interview/<int:interview_id>/user_ids')
@login_required
@admin_required
def get_interview_user_ids(interview_id):
    """获取某个访谈的所有用户ID"""
    db = get_supabase()
    res = db.table("interview_results").select("user_id").eq("interview_id", interview_id).execute()
    ids = list(set(r['user_id'] for r in (res.data or [])))
    return jsonify({"user_ids": ids})

@app.route('/admin/interview/<int:interview_id>')
@login_required
@admin_required
def admin_interview_detail_page(interview_id):
    """3. 访谈二级菜单数据接口 访谈详情页面"""
    return render_template('admin/list_inspection_details.html', interview_id=interview_id)

@app.route('/api/admin/interview/<int:interview_id>/results')
@login_required
@admin_required
def api_interview_results(interview_id):
    """获取某个访谈的所有结果，支持筛选"""
    db = get_supabase()
    search = request.args.get('search', '')
    country = request.args.get('country', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # 基本信息
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    interview = inv_res.data

    # 查询访谈结果并关联用户信息
    query = db.table("interview_results").select("*, users(name_cn, name_en, country, wh_id)").eq("interview_id", interview_id)
    if country:
        query = query.eq("users.country", country)
    if search:
        query = query.or_(f"users.name_cn.ilike.%{search}%,users.name_en.ilike.%{search}%")
    
    # 分页
    start = (page - 1) * per_page
    end = start + per_page - 1
    res = query.range(start, end).order("user_id").execute()
    total = res.count if hasattr(res, 'count') else len(res.data or [])

    # 组装数据：按用户聚合
    user_results = {}
    for row in (res.data or []):
        uid = row['user_id']
        if uid not in user_results:
            user_info = row.get('users', {})
            user_results[uid] = {
                "user_id": uid,
                "name": user_info.get('name_cn') or user_info.get('name_en', ''),
                "country": user_info.get('country', ''),
                "wh_id": user_info.get('wh_id', ''),
                "results": [],
                "submitted_at": None
            }
        user_results[uid]["results"].append({
            "question_id": row['question_id'],
            "answer": row['answer'],
            "is_correct": row['is_correct'],
            "feedback": row['feedback']
        })
        # 提交时间取该用户最近一条有答案的记录时间
        if row.get('submitted_at') and (not user_results[uid]["submitted_at"] or row['submitted_at'] > user_results[uid]["submitted_at"]):
            user_results[uid]["submitted_at"] = row['submitted_at']

    # 转换为列表并计算答对数量
    result_list = []
    for uid, data in user_results.items():
        correct_count = sum(1 for r in data['results'] if r['is_correct'])
        total_questions = len(data['results'])
        result_list.append({
            "user_id": uid,
            "name": data['name'],
            "country": data['country'],
            "wh_id": data['wh_id'],
            "total_questions": total_questions,
            "correct_count": correct_count,
            "submitted_at": data['submitted_at'],
            "feedback": "",  # 可后续合并
            "results": data['results']
        })

    return jsonify({
        "interview": interview,
        "results": result_list,
        "total": total,
        "page": page,
        "per_page": per_page
    })

@app.route('/api/admin/interview/<int:interview_id>/resample/<user_id>', methods=['POST'])
@login_required
@admin_required
def resample_interview(interview_id, user_id):
    """4. 重新访谈接口 重新为指定用户抽题，保留历史记录"""
    db = get_supabase()
    inv_res = db.table("interviews").select("exam_id, question_count").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"success": False, "message": "访谈不存在"}), 404
    exam_id = inv_res.data['exam_id']
    count = inv_res.data['question_count']
    
    # 删除该用户在此访谈中的旧题目（仅删除未作答的？根据需求保留历史记录，这里简单起见先删除所有旧记录再插入新题）
    db.table("interview_results").delete().eq("interview_id", interview_id).eq("user_id", user_id).execute()
    
    questions = random_pick_questions(exam_id, count)
    for q in questions:
        db.table("interview_results").insert({
            "interview_id": interview_id,
            "user_id": user_id,
            "question_id": q['id']
        }).execute()
    return jsonify({"success": True})

@app.route('/api/admin/interview/preview', methods=['POST'])
@login_required
@admin_required
def interview_preview():
    """5. 预览接口（用于模态框 → 预览页） 预览访谈：返回每个被选中学员的抽题情况"""
    db = get_supabase()
    data = request.json
    exam_id = data.get('exam_id')
    
    # 检查题库是否存在
    q_check = db.table("questions").select("id").eq("exam_id", exam_id).limit(1).execute()
    if not q_check.data:
        return jsonify({"error": "该考试没有题目"}), 400

    user_ids = data.get('user_ids', [])
    question_count = data.get('question_count', 5)

    if not exam_id:
        return jsonify({"error": "exam_id 不能为空"}), 400

    logger.info(f"预览访谈: exam_id={exam_id}, users={len(user_ids)}, count={question_count}")
    
    exam_res = db.table("exams").select("title").eq("id", exam_id).maybe_single().execute()
    exam_title = exam_res.data['title'] if exam_res.data else ''
    
    preview = []
    for uid in user_ids:
        user_res = db.table("users").select("name_cn, name_en").eq("id", uid).maybe_single().execute()
        user_name = ''
        if user_res.data:
            user_name = user_res.data.get('name_cn') or user_res.data.get('name_en', '')
        # 随机抽取题目
        questions = random_pick_questions(exam_id, question_count)
        # 处理题目数据，确保 options 为字典，并筛选必要字段
        questions_light = []
        for q in questions:
            # 解析 options（可能为字符串 JSON）
            opts = q.get('options', {})
            if isinstance(opts, str):
                try:
                    opts = json.loads(opts)
                except:
                    opts = {}
            # 过滤空选项
            if opts:
                opts = {k: v for k, v in opts.items() if v and v.strip()}
            questions_light.append({
                'num': q.get('num'),
                'content': q.get('content_cn') or q.get('content') or q.get('content_raw', '无题目内容'),
                'type': q.get('type', 'single'),
                'options': opts   # 前端可能需要展示选项
            })
        preview.append({
            "user_id": uid,
            "user_name": user_name,
            "questions": questions_light
        })
    
    return jsonify({"exam_title": exam_title, "preview": preview})

@app.route('/api/my/interviews')
@login_required
def my_interviews():
    """① 后端新增接口：获取学员的活跃访谈"""
    user_id = session['user_id']
    db = get_supabase()
    res = db.table("interview_results").select("interview_id").eq("user_id", user_id).execute()
    interview_ids = list(set(r['interview_id'] for r in (res.data or [])))
    if not interview_ids:
        return jsonify([])
    inv_res = db.table("interviews").select("*").in_("id", interview_ids).execute()
    now = datetime.now(timezone.utc).isoformat()
    active = []
    for inv in (inv_res.data or []):
        start = inv.get('start_time')
        end = inv.get('end_time')
        if not start or not end or not (start < now < end):
            continue   # 不在有效期内的不显示
        # 检查是否全部作答
        answers = db.table("interview_results").select("answer").eq("interview_id", inv['id']).eq("user_id", user_id).execute()
        all_answered = all(row.get('answer') for row in (answers.data or []))
        inv['is_completed'] = all_answered
        active.append(inv)
    return jsonify(active)

@app.route('/interview/take/<int:interview_id>')
@login_required
def take_interview(interview_id):
    """① 后端新增接口：学员进入访谈，检查是否在名单中，获取访谈基本信息和题目"""
    user_id = session['user_id']
    db = get_supabase()
    # 检查用户是否属于该访谈
    result = db.table("interview_results").select("interview_id").eq("interview_id", interview_id).eq("user_id", user_id).limit(1).execute()
    if not result.data:
        flash("您不在本次访谈名单中", "danger")
        return redirect(url_for('dashboard'))

    inv = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv.data:
        flash("访谈不存在", "danger")
        return redirect(url_for('dashboard'))

    questions_res = db.table("interview_results").select("*, questions(*)").eq("interview_id", interview_id).eq("user_id", user_id).execute()
    questions = []
    for row in (questions_res.data or []):
        q = row.get('questions', {})
        # 解析 options
        opts = q.get('options', {})
        if isinstance(opts, str):
            try:
                q['options'] = json.loads(opts)
            except:
                q['options'] = {}
        # 判断题默认选项
        if q.get('type') == 'judge' and (not q['options']):
            q['options'] = {"A": "正确 True", "B": "错误 False"}

        # 确保选项是字典
        if not isinstance(q['options'], dict):
            q['options'] = {}

        q['interview_result_id'] = row['id']
        # 确保答案不是 None，避免模板错误
        q['user_answer'] = row.get('answer') or ''
        questions.append(q)

    # ✅ 按 num 排序
    questions.sort(key=lambda x: x.get('num', 0))
    for idx, q in enumerate(questions, 1):
        q['num'] = idx
        if q.get('options'):
            q['options'] = {k: v for k, v in q['options'].items() if v.strip()}
        else:
            q['options'] = {}

    return render_template('exam/take_interview.html', interview=inv.data, questions=questions)

@app.route('/api/interview/<int:interview_id>/submit', methods=['POST'])
@login_required
def submit_interview(interview_id):
    """② 后端新增接口：提交学员的答案"""
    user_id = session['user_id']
    answers = request.json.get('answers', {})  # {result_id: answer}
    db = get_supabase()
    for rid, ans in answers.items():
        # 获取关联题目 ID
        result = db.table("interview_results").select("question_id").eq("id", rid).eq("user_id", user_id).maybe_single().execute()
        if not result.data:
            continue
        qid = result.data['question_id']
        # 获取标准答案和题型
        q = db.table("questions").select("answer, type").eq("id", qid).maybe_single().execute()
        is_correct = False
        if q.data:
            correct_ans = q.data['answer'].upper()
            q_type = q.data['type']
            if q_type == 'multi':
                u_set = set(ans.upper().replace(' ', ''))
                c_set = set(correct_ans.replace(' ', ''))
                is_correct = (u_set == c_set)
            elif q_type == 'judge':
                norm = ans.upper()
                if norm in ('A', 'T', '√', '正确', '对'): norm = 'T'
                elif norm in ('B', 'F', '×', '错误', '错'): norm = 'F'
                correct_std = correct_ans.replace('√', 'T').replace('×', 'F')
                is_correct = (norm == correct_std)
            else:
                is_correct = (ans.strip().upper() == correct_ans)
        db.table("interview_results").update({
            "answer": ans,
            "is_correct": is_correct,
            "submitted_at": datetime.now(timezone.utc).isoformat()
        }).eq("id", rid).eq("user_id", user_id).execute()
    return jsonify({"success": True, "message": "您的回答已提交"})

@app.route('/api/interview/<int:interview_id>/submit', methods=['POST'])
@login_required
def submit_interview_answer(interview_id):
    """② 后端新增接口：提交学员的答案"""
    user_id = session['user_id']
    answers = request.json.get('answers', {})  # {question_id: answer}
    db = get_supabase()
    for qid, ans in answers.items():
        db.table("interview_results").update({"answer": ans, "submitted_at": datetime.now(timezone.utc).isoformat()}).eq("interview_id", interview_id).eq("user_id", user_id).eq("question_id", qid).execute()
    return jsonify({"success": True})

@app.route('/admin/interview/<int:interview_id>/details')
@login_required
@admin_required
def admin_interview_details_page(interview_id):
    """③ 后端新增接口：管理员查看访谈详情"""
    return render_template('admin/list_inspection_details.html', interview_id=interview_id)

@app.route('/api/admin/interview/<int:interview_id>/details')
@login_required
@admin_required
def api_interview_details(interview_id):
    """获取访谈详情数据（支持筛选和分页）"""
    db = get_supabase()
    search = request.args.get('search', '')
    country = request.args.get('country', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # 获取访谈信息
    inv_res = db.table("interviews").select("*").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    interview = inv_res.data

    # 获取考试标题
    exam_title = ''
    if interview.get('exam_id'):
        exam_res = db.table("exams").select("title").eq("id", interview['exam_id']).maybe_single().execute()
        if exam_res.data:
            exam_title = exam_res.data['title']

    # 查询该访谈的所有答题记录（按用户聚合）
    query = db.table("interview_results").select("*").eq("interview_id", interview_id)
    # 先获取所有记录，然后在 Python 中处理筛选和聚合
    all_res = query.execute()
    all_data = all_res.data or []

    # ===== 国家权限过滤 =====
    allowed = get_allowed_countries()
    if allowed is not None:
        if not allowed:
            # 没有允许的国家，直接返回空
            return jsonify({
                "interview": {"id": interview_id, "title": interview.get('title'), "exam_title": exam_title, "reviewer": interview.get('reviewer', '')},
                "data": [],
                "total": 0,
                "page": page,
                "per_page": per_page
            })
        # 收集所有 user_id
        all_user_ids = list(set(r['user_id'] for r in all_data))
        # 查询这些用户的国家
        users_country_res = db.table("users").select("id, country").in_("id", all_user_ids).execute()
        allowed_ids = {u['id'] for u in (users_country_res.data or []) if u.get('country') in allowed}
        # 过滤数据
        all_data = [r for r in all_data if r['user_id'] in allowed_ids]
    # ======================
    # 批量获取用户信息
    user_ids = list(set(r['user_id'] for r in all_data))
    users_map = {}
    if user_ids:
        users_res = db.table("users").select("id, name_cn, name_en, country, wh_id, department").in_("id", user_ids).execute()
        for u in (users_res.data or []):
            users_map[u['id']] = u

    # 按用户聚合
    user_results = {}
    for row in all_data:
        uid = row['user_id']
        user_info = users_map.get(uid, {})
        # 筛选：姓名搜索
        if search:
            name = user_info.get('name_cn') or user_info.get('name_en', '')
            if search.lower() not in name.lower():
                continue
        # 筛选：国家
        if country and user_info.get('country') != country:
            continue

        if uid not in user_results:
            user_results[uid] = {
                "user_id": uid,
                "name": user_info.get('name_cn') or user_info.get('name_en', ''),
                "country": user_info.get('country', ''),
                "wh_id": user_info.get('wh_id', ''),
                "department": user_info.get('department', ''),
                "submitted_at": None,
                "total_questions": 0,
                "correct_count": 0,
                "feedback": "",   # 可暂合并所有 feedback
                "reviewer": interview.get('reviewer', ''),
                "results": []
            }
        user_results[uid]["total_questions"] += 1
        if row.get('is_correct'):
            user_results[uid]["correct_count"] += 1
        if row.get('submitted_at') and (not user_results[uid]["submitted_at"] or row['submitted_at'] > user_results[uid]["submitted_at"]):
            user_results[uid]["submitted_at"] = row['submitted_at']

    # 转换为列表
    detail_list = []
    for uid, data in user_results.items():
        detail_list.append({
            "user_id": data["user_id"],
            "name": data["name"],
            "country": data["country"],
            "wh_id": data["wh_id"],
            "department": data["department"],
            "submitted_at": data["submitted_at"],
            "reviewer": data["reviewer"],
            "total_questions": data["total_questions"],
            "correct_count": data["correct_count"],
            "feedback": data["feedback"],
            "has_submitted": bool(data["submitted_at"])
        })

    # 分页
    total = len(detail_list)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = detail_list[start:end]

    return jsonify({
        "interview": {
            "id": interview_id,
            "title": interview.get('title'),
            "exam_title": exam_title,
            "reviewer": interview.get('reviewer', '')
        },
        "data": paginated,
        "total": total,
        "page": page,
        "per_page": per_page
    })

@app.route('/api/admin/interview/<int:interview_id>/user/<user_id>/answers')
@login_required
@admin_required
def get_interview_user_answers(interview_id, user_id):
    """获取指定访谈中指定用户的答题详情"""
    db = get_supabase()
    # 获取访谈信息
    inv_res = db.table("interviews").select("id, title").eq("id", interview_id).maybe_single().execute()
    if not inv_res.data:
        return jsonify({"error": "访谈不存在"}), 404
    # 获取该用户的答题记录，关联题目
    results = db.table("interview_results").select("*, questions(*)") \
        .eq("interview_id", interview_id).eq("user_id", user_id).execute()
    data = []
    for row in (results.data or []):
        q = row.get('questions', {})
        # 解析 options
        opts = q.get('options', {})
        if isinstance(opts, str):
            try:
                q['options'] = json.loads(opts)
            except:
                q['options'] = {}
        if q.get('type') == 'judge' and not q['options']:
            q['options'] = {"A": "正确 True", "B": "错误 False"}
        data.append({
            "question_num": q.get('num'),
            "question_content": q.get('content_cn') or q.get('content') or q.get('content_raw', ''),
            "question_type": q.get('type'),
            "options": q.get('options', {}),
            "user_answer": row.get('answer') or '',
            "is_correct": row.get('is_correct'),
            "result_id": row['id']
        })
    # 按题目编号排序
    data.sort(key=lambda x: x.get('question_num', 0))
    return jsonify({"answers": data})

#----------1. 用户管理 API---------
# ================= 用户管理 API =================
def can_modify_user(target_user, current_user, action='edit'):
    """检查当前用户是否有权限修改目标用户
    target_user: 目标用户的字典或ID（需包含 is_protected, id）
    current_user: 当前登录用户信息（如从 session 中获取）
    """
    dev_id = os.environ.get('DEVELOPER_USER_ID')
    print(f"DEBUG: DEVELOPER_USER_ID = {dev_id}")
    # 开发者凭据：可以从环境变量读取一个特定的用户ID或邮箱
    DEVELOPER_USER_ID = os.environ.get('DEVELOPER_USER_ID')
    if DEVELOPER_USER_ID and current_user['id'] == DEVELOPER_USER_ID:
        return True   # 开发者本人可以操作任何账号
    # 如果目标用户是受保护的，则禁止任何其他超管操作
    if target_user.get('is_protected'):
        return False
    return True

def get_current_user():
    """从 session 获取当前用户信息"""
    return {
        'id': session.get('user_id'),
        'role': session.get('role'),
        'is_protected': False   # 当前用户不可能是保护账号（因为保护账号只能在数据库设）
    }

@app.route('/admin/users')
@login_required
@admin_required
def admin_user_list():
    """用户管理页面,确保将开发者信息传递给模板"""
    return render_template(
        'admin/list_users.html',
        is_developer=is_developer()  # 传递开发者标志
    )

@app.route('/api/admin/users/<user_id>')
@login_required
@admin_required
def api_admin_user_detail(user_id):
    """后端单用户查询接口"""
    db = get_supabase()
    res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
    if not res.data:
        return jsonify({"error": "用户不存在"}), 404
    return jsonify(res.data)

@app.route('/api/admin/users')
@login_required
@admin_required
def api_admin_users():
    """获取用户列表（带完整权限控制）"""
    logger.info("=" * 50)
    logger.info("调用 api_admin_users")
    logger.info(f"当前用户 role={session.get('role')}, user_country={session.get('user_country')}")

    db = get_supabase()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    country = request.args.get('country', '')
    user_status = request.args.get('status', '')
    
    # 批量查询支持
    if 'ids' in request.args:
        ids = request.args.get('ids').split(',')
        if ids:
            query = db.table("users").select("*").in_("id", ids)
            # 应用权限过滤
            query = apply_country_filter(query, 'country')
            res = query.execute()
            # 过滤掉无权限查看的用户
            users = [u for u in (res.data or []) if can_view_user(u)]
            return jsonify({"data": users, "total": len(users)})
    
    # ✅ 调试：检查权限范围
    allowed_countries = get_admin_allowed_countries()
    logger.info(f"allowed_countries = {allowed_countries}")

    # 基础查询
    query = db.table("users").select("*", count="exact").is_("deleted_at", "null")
    
    # 非开发者不能看到受保护账号
    if not is_developer():
        query = query.eq("is_protected", False)
    
    # 应用国家权限过滤
    query = apply_country_filter(query, 'country')
    
    # ✅ 调试：打印最终查询条件（Supabase 不直接支持，但可以记录）
    logger.info(f"国家过滤已应用: allowed={allowed_countries}")

    # 角色过滤：非超管不能看到超管和开发者
    current_role = session.get('role')
    if current_role != 'super_admin' and not is_developer():
        query = query.neq("role", "super_admin").neq("role", "developer")
    
    if user_status:
        query = query.eq("user_status", user_status)
    
    if country and country != '___NONE___':
        if ',' in country:
            codes = [c.strip() for c in country.split(',') if c.strip()]
            query = query.in_("country", codes)
        else:
            query = query.eq("country", country)
    
    # 执行查询
    full_res = query.order("created_at", desc=True).execute()
    all_users = full_res.data or []
    
    # 内存搜索和二次权限过滤
    if search:
        search_lower = search.lower()
        all_users = [u for u in all_users if 
                     (u.get('email') and search_lower in u['email'].lower()) or
                     (u.get('name_cn') and search_lower in u['name_cn'].lower()) or
                     (u.get('name_en') and search_lower in u['name_en'].lower())]
    
    # 再次过滤可查看的用户
    all_users = [u for u in all_users if can_view_user(u)]
    
    total = len(all_users)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = all_users[start:end]
    
    return jsonify({
        "data": paginated,
        "total": total,
        "page": page,
        "per_page": per_page
    })

@app.route('/api/admin/users', methods=['POST'])
@login_required
@admin_required
def api_admin_add_user():
    """管理员添加用户（带角色权限控制）"""
    data = request.json
    email = data.get('email', '').strip().lower() or None
    name_en = data.get('name_en', '').strip()
    role = data.get('role', 'user')
    admin_countries = data.get('admin_countries', '[]')
    user_status = data.get('user_status', 'imported')
    
    db = get_supabase()
    
    # ========== 角色权限校验 ==========
    current_role = session.get('role')
    
    # 开发者可以创建任何角色
    if not is_developer():
        # 非开发者不能创建开发者角色
        if role == 'developer':
            return jsonify({"success": False, "message": "cannot_create_developer", "params": []}), 403
        
        # 超管可以创建超管、管理员、用户
        if current_role == 'super_admin':
            if role not in ['super_admin', 'admin', 'user']:
                return jsonify({"success": False, "message": "invalid_role", "params": []}), 400
        # 管理员只能创建用户
        elif current_role == 'admin':
            if role != 'user':
                return jsonify({"success": False, "message": "admin_can_only_create_user", "params": []}), 403
        else:
            return jsonify({"success": False, "message": "permission_denied", "params": []}), 403
    
    # ========== 超管/管理员权限范围校验 ==========
    if role in ['super_admin', 'admin']:
        # 解析权限范围
        if isinstance(admin_countries, str):
            try:
                countries_list = json.loads(admin_countries)
            except:
                countries_list = parse_countries_input(admin_countries)
        else:
            countries_list = admin_countries
        
        # 权限范围不能为空
        if not countries_list or len(countries_list) == 0:
            return jsonify({
                "success": False, 
                "message": "admin_countries_required",
                "params": [role]
            }), 400
        
        # 存储为 JSON 字符串
        admin_countries_json = json.dumps(countries_list)
    else:
        admin_countries_json = None
    
    # ========== 姓名必填校验 ==========
    if not name_en:
        return jsonify({"success": False, "message": "name_cannot_empty", "params": []}), 400
    
    # ========== 邮箱条件校验 ==========
    if not email and user_status != 'imported':
        return jsonify({"success": False, "message": "mail_cannot_empty", "params": []}), 400
    
    # ========== 检查用户是否已存在 ==========
    existing_users = db.table("users").select("*").eq("name_en", name_en).execute()
    existing_users_list = existing_users.data or []
    
    # 分离活跃用户和已删除用户
    active_users = [u for u in existing_users_list if u.get('deleted_at') is None]
    deleted_users = [u for u in existing_users_list if u.get('deleted_at') is not None]
    
    # 检查活跃用户重复
    birthday = data.get('birthday', '') or None
    employee_id = data.get('employee_id', '').strip() or None
    
    for active in active_users:
        active_birthday = active.get('birthday')
        active_employee_id = active.get('employee_id')
        
        if (birthday and active_birthday == birthday) or (employee_id and active_employee_id == employee_id):
            return jsonify({"success": False, "message": "duplicate_user_found", "params": []}), 400
        if not birthday and not employee_id:
            return jsonify({"success": False, "message": "duplicate_user_found", "params": []}), 400
    
    # 处理已删除用户的恢复
    if deleted_users:
        existing_deleted = deleted_users[0]
        update_data = {
            "email": email,
            "user_status": user_status,
            "is_active": False if user_status == 'imported' else True,
            "deleted_at": None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "role": role,
            "admin_countries": admin_countries_json,
            "created_by": session['user_id']
        }
        
        # 更新其他字段（如果提供）
        if birthday:
            update_data["birthday"] = birthday
        if employee_id:
            update_data["employee_id"] = employee_id
        if data.get('company'):
            update_data["company"] = data.get('company')
        if data.get('department'):
            update_data["department"] = data.get('department')
        if data.get('country'):
            update_data["country"] = data.get('country')
        if data.get('phone'):
            update_data["phone"] = data.get('phone')
        if data.get('role'):
            update_data["role"] = data.get('role')
        if data.get('wh_type'):
            update_data["wh_type"] = data.get('wh_type')
        if data.get('wh_id'):
            update_data["wh_id"] = data.get('wh_id')
        if data.get('wh_name_en'):
            update_data["wh_name_en"] = data.get('wh_name_en')
        if data.get('is_partner'):
            update_data["is_partner"] = data.get('is_partner') == 'Y'
        
        try:
            db.table("users").update(update_data).eq("id", existing_deleted['id']).execute()
            logger.info(f"恢复已删除用户: {name_en}, ID: {existing_deleted['id']}")
            
            # 如果需要发送邮件通知（仅当有邮箱时）
            if email:
                import secrets, string
                temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
                password_hash = auth.hash_password(temp_password)
                db.table("users").update({"password_hash": password_hash}).eq("id", existing_deleted['id']).execute()

                # 邮件主题 / Email Subject
                subject = "您的考试系统账号已创建Your exam system account has been created"
                
                # 邮件正文 / Email Body
                body = f"""
                # ---------------------------- 中文版 ----------------------------
                尊敬的 {name_en or email}：
                
                您的在线考试系统账号已由管理员创建。
                
                登录邮箱：{email}
                临时密码：{temp_password}
                
                请尽快登录系统并修改密码。
                
                登录地址：{request.host_url}
                
                # ---------------------------- English Version ----------------------------
                Dear {name_en or email},
                
                Your online exam system account has been created by the administrator.
                
                Login Email: {email}
                Temporary Password: {temp_password}
                
                Please log in and change your password as soon as possible.
                
                Login URL: {request.host_url}
                """
                try:
                    auth.send_email(email, subject, body)
                except Exception as e:
                    logger.warning(f"发送邮件失败: {e}")
            
            return jsonify({"success": True, "user_id": existing_deleted['id'], "restored": True})
        except Exception as e:
            logger.error(f"恢复用户失败: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    # 邮箱唯一性检查
    if email:
        exist = db.table("users").select("id").eq("email", email).is_("deleted_at", "null").execute()
        if exist.data:
            return jsonify({"success": False, "message": "email_already_registered", "params": []}), 400

    # 生成用户ID和密码
    user_id = str(uuid.uuid4())
    temp_password = ''
    password_hash = ''
    if email:
        import secrets, string
        alphabet = string.ascii_letters + string.digits
        temp_password = ''.join(secrets.choice(alphabet) for _ in range(10))
        password_hash = auth.hash_password(temp_password)

    # 准备插入数据
    insert_data = {
        "id": user_id,
        "email": email,
        "password_hash": password_hash,
        "name_en": name_en,
        "company": data.get('company', ''),
        "department": data.get('department', ''),
        "employee_id": employee_id,
        "birthday": birthday,
        "country": data.get('country', ''),
        "phone": data.get('phone', ''),
        "role": role,
        "admin_countries": admin_countries_json,
        "user_status": user_status,
        "is_partner": data.get('is_partner', 'N') == 'Y',
        "wh_type": data.get('wh_type', ''),
        "wh_id": data.get('wh_id', ''),
        "wh_name_en": data.get('wh_name_en', ''),
        "is_active": False if user_status == 'imported' else True,
        "created_by": session['user_id'],
        "is_protected": (role == 'developer')  # 开发者账号自动保护
    }
    try:
        db.table("users").insert(insert_data).execute()
        # 发送邮件通知（仅当邮箱存在）
        if email:
            try:
                # 邮件主题 / Email Subject
                subject = "您的考试系统账号已创建Your exam system account has been created"
                
                # 邮件正文 / Email Body
                body = f"""
                # ---------------------------- 中文版 ----------------------------
                尊敬的 {name_en or email}：
                
                您的在线考试系统账号已由管理员创建。
                
                登录邮箱：{email}
                临时密码：{temp_password}
                
                请尽快登录系统并修改密码。
                
                登录地址：{request.host_url}
                
                # ---------------------------- English Version ----------------------------
                Dear {name_en or email},
                
                Your online exam system account has been created by the administrator.
                
                Login Email: {email}
                Temporary Password: {temp_password}
                
                Please log in and change your password as soon as possible.
                
                Login URL: {request.host_url}
                """
                
                auth.send_email(email, subject, body)
            except Exception as e:
                logger.warning(f"发送邮件失败: {e}")
        
        logger.info(f"管理员添加用户: {email or '无邮箱'}, 状态: {user_status}")
        return jsonify({"success": True, "user_id": user_id, "temp_password": temp_password})
    except Exception as e:
        logger.error(f"添加用户失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/admin/users/<user_id>', methods=['PUT'])
@login_required
@admin_required
def api_admin_edit_user(user_id):
    """编辑用户信息（带完整权限控制）"""
    data = request.json
    db = get_supabase()
    
    # 获取目标用户信息
    target_user_res = db.table("users").select("*").eq("id", user_id).maybe_single().execute()
    if not target_user_res.data:
        return jsonify({"success": False, "message": "user_not_found", "params": []}), 404
    target_user = target_user_res.data
    
    # ========== 权限检查 ==========
    # 开发者可以编辑任何用户
    if not is_developer():
        # 受保护账号只能本人编辑
        if target_user.get('is_protected') and user_id != session['user_id']:
            return jsonify({"success": False, "message": "protected_account", "params": []}), 403
        
        current_role = session.get('role')
        target_role = target_user.get('role', 'user')
        
        # 超管不能编辑开发者
        if target_role == 'developer':
            return jsonify({"success": False, "message": "cannot_edit_developer", "params": []}), 403
        
        # 管理员不能编辑超管
        if current_role == 'admin' and target_role in ['super_admin', 'admin']:
            return jsonify({"success": False, "message": "cannot_edit_admin", "params": []}), 403
        
        # 管理员不能修改角色字段
        if current_role == 'admin' and 'role' in data and data['role'] != target_role:
            return jsonify({"success": False, "message": "cannot_change_role", "params": []}), 403
    
    # ========== 构建更新数据 ==========
    update_data = {}
    
    # 基础字段（所有角色可修改自己）
    allowed_fields = ['name_en', 'company', 'department', 'employee_id', 'birthday', 
                      'country', 'phone', 'wh_type', 'wh_id', 'wh_name_en', 'user_status', 'is_partner']
    
    # 角色和权限范围字段（需要更高权限）
    if 'role' in data:
        new_role = data['role']
        # 开发者可以修改任何角色
        if is_developer():
            update_data['role'] = new_role
            # 如果改为开发者，自动设置保护
            if new_role == 'developer':
                update_data['is_protected'] = True
        # 超管可以修改角色（但不能改为开发者）
        elif session.get('role') == 'super_admin':
            if new_role != 'developer':
                update_data['role'] = new_role
            else:
                return jsonify({"success": False, "message": "cannot_set_developer", "params": []}), 403
    
    # 权限范围字段
    if 'admin_countries' in data:
        if is_developer() or session.get('role') == 'super_admin':
            countries_input = data['admin_countries']
            if isinstance(countries_input, str):
                try:
                    countries_list = json.loads(countries_input)
                except:
                    countries_list = parse_countries_input(countries_input)
            else:
                countries_list = countries_input
            
            if countries_list:
                update_data['admin_countries'] = json.dumps(countries_list)
            else:
                update_data['admin_countries'] = None
    
    # 邮箱字段（需要唯一性检查）
    if 'email' in data:
        email_val = data['email'].strip().lower() or None
        if email_val:
            conflict = db.table("users").select("id").eq("email", email_val).neq("id", user_id).execute()
            if conflict.data:
                return jsonify({"success": False, "message": "email_already_used", "params": []}), 400
        update_data['email'] = email_val
    
    # 普通字段
    for field in allowed_fields:
        if field in data:
            val = data[field]
            if field == 'birthday':
                val = val if val else None
            elif field == 'is_partner':
                val = True if val in ('Y', 'true', True, '1') else False
            update_data[field] = val
    
    if not update_data:
        return jsonify({"success": False, "message": "no_fields_to_update", "params": []}), 400
    
    try:
        db.table("users").update(update_data).eq("id", user_id).execute()
        
        # 如果更新的是当前用户，同步 session
        if user_id == session.get('user_id'):
            if 'role' in update_data:
                session['role'] = update_data['role']
            if 'admin_countries' in update_data:
                session['admin_countries'] = update_data['admin_countries']
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/admin/users/<user_id>', methods=['DELETE'])
@login_required
@admin_required
def api_admin_delete_user(user_id):
    """删除用户：已导入 → 硬删除；已注册 → 软删除"""
    if user_id == session['user_id']:
        return jsonify({"success": False, "message": "不能删除自己的账号"}), 400
    db = get_supabase()

    # 获取目标用户信息（包括 is_protected）
    target_res = db.table("users").select("role, user_status, is_protected").eq("id", user_id).maybe_single().execute()
    if not target_res.data:
        return jsonify({"success": False, "message": "用户不存在"}), 404
    target = target_res.data

    if target.get('is_protected') and user_id != session['user_id']:
        return jsonify({"success": False, "message": "该账号被保护，无法删除"}), 403

    # 权限检查（开发者保护）
    current_user = get_current_user()
    if not can_modify_user(target, current_user, 'delete'):
        return jsonify({"success": False, "message": "该账号被保护，无法删除"}), 403

    target_role = target.get('role', 'user')
    current_role = session.get('role')

    # 权限控制：普通管理员不能删除管理员或超管
    if current_role != 'super_admin' and target_role in ('admin', 'super_admin'):
        return jsonify({"success": False, "message": "权限不足，无法删除管理员"}), 403

    user_status = target.get('user_status', 'registered')

    # ========= 已导入用户：硬删除 =========
    if user_status == 'imported':
        try:
            # 清理可能存在的考试分配记录（防止外键冲突）
            db.table("exam_assignments").delete().eq("user_id", user_id).execute()
            # 物理删除用户
            db.table("users").delete().eq("id", user_id).execute()
            logger.info(f"硬删除已导入用户 {user_id}")
            return jsonify({"success": True, "hard_delete": True})
        except Exception as e:
            # 如果硬删除失败（例如仍有其他关联记录），回退为软删除
            logger.error(f"硬删除失败，尝试软删除: {e}")
            try:
                db.table("users").update({"deleted_at": datetime.now(timezone.utc).isoformat()}).eq("id", user_id).execute()
                return jsonify({"success": True, "hard_delete": False, "fallback": True, "message": "用户存在关联记录，已转为软删除"})
            except Exception as soft_err:
                logger.error(f"软删除也失败: {soft_err}")
                return jsonify({"success": False, "message": str(soft_err)}), 500

    # ========= 已注册用户：软删除 =========
    try:
        db.table("users").update({"deleted_at": datetime.now(timezone.utc).isoformat()}).eq("id", user_id).execute()
        logger.info(f"软删除用户 {user_id}")
        return jsonify({"success": True, "hard_delete": False})
    except Exception as e:
        logger.error(f"软删除失败: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/admin/users/<user_id>/reset_password', methods=['POST'])
@login_required
@admin_required
def api_admin_reset_user_password(user_id):
    """重置用户密码（生成新密码并发送邮件）"""
    import secrets, string
    db = get_supabase()
    # 获取用户信息（包括 is_protected）
    user_res = db.table("users").select("email, name_en, is_protected").eq("id", user_id).maybe_single().execute()
    if not user_res.data:
        return jsonify({"success": False, "message": "用户不存在"}), 404
    target_user = user_res.data
    if target_user.get('is_protected') and user_id != session['user_id']:
        return jsonify({"success": False, "message": "该账号被保护，无法重置密码"}), 403

    # 权限检查（开发者保护）
    current_user = get_current_user()
    if not can_modify_user(target_user, current_user, 'reset_password'):
        return jsonify({"success": False, "message": "该账号被保护，无法重置密码"}), 403

    email = target_user.get('email')
    if not email:
        return jsonify({"success": False, "message": "该用户没有邮箱，无法重置密码"}), 400

    new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
    password_hash = auth.hash_password(new_password)
    
    try:
        db.table("users").update({"password_hash": password_hash}).eq("id", user_id).execute()
        # 发送邮件
        subject = "您的考试系统密码已重置"
        body = f"""
        尊敬的 {user.get('name_en') or user.get('email')}：
        
        您的考试系统密码已被管理员重置。
        
        新密码：{new_password}
        
        请登录后尽快修改密码。
        
        登录地址：{request.host_url}
        """
        auth.send_email(user['email'], subject, body)
        return jsonify({"success": True, "new_password": new_password})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

#----------2. 考试统计与清单 API---------
@app.route('/admin/exams')
@login_required
@admin_required
def admin_exams_page():
    """考试清单页面（一级菜单）"""
    return render_template('admin/list_exams.html')

# 2.1 考试统计（用于卡片显示）
@app.route('/api/admin/exams/stats')
@login_required
@admin_required
def api_admin_exams_stats():
    db = get_supabase()
    status_counts = {status: 0 for status in ['draft', 'active', 'closed']}
    for status in status_counts.keys():
        res = db.table("exams").select("id", count="exact").eq("status", status).execute()
        status_counts[status] = res.count or 0
    return jsonify(status_counts)

@app.route('/api/admin/exam/<int:exam_id>')
@login_required
@admin_required
def api_admin_exam_detail(exam_id):
    """获取单个考试信息接口（用于模态框回显）"""
    db = get_supabase()
    exam = db.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
    if not exam.data:
        return jsonify({"error": "考试不存在"}), 404
    return jsonify({
        "start_time": exam.data.get('start_time'),
        "end_time": exam.data.get('end_time'),
        "duration": exam.data.get('duration', 60)
    })

@app.route('/api/admin/exam/<int:exam_id>/assignments')
@login_required
@admin_required
def api_admin_exam_assignments(exam_id):
    """获取考试已分配考生列表"""
    db = get_supabase()
    res = db.table("exam_assignments").select("user_id").eq("exam_id", exam_id).execute()
    user_ids = [row['user_id'] for row in (res.data or [])]
    return jsonify({"user_ids": user_ids})

@app.route('/api/admin/exam/<int:exam_id>/update', methods=['PUT'])
@login_required
@admin_required
def api_admin_exam_update(exam_id):
    data = request.json
    update_data = {}
    if 'start_time' in data:
        update_data['start_time'] = data['start_time'] if data['start_time'] else None
    if 'end_time' in data:
        update_data['end_time'] = data['end_time'] if data['end_time'] else None
    if 'duration' in data:
        update_data['duration'] = data['duration']

    # 根据是否有完整有效期，同步 is_active 和 status
    if update_data.get('start_time') and update_data.get('end_time'):
        update_data['is_active'] = True
        update_data['status'] = 'active'
    elif 'start_time' in update_data or 'end_time' in update_data:
        # 只清空了某一端，视为无效，设为草稿
        update_data['is_active'] = False
        update_data['status'] = 'draft'

    if not update_data:
        return jsonify({"success": False, "message": "无更新内容"}), 400
    db = get_supabase()
    db.table("exams").update(update_data).eq("id", exam_id).execute()
    return jsonify({"success": True})

@app.route('/api/admin/exam/<int:exam_id>/push_with_settings', methods=['POST'])
@admin_required
def admin_push_exam_with_settings(exam_id):
    data = request.json
    logger.info(f"接收到的推送数据: {data}")  # ✅ 添加这行，查看前端是否传递了 reviewer

    start_time_local = data.get('start_time')
    end_time_local = data.get('end_time')
    duration = data.get('duration')
    user_ids = data.get('user_ids', [])
    reviewer = data.get('reviewer', '')  # ✅ 新增：获取阅卷人

    logger.info(f"获取到的 reviewer 值: {reviewer}")  # ✅ 添加这行

    db = get_supabase()

    # 获取考试信息（用于国家和标题）
    exam_info = db.table("exams").select("country, title").eq("id", exam_id).maybe_single().execute()
    if not exam_info.data:
        return jsonify({"success": False, "message": "考试不存在"}), 404
    exam_data = exam_info.data

    # 更新考试有效期和时长
    update_data = {}
    if start_time_local is not None:
        update_data['start_time'] = start_time_local
        logger.info(f"本地开始时间: {start_time_local}, UTC: {update_data['start_time']}")
    if end_time_local is not None:
        update_data['end_time'] = end_time_local
    if duration is not None:
        update_data['duration'] = duration
    if start_time_local and end_time_local:
        update_data['status'] = 'active'
        update_data['is_active'] = True   # ✅ 关键：确保考试对学员可见
    else:
        # 如果没有有效期，视为草稿，关闭激活状态
        update_data['is_active'] = False
        update_data['status'] = 'draft'

    # ✅ 新增：更新阅卷人（如果有值）
    if reviewer and reviewer.strip():
        update_data['reviewer'] = reviewer
        logger.info(f"使用前端传递的阅卷人: {reviewer}")
    elif not exam_data.get('reviewer'):
        # 如果没有指定阅卷人，尝试根据国家自动获取默认阅卷人
        default_reviewer = get_default_reviewer_by_country(
            user_country=exam_data.get('country'),
            exam_reviewer=None,
            url_reviewer=None
            )
        if default_reviewer and default_reviewer != "Administrator":
            update_data['reviewer'] = default_reviewer
            logger.info(f"使用默认阅卷人: {default_reviewer}")
        else:
            # 最后的保底
            update_data['reviewer'] = "Administrator"
            logger.info(f"使用保底阅卷人: Administrator")
    else:
        # 保留考试表原有的 reviewer
        logger.info(f"保留原有阅卷人: {exam_data.get('reviewer')}")

    if update_data:
        db.table("exams").update(update_data).eq("id", exam_id).execute()
        logger.info(f"更新考试 {exam_id} 的数据: {update_data}")

    # 可选：按国家过滤考生（如果不需要，可以注释掉整个块）
    def get_user_country(uid):
        user_res = db.table("users").select("country").eq("id", uid).maybe_single().execute()
        return user_res.data.get('country') if user_res.data else None

    country = exam_data.get('country')
    if country and user_ids:
        filtered_ids = [uid for uid in user_ids if get_user_country(uid) == country]
        if not filtered_ids:
            return jsonify({"success": False, "message": "没有符合国家条件的考生"}), 400
        user_ids = filtered_ids

    # 更新考生分配
    # 不再删除全部，而是获取现有分配，然后合并
    if user_ids:
        # 获取现有分配
        existing_res = db.table("exam_assignments").select("user_id").eq("exam_id", exam_id).execute()
        existing_ids = {row['user_id'] for row in (existing_res.data or [])}
        # 需要新增的用户
        to_add = [uid for uid in user_ids if uid not in existing_ids]
        # 需要删除的用户（如果前端传递了完整列表，则删除不在新列表中的用户）
        to_remove = [uid for uid in existing_ids if uid not in user_ids]
        if to_remove:
            db.table("exam_assignments").delete().eq("exam_id", exam_id).in_("user_id", to_remove).execute()
        for uid in to_add:
            db.table("exam_assignments").insert({"exam_id": exam_id, "user_id": uid}).execute()

        # 发送邮件通知
        exam_title = exam_data.get('title', '考试')
        for uid in user_ids:
            user_res = db.table("users").select("email, name_en").eq("id", uid).execute()
            if user_res.data:
                email = user_res.data[0]['email']
                name = user_res.data[0].get('name_en', '用户')
                subject = f"考试通知：{exam_title}"
                body = f"您好 {name}，考试《{exam_title}》已设置，有效期从 {start_time_local} 到 {end_time_local}，时长为 {duration} 分钟。请登录系统参加。"
                try:
                    auth.send_email(email, subject, body)
                except Exception as e:
                    logger.warning(f"发送邮件失败: {e}")

    return jsonify({"success": True})

# 2.2 考试列表（一级菜单，支持过滤）

@app.route('/api/admin/exams/list')
@login_required
@admin_required
def api_admin_exams_list():
    db = get_supabase()
    include_deleted = request.args.get('include_deleted', 'false').lower() == 'true'
    
    # 获取过滤参数
    country_input = request.args.get('country', '')
    name = request.args.get('name', '')
    target_status = request.args.get('status', '')
    quarter = request.args.get('quarter', '')
    creator = request.args.get('creator', '')
    reviewer = request.args.get('reviewer', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # 管理员权限范围
    allowed = get_allowed_countries()
    # 用户选择的国家筛选（用于前端过滤）
    filter_country_code = match_country_code(country_input) if country_input else None
    q_start, q_end = quarter_to_date_range(quarter) if quarter else (None, None)

    # 1. 基础查询（不应用国家过滤，后面内存处理）
    query = db.table("exams").select("*", count="exact")
    query = apply_country_filter(query, 'country')  # ✅ 添加过滤
    if not include_deleted:
        query = query.is_("deleted_at", "null")
    if name:
        query = query.ilike("title", f"%{name}%")
    if creator:
        users_res = db.table("users").select("id").ilike("name_en", f"%{creator}%").execute()
        creator_ids = [u['id'] for u in (users_res.data or [])]
        if creator_ids:
            query = query.in_("created_by", creator_ids)
        else:
            return jsonify({"data": [], "total": 0})
    if reviewer:
        query = query.ilike("reviewer", f"%{reviewer}%")

    res = query.execute()
    all_exams = res.data or []

    # 2. 管理员权限过滤（如果 allowed 不为 None）
    if allowed is not None:
        if not allowed:
            return jsonify({"data": [], "total": 0, "page": page, "per_page": per_page})
        
        # 获取允许国家的所有用户ID
        users_in_allowed = db.table("users").select("id").in_("country", allowed).execute()
        allowed_user_ids = [u['id'] for u in (users_in_allowed.data or [])] if users_in_allowed.data else []
        
        # 查询分配了至少一个允许国家考生的考试ID
        allowed_exam_ids = set()
        if allowed_user_ids:
            assign_res = db.table("exam_assignments").select("exam_id").in_("user_id", allowed_user_ids).execute()
            allowed_exam_ids = {a['exam_id'] for a in (assign_res.data or [])}
        
        # 过滤考试：考试自身 country 在 allowed 中 或 考试ID 在 allowed_exam_ids 中
        filtered = []
        for exam in all_exams:
            if exam.get('country') in allowed or exam['id'] in allowed_exam_ids:
                filtered.append(exam)
        all_exams = filtered

    # 3. 前端选择的额外国家筛选（如果有）
    if filter_country_code:
        # 同样基于考试自身 country 或考生国家
        if allowed is not None:
            # 已经限制了 allowed，现在进一步限制
            pass
        # 简单做法：重新过滤 all_exams（因为数据量不大）
        # 获取该国家的用户ID
        users_in_filter = db.table("users").select("id").eq("country", filter_country_code).execute()
        filter_user_ids = [u['id'] for u in (users_in_filter.data or [])] if users_in_filter.data else []
        filter_exam_ids = set()
        if filter_user_ids:
            assign_filter = db.table("exam_assignments").select("exam_id").in_("user_id", filter_user_ids).execute()
            filter_exam_ids = {a['exam_id'] for a in (assign_filter.data or [])}
        all_exams = [exam for exam in all_exams 
                    if exam.get('country') == filter_country_code or exam['id'] in filter_exam_ids]

    # 4. 季度过滤（内存）
    if quarter and q_start and q_end:
        q_start_dt = datetime.fromisoformat(q_start)
        q_end_dt = datetime.fromisoformat(q_end)
        filtered = []
        for exam in all_exams:
            start, end = exam.get('start_time'), exam.get('end_time')
            if start and end:
                try:
                    start_dt = datetime.fromisoformat(start)
                    end_dt = datetime.fromisoformat(end)
                    if start_dt <= q_end_dt and end_dt >= q_start_dt:
                        filtered.append(exam)
                except:
                    pass
        all_exams = filtered

    # 5. 构建返回数据（动态状态、题目数、人数等，与原来相同）
    now = datetime.now(timezone.utc)
    exams_with_status = []
    for exam in all_exams:
        exam_id = exam['id']
        dynamic_status = get_exam_status(exam)
        if target_status and dynamic_status != target_status:
            continue
        # 其他字段计算（题目数、人数、最高分等）
        q_count = db.table("questions").select("id", count="exact").eq("exam_id", exam_id).execute().count or 0
        assigned_count = db.table("exam_assignments").select("user_id", count="exact").eq("exam_id", exam_id).execute().count or 0
        submitted_res = db.table("exam_results").select("user_id", count="exact").eq("exam_id", exam_id).execute()
        submitted_count = submitted_res.count or 0
        max_res = db.table("exam_results").select("total_score").eq("exam_id", exam_id).order("total_score", desc=True).limit(1).execute()
        max_score = max_res.data[0]['total_score'] if max_res.data else None
        min_res = db.table("exam_results").select("total_score").eq("exam_id", exam_id).order("total_score", desc=False).limit(1).execute()
        min_score = min_res.data[0]['total_score'] if min_res.data else None
        creator_name = ''
        if exam.get('created_by'):
            creator_res = db.table("users").select("name_en").eq("id", exam['created_by']).maybe_single().execute()
            creator_name = creator_res.data.get('name_en', '') if creator_res.data else ''
        exams_with_status.append({
            "id": exam_id, "title": exam['title'], "status": dynamic_status,
            "start_time": exam.get('start_time'), "end_time": exam.get('end_time'),
            "duration": exam.get('duration', 60), "question_count": q_count,
            "assigned_count": assigned_count, "submitted_count": submitted_count,
            "max_score": max_score, "min_score": min_score, "retake_count": 0,
            "reviewer": exam.get('reviewer', ''), "created_by_name": creator_name,
            "quarter": exam.get('quarter', ''), "deleted_at": exam.get('deleted_at'), 
            "country": exam.get('country', '')
        })

    total = len(exams_with_status)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = exams_with_status[start:end]
    return jsonify({"data": paginated, "total": total, "page": page, "per_page": per_page})

@app.route('/api/common/quarters')
def api_quarters():
    """根据数据库中考试的有效期动态生成季度选项列表"""
    db = get_supabase()
    # 查询所有考试的 start_time 和 end_time（只取非空）
    res = db.table("exams").select("start_time, end_time").execute()
    quarters = set()
    from datetime import datetime
    for exam in res.data or []:
        for field in ['start_time', 'end_time']:
            time_str = exam.get(field)
            if not time_str:
                continue
            try:
                # 解析 ISO 8601 字符串（可能带时区，如 2026-04-28T15:00:00+00:00）
                # 使用 dateutil.parser 或简单切片
                # 这里我们直接取年份和月份
                if 'T' in time_str:
                    # 格式如 "2026-04-28T15:00:00+00:00" 或 "2026-04-28T15:00:00Z"
                    date_part = time_str.split('T')[0]
                else:
                    date_part = time_str[:10]  # 假设是 YYYY-MM-DD
                year, month = int(date_part[:4]), int(date_part[5:7])
                quarter = (month - 1) // 3 + 1
                quarters.add(f"{year}Q{quarter}")
            except Exception as e:
                print(f"解析时间出错: {time_str}, {e}")
                continue
    # 按年份和季度排序
    sorted_quarters = sorted(list(quarters), key=lambda x: (int(x[:4]), int(x[5])))
    return jsonify(sorted_quarters)

# 2.3 二级菜单页面路由
@app.route('/admin/exam/<int:exam_id>/scores')
@login_required
@admin_required
def admin_exam_scores_page(exam_id):
    db = get_supabase()
    
    # 🔧 获取考试标题
    exam_res = db.table("exams") \
        .select("title") \
        .eq("id", exam_id) \
        .maybe_single() \
        .execute()
    
    exam_title = exam_res.data.get('title', f'考试 #{exam_id}') if exam_res.data else f'考试 #{exam_id}'
    
    return render_template(
        'admin/list_exams_scores.html',
        exam_id=exam_id,
        exam_title=exam_title  # 🔧 传递标题
    )

# 2.3.1 获取某考试的考生成绩列表（二级菜单）

@app.route('/api/admin/exam/<int:exam_id>/scores')
@admin_required
def api_admin_exam_scores(exam_id):
    db = get_supabase()
    allowed = get_allowed_countries()
    search = request.args.get('search', '').strip()
    submit_method = request.args.get('submit_method', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # 1. 获取该考试的所有成绩记录
    query = db.table("exam_results").select("*").eq("exam_id", exam_id)
    if submit_method:
        query = query.eq("submit_method", submit_method)
    results_all = query.execute().data or []

    if not results_all:
        return jsonify({"data": [], "total": 0, "page": page, "per_page": per_page})

    # 2. 获取所有相关用户的详细信息（避免多次请求）
    user_ids = list(set(r['user_id'] for r in results_all))
    users_res = db.table("users").select("id, name_cn, name_en, email, country, wh_id, is_partner").in_("id", user_ids).execute()
    users_dict = {u['id']: u for u in (users_res.data or [])}

    # 3. 国家权限过滤（基于用户的国家）
    if allowed is not None:
        if not allowed:
            return jsonify({"data": [], "total": 0, "page": page, "per_page": per_page})
        filtered_results = []
        for r in results_all:
            user = users_dict.get(r['user_id'])
            if user and user.get('country') in allowed:
                filtered_results.append(r)
        results_all = filtered_results

    # 4. 搜索过滤（基于用户姓名或邮箱）
    if search:
        search_lower = search.lower()
        filtered_results = []
        for r in results_all:
            user = users_dict.get(r['user_id'])
            if user:
                name = user.get('name_cn') or user.get('name_en') or ''
                email = user.get('email') or ''
                if search_lower in name.lower() or search_lower in email.lower():
                    filtered_results.append(r)
        results_all = filtered_results

    # 5. 按提交时间倒序排序
    results_all.sort(key=lambda x: x.get('created_at', ''), reverse=True)

    # 6. 内存分页
    total = len(results_all)
    start = (page - 1) * per_page
    end = start + per_page
    paginated = results_all[start:end]

    # 7. 构造返回数据
    scores = []
    for r in paginated:
        user = users_dict.get(r['user_id'], {})
        scores.append({
            "user_id": r['user_id'],
            "name": user.get('name_cn') or user.get('name_en') or user.get('email', ''),
            "email": user.get('email', ''),
            "country": user.get('country', ''),
            "status": "Submitted",
            "submitted_at": r.get('created_at'),
            "submit_method": r.get('submit_method', 'manual'),
            "time_used": r.get('time_used'),      # 如果存在该字段
            "score": r.get('total_score', 0),
            "result_id": r['id']
        })

    return jsonify({"data": scores, "total": total, "page": page, "per_page": per_page})

# 2.4 批量导出考生试卷（ZIP）
@app.route('/api/admin/exam/<int:exam_id>/batch_export', methods=['POST'])
@admin_required
def admin_batch_export_pdf(exam_id):
    data = request.json
    db = get_supabase()
    user_ids = data.get('user_ids', [])
    if not user_ids:
        return jsonify({"error": "未选择考生"}), 400
    # 去重
    user_ids = list(set(user_ids))
    logger.info(f"批量导出考试 {exam_id}，考生数量: {len(user_ids)}，IDs: {user_ids}")

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for uid in user_ids:
            # 调用生成单个考生PDF的函数，返回字节流
            pdf_bytes = generate_single_user_pdf(exam_id, uid)
            if pdf_bytes:
                # 获取考生姓名
                user_res = db.table("users").select("name_en, email").eq("id", uid).execute()
                name = user_res.data[0].get('name_en', uid) if user_res.data else uid
                result_id = get_latest_result_id(exam_id, uid)  # 需实现该函数
                filename = f"{name}_{exam_id}_{result_id}.pdf"
                zf.writestr(filename, pdf_bytes)
                logger.info(f"  已添加文件: {filename}")
    zip_buffer.seek(0)
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=f"exam_{exam_id}_scores.zip")

def get_latest_result_id(exam_id, user_id):
    db = get_supabase()
    res = db.table("exam_results").select("id").eq("exam_id", exam_id).eq("user_id", user_id).order("created_at", desc=True).limit(1).execute()
    return res.data[0]['id'] if res.data else None

# 2.5 推送考试（设置有效期和考生）
@app.route('/api/admin/exam/<int:exam_id>/push', methods=['POST'])
@admin_required
def admin_push_exam(exam_id):
    data = request.json
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    user_ids = data.get('user_ids', [])  # 选中的考生ID列表

    if not start_time or not end_time or not user_ids:
        return jsonify({"success": False, "message": "缺少参数"}), 400

    db = get_supabase()
    # 更新考试的起止时间和状态
    update_data = {
        "start_time": start_time,
        "end_time": end_time,
        "status": "active"
    }
    db.table("exams").update(update_data).eq("id", exam_id).execute()

    # 插入考试-用户关联记录（清除旧的再插入，或使用upsert）
    # 先删除旧的关联
    db.table("exam_assignments").delete().eq("exam_id", exam_id).execute()
    # 批量插入
    for uid in user_ids:
        db.table("exam_assignments").insert({"exam_id": exam_id, "user_id": uid}).execute()

    # 发送邮件通知（异步或同步）
    exam_res = db.table("exams").select("title").eq("id", exam_id).execute()
    exam_title = exam_res.data[0]['title'] if exam_res.data else "考试"
    for uid in user_ids:
        user_res = db.table("users").select("email, name_en").eq("id", uid).execute()
        if user_res.data:
            email = user_res.data[0]['email']
            name = user_res.data[0].get('name_en', '用户')
            subject = f"新考试通知：{exam_title}"
            body = f"您好 {name}，您有一场考试《{exam_title}》已开放，有效期从 {start_time} 到 {end_time}，请登录系统参加。"
            # 调用邮件发送函数（已有的 auth.send_email）
            try:
                auth.send_email(email, subject, body)
            except Exception as e:
                logger.warning(f"发送邮件失败: {e}")

    return jsonify({"success": True})

def generate_single_user_pdf(exam_id, user_id):
    """为指定考试和考生生成PDF字节流"""
    logger.info(f"[批量导出] 开始生成 PDF: exam_id={exam_id}, user_id={user_id}")
    db = get_supabase()
    # 1. 获取该考生在该考试的最新成绩记录
    result_res = db.table("exam_results") \
        .select("*") \
        .eq("exam_id", exam_id) \
        .eq("user_id", user_id) \
        .order("created_at", desc=True) \
        .limit(1) \
        .execute()
    if not result_res.data:
        logger.warning(f"未找到考试 {exam_id} 用户 {user_id} 的成绩记录")
        return None
    result = result_res.data[0]
    logger.info(f"  找到成绩记录 result_id={result['id']}, score={result.get('total_score')}")

    # 2. 获取考试信息
    exam_res = db.table("exams").select("*").eq("id", exam_id).execute()
    if not exam_res.data:
        return None
    exam_data = exam_res.data[0]

    # 3. 获取考生信息
    user_res = db.table("users").select("*").eq("id", user_id).execute()
    if not user_res.data:
        return None
    user_data = user_res.data[0]

    # 4. 解析 answers, details（递归）
    def robust_parse_json(value):
        if not value:
            return {}
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
                if isinstance(parsed, str):
                    return robust_parse_json(parsed)
                return parsed
            except:
                return {}
        return {}
    answers = robust_parse_json(result.get('answers'))
    details = robust_parse_json(result.get('details'))

    # 5. 获取题目列表
    questions_res = db.table("questions").select("*").eq("exam_id", exam_id).order("num").execute()
    questions = questions_res.data or []

    # 6. 阅卷人（默认）
    # ---------- 获取阅卷人（多级优先级）----------
    reviewer = get_reviewer_by_country(
        user_country=user_data.get('country'),
        exam_reviewer=exam_data.get('reviewer'),
        url_reviewer=None  # 批量导出时没有 URL 参数
    )

    # ✅ 添加调试日志
    logger.info(f"========== generate_single_user_pdf 阅卷人 ==========")
    logger.info(f"考生国家: {user_data.get('country')}")
    logger.info(f"考试表 reviewer: {exam_data.get('reviewer')}")
    logger.info(f"最终 reviewer: {reviewer}")
    logger.info(f"=================================================")

    # 7. 生成PDF字节流
    from services import export
    pdf_buffer = export.generate_user_pdf(
        user_name=user_data.get('name_cn') or user_data.get('name_en', '未知考生'),
        user_email=user_data.get('email', ''),
        exam_title=exam_data.get('title', '未命名考试'),
        score=result.get('total_score', 0),
        questions=questions,
        answers=answers,
        details=details,
        submitted_at=result.get('created_at', ''),
        reviewer=reviewer
    )
    # 返回字节数据（BytesIO 需要 .getvalue()）
    return pdf_buffer.getvalue()   # 注意：generate_user_pdf 返回 BytesIO，需提取内容

@app.route('/api/search/trainings')
@login_required
@admin_required
def search_trainings():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    db = get_supabase()
    res = db.table("trainings").select("name").ilike("name", f"%{q}%").limit(10).execute()
    names = [row['name'] for row in (res.data or [])]
    return jsonify(names)

@app.route('/api/search/exams')
@login_required
@admin_required
def search_exams():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    db = get_supabase()
    res = db.table("exams").select("title").ilike("title", f"%{q}%").limit(10).execute()
    names = [row['title'] for row in (res.data or [])]
    return jsonify(names)

@app.route('/api/search/warehouses')
@login_required
@admin_required
def search_warehouses():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    db = get_supabase()
    # 分别模糊查询 wh_id 和 wh_name_en 字段
    r1 = db.table("users").select("wh_id, wh_name_en").ilike("wh_id", f"%{q}%").limit(5).execute()
    r2 = db.table("users").select("wh_id, wh_name_en").ilike("wh_name_en", f"%{q}%").limit(5).execute()

    seen = set()
    suggestions = []
    for row in (r1.data or []) + (r2.data or []):
        wh_id = row.get('wh_id', '')
        wh_name = row.get('wh_name_en', '')
        label = f"{wh_id} ({wh_name})" if wh_name else wh_id
        if label and label not in seen:
            seen.add(label)
            suggestions.append(label)
    return jsonify(suggestions[:10])
    
#----------3. 培训签到管理 API（类似考试清单，复用现有逻辑）---------
# ================= 管理员培训管理路由 =================
@app.route('/admin/trainings')
@login_required
@admin_required
def admin_trainings():
    """培训管理页面一级菜单可新增培训"""
    return render_template('admin/list_trainings.html')

@app.route('/api/admin/trainings', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
@admin_required
def api_admin_trainings():
    db = get_supabase()
    
    if request.method == 'GET':
        # 获取过滤参数
        country_filter = request.args.get('country', '')   # 前端选择的国家筛选
        name = request.args.get('name', '')
        quarter = request.args.get('quarter', '')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        # ✅ 添加调试日志
        logger.info("=" * 50)
        logger.info("调用 api_admin_trainings (GET)")
        logger.info(f"当前用户: role={session.get('role')}, user_country={session.get('user_country')}")
        logger.info(f"admin_countries={session.get('admin_countries')}")
        
        # ✅ 获取管理员的权限范围
        
        # 1. 获取所有培训（先不应用权限过滤，因为培训表的 country 字段可能为空）
        query = db.table("trainings").select("*")
        
        if name:
            query = query.ilike("name", f"%{name}%")

        res = query.execute()
        all_trainings = res.data or []
        
        logger.info(f"初始查询到 {len(all_trainings)} 条培训记录")
        
        # ✅ 2. 获取管理员允许的国家列表
        allowed_countries = get_admin_allowed_countries()
        logger.info(f"allowed_countries = {allowed_countries}")
        
        # ✅ 3. 根据权限范围过滤培训记录
        if allowed_countries is not None:
            if not allowed_countries:
                # 没有权限，返回空
                logger.warning("allowed_countries 为空，返回空列表")
                filtered_trainings = []
            else:
                # 获取允许国家下的所有用户ID（用于通过学员签到关联的培训）
                users_in_allowed = db.table("users").select("id").in_("country", allowed_countries).execute()
                allowed_user_ids = [u['id'] for u in (users_in_allowed.data or [])] if users_in_allowed.data else []
                logger.info(f"允许国家下的用户ID数量: {len(allowed_user_ids)}")
                
                # 查询存在允许国家学员签到的培训ID
                allowed_training_ids = set()
                if allowed_user_ids:
                    attend_res = db.table("training_attendances").select("training_id").in_("user_id", allowed_user_ids).execute()
                    allowed_training_ids = {a['training_id'] for a in (attend_res.data or [])}
                    logger.info(f"通过学员签到关联的培训ID: {allowed_training_ids}")
                
                # 过滤：培训自身 country 在允许列表中 或 培训ID在 allowed_training_ids 中
                filtered_trainings = []
                for training in all_trainings:
                    training_country = training.get('country')
                    if training_country and training_country in allowed_countries:
                        filtered_trainings.append(training)
                    elif training['id'] in allowed_training_ids:
                        filtered_trainings.append(training)
                
                logger.info(f"权限过滤后剩余 {len(filtered_trainings)} 条培训记录")
        else:
            # 无限制（开发者或超管）
            filtered_trainings = all_trainings
            logger.info("无权限限制，返回所有培训")
        
        # 4. 前端选择的额外国家筛选（与权限取交集）
        if country_filter:
            from utils.common import match_country_code
            filter_country_code = match_country_code(country_filter) if country_filter else None
            if filter_country_code:
                # 获取该国家下的用户ID
                users_in_filter = db.table("users").select("id").eq("country", filter_country_code).execute()
                filter_user_ids = [u['id'] for u in (users_in_filter.data or [])] if users_in_filter.data else []
                filter_training_ids = set()
                if filter_user_ids:
                    attend_filter = db.table("training_attendances").select("training_id").in_("user_id", filter_user_ids).execute()
                    filter_training_ids = {a['training_id'] for a in (attend_filter.data or [])}
                
                # 过滤
                temp_filtered = []
                for training in filtered_trainings:
                    if training.get('country') == filter_country_code or training['id'] in filter_training_ids:
                        temp_filtered.append(training)
                filtered_trainings = temp_filtered
                logger.info(f"前端国家筛选后剩余 {len(filtered_trainings)} 条培训记录")

        # 5. 季度过滤（内存中）
        if quarter:
            from utils.common import quarter_to_date_range
            q_start, q_end = quarter_to_date_range(quarter)
            if q_start and q_end:
                q_start_dt = datetime.fromisoformat(q_start)
                q_end_dt = datetime.fromisoformat(q_end)
                temp_filtered = []
                for training in filtered_trainings:
                    start = training.get('start_time')
                    end = training.get('end_time')
                    if start and end:
                        try:
                            start_dt = datetime.fromisoformat(start)
                            end_dt = datetime.fromisoformat(end)
                            if start_dt <= q_end_dt and end_dt >= q_start_dt:
                                temp_filtered.append(training)
                        except:
                            pass
                filtered_trainings = temp_filtered
                logger.info(f"季度过滤后剩余 {len(filtered_trainings)} 条培训记录")

        # 6. 按创建时间倒序排序
        filtered_trainings.sort(key=lambda x: x.get('created_at', ''), reverse=True)

        # 7. 手动分页
        total = len(filtered_trainings)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated = filtered_trainings[start_idx:end_idx]

        # 8. 补充签到人数和动态状态
        for t in paginated:
            signed_count = db.table("training_attendances").select("id", count="exact").eq("training_id", t['id']).execute().count or 0
            t['signed_count'] = signed_count
            t['dynamic_status'] = get_training_status(t)
            
            # 检查是否有多个国家的模板（用于前端禁用表头录入按钮）
            if not t.get('country'):
                att_users = db.table("training_attendances").select("user_id").eq("training_id", t['id']).execute()
                user_ids = [u['user_id'] for u in (att_users.data or [])] if att_users.data else []
                countries = set()
                if user_ids:
                    users_res = db.table("users").select("country").in_("id", user_ids).execute()
                    countries = {u['country'] for u in (users_res.data or []) if u.get('country')}
                
                if len(countries) > 1:
                    t['has_inconsistent_templates'] = True
                else:
                    ct_res = db.table("training_country_templates").select("country, header_template").eq("training_id", t['id']).execute()
                    templates = ct_res.data or []
                    if not templates:
                        t['has_inconsistent_templates'] = False
                    else:
                        unique_templates = set()
                        for ct in templates:
                            tpl = ct.get('header_template', {})
                            unique_templates.add(json.dumps(tpl, sort_keys=True))
                        t['has_inconsistent_templates'] = len(unique_templates) > 1
            else:
                t['has_inconsistent_templates'] = False

        logger.info(f"最终返回 {len(paginated)} 条培训记录")
        return jsonify({
            "data": paginated,
            "total": total,
            "page": page,
            "per_page": per_page
        })

    elif request.method == 'POST':
        # 创建培训（也需要权限校验）
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({"success": False, "message": "培训名称不能为空"}), 400
        
        # ✅ 管理员创建培训时，检查国家权限
        training_country = data.get('country')
        if training_country:
            allowed = get_admin_allowed_countries()
            if allowed is not None and training_country not in allowed:
                return jsonify({"success": False, "message": "无权在该国家创建培训"}), 403
        
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        if not start_time:
            start_time = datetime.now(timezone.utc).isoformat()
        if not end_time:
            end_time = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        
        res = db.table("trainings").insert({
            "name": name,
            "start_time": start_time,
            "end_time": end_time,
            "header_template": data.get('header_template', {}),
            "country": data.get('country', ''),
            "quarter": data.get('quarter', '')
        }).execute()
        
        logger.info(f"创建培训成功: id={res.data[0]['id']}, name={name}")
        return jsonify({"success": True, "id": res.data[0]['id']})
    
    elif request.method == 'PUT':
        # 更新培训（需要权限校验）
        data = request.json
        tid = data.get('id')
        if tid is None or tid == 'None' or str(tid).lower() == 'null':
            return jsonify({"success": False, "message": "无效的培训ID"}), 400
        
        try:
            tid = int(tid)
        except (ValueError, TypeError):
            return jsonify({"success": False, "message": "培训ID必须是整数"}), 400
        
        # ✅ 获取原培训信息，检查权限
        original = db.table("trainings").select("country").eq("id", tid).maybe_single().execute()
        if original.data:
            allowed = get_admin_allowed_countries()
            original_country = original.data.get('country')
            if allowed is not None and original_country and original_country not in allowed:
                return jsonify({"success": False, "message": "无权修改该培训"}), 403
        
        # 处理 header_template 保存
        country_code = data.get('country_code')
        header_template = data.get('header_template')
        if header_template is not None:
            if country_code:
                # 检查国家权限
                allowed = get_admin_allowed_countries()
                if allowed is not None and country_code not in allowed:
                    return jsonify({"success": False, "message": "无权在该国家设置表头模板"}), 403
                set_training_country_template(tid, country_code, header_template)
            else:
                db.table("trainings").update({"header_template": header_template}).eq("id", tid).execute()
            return jsonify({"success": True})
        
        update_data = {}
        if 'name' in data:
            update_data['name'] = data['name']
        if 'header_template' in data:
            update_data['header_template'] = data['header_template']
        if 'start_time' in data and data['start_time'] is not None:
            update_data['start_time'] = data['start_time']
        if 'end_time' in data and data['end_time'] is not None:
            update_data['end_time'] = data['end_time']
        if 'is_active' in data:
            update_data['is_active'] = data['is_active']
        if 'country' in data:
            # 检查新国家权限
            new_country = data['country']
            allowed = get_admin_allowed_countries()
            if allowed is not None and new_country and new_country not in allowed:
                return jsonify({"success": False, "message": "无权将培训改到该国家"}), 403
            update_data['country'] = new_country
        if 'quarter' in data:
            update_data['quarter'] = data['quarter']
        
        if update_data:
            db.table("trainings").update(update_data).eq("id", tid).execute()
            logger.info(f"更新培训成功: id={tid}, 更新字段={list(update_data.keys())}")
        
        return jsonify({"success": True})
    
    elif request.method == 'DELETE':
        tid = request.args.get('id')
        if not tid:
            return jsonify({"success": False, "message": "缺少培训ID"}), 400
        
        # ✅ 删除前检查权限
        original = db.table("trainings").select("country").eq("id", tid).maybe_single().execute()
        if original.data:
            allowed = get_admin_allowed_countries()
            original_country = original.data.get('country')
            if allowed is not None and original_country and original_country not in allowed:
                return jsonify({"success": False, "message": "无权删除该培训"}), 403
        
        db.table("trainings").delete().eq("id", tid).execute()
        logger.info(f"删除培训成功: id={tid}")
        return jsonify({"success": True})

@app.route('/api/admin/training/<int:training_id>/attendance_by_country')
@login_required
@admin_required
def api_training_attendance_by_country(training_id):
    db = get_supabase()
    allowed = get_allowed_countries()
    
    # 获取该培训的所有签到记录，并关联用户国家
    att_res = db.table("training_attendances") \
        .select("id, user_id, signed_name, signature_url, sign_time, users(country, name_cn, name_en, department, employee_id)") \
        .eq("training_id", training_id) \
        .execute()
    records = att_res.data or []
    
    # 国家权限过滤（仅保留允许国家的记录）
    if allowed is not None:
        if not allowed:
            return jsonify([])
        records = [r for r in records if r.get('users', {}).get('country') in allowed]
    
    # 按国家分组
    groups = {}
    for rec in records:
        user = rec.get('users', {})
        country = user.get('country') or '未指定'
        if country not in groups:
            groups[country] = {
                'country': country,
                'count': 0,
                'attendances': []
            }
        groups[country]['count'] += 1
        groups[country]['attendances'].append({
            'user_id': rec['user_id'],
            'department': user.get('department', ''),
            'name_cn': user.get('name_cn', ''),
            'name_en': user.get('name_en', ''),
            'employee_id': user.get('employee_id', ''),
            'signed_name': rec.get('signed_name', ''),
            'signature_url': rec.get('signature_url', ''),
            'sign_time': rec.get('sign_time')
        })
    return jsonify(list(groups.values()))

@app.route('/admin/training/<int:training_id>/attendance')
@login_required
@admin_required
def admin_training_attendance(training_id):
    """培训签到统计的二级菜单可导出签到记录"""
    return render_template('admin/list_training_attendance.html', training_id=training_id)

#----------4. 访谈结果统计 API---------
@app.route('/api/admin/interviewee/stats')
@admin_required
def admin_interviewee_stats():
    return render_template('admin/list_inspection.html')

#----------5. 题库统计 API---------
@app.route('/api/admin/questions/stats')
@admin_required
def admin_questions_stats():
    # 返回 { total_questions, total_exams, total_participants }
    pass

# ================= 启动入口 =================
if __name__ == "__main__":
    app.run(
        host=host,
        port=port,
        debug=DEBUG,
        use_reloader=False,  # 生产环境禁用重载器
        threaded=True
    )
